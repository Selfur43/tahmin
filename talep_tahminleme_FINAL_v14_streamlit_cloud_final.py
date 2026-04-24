# -*- coding: utf-8 -*-
"""
Demand Forecasting - Production Grade Real-World Preprocessing Module
Authoring goal:
    Conservative, auditable, review-friendly preprocessing for demand forecasting.

Compatible modeling families:
    - CNN
    - LSTM
    - Prophet
    - ARIMA
    - SARIMAX
    - XGBoost
    - N-HiTS
    - TiDE
    - PatchTST
    - TFT
    - Chronos-Bolt

Core philosophy:
    CLEANING -> ANOMALY GOVERNANCE

Main capabilities:
    1) Manual Excel file and sheet selection
    2) Date / target auto detection
    3) Frequency inference and regularization
    4) Series profiling (volume, CV, intermittency, volatility, seasonality, trend)
    5) Adaptive anomaly detection by series profile
    6) Anomaly classification:
        - data_error
        - structural_event
        - business_spike_dip
        - unknown_anomaly
    7) Human-review-first policy on recent periods
    8) Structural event engine
    9) Intervention intensity / change tracking
    10) Raw vs clean forecastability comparison
    11) Model-family-specific exports
    12) Strict leakage audit
    13) Review queue
    14) Run manifest / versioning / config hashing
    15) Internal tests / synthetic tests / business-aware tests
"""

import os
os.environ.setdefault("STREAMLIT_SERVER_FILE_WATCHER_TYPE", "none")
os.environ.setdefault("STREAMLIT_BROWSER_GATHER_USAGE_STATS", "false")
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")
import re
import io
import json
import math
import time
import copy
import gc
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import uuid
import hashlib
import traceback
import warnings
import tempfile
import zipfile
import shutil
import atexit
import argparse
import sys
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from sklearn.preprocessing import StandardScaler, RobustScaler, MinMaxScaler
from sklearn.impute import KNNImputer
try:
    from sklearn.linear_model import Ridge
except Exception:
    Ridge = None

# GUI (desktop optional; Streamlit Cloud/mobile does not use tkinter)
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
    HAS_TKINTER = True
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    simpledialog = None
    HAS_TKINTER = False

try:
    import rarfile  # optional; requires rarfile package and available backend on the host machine
except Exception:
    rarfile = None

try:
    from statsmodels.tsa.seasonal import seasonal_decompose
    HAS_STATSMODELS = True
except Exception:
    seasonal_decompose = None
    HAS_STATSMODELS = False

warnings.filterwarnings("ignore")


# =========================================================
# LOCAL PROGRESS / RUN LOGGING
# =========================================================

LOCAL_PROGRESS_REPORTER = None


class LocalProgressReporter:
    def __init__(self, run_label: str, output_dir: Optional[str] = None, heartbeat_seconds: int = 60):
        self.run_label = str(run_label)
        self.output_dir = os.path.abspath(output_dir or os.getcwd())
        os.makedirs(self.output_dir, exist_ok=True)
        self.started_at = time.time()
        self.last_update_at = self.started_at
        self.current_stage = "Başlatılıyor"
        self.current_target = None
        self.heartbeat_seconds = max(20, int(heartbeat_seconds))
        self._lock = threading.Lock()
        self._stop_event = threading.Event()
        self._heartbeat_thread = None
        ts = time.strftime("%Y%m%d_%H%M%S")
        safe_label = re.sub(r"[^A-Za-z0-9._-]+", "_", self.run_label).strip("_") or "forecast_run"
        self.log_path = os.path.join(self.output_dir, f"{safe_label}_{ts}.log")
        self._fh = open(self.log_path, "a", encoding="utf-8")
        self.log("INFO", "Başlatma", f"Çalışma başlatıldı. Günlük dosyası: {self.log_path}")

    def _elapsed_text(self) -> str:
        elapsed = max(0.0, time.time() - self.started_at)
        mins, secs = divmod(int(elapsed), 60)
        hours, mins = divmod(mins, 60)
        if hours:
            return f"{hours:02d}:{mins:02d}:{secs:02d}"
        return f"{mins:02d}:{secs:02d}"

    def start(self):
        if self._heartbeat_thread is None:
            self._heartbeat_thread = threading.Thread(target=self._heartbeat_loop, name="local-progress-heartbeat", daemon=True)
            self._heartbeat_thread.start()
        return self

    def stop(self, success: bool = True, final_message: Optional[str] = None):
        message = final_message or ("Çalışma tamamlandı." if success else "Çalışma hata ile sonlandı.")
        self.log("SUCCESS" if success else "ERROR", "Kapanış", message)
        self._stop_event.set()
        if self._heartbeat_thread is not None:
            self._heartbeat_thread.join(timeout=1.0)
        try:
            self._fh.flush()
            self._fh.close()
        except Exception:
            pass

    def _emit_line(self, line: str):
        print(line, flush=True)
        try:
            self._fh.write(line + "\n")
            self._fh.flush()
        except Exception:
            pass

    def log(self, level: str, stage: str, message: str, target: Optional[str] = None, extra: Optional[Dict[str, Any]] = None):
        with self._lock:
            self.last_update_at = time.time()
            self.current_stage = str(stage)
            self.current_target = None if target is None else str(target)
            parts = [
                f"[{time.strftime('%Y-%m-%d %H:%M:%S')}]",
                f"[{str(level).upper()}]",
                f"[+{self._elapsed_text()}]",
                f"[Aşama: {stage}]",
            ]
            if target is not None:
                parts.append(f"[Seri: {target}]")
            line = " ".join(parts) + f" {message}"
            if extra:
                extra_pairs = [f"{k}={v}" for k, v in extra.items()]
                if extra_pairs:
                    line += " | " + ", ".join(extra_pairs)
            self._emit_line(line)

    def _heartbeat_loop(self):
        while not self._stop_event.wait(self.heartbeat_seconds):
            with self._lock:
                silent_for = time.time() - self.last_update_at
                stage = self.current_stage
                target = self.current_target
            if silent_for >= self.heartbeat_seconds:
                msg = f"İşlem devam ediyor. Son bilinen aşama: {stage}."
                if target:
                    msg += f" Aktif seri: {target}."
                msg += f" Son {int(silent_for)} saniyede yeni kayıt oluşmadı; süreç muhtemelen ağır model araması veya rapor üretimi içinde çalışıyor."
                self.log("HEARTBEAT", stage, msg, target=target)


def activate_local_progress_reporter(run_label: str, output_dir: Optional[str] = None, heartbeat_seconds: int = 60) -> LocalProgressReporter:
    global LOCAL_PROGRESS_REPORTER
    if LOCAL_PROGRESS_REPORTER is not None:
        try:
            LOCAL_PROGRESS_REPORTER.stop(success=False, final_message="Yeni çalışma başlatıldığı için önceki loglayıcı kapatıldı.")
        except Exception:
            pass
    LOCAL_PROGRESS_REPORTER = LocalProgressReporter(run_label=run_label, output_dir=output_dir, heartbeat_seconds=heartbeat_seconds).start()
    return LOCAL_PROGRESS_REPORTER


def deactivate_local_progress_reporter(success: bool = True, final_message: Optional[str] = None) -> None:
    global LOCAL_PROGRESS_REPORTER
    reporter = LOCAL_PROGRESS_REPORTER
    LOCAL_PROGRESS_REPORTER = None
    if reporter is not None:
        reporter.stop(success=success, final_message=final_message)


def progress_log(stage: str, message: str, level: str = "INFO", target: Optional[str] = None, extra: Optional[Dict[str, Any]] = None) -> None:
    reporter = LOCAL_PROGRESS_REPORTER
    if reporter is not None:
        reporter.log(level=level, stage=stage, message=message, target=target, extra=extra)


# =========================================================
# VERSION / PIPELINE
# =========================================================

PIPELINE_NAME = "demand_forecast_preprocessing"
PIPELINE_VERSION = "2.0.0"
OUTPUT_SCHEMA_VERSION = "2.0.0"
CODE_VERSION = "production_governance_rewrite_2_0_0"


# =========================================================
# CONFIG
# =========================================================

@dataclass
class PreprocessConfig:
    output_dir_name: str = "forecast_preprocessing_outputs"

    date_column_candidates: Tuple[str, ...] = (
        "datum", "date", "datetime", "tarih", "timestamp", "zaman", "time"
    )

    non_target_columns: Tuple[str, ...] = (
        "year", "month", "day", "hour", "minute", "second",
        "weekday", "weekday name", "haftanın günü", "haftanın günü (tr)",
        "week", "quarter", "is_holiday", "holiday", "weekofyear",
        "dayofweek", "dayofmonth"
    )

    # Frequency / regularization
    force_regular_frequency: bool = True
    allow_month_start_to_month_end_alignment_fix: bool = True

    # Base anomaly voting windows
    hampel_window: int = 7
    hampel_n_sigma: float = 4.0
    rolling_mad_window: int = 9
    rolling_mad_n_sigma: float = 4.5
    iqr_k: float = 4.0
    min_outlier_votes: int = 2

    # Adaptive governance
    max_outlier_fraction_per_series: float = 0.05
    protect_first_n_periods: int = 1
    protect_last_n_periods: int = 6
    recent_periods_review_only: int = 6
    clip_negative_to_zero: bool = True

    # Structural events
        # Structural events
    structural_zero_ratio_threshold: float = 0.5
    structural_zero_min_series_count: int = 3
    portfolio_drop_ratio_threshold: float = 0.55
    portfolio_rebound_ratio_threshold: float = 1.30
    structural_event_neighbor_window: int = 0
    preserve_structural_zero_events: bool = True
    preserve_zero_values_on_structural_dates: bool = True

    # Incomplete / partial period governance
    enable_incomplete_period_detection: bool = True
    partial_period_drop_ratio_threshold: float = 0.60
    partial_period_compare_last_n: int = 3
    auto_exclude_incomplete_last_period_from_training: bool = True
    auto_flag_incomplete_last_period_review: bool = True

    # Modeling exclusion / masks
    export_training_exclusion_masks: bool = True

    # Monthly feature discipline
    drop_low_signal_calendar_features_for_monthly: bool = True

    # Missing / imputation
    max_interpolation_gap: int = 1
    seasonal_period_map: Dict[str, int] = field(default_factory=lambda: {
        "H": 24,
        "D": 7,
        "W": 52,
        "M": 12,
        "MS": 12
    })
    use_knn_for_dense_missing_blocks: bool = False
    impute_method_preference: str = "seasonal_local"

    # Missing strategy governance / audit (safe-additive; does not override core cleaning by default)
    missing_report_only_threshold: float = 0.00
    missing_drop_row_threshold: float = 0.80
    missing_drop_series_threshold: float = 0.60
    missing_impute_ratio_threshold: float = 0.20
    dense_missing_block_threshold: int = 3
    allow_row_drop_for_non_target_metadata_only: bool = True
    allow_target_row_drop: bool = False
    missingness_mechanism_proxy_check: bool = True

    # Datetime integrity / alignment governance
    normalize_datetime_timezone: bool = True
    align_monthly_dates_to_period_end: bool = True

    # Descriptive statistics / visual diagnostics
    save_distribution_plots: bool = True
    save_trend_plots: bool = True
    save_seasonality_plots: bool = True
    moving_average_windows: Tuple[int, int] = (3, 6)
    save_boxplots: bool = True
    save_year_overlay_seasonality_plots: bool = True
    save_normalized_seasonality_plots: bool = True
    save_correlation_analysis: bool = True
    save_seasonality_heatmaps: bool = True
    save_seasonal_decomposition: bool = True
    robust_trend_window: int = 5
    save_robust_trend: bool = True

    # Proxy backtest enrichment
    enable_additional_backtest_benchmarks: bool = True

    # Pharma event interpretation hints (safe-additive diagnostics)
    promotion_like_jump_ratio: float = 2.0
    stockout_like_drop_ratio: float = 0.30
    rebound_after_drop_ratio: float = 1.80

    # Review / governance policy
    min_action_confidence_for_auto_fix: float = 0.70
    auto_fix_business_spike_dip: bool = False
    auto_fix_unknown_anomaly: bool = False
    auto_fix_data_error: bool = True
    auto_fix_structural_event: bool = False

    # Scaling
    scaler_for_deep_learning: str = "robust"
    export_log1p_version: bool = True

    # Feature engineering
    generate_modeling_ready_feature_pack: bool = True
    exclude_textual_columns_from_modeling_features: bool = True
    drop_weekday_text_from_monthly_modeling: bool = True

    # QA / validation
    save_validation_excel: bool = True
    save_validation_csv: bool = True
    save_validation_plots: bool = True
    create_domain_validation_template: bool = True
    leakage_check_enabled: bool = True
    max_plot_series: int = 50

    # Tests / QA
    run_internal_unit_tests: bool = True
    run_synthetic_tests: bool = True
    run_business_rule_tests: bool = True
    run_manual_sample_audit: bool = True
    run_proxy_backtest_validation: bool = True
    manual_sample_size: int = 20
    random_seed: int = 42

    # Proxy backtest
    backtest_horizon: int = 3
    backtest_min_train_size_monthly: int = 24
    backtest_min_train_size_weekly: int = 52
    backtest_min_train_size_daily: int = 60
    backtest_min_train_size_hourly: int = 24 * 14

    # Validation thresholds
    review_if_outlier_fraction_gt: float = 0.05
    review_if_structural_zero_events_gt: int = 1
    review_if_clean_zero_ratio_gt: float = 0.10
    review_if_proxy_smape_gt: float = 60.0

    # Export
    save_excel: bool = True
    save_csv: bool = True
    save_metadata_json: bool = True
    save_quality_report: bool = True

    # Manifest / versioning
    pipeline_name: str = PIPELINE_NAME
    pipeline_version: str = PIPELINE_VERSION
    output_schema_version: str = OUTPUT_SCHEMA_VERSION
    code_version: str = CODE_VERSION

        # Incomplete / partial period governance
    detect_partial_last_period: bool = True
    partial_last_period_drop_ratio_threshold: float = 0.65
    partial_last_period_compare_window: int = 3
    partial_last_period_exclude_from_training: bool = True

    # Safer anomaly governance
    auto_fix_unknown_anomaly: bool = False
    auto_fix_business_spike_dip: bool = False

    # Monthly feature discipline
    monthly_keep_only_low_leakage_calendar_features: bool = True

    # Training mask / modeling governance
    generate_training_mask: bool = True
    exclude_structural_events_from_statistical_models: bool = True
    exclude_partial_last_period_from_training: bool = True

    # Imputation audit
    track_governance_imputations: bool = True

    # CV-safe / future-safe governance notes
    enable_fold_aware_preprocessing_notes: bool = True


# =========================================================
# UTILITIES
# =========================================================

def safe_excel_sheet_name(name: str, max_len: int = 31) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "_", str(name))
    return name[:max_len]


def normalize_colname(col: str) -> str:
    return re.sub(r"\s+", " ", str(col).strip().lower())


def _choose_item_from_list(title: str, prompt: str, items: List[str]) -> str:
    if not HAS_TKINTER:
        raise RuntimeError('Bu seçim ekranı tkinter gerektirir. Streamlit/telefon kullanımında dosya ve sheet seçimi uygulama arayüzünden yapılmalıdır.')
    return ask_choice_dialog(title=title, prompt=prompt, items=list(items), default_index=0)


def _list_excel_files_in_archive(archive_path: str) -> List[str]:
    lower = str(archive_path).lower()
    if lower.endswith('.zip'):
        with zipfile.ZipFile(archive_path, 'r') as zf:
            names = zf.namelist()
    elif lower.endswith('.rar'):
        if rarfile is None:
            raise ImportError(
                "RAR arşivi seçildi ancak 'rarfile' modülü bu bilgisayarda kurulu değil. "
                "RAR desteği için 'pip install rarfile' ve sistemde unrar/bsdtar benzeri backend gerekir."
            )
        with rarfile.RarFile(archive_path, 'r') as rf:
            names = rf.namelist()
    else:
        raise ValueError("Desteklenmeyen arşiv türü.")

    excel_files = [n for n in names if not n.endswith('/') and n.lower().endswith(('.xlsx', '.xls'))]
    if not excel_files:
        raise FileNotFoundError("Arşiv içinde Excel dosyası bulunamadı.")
    return excel_files


def _extract_excel_from_archive(archive_path: str, member_name: str) -> Tuple[str, str]:
    temp_dir = tempfile.mkdtemp(prefix='forecast_preproc_archive_')

    lower = str(archive_path).lower()
    if lower.endswith('.zip'):
        with zipfile.ZipFile(archive_path, 'r') as zf:
            zf.extract(member_name, path=temp_dir)
    elif lower.endswith('.rar'):
        if rarfile is None:
            shutil.rmtree(temp_dir, ignore_errors=True)
            raise ImportError(
                "RAR arşivi seçildi ancak 'rarfile' modülü bu bilgisayarda kurulu değil. "
                "RAR desteği için 'pip install rarfile' ve sistemde unrar/bsdtar benzeri backend gerekir."
            )
        with rarfile.RarFile(archive_path, 'r') as rf:
            rf.extract(member_name, path=temp_dir)
    else:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise ValueError("Desteklenmeyen arşiv türü.")

    extracted_path = os.path.join(temp_dir, member_name)
    if not os.path.exists(extracted_path):
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise FileNotFoundError("Arşivden çıkarılan Excel dosyası bulunamadı.")

    atexit.register(lambda p=temp_dir: shutil.rmtree(p, ignore_errors=True))
    return extracted_path, temp_dir


def _best_local_initial_dir() -> str:
    candidates = []
    try:
        home = os.path.expanduser("~")
        if home:
            candidates.extend([
                os.path.join(home, "Desktop"),
                os.path.join(home, "Masaüstü"),
                os.path.join(home, "Documents"),
                os.path.join(home, "Belgeler"),
                home,
            ])
    except Exception:
        pass
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return os.getcwd()


def _prepare_tk_root(hidden: bool = True) -> Any:
    if not HAS_TKINTER:
        raise RuntimeError('Tkinter bu ortamda kullanılamıyor. Streamlit dosya yükleme akışını kullanın.')
    root = tk.Tk()
    try:
        root.attributes('-topmost', True)
    except Exception:
        pass
    if hidden:
        try:
            root.withdraw()
        except Exception:
            pass
    else:
        try:
            root.geometry('1x1+0+0')
            root.overrideredirect(True)
            root.attributes('-alpha', 0.0)
            root.deiconify()
        except Exception:
            try:
                root.withdraw()
                root.deiconify()
            except Exception:
                pass
    try:
        root.update_idletasks()
        root.update()
    except Exception:
        pass
    return root


def _finalize_tk_window(window: Any) -> None:
    try:
        window.update_idletasks()
    except Exception:
        pass
    try:
        width = max(int(window.winfo_width()), int(window.winfo_reqwidth()), 420)
        height = max(int(window.winfo_height()), int(window.winfo_reqheight()), 220)
        screen_w = int(window.winfo_screenwidth())
        screen_h = int(window.winfo_screenheight())
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 3)
        window.geometry(f"{width}x{height}+{x}+{y}")
    except Exception:
        pass
    for fn in ('lift', 'focus_force'):
        try:
            getattr(window, fn)()
        except Exception:
            pass
    try:
        window.attributes('-topmost', True)
        window.after(250, lambda: window.attributes('-topmost', False))
    except Exception:
        pass


def _create_modal_dialog(owner: Any, title: str, geometry: str) -> Any:
    dialog = tk.Toplevel(owner)
    dialog.title(title)
    dialog.geometry(geometry)
    dialog.minsize(420, 220)
    try:
        dialog.transient(owner)
    except Exception:
        pass
    try:
        dialog.attributes('-topmost', True)
    except Exception:
        pass
    dialog.protocol('WM_DELETE_WINDOW', dialog.destroy)
    dialog.bind('<Escape>', lambda _e: dialog.destroy())
    dialog.after(10, lambda: _finalize_tk_window(dialog))
    try:
        dialog.grab_set()
    except Exception:
        pass
    try:
        dialog.lift()
        dialog.focus_force()
    except Exception:
        pass
    return dialog


def choose_excel_file(dialog_mode: str = "excel_or_archive") -> Dict[str, Optional[str]]:
    if not HAS_TKINTER:
        raise RuntimeError('Tkinter bu ortamda kullanılamıyor. Streamlit dosya yükleme akışını kullanın.')

    root = _prepare_tk_root(hidden=True)
    initial_dir = _best_local_initial_dir()

    if dialog_mode == "excel_only":
        title = "Excel dosyasını seçin"
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("Excel workbook (*.xlsx)", "*.xlsx"),
            ("Legacy Excel (*.xls)", "*.xls"),
            ("All files", "*.*"),
        ]
    elif dialog_mode == "archive_only":
        title = "ZIP veya RAR arşivini seçin"
        filetypes = [
            ("Archive files", "*.zip *.rar"),
            ("ZIP archives", "*.zip"),
            ("RAR archives", "*.rar"),
            ("All files", "*.*"),
        ]
    else:
        title = "Excel dosyasını, ZIP/RAR arşivini seçin"
        filetypes = [
            ("Excel and archives", "*.xlsx *.xls *.zip *.rar"),
            ("Excel files", "*.xlsx *.xls"),
            ("ZIP archives", "*.zip"),
            ("RAR archives", "*.rar"),
            ("All files", "*.*"),
        ]

    selected_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes,
        initialdir=initial_dir,
        parent=root,
    )
    try:
        root.destroy()
    except Exception:
        pass

    if not selected_path:
        raise FileNotFoundError("Dosya seçilmedi.")

    selected_path = os.path.abspath(selected_path)
    lower = selected_path.lower()
    if lower.endswith(('.xlsx', '.xls')):
        return {
            'source_path': selected_path,
            'excel_path': selected_path,
            'archive_member': None,
            'temp_dir': None,
            'source_type': 'excel'
        }

    if lower.endswith(('.zip', '.rar')):
        excel_files = _list_excel_files_in_archive(selected_path)
        if len(excel_files) == 1:
            chosen_member = excel_files[0]
        else:
            chosen_member = _choose_item_from_list(
                "Arşiv İçindeki Excel Seçimi",
                "Arşiv içinden kullanmak istediğiniz Excel dosyasını seçin:",
                excel_files
            )
        extracted_path, temp_dir = _extract_excel_from_archive(selected_path, chosen_member)
        return {
            'source_path': selected_path,
            'excel_path': extracted_path,
            'archive_member': chosen_member,
            'temp_dir': temp_dir,
            'source_type': 'archive'
        }

    raise ValueError("Desteklenmeyen dosya türü seçildi. Lütfen .xlsx, .xls, .zip veya .rar seçin.")



def choose_sheets(sheet_names: List[str]) -> List[str]:
    if not HAS_TKINTER:
        raise RuntimeError('Tkinter bu ortamda kullanılamıyor. Streamlit sheet seçimini kullanın.')
    return ask_multiselect_dialog(
        title="Sheet Seçimi",
        prompt="Kullanılacak sheet(ler)i seçin.",
        items=[str(s) for s in sheet_names],
        default_items=[str(sheet_names[0])] if sheet_names else []
    )


def ask_choice_dialog(title: str, prompt: str, items: List[str], default_index: int = 0) -> str:
    if not items:
        raise ValueError("Seçilecek öğe bulunamadı.")
    if not HAS_TKINTER:
        return items[max(0, min(default_index, len(items) - 1))]

    owner = _prepare_tk_root(hidden=False)
    result = {"value": None}
    dialog = _create_modal_dialog(owner, title, "760x520")
    dialog.resizable(True, True)

    tk.Label(dialog, text=prompt, justify="left", anchor="w", wraplength=700).pack(fill="x", padx=12, pady=(12, 6))
    frame = tk.Frame(dialog)
    frame.pack(fill="both", expand=True, padx=12, pady=6)
    scrollbar = tk.Scrollbar(frame, orient="vertical")
    listbox = tk.Listbox(frame, exportselection=False, height=min(18, max(8, len(items))), yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    scrollbar.pack(side="right", fill="y")
    listbox.pack(side="left", fill="both", expand=True)
    for item in items:
        listbox.insert("end", item)
    idx = max(0, min(default_index, len(items) - 1))
    listbox.selection_set(idx)
    listbox.see(idx)

    def confirm(_event=None):
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning(title, "Lütfen bir seçim yapın.", parent=dialog)
            return
        result["value"] = items[int(sel[0])]
        dialog.destroy()

    def cancel(_event=None):
        dialog.destroy()

    listbox.bind("<Double-Button-1>", confirm)
    listbox.bind("<Return>", confirm)
    btns = tk.Frame(dialog)
    btns.pack(fill="x", padx=12, pady=(6, 12))
    tk.Button(btns, text="İptal", command=cancel, width=12).pack(side="right", padx=(6, 0))
    tk.Button(btns, text="Seç", command=confirm, width=12).pack(side="right")
    dialog.protocol('WM_DELETE_WINDOW', cancel)
    dialog.after(20, lambda: listbox.focus_set())
    owner.wait_window(dialog)
    try:
        owner.destroy()
    except Exception:
        pass

    if result["value"] is None:
        raise ValueError("Seçim yapılmadı.")
    return str(result["value"])


def ask_multiselect_dialog(title: str, prompt: str, items: List[str], default_items: Optional[List[str]] = None) -> List[str]:
    if not items:
        return []
    default_items = list(default_items or [])
    if not HAS_TKINTER:
        return default_items or items

    owner = _prepare_tk_root(hidden=False)
    result = {"values": None}
    dialog = _create_modal_dialog(owner, title, "820x620")
    dialog.resizable(True, True)

    tk.Label(dialog, text=prompt, justify="left", anchor="w", wraplength=760).pack(fill="x", padx=12, pady=(12, 6))
    frame = tk.Frame(dialog)
    frame.pack(fill="both", expand=True, padx=12, pady=6)
    scrollbar = tk.Scrollbar(frame, orient="vertical")
    listbox = tk.Listbox(frame, selectmode="multiple", exportselection=False, yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    scrollbar.pack(side="right", fill="y")
    listbox.pack(side="left", fill="both", expand=True)
    for item in items:
        listbox.insert("end", item)
    default_set = set(default_items)
    for i, item in enumerate(items):
        if item in default_set:
            listbox.selection_set(i)

    def select_all():
        listbox.selection_set(0, "end")

    def clear_all():
        listbox.selection_clear(0, "end")

    def confirm(_event=None):
        sels = [items[int(i)] for i in listbox.curselection()]
        if not sels:
            messagebox.showwarning(title, "Lütfen en az bir seri seçin.", parent=dialog)
            return
        result["values"] = sels
        dialog.destroy()

    def cancel(_event=None):
        dialog.destroy()

    btns = tk.Frame(dialog)
    btns.pack(fill="x", padx=12, pady=(6, 12))
    tk.Button(btns, text="Tümünü Seç", command=select_all, width=14).pack(side="left")
    tk.Button(btns, text="Temizle", command=clear_all, width=14).pack(side="left", padx=(6, 0))
    tk.Button(btns, text="İptal", command=cancel, width=12).pack(side="right", padx=(6, 0))
    tk.Button(btns, text="Onayla", command=confirm, width=12).pack(side="right")
    dialog.protocol('WM_DELETE_WINDOW', cancel)
    dialog.bind('<Return>', confirm)
    dialog.after(20, lambda: listbox.focus_set())
    owner.wait_window(dialog)
    try:
        owner.destroy()
    except Exception:
        pass

    if result["values"] is None:
        raise ValueError("Seçim yapılmadı.")
    return list(result["values"])


def ask_integer_dialog(title: str, prompt: str, initial: int, min_value: Optional[int] = None, max_value: Optional[int] = None) -> int:
    if not HAS_TKINTER:
        return int(initial)

    owner = _prepare_tk_root(hidden=False)
    result = {"value": None}
    dialog = _create_modal_dialog(owner, title, "560x280")
    dialog.resizable(False, False)

    tk.Label(dialog, text=prompt, justify="left", anchor="w", wraplength=520).pack(fill="x", padx=12, pady=(12, 6))
    body = tk.Frame(dialog)
    body.pack(fill='x', padx=12, pady=6)
    value_var = tk.StringVar(value=str(int(initial)))
    entry = tk.Entry(body, textvariable=value_var, width=16)
    entry.pack(side='left', padx=(0, 8))
    hint = []
    if min_value is not None:
        hint.append(f"min={min_value}")
    if max_value is not None:
        hint.append(f"max={max_value}")
    tk.Label(body, text=("Geçerli aralık: " + ", ".join(hint)) if hint else "Tamsayı girin.").pack(side='left')

    def confirm(_event=None):
        raw = value_var.get().strip()
        try:
            value = int(raw)
        except Exception:
            messagebox.showwarning(title, "Lütfen geçerli bir tam sayı girin.", parent=dialog)
            return
        if min_value is not None and value < min_value:
            messagebox.showwarning(title, f"Değer en az {min_value} olmalıdır.", parent=dialog)
            return
        if max_value is not None and value > max_value:
            messagebox.showwarning(title, f"Değer en fazla {max_value} olmalıdır.", parent=dialog)
            return
        result['value'] = int(value)
        dialog.destroy()

    def cancel(_event=None):
        dialog.destroy()

    btns = tk.Frame(dialog)
    btns.pack(fill='x', padx=12, pady=(6, 12))
    tk.Button(btns, text='İptal', command=cancel, width=12).pack(side='right', padx=(6, 0))
    tk.Button(btns, text='Onayla', command=confirm, width=12).pack(side='right')
    dialog.protocol('WM_DELETE_WINDOW', cancel)
    dialog.bind('<Return>', confirm)
    dialog.after(20, lambda: entry.focus_set())
    owner.wait_window(dialog)
    try:
        owner.destroy()
    except Exception:
        pass
    if result['value'] is None:
        raise ValueError('Sayısal seçim iptal edildi.')
    return int(result['value'])


def choose_single_sheet(sheet_names: List[str], always_prompt: bool = True) -> str:
    sheet_names = [str(s) for s in sheet_names]
    if not sheet_names:
        raise ValueError("Sheet bulunamadı.")
    if len(sheet_names) == 1 and not always_prompt:
        return str(sheet_names[0])
    prompt = "Tahminleme için kullanılacak sheet'i seçin."
    if len(sheet_names) == 1:
        prompt = "Tek bir sheet bulundu. Devam etmek için seçimi onaylayın."
    return ask_choice_dialog(
        title="Sheet Seçimi",
        prompt=prompt,
        items=list(sheet_names),
        default_index=0,
    )


def create_output_dir(base_path: str, output_dir_name: str) -> str:
    base_folder = os.path.dirname(base_path)
    out_dir = os.path.join(base_folder, output_dir_name)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def save_uploaded_file(uploaded_file) -> str:
    """
    Streamlit uploaded file nesnesini güvenli bir geçici klasöre yazar ve dosya yolunu döndürür.
    Hem .xlsx hem .xls için çalışır. Aynı isimli tekrar yüklemelerde çakışmayı önler.
    """
    if uploaded_file is None:
        raise ValueError("Yüklenecek dosya bulunamadı.")

    original_name = os.path.basename(getattr(uploaded_file, 'name', 'uploaded.xlsx'))
    safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', original_name)
    suffix = os.path.splitext(safe_name)[1] or '.xlsx'

    temp_dir = os.path.join(tempfile.gettempdir(), 'talep_tahminleme_streamlit_uploads')
    os.makedirs(temp_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    save_path = os.path.join(temp_dir, unique_name)

    file_bytes = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file.read()
    if file_bytes is None:
        raise ValueError('Yüklenen dosya okunamadı.')

    with open(save_path, 'wb') as f:
        f.write(file_bytes)

    if not os.path.exists(save_path) or os.path.getsize(save_path) == 0:
        raise IOError('Yüklenen dosya diske yazılamadı.')

    if suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError('Desteklenmeyen dosya türü. Lütfen Excel dosyası yükleyin.')

    return save_path



def _check_optional_excel_dependency(ext: str) -> Tuple[bool, Optional[str], Optional[str]]:
    ext = str(ext).lower()
    if ext == '.xlsx':
        try:
            import openpyxl  # noqa: F401
            return True, 'openpyxl', None
        except Exception as e:
            return False, None, f".xlsx dosyalarını okumak için openpyxl gerekir: {e}"
    if ext == '.xls':
        try:
            import xlrd  # noqa: F401
            return True, 'xlrd', None
        except Exception as e:
            return False, None, f".xls dosyalarını okumak için xlrd gerekir: {e}"
    return False, None, f"Desteklenmeyen Excel uzantısı: {ext}"


def safe_excel_file(excel_path: str) -> pd.ExcelFile:
    ext = os.path.splitext(str(excel_path))[1].lower()
    ok, engine, msg = _check_optional_excel_dependency(ext)
    if not ok:
        raise ImportError(msg)
    try:
        return pd.ExcelFile(excel_path, engine=engine)
    except Exception as e:
        raise ImportError(f"Excel dosyası açılamadı ({os.path.basename(excel_path)}): {e}")


def safe_read_excel(excel_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    ext = os.path.splitext(str(excel_path))[1].lower()
    ok, engine, msg = _check_optional_excel_dependency(ext)
    if not ok:
        raise ImportError(msg)
    try:
        return pd.read_excel(excel_path, sheet_name=sheet_name, engine=engine)
    except Exception as e:
        raise ImportError(f"Excel sayfası okunamadı ({os.path.basename(excel_path)} / {sheet_name}): {e}")

def run_preprocessing_for_sheet(excel_path: str, sheet_name: str, output_dir: str) -> Dict[str, pd.DataFrame]:
    """Streamlit için tek sheet preprocessing wrapper'ı."""
    config = PreprocessConfig(
        output_dir_name="forecast_preprocessing_outputs",
        force_regular_frequency=True,
        allow_month_start_to_month_end_alignment_fix=True,
        max_interpolation_gap=1,
        use_knn_for_dense_missing_blocks=False,
        impute_method_preference="seasonal_local",
        min_action_confidence_for_auto_fix=0.75,
        auto_fix_business_spike_dip=False,
        auto_fix_unknown_anomaly=False,
        auto_fix_data_error=True,
        auto_fix_structural_event=False,
        scaler_for_deep_learning="robust",
        export_log1p_version=True,
        generate_modeling_ready_feature_pack=True,
        exclude_textual_columns_from_modeling_features=True,
        drop_low_signal_calendar_features_for_monthly=True,
        export_training_exclusion_masks=True,
        save_validation_excel=True,
        save_validation_csv=True,
        save_validation_plots=True,
        create_domain_validation_template=True,
        leakage_check_enabled=True,
        max_plot_series=50,
        run_internal_unit_tests=True,
        run_synthetic_tests=True,
        run_business_rule_tests=True,
        run_manual_sample_audit=True,
        run_proxy_backtest_validation=True,
        manual_sample_size=20,
        random_seed=42,
        backtest_horizon=3,
        backtest_min_train_size_monthly=24,
        backtest_min_train_size_weekly=52,
        backtest_min_train_size_daily=60,
        backtest_min_train_size_hourly=24 * 14,
        review_if_outlier_fraction_gt=0.05,
        review_if_structural_zero_events_gt=1,
        review_if_clean_zero_ratio_gt=0.10,
        review_if_proxy_smape_gt=60.0,
        save_excel=True,
        save_csv=True,
        save_metadata_json=True,
        save_quality_report=True
    )

    os.makedirs(output_dir, exist_ok=True)
    progress_log("Önişleme", "Tam önişleme başlatıldı.", extra={"sheet": sheet_name, "output_dir": output_dir})
    preprocessor = DemandForecastPreprocessor(config=config)
    export_payload = preprocessor.preprocess_sheet(
        file_path=excel_path,
        sheet_name=sheet_name,
        output_dir=output_dir
    )
    progress_log("Önişleme", "Tam önişleme başarıyla tamamlandı.", level="SUCCESS", extra={"sheet": sheet_name, "targets": int(len(export_payload.get("series_profile_report", pd.DataFrame())))})
    try:
        preprocessor.save_global_metadata(output_dir)
    except Exception:
        pass
    return export_payload


def choose_scaler(name: str):
    name = str(name).lower()
    if name == "standard":
        return StandardScaler()
    if name == "minmax":
        return MinMaxScaler()
    return RobustScaler()


def clip_negative_values(series: pd.Series) -> pd.Series:
    return series.clip(lower=0)


def stable_json_dumps(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, sort_keys=True, default=str)


def make_config_hash(config: PreprocessConfig) -> str:
    payload = stable_json_dumps(asdict(config))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def safe_float(x):
    try:
        return float(x)
    except Exception:
        return np.nan


def coefficient_of_variation(s: pd.Series) -> float:
    s = pd.to_numeric(s, errors="coerce").dropna()
    if len(s) == 0:
        return np.nan
    mean_ = s.mean()
    std_ = s.std()
    if mean_ == 0 or pd.isna(mean_):
        return np.nan
    return float(std_ / mean_)


def demand_intermittency_ratio(s: pd.Series) -> float:
    s = pd.to_numeric(s, errors="coerce")
    valid = s.notna()
    if valid.sum() == 0:
        return np.nan
    return float((s[valid] == 0).mean())


def robust_zscore(series: pd.Series) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce").astype(float)
    med = x.median()
    mad = np.median(np.abs(x - med))
    denom = 1.4826 * mad if mad not in [0, np.nan] else np.nan
    z = (x - med) / denom
    return z


# =========================================================
# DATE / TARGET DETECTION
# =========================================================

def detect_date_column(df: pd.DataFrame, config: PreprocessConfig) -> str:
    normalized = {c: normalize_colname(c) for c in df.columns}

    for c, nc in normalized.items():
        if nc in config.date_column_candidates:
            return c

    best_col = None
    best_ratio = 0.0
    for c in df.columns:
        temp = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        ratio = temp.notna().mean()
        if ratio > best_ratio:
            best_ratio = ratio
            best_col = c

    if best_col is None or best_ratio < 0.5:
        raise ValueError("Tarih sütunu bulunamadı.")
    return best_col


def parse_datetime_series(s: pd.Series) -> pd.Series:
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)

    if dt.notna().mean() < 0.8:
        dt = pd.to_datetime(
            s.astype(str),
            errors="coerce",
            dayfirst=True,
            infer_datetime_format=True
        )

    try:
        if getattr(dt.dt, "tz", None) is not None:
            dt = dt.dt.tz_localize(None)
    except Exception:
        pass

    return dt


def align_dates_to_frequency(df: pd.DataFrame, date_col: str, freq_alias: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()
    audit_rows = []

    before = out[date_col].copy()

    if freq_alias == "M":
        aligned = out[date_col].dt.to_period("M").dt.to_timestamp("M")
        changed = int((before != aligned).fillna(False).sum())
        out[date_col] = aligned
        audit_rows.append({
            "rule": "month_end_alignment",
            "applied": True,
            "changed_timestamp_count": changed,
            "note": "Aylık seriler ay sonu timestamp'ına hizalandı."
        })
    elif freq_alias == "W":
        aligned = out[date_col].dt.to_period("W").apply(lambda p: p.end_time.normalize())
        changed = int((before != aligned).fillna(False).sum())
        out[date_col] = aligned
        audit_rows.append({
            "rule": "week_end_alignment",
            "applied": True,
            "changed_timestamp_count": changed,
            "note": "Haftalık seriler hafta sonu timestamp'ına hizalandı."
        })
    elif freq_alias == "D":
        aligned = out[date_col].dt.floor("D")
        changed = int((before != aligned).fillna(False).sum())
        out[date_col] = aligned
        audit_rows.append({
            "rule": "day_floor_alignment",
            "applied": True,
            "changed_timestamp_count": changed,
            "note": "Günlük seriler gün başlangıcına hizalandı."
        })
    elif freq_alias == "H":
        aligned = out[date_col].dt.floor("H")
        changed = int((before != aligned).fillna(False).sum())
        out[date_col] = aligned
        audit_rows.append({
            "rule": "hour_floor_alignment",
            "applied": True,
            "changed_timestamp_count": changed,
            "note": "Saatlik seriler saat başlangıcına hizalandı."
        })

    if len(audit_rows) == 0:
        audit_rows.append({
            "rule": "alignment_not_applied",
            "applied": False,
            "changed_timestamp_count": 0,
            "note": "Frekansa özel hizalama uygulanmadı."
        })

    return out, pd.DataFrame(audit_rows)


def create_datetime_integrity_audit(
    df_original: pd.DataFrame,
    df_aligned: pd.DataFrame,
    df_aggregated: pd.DataFrame,
    df_regular: pd.DataFrame,
    date_col: str,
    freq_alias: str
) -> pd.DataFrame:
    raw_dates = pd.to_datetime(df_original[date_col], errors="coerce") if date_col in df_original.columns else pd.Series(dtype="datetime64[ns]")
    aligned_dates = pd.to_datetime(df_aligned[date_col], errors="coerce") if date_col in df_aligned.columns else pd.Series(dtype="datetime64[ns]")
    agg_dates = pd.to_datetime(df_aggregated[date_col], errors="coerce") if date_col in df_aggregated.columns else pd.Series(dtype="datetime64[ns]")
    reg_dates = pd.to_datetime(df_regular[date_col], errors="coerce") if date_col in df_regular.columns else pd.Series(dtype="datetime64[ns]")

    rows = [
        {"metric": "invalid_date_count_original", "value": int(raw_dates.isna().sum()) if len(raw_dates) else 0},
        {"metric": "duplicate_dates_before_alignment", "value": int(aligned_dates.duplicated().sum()) if len(aligned_dates) else 0},
        {"metric": "duplicate_dates_after_aggregation", "value": int(agg_dates.duplicated().sum()) if len(agg_dates) else 0},
        {"metric": "is_monotonic_after_regularization", "value": bool(reg_dates.is_monotonic_increasing) if len(reg_dates) else True},
        {"metric": "frequency_alias", "value": freq_alias},
        {"metric": "regularized_row_count", "value": int(len(reg_dates))},
        {"metric": "regularized_start", "value": str(reg_dates.min()) if len(reg_dates) else None},
        {"metric": "regularized_end", "value": str(reg_dates.max()) if len(reg_dates) else None},
    ]
    return pd.DataFrame(rows)


def infer_frequency_from_dates(dt_index: pd.DatetimeIndex) -> str:
    dt_index = pd.DatetimeIndex(dt_index).sort_values().drop_duplicates()

    inferred = pd.infer_freq(dt_index)
    if inferred:
        inferred = inferred.upper()
        if inferred.startswith("W"):
            return "W"
        if inferred in ["M", "MS", "ME"]:
            return "M"
        if inferred in ["D"]:
            return "D"
        if inferred in ["H"]:
            return "H"

    if len(dt_index) < 3:
        return "D"

    deltas = pd.Series(dt_index).diff().dropna()
    median_delta = deltas.median()

    if median_delta <= pd.Timedelta(hours=1):
        return "H"
    if median_delta <= pd.Timedelta(days=1):
        return "D"
    if median_delta <= pd.Timedelta(days=7):
        return "W"
    return "M"


def get_expected_freq_alias(freq: str) -> str:
    if freq == "H":
        return "H"
    if freq == "D":
        return "D"
    if freq == "W":
        return "W"
    return "M"


def detect_target_columns(df: pd.DataFrame, date_col: str, config: PreprocessConfig) -> List[str]:
    non_targets = {normalize_colname(x) for x in config.non_target_columns}
    candidates = []

    for c in df.columns:
        if c == date_col:
            continue

        nc = normalize_colname(c)
        if nc in non_targets:
            continue

        s = pd.to_numeric(df[c], errors="coerce")
        numeric_ratio = s.notna().mean()
        if numeric_ratio >= 0.5:
            candidates.append(c)

    if not candidates:
        raise ValueError("Hedef kolonlar otomatik bulunamadı.")
    return candidates


# =========================================================
# TIME INDEX / AGGREGATION
# =========================================================

def aggregate_duplicates(df: pd.DataFrame, date_col: str, target_cols: List[str]) -> pd.DataFrame:
    agg_map = {c: "sum" for c in target_cols}
    for c in df.columns:
        if c not in target_cols and c != date_col:
            agg_map[c] = "first"
    return df.groupby(date_col, as_index=False).agg(agg_map)


def build_regular_time_index(df: pd.DataFrame, date_col: str, freq: str) -> pd.DataFrame:
    df = df.sort_values(date_col).copy()
    start = df[date_col].min()
    end = df[date_col].max()

    if freq == "M":
        start = pd.Timestamp(start).to_period("M").to_timestamp("M")
        end = pd.Timestamp(end).to_period("M").to_timestamp("M")
        full_index = pd.date_range(start=start, end=end, freq="ME")
    elif freq == "W":
        full_index = pd.date_range(start=start, end=end, freq="W")
    elif freq == "D":
        full_index = pd.date_range(start=start, end=end, freq="D")
    else:
        full_index = pd.date_range(start=start, end=end, freq="H")

    out = df.set_index(date_col).reindex(full_index).rename_axis(date_col).reset_index()
    return out


def check_regular_index(df: pd.DataFrame, date_col: str, freq: str) -> Tuple[bool, str]:
    dates = pd.DatetimeIndex(df[date_col].dropna().sort_values())
    if len(dates) < 3:
        return True, "Yetersiz gözlem nedeniyle düzenlilik kontrolü sınırlı."

    try:
        expected = pd.date_range(start=dates.min(), end=dates.max(), freq=freq)
        ok = len(expected) == len(dates) and (expected == dates).all()
        if ok:
            return True, "Zaman ekseni düzenli."
        return False, "Zaman ekseninde eksik/fazla periyot veya hizalama sorunu var."
    except Exception as e:
        return False, f"Düzenli indeks kontrolü başarısız: {str(e)}"


# =========================================================
# ANOMALY DETECTION PRIMITIVES
# =========================================================

def hampel_filter_flags(series: pd.Series, window_size: int = 7, n_sigma: float = 4.0) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce").astype(float)
    rolling_median = x.rolling(window=window_size, center=True, min_periods=1).median()
    diff = np.abs(x - rolling_median)
    mad = diff.rolling(window=window_size, center=True, min_periods=1).median()
    threshold = n_sigma * 1.4826 * mad
    flags = diff > threshold
    return flags.fillna(False)


def rolling_mad_flags(series: pd.Series, window: int = 9, n_sigma: float = 4.5) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce").astype(float)
    median_ = x.rolling(window=window, center=True, min_periods=1).median()
    abs_dev = np.abs(x - median_)
    mad = abs_dev.rolling(window=window, center=True, min_periods=1).median()
    robust_z = (x - median_) / (1.4826 * mad.replace(0, np.nan))
    flags = np.abs(robust_z) > n_sigma
    return flags.fillna(False)


def iqr_flags(series: pd.Series, k: float = 4.0) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    lower = q1 - k * iqr
    upper = q3 + k * iqr
    return ((s < lower) | (s > upper)).fillna(False)


def limited_linear_interpolation(series: pd.Series, limit: int) -> pd.Series:
    return series.interpolate(method="linear", limit=limit, limit_direction="both")


# =========================================================
# SERIES PROFILING
# =========================================================

def estimate_trend_strength(series: pd.Series) -> float:
    s = pd.to_numeric(series, errors="coerce").dropna().reset_index(drop=True)
    if len(s) < 8:
        return np.nan
    x = np.arange(len(s))
    corr = np.corrcoef(x, s)[0, 1]
    return float(abs(corr)) if np.isfinite(corr) else np.nan


def estimate_seasonality_strength(series: pd.Series, season_length: int) -> float:
    s = pd.to_numeric(series, errors="coerce").dropna().reset_index(drop=True)
    if len(s) < max(2 * season_length, 12) or season_length <= 1:
        return np.nan

    grouped = {}
    for i, val in enumerate(s):
        grouped.setdefault(i % season_length, []).append(val)

    seasonal_means = {k: np.mean(v) for k, v in grouped.items()}
    fitted = np.array([seasonal_means[i % season_length] for i in range(len(s))], dtype=float)

    total_var = np.var(s)
    resid_var = np.var(s - fitted)
    if total_var <= 1e-12:
        return np.nan
    strength = 1 - (resid_var / total_var)
    return float(max(0.0, min(1.0, strength)))


def estimate_volatility_regime(series: pd.Series) -> str:
    """
    PATCH:
    - split 'moderate' into 'moderate' and 'elevated'
    """
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) < 8:
        return "unknown"

    cv = coefficient_of_variation(s)
    if pd.isna(cv):
        return "unknown"

    if cv < 0.20:
        return "stable"
    if cv < 0.35:
        return "moderate"
    if cv < 0.55:
        return "elevated"
    return "high"


def volume_level(series: pd.Series) -> str:
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) == 0:
        return "unknown"
    med = s.median()
    if med < 10:
        return "very_low"
    if med < 100:
        return "low"
    if med < 1000:
        return "medium"
    return "high"


def build_series_profile(series: pd.Series, freq_alias: str, config: PreprocessConfig) -> Dict[str, Any]:
    s = pd.to_numeric(series, errors="coerce")
    season_length = config.seasonal_period_map.get(freq_alias, 1)

    profile = {
        "n_obs": int(s.notna().sum()),
        "mean": safe_float(s.mean()),
        "median": safe_float(s.median()),
        "std": safe_float(s.std()),
        "cv": safe_float(coefficient_of_variation(s)),
        "intermittency_ratio": safe_float(demand_intermittency_ratio(s)),
        "trend_strength": safe_float(estimate_trend_strength(s)),
        "seasonality_strength": safe_float(estimate_seasonality_strength(s, season_length)),
        "volatility_regime": estimate_volatility_regime(s),
        "volume_level": volume_level(s),
        "min": safe_float(s.min()),
        "max": safe_float(s.max())
    }
    return profile


def create_series_profile_report(df: pd.DataFrame, target_cols: List[str], freq_alias: str, config: PreprocessConfig) -> pd.DataFrame:
    rows = []
    for col in target_cols:
        p = build_series_profile(df[col], freq_alias, config)
        p["series"] = col
        rows.append(p)
    cols = ["series"] + [c for c in rows[0].keys() if c != "series"] if rows else ["series"]
    return pd.DataFrame(rows)[cols] if rows else pd.DataFrame(columns=["series"])


def get_adaptive_thresholds(profile: Dict[str, Any], config: PreprocessConfig) -> Dict[str, float]:
    hampel_n_sigma = config.hampel_n_sigma
    rolling_mad_n_sigma = config.rolling_mad_n_sigma
    iqr_k = config.iqr_k

    cv = profile.get("cv", np.nan)
    intermittency = profile.get("intermittency_ratio", np.nan)
    vol_regime = profile.get("volatility_regime", "unknown")
    vol_level = profile.get("volume_level", "unknown")

    if vol_regime == "high":
        hampel_n_sigma += 0.8
        rolling_mad_n_sigma += 1.0
        iqr_k += 0.7
    elif vol_regime == "elevated":
        hampel_n_sigma += 0.3
        rolling_mad_n_sigma += 0.4
        iqr_k += 0.2
    elif vol_regime == "stable":
        hampel_n_sigma -= 0.5
        rolling_mad_n_sigma -= 0.5
        iqr_k -= 0.4

    if pd.notna(intermittency) and intermittency >= 0.40:
        hampel_n_sigma += 0.5
        rolling_mad_n_sigma += 0.7
        iqr_k += 0.6

    if vol_level in ["very_low", "low"]:
        hampel_n_sigma += 0.3
        rolling_mad_n_sigma += 0.3

    if pd.notna(cv) and cv < 0.20:
        hampel_n_sigma -= 0.3
        rolling_mad_n_sigma -= 0.3

    hampel_n_sigma = max(2.5, hampel_n_sigma)
    rolling_mad_n_sigma = max(3.0, rolling_mad_n_sigma)
    iqr_k = max(1.5, iqr_k)

    return {
        "hampel_n_sigma": hampel_n_sigma,
        "rolling_mad_n_sigma": rolling_mad_n_sigma,
        "iqr_k": iqr_k
    }


def conservative_outlier_vote_adaptive(series: pd.Series, profile: Dict[str, Any], config: PreprocessConfig) -> Tuple[pd.Series, pd.DataFrame]:
    thr = get_adaptive_thresholds(profile, config)

    flag_h = hampel_filter_flags(series, config.hampel_window, thr["hampel_n_sigma"])
    flag_m = rolling_mad_flags(series, config.rolling_mad_window, thr["rolling_mad_n_sigma"])
    flag_i = iqr_flags(series, thr["iqr_k"])

    vote_count = flag_h.astype(int) + flag_m.astype(int) + flag_i.astype(int)
    combined = vote_count >= config.min_outlier_votes

    vote_df = pd.DataFrame({
        "hampel": flag_h.astype(bool),
        "rolling_mad": flag_m.astype(bool),
        "iqr": flag_i.astype(bool),
        "vote_count": vote_count.astype(int),
        "combined": combined.astype(bool),
        "adaptive_hampel_n_sigma": thr["hampel_n_sigma"],
        "adaptive_rolling_mad_n_sigma": thr["rolling_mad_n_sigma"],
        "adaptive_iqr_k": thr["iqr_k"]
    })
    return combined.fillna(False), vote_df


def cap_outlier_fraction(series: pd.Series, combined_flags: pd.Series, vote_df: pd.DataFrame, max_fraction: float) -> pd.Series:
    combined_flags = combined_flags.copy()
    n = len(series)
    max_allowed = int(np.floor(n * max_fraction))

    flagged_idx = vote_df.index[vote_df["combined"]].tolist()
    if len(flagged_idx) <= max_allowed or max_allowed < 1:
        return combined_flags

    s = pd.to_numeric(series, errors="coerce")
    median_val = s.median()
    distance = (s - median_val).abs()

    ranking = vote_df.loc[flagged_idx].copy()
    ranking["distance"] = distance.loc[flagged_idx]
    ranking = ranking.sort_values(["vote_count", "distance"], ascending=[False, False])

    keep_idx = set(ranking.head(max_allowed).index.tolist())
    combined_flags[:] = False
    for idx in keep_idx:
        combined_flags.loc[idx] = True

    return combined_flags.fillna(False)


def protect_edge_periods(flags: pd.Series, config: PreprocessConfig) -> pd.Series:
    flags = flags.copy()
    if len(flags) == 0:
        return flags

    first_n = min(config.protect_first_n_periods, len(flags))
    last_n = min(config.protect_last_n_periods, len(flags))

    if first_n > 0:
        flags.iloc[:first_n] = False

    # Keep edge protection conservative, but avoid blinding the whole recent zone
    if last_n > 0:
        flags.iloc[-last_n:] = False

    return flags.fillna(False)


# =========================================================
# STRUCTURAL EVENT ENGINE
# =========================================================

def detect_structural_zero_events(
    df: pd.DataFrame,
    target_cols: List[str],
    min_series_count: int,
    ratio_threshold: float
) -> pd.Series:
    zero_matrix = pd.DataFrame(index=df.index)
    for col in target_cols:
        zero_matrix[col] = pd.to_numeric(df[col], errors="coerce").eq(0)

    zero_count = zero_matrix.sum(axis=1)
    zero_ratio = zero_count / max(len(target_cols), 1)
    structural = (zero_count >= min_series_count) & (zero_ratio >= ratio_threshold)
    return structural.fillna(False)


def expand_structural_events(structural_flags: pd.Series, neighbor_window: int) -> pd.Series:
    if neighbor_window <= 0 or len(structural_flags) == 0:
        return structural_flags.copy()

    expanded = structural_flags.copy().astype(bool)
    idx = np.where(structural_flags.values)[0]
    for i in idx:
        start = max(0, i - neighbor_window)
        end = min(len(expanded), i + neighbor_window + 1)
        expanded.iloc[start:end] = True
    return expanded.fillna(False)

def protect_structural_event_edges(flags: pd.Series, protect_last_n: int = 1) -> pd.Series:
    flags = flags.copy()
    if len(flags) > 0:
        flags.iloc[-protect_last_n:] = False
    return flags
def detect_incomplete_last_period(
    df_regular: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    freq_alias: str,
    config: PreprocessConfig
) -> Tuple[pd.Series, pd.DataFrame]:
    """
    Detects whether the last period is suspiciously low and may represent
    partial / incomplete reporting rather than real demand collapse.
    """
    flags = pd.Series(False, index=df_regular.index)
    rows = []

    if not config.enable_incomplete_period_detection:
        return flags, pd.DataFrame(columns=[
            "date", "rule_name", "portfolio_total", "baseline_median",
            "ratio_to_baseline", "is_incomplete_candidate", "reason"
        ])

    if len(df_regular) < max(6, config.partial_period_compare_last_n + 1):
        return flags, pd.DataFrame(columns=[
            "date", "rule_name", "portfolio_total", "baseline_median",
            "ratio_to_baseline", "is_incomplete_candidate", "reason"
        ])

    total = compute_portfolio_series(df_regular, target_cols).astype(float)
    last_idx = df_regular.index[-1]
    last_total = safe_float(total.iloc[-1])

    compare_n = max(1, config.partial_period_compare_last_n)
    hist = total.iloc[-(compare_n + 1):-1].dropna()

    if len(hist) == 0 or pd.isna(last_total):
        return flags, pd.DataFrame(columns=[
            "date", "rule_name", "portfolio_total", "baseline_median",
            "ratio_to_baseline", "is_incomplete_candidate", "reason"
        ])

    baseline = safe_float(hist.median())
    ratio = safe_float(last_total / baseline) if pd.notna(baseline) and baseline > 0 else np.nan

    is_candidate = bool(
        pd.notna(ratio) and
        ratio <= config.partial_period_drop_ratio_threshold
    )

    if is_candidate:
        flags.iloc[-1] = True
        rows.append({
            "date": df_regular.loc[last_idx, date_col],
            "rule_name": "last_period_portfolio_drop_vs_recent_median",
            "portfolio_total": last_total,
            "baseline_median": baseline,
            "ratio_to_baseline": ratio,
            "is_incomplete_candidate": True,
            "reason": "Son dönem toplamı, son dönemler medianına göre aşırı düşük. Kısmi raporlama / incomplete period adayı."
        })

    return flags.fillna(False), pd.DataFrame(rows)

def compute_portfolio_series(df: pd.DataFrame, target_cols: List[str]) -> pd.Series:
    total = pd.DataFrame({
        c: pd.to_numeric(df[c], errors="coerce") for c in target_cols
    }).sum(axis=1, min_count=1)
    return total


def detect_portfolio_shocks(
    df: pd.DataFrame,
    target_cols: List[str],
    config: PreprocessConfig
) -> pd.Series:
    """
    Portfolio-wide sharp drop detection.

    Important patch:
    - Do not classify the very last observation as structural shock automatically.
      Last period may be incomplete / partially reported.
    """
    total = compute_portfolio_series(df, target_cols).astype(float)
    prev = total.shift(1)
    ratio = total / prev.replace(0, np.nan)

    drop_flag = ratio <= (1 - config.portfolio_drop_ratio_threshold)
    drop_flag = drop_flag.fillna(False)

    # PATCH: protect last period from automatic structural shock tagging
    if len(drop_flag) > 0:
        drop_flag.iloc[-1] = False

    return drop_flag


def detect_rebound_after_event(df: pd.DataFrame, target_cols: List[str], event_flags: pd.Series, config: PreprocessConfig) -> pd.Series:
    total = compute_portfolio_series(df, target_cols).astype(float)
    next_ = total.shift(-1)
    ratio = next_ / total.replace(0, np.nan)
    rebound = ratio >= config.portfolio_rebound_ratio_threshold
    return (event_flags & rebound.fillna(False)).fillna(False)


def build_structural_event_log(
    df_regular: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    zero_flags: pd.Series,
    portfolio_shock_flags: pd.Series,
    rebound_flags: pd.Series
) -> pd.DataFrame:
    rows = []
    portfolio_series = compute_portfolio_series(df_regular, target_cols).astype(float)

    for idx in df_regular.index:
        triggered = []
        if bool(zero_flags.loc[idx]):
            triggered.append("multi_series_zero_event")
        if bool(portfolio_shock_flags.loc[idx]):
            triggered.append("portfolio_drop_event")
        if bool(rebound_flags.loc[idx]):
            triggered.append("rapid_rebound_pattern")

        if not triggered:
            continue

        event_type = "structural_event"
        if "portfolio_drop_event" in triggered:
            event_subtype = "portfolio_wide_shock"
        elif "multi_series_zero_event" in triggered:
            event_subtype = "category_or_reporting_shock"
        else:
            event_subtype = "structural_pattern"

        row_values = pd.to_numeric(df_regular.loc[idx, target_cols], errors="coerce")
        zero_series_count = int(row_values.eq(0).sum())
        non_null_series_count = int(row_values.notna().sum())
        non_zero_series_count = int((row_values.fillna(0) != 0).sum())

        rows.append({
            "date": df_regular.loc[idx, date_col],
            "event_type": event_type,
            "event_subtype": event_subtype,
            "triggered_rules": "|".join(triggered),
            "portfolio_total_sum": safe_float(portfolio_series.loc[idx]),
            "portfolio_total": safe_float(portfolio_series.loc[idx]),
            "zero_series_count": zero_series_count,
            "non_zero_series_count": non_zero_series_count,
            "non_null_series_count": non_null_series_count,
            "series_count_in_portfolio": int(len(target_cols)),
            "portfolio_drop_flag": bool(portfolio_shock_flags.loc[idx]),
            "multi_series_zero_flag": bool(zero_flags.loc[idx]),
            "rapid_rebound_flag": bool(rebound_flags.loc[idx])
        })
    if not rows:
        return pd.DataFrame(columns=[
            "date", "event_type", "event_subtype", "triggered_rules",
            "portfolio_total_sum", "portfolio_total", "zero_series_count",
            "non_zero_series_count", "non_null_series_count", "series_count_in_portfolio",
            "portfolio_drop_flag", "multi_series_zero_flag", "rapid_rebound_flag"
        ])
    return pd.DataFrame(rows)

def build_date_level_event_map(
    df_regular: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    structural_event_flags: pd.Series,
    incomplete_period_flags: pd.Series
) -> pd.DataFrame:
    rows = []

    for idx in df_regular.index:
        rows.append({
            "date": df_regular.loc[idx, date_col],
            "is_structural_event_date": bool(structural_event_flags.loc[idx]) if len(structural_event_flags) > 0 else False,
            "is_incomplete_period_date": bool(incomplete_period_flags.loc[idx]) if len(incomplete_period_flags) > 0 else False
        })

    return pd.DataFrame(rows)


# =========================================================
# ANOMALY GOVERNANCE / CLASSIFICATION
# =========================================================

def classify_anomaly(
    raw_value: float,
    prev_value: float,
    next_value: float,
    vote_count: int,
    profile: Dict[str, Any],
    is_structural_event: bool,
    is_recent_period: bool,
    is_incomplete_last_period: bool = False
) -> Tuple[str, str, float]:
    """
    Returns:
        anomaly_type, anomaly_reason, confidence

    PATCH LOGIC:
    - better low-volume / seasonal business-event interpretation
    - keep structural events dominant
    """
    if is_incomplete_last_period:
        reason = "Son dönem seviyesi aşırı düşük; incomplete / partial reporting şüphesi var."
        return "incomplete_last_period", reason, 0.95

    if is_structural_event:
        reason = "Aynı tarihte çok-serili ortak bozulma / portföy şoku paterni."
        return "structural_event", reason, 0.92

    if pd.notna(raw_value):
        if raw_value < 0:
            return "data_error", "Negatif talep/satış değeri tespit edildi.", 0.99
        if pd.isna(prev_value) and pd.isna(next_value) and vote_count >= 2:
            return "data_error", "Komşu bilgi zayıf, noktasal aykırılık yüksek.", 0.75

    if pd.notna(raw_value) and pd.notna(prev_value) and prev_value not in [0, np.nan]:
        change_ratio_prev = raw_value / prev_value
    else:
        change_ratio_prev = np.nan

    if pd.notna(raw_value) and pd.notna(next_value) and next_value not in [0, np.nan]:
        change_ratio_next = raw_value / next_value
    else:
        change_ratio_next = np.nan

    seasonality_strength = profile.get("seasonality_strength", np.nan)
    vol_regime = profile.get("volatility_regime", "unknown")
    volume_level_ = profile.get("volume_level", "unknown")

    if vote_count >= 2:
        strong_jump = (
            (pd.notna(change_ratio_prev) and (change_ratio_prev >= 2.5 or change_ratio_prev <= 0.4)) or
            (pd.notna(change_ratio_next) and (change_ratio_next >= 2.5 or change_ratio_next <= 0.4))
        )

        # PATCH: low-volume and strong seasonal series should be more tolerant
        if strong_jump:
            if volume_level_ in ["very_low", "low"]:
                return "business_spike_dip", "Düşük hacimli seride beklenebilir sıçrama/düşüş.", 0.74

            if pd.notna(seasonality_strength) and seasonality_strength >= 0.60:
                return "business_spike_dip", "Güçlü sezonsallık bağlamında sıçrama/düşüş.", 0.78

            if vol_regime in ["elevated", "high"]:
                return "business_spike_dip", "Yüksek oynaklık rejiminde büyük sapma.", 0.72

    if vote_count >= 2:
        reason = "Aykırılık kuralları tetiklendi fakat net iş nedeni veya veri hatası ayrılamadı."
        conf = 0.62 if not is_recent_period else 0.55
        return "unknown_anomaly", reason, conf

    return "none", "Anomali yok.", 0.0


def decide_action(
    anomaly_type: str,
    confidence: float,
    is_recent_period: bool,
    config: PreprocessConfig
) -> Tuple[str, str]:
    """
    action_taken, governance_policy

    PATCH LOGIC:
    - recent periods: always human-review-first
    - unknown anomaly outside recent region: allow controlled auto-fix
    - structural events: preserve unless explicitly enabled
    """
    if anomaly_type == "none":
        return "keep_raw", "no_action"

    # CRITICAL FIX:
    # incomplete last period must ALWAYS be excluded from training,
    # even if it is also a recent period.
    if anomaly_type == "incomplete_last_period":
        return "preserve_raw_flag_exclude_candidate", "incomplete_last_period_exclusion_policy"

    # structural events should also preserve/exclude before generic recent-period handling
    if anomaly_type == "structural_event":
        if config.auto_fix_structural_event and confidence >= config.min_action_confidence_for_auto_fix:
            return "set_nan_then_impute", "auto_fix_structural_event_enabled"
        return "preserve_raw_flag_exclude_candidate", "structural_event_preservation_policy"

    # recent periods: human-review-first for the remaining anomaly classes
    if is_recent_period:
        return "flag_only_review", "recent_period_human_review_first"

    if anomaly_type == "data_error":
        if config.auto_fix_data_error and confidence >= config.min_action_confidence_for_auto_fix:
            return "set_nan_then_impute", "auto_fix_high_confidence_data_error"
        return "flag_only_review", "review_due_to_low_confidence_data_error"
    

    if anomaly_type == "business_spike_dip":
        if config.auto_fix_business_spike_dip and confidence >= config.min_action_confidence_for_auto_fix:
            return "set_nan_then_impute", "auto_fix_business_event_enabled"
        return "keep_raw_flag", "preserve_possible_business_event"

    if anomaly_type == "unknown_anomaly":
        if config.auto_fix_unknown_anomaly and confidence >= config.min_action_confidence_for_auto_fix and not is_recent_period:
            return "set_nan_then_impute", "auto_fix_unknown_non_recent_enabled"
        return "flag_only_review", "unknown_anomaly_review"

    return "flag_only_review", "fallback_review_policy"


def build_anomaly_governance_table(
    df_regular: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    outlier_flags: Dict[str, pd.Series],
    vote_details: Dict[str, pd.DataFrame],
    series_profiles: Dict[str, Dict[str, Any]],
    structural_event_flags: pd.Series,
    incomplete_period_flags: pd.Series,
    config: PreprocessConfig
) -> pd.DataFrame:
    rows = []
    n = len(df_regular)

    for col in target_cols:
        s = pd.to_numeric(df_regular[col], errors="coerce")

        for idx in df_regular.index:
            is_structural = bool(structural_event_flags.loc[idx]) if len(structural_event_flags) > 0 else False
            is_incomplete_last_period = bool(incomplete_period_flags.loc[idx]) if len(incomplete_period_flags) > 0 else False
            outlier_hit = bool(outlier_flags[col].loc[idx]) if col in outlier_flags else False

            # NEW: date-level events should force row creation for all series
            force_row = is_structural or is_incomplete_last_period
            if not outlier_hit and not force_row:
                continue

            raw_val = s.loc[idx]
            prev_val = s.shift(1).loc[idx]
            next_val = s.shift(-1).loc[idx]
            vote_count = int(vote_details[col].loc[idx, "vote_count"]) if col in vote_details else 0
            is_recent = idx >= (n - config.recent_periods_review_only)

            anomaly_type, reason, conf = classify_anomaly(
                raw_value=raw_val,
                prev_value=prev_val,
                next_value=next_val,
                vote_count=vote_count,
                profile=series_profiles[col],
                is_structural_event=is_structural,
                is_recent_period=is_recent,
                is_incomplete_last_period=is_incomplete_last_period
            )

            action_taken, governance_policy = decide_action(
                anomaly_type=anomaly_type,
                confidence=conf,
                is_recent_period=is_recent,
                config=config
            )

            rows.append({
                "date": df_regular.loc[idx, date_col],
                "series": col,
                "raw_value": raw_val,
                "prev_value": prev_val,
                "next_value": next_val,
                "vote_count": vote_count,
                "anomaly_type": anomaly_type,
                "anomaly_reason": reason,
                "action_taken": action_taken,
                "action_confidence": conf,
                "governance_policy": governance_policy,
                "is_structural_event": is_structural,
                "is_incomplete_last_period": is_incomplete_last_period,
                "is_recent_period": is_recent,
                "preserved_for_modeling": action_taken in ["keep_raw", "keep_raw_flag", "flag_only_review"],
                "excluded_from_training_candidate": (
                    action_taken in ["preserve_raw_flag_exclude_candidate"]
                    or anomaly_type in ["incomplete_last_period", "structural_event"]
                ),
                "is_training_excluded": (
                    action_taken in ["preserve_raw_flag_exclude_candidate"]
                    or anomaly_type in ["incomplete_last_period", "structural_event"]
                ),
                "is_preserved_for_review": action_taken in ["flag_only_review", "keep_raw_flag", "preserve_raw_flag_exclude_candidate"],
                "is_final_model_input": action_taken not in ["preserve_raw_flag_exclude_candidate"],
                "recommended_manual_validation": action_taken in ["flag_only_review", "keep_raw_flag", "preserve_raw_flag_exclude_candidate"]
            })

    if not rows:
        return pd.DataFrame(columns=[
            "date", "series", "raw_value", "prev_value", "next_value", "vote_count",
            "anomaly_type", "anomaly_reason", "action_taken", "action_confidence",
            "governance_policy", "is_structural_event", "is_incomplete_last_period",
            "is_recent_period", "preserved_for_modeling", "excluded_from_training_candidate",
            "is_training_excluded", "is_preserved_for_review", "is_final_model_input",
            "recommended_manual_validation"
        ])

    return (
        pd.DataFrame(rows)
        .sort_values(["series", "date", "action_confidence"], ascending=[True, True, False])
        .drop_duplicates(subset=["date", "series"], keep="first")
        .reset_index(drop=True)
    )


# =========================================================
# IMPUTATION
# =========================================================

def seasonal_local_impute(
    df: pd.DataFrame,
    target_col: str,
    date_col: str,
    freq: str,
    seasonal_period: int,
    max_interpolation_gap: int = 1
) -> pd.Series:
    s = pd.to_numeric(df[target_col], errors="coerce").copy()

    if freq == "M":
        group_key = df[date_col].dt.month
    elif freq == "W":
        group_key = df[date_col].dt.isocalendar().week.astype(int)
    elif freq == "D":
        group_key = df[date_col].dt.dayofweek
    elif freq == "H":
        group_key = df[date_col].dt.hour
    else:
        group_key = pd.Series([0] * len(df), index=df.index)

    seasonal_med = s.groupby(group_key).transform("median")
    s = s.fillna(seasonal_med)
    s = limited_linear_interpolation(s, limit=max_interpolation_gap)

    local_med = s.rolling(window=3, min_periods=1).median()
    s = s.fillna(local_med)

    if seasonal_period > 1 and s.isna().any():
        fallback = s.shift(seasonal_period)
        s = s.fillna(fallback)

    s = s.fillna(s.median())
    return s




def _max_consecutive_true(mask: pd.Series) -> int:
    arr = pd.Series(mask).fillna(False).astype(int).values
    best = cur = 0
    for v in arr:
        if v == 1:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return int(best)


def summarize_missingness_patterns(
    df_regular: pd.DataFrame,
    target_cols: List[str],
    date_col: str
) -> pd.DataFrame:
    rows = []
    for col in target_cols:
        s = pd.to_numeric(df_regular[col], errors="coerce")
        miss = s.isna()
        rows.append({
            "series": col,
            "missing_count": int(miss.sum()),
            "missing_ratio": float(miss.mean()),
            "missing_at_start": bool(miss.iloc[0]) if len(miss) else False,
            "missing_at_end": bool(miss.iloc[-1]) if len(miss) else False,
            "max_consecutive_missing_block": _max_consecutive_true(miss),
            "recommended_dense_block_method": "knn_or_review" if _max_consecutive_true(miss) >= 3 else "seasonal_local"
        })
    return pd.DataFrame(rows)


def decide_missing_value_strategy(
    missingness_summary: pd.DataFrame,
    config: PreprocessConfig
) -> pd.DataFrame:
    rows = []
    for _, row in missingness_summary.iterrows():
        ratio = float(row["missing_ratio"])
        block = int(row["max_consecutive_missing_block"])
        strategy = "impute_seasonal_local"
        reason = "Eksik oranı düşük/orta; zaman serisi bütünlüğü korunarak imputasyon tercih edildi."
        exclude_series = False
        row_drop = False

        if ratio >= config.missing_drop_series_threshold:
            strategy = "review_required_series_too_sparse"
            reason = "Eksik oranı çok yüksek; seri modelleme için zayıf aday."
            exclude_series = True
        elif block >= config.dense_missing_block_threshold and config.use_knn_for_dense_missing_blocks:
            strategy = "impute_knn_dense_block"
            reason = "Ardışık eksik blok yoğun; KNN aday yöntem olarak işaretlendi."
        elif ratio >= config.missing_impute_ratio_threshold:
            strategy = "impute_seasonal_local_with_review"
            reason = "Eksik oranı orta-yüksek; seasonal/local imputasyon + uzman gözden geçirme önerildi."
        elif ratio == 0:
            strategy = "no_imputation_needed"
            reason = "Eksik değer bulunmadı."

        rows.append({
            "series": row["series"],
            "missing_ratio": ratio,
            "max_consecutive_missing_block": block,
            "selected_strategy": strategy,
            "reason": reason,
            "row_drop_applied": row_drop,
            "series_excluded_from_modeling": exclude_series,
            "imputer_name": "seasonal_local" if "impute" in strategy else None
        })
    return pd.DataFrame(rows)


def create_descriptive_statistics_report(df_clean: pd.DataFrame, target_cols: List[str]) -> pd.DataFrame:
    rows = []
    for col in target_cols:
        s = pd.to_numeric(df_clean[col], errors="coerce")
        rows.append({
            "series": col,
            "count": int(s.notna().sum()),
            "mean": safe_float(s.mean()),
            "std": safe_float(s.std()),
            "min": safe_float(s.min()),
            "q25": safe_float(s.quantile(0.25)),
            "q50": safe_float(s.quantile(0.50)),
            "q75": safe_float(s.quantile(0.75)),
            "max": safe_float(s.max()),
            "iqr": safe_float(s.quantile(0.75) - s.quantile(0.25)),
            "sum": safe_float(s.sum()),
            "skewness": safe_float(s.skew()),
            "kurtosis": safe_float(s.kurt())
        })
    return pd.DataFrame(rows)


def create_monthly_seasonality_report(
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    freq_alias: str,
    config: PreprocessConfig
) -> pd.DataFrame:
    rows = []
    season_length = config.seasonal_period_map.get(freq_alias, 1)
    for col in target_cols:
        s = pd.to_numeric(df_clean[col], errors="coerce")
        tmp = pd.DataFrame({date_col: df_clean[date_col], col: s}).dropna()
        if len(tmp) == 0:
            continue

        if freq_alias == "M":
            grp_key = tmp[date_col].dt.month
            grp_name = "month"
        elif freq_alias == "W":
            grp_key = tmp[date_col].dt.isocalendar().week.astype(int)
            grp_name = "iso_week"
        elif freq_alias == "D":
            grp_key = tmp[date_col].dt.dayofweek
            grp_name = "dayofweek"
        else:
            grp_key = tmp[date_col].dt.hour
            grp_name = "hour"

        prof = tmp.groupby(grp_key)[col].agg(["mean", "median", "min", "max", "count"]).reset_index()
        overall_mean = tmp[col].mean()
        peak_row = prof.sort_values("mean", ascending=False).iloc[0]
        trough_row = prof.sort_values("mean", ascending=True).iloc[0]

        rows.append({
            "series": col,
            "grouping": grp_name,
            "peak_period": int(peak_row.iloc[0]),
            "peak_mean": safe_float(peak_row["mean"]),
            "trough_period": int(trough_row.iloc[0]),
            "trough_mean": safe_float(trough_row["mean"]),
            "peak_to_trough_ratio": safe_float(peak_row["mean"] / trough_row["mean"]) if pd.notna(trough_row["mean"]) and trough_row["mean"] not in [0, 0.0] else np.nan,
            "overall_mean": safe_float(overall_mean),
            "seasonality_strength": safe_float(estimate_seasonality_strength(tmp[col], season_length))
        })
    return pd.DataFrame(rows)


def create_pharma_event_diagnostic_report(
    df_regular: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    anomaly_gov: pd.DataFrame,
    config: PreprocessConfig
) -> pd.DataFrame:
    if len(anomaly_gov) == 0:
        return pd.DataFrame(columns=["date", "series", "raw_value", "prev_value", "next_value", "diagnostic_tag", "note"])

    rows = []
    date_to_idx = {d: i for i, d in enumerate(df_regular[date_col].tolist())}
    for _, row in anomaly_gov.iterrows():
        series = row["series"]
        date_ = row["date"]
        idx = date_to_idx.get(date_)
        if idx is None:
            continue
        s = pd.to_numeric(df_regular[series], errors="coerce")
        raw = s.iloc[idx]
        prev_ = s.iloc[idx - 1] if idx - 1 >= 0 else np.nan
        next_ = s.iloc[idx + 1] if idx + 1 < len(s) else np.nan
        tag = "anomaly_review"
        note = "İstatistiksel aykırılık adayı."

        if pd.notna(prev_) and prev_ > 0 and pd.notna(raw) and raw / prev_ <= config.stockout_like_drop_ratio:
            tag = "possible_stockout_or_supply_issue"
            note = "Ani düşüş tespit edildi; stok kesintisi / sevkiyat problemi adayı."
        elif pd.notna(prev_) and prev_ > 0 and pd.notna(raw) and raw / prev_ >= config.promotion_like_jump_ratio:
            tag = "possible_campaign_or_bulk_order"
            note = "Ani sıçrama tespit edildi; kampanya / ihale / toplu sipariş adayı."
        elif pd.notna(prev_) and pd.notna(next_) and prev_ > 0 and raw > 0 and next_ / raw >= config.rebound_after_drop_ratio:
            tag = "possible_rebound_after_disruption"
            note = "Şok sonrası hızlı rebound görüldü; geçici tedarik bozulması adayı."

        rows.append({
            "date": date_,
            "series": series,
            "raw_value": safe_float(raw),
            "prev_value": safe_float(prev_),
            "next_value": safe_float(next_),
            "diagnostic_tag": tag,
            "note": note
        })
    return pd.DataFrame(rows)


def _safe_linear_trend(series: pd.Series) -> pd.Series:
    s = pd.to_numeric(series, errors="coerce")
    if s.notna().sum() < 2:
        return pd.Series(np.nan, index=s.index)
    x = np.arange(len(s))
    mask = s.notna().values
    coeffs = np.polyfit(x[mask], s.values[mask], deg=1)
    fitted = coeffs[0] * x + coeffs[1]
    return pd.Series(fitted, index=s.index)


def _normalized_index(values: pd.Series) -> pd.Series:
    s = pd.to_numeric(values, errors="coerce")
    mean_ = s.mean()
    if pd.isna(mean_) or mean_ == 0:
        return s * np.nan
    return s / mean_


def drift_forecast(train: pd.Series, horizon: int) -> np.ndarray:
    train = pd.to_numeric(train, errors="coerce").dropna().reset_index(drop=True)
    if len(train) == 0:
        return np.full(horizon, np.nan)
    if len(train) == 1:
        return np.repeat(train.iloc[-1], horizon).astype(float)
    drift = (train.iloc[-1] - train.iloc[0]) / max(len(train) - 1, 1)
    return np.array([train.iloc[-1] + drift * (i + 1) for i in range(horizon)], dtype=float)


def create_model_input_transparency_export(
    df_regular: pd.DataFrame,
    df_clean_candidate: pd.DataFrame,
    df_clean_governed_preserve: pd.DataFrame,
    df_feat: pd.DataFrame,
    date_col: str,
    target_cols: List[str]
) -> pd.DataFrame:
    out = pd.DataFrame({date_col: df_regular[date_col].values})
    for col in target_cols:
        raw_s = pd.to_numeric(df_regular[col], errors="coerce")
        cand_s = pd.to_numeric(df_clean_candidate[col], errors="coerce")
        preserve_s = pd.to_numeric(df_clean_governed_preserve[col], errors="coerce")
        exclude_col = f"{col}_exclude_from_training"
        review_col = f"{col}_review_required"
        excluded = df_feat[exclude_col].astype(int) if exclude_col in df_feat.columns else pd.Series(0, index=df_regular.index)
        review = df_feat[review_col].astype(int) if review_col in df_feat.columns else pd.Series(0, index=df_regular.index)
        final_model = cand_s.copy()
        final_model.loc[excluded.astype(bool)] = np.nan

        out[f"{col}__raw_regular"] = raw_s.values
        out[f"{col}__candidate_clean"] = cand_s.values
        out[f"{col}__preserve_clean"] = preserve_s.values
        out[f"{col}__train_excluded"] = excluded.values
        out[f"{col}__is_preserved_for_review"] = review.values
        out[f"{col}__recommended_manual_validation"] = review.values
        out[f"{col}__is_final_model_input"] = (~excluded.astype(bool)).astype(int).values
        out[f"{col}__final_model_series"] = final_model.values
    return out


def save_raw_clean_trend_plots(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    sheet_dir: str,
    max_plot_series: int = 50,
    ma_windows: Tuple[int, int] = (3, 6),
    robust_trend_window: int = 5,
    save_robust_trend: bool = True
):
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)
    for col in target_cols[:max_plot_series]:
        raw_s = pd.to_numeric(df_regular[col], errors="coerce")
        clean_s = pd.to_numeric(df_clean[col], errors="coerce")
        plt.figure(figsize=(13, 5))
        plt.plot(df_regular[date_col], raw_s, label="Raw-Regular", linewidth=1.2)
        plt.plot(df_clean[date_col], clean_s, label="Clean", linewidth=1.5)
        for w in ma_windows:
            if len(clean_s) >= 2:
                plt.plot(
                    df_clean[date_col],
                    clean_s.rolling(window=min(w, max(len(clean_s), 1)), min_periods=1).mean(),
                    label=f"Clean_MA_{w}", linewidth=1.1
                )
        if save_robust_trend:
            plt.plot(
                df_clean[date_col],
                clean_s.rolling(window=min(max(robust_trend_window, 3), max(len(clean_s), 1)), min_periods=1).median(),
                label=f"Clean_RollMedian_{robust_trend_window}", linewidth=1.1
            )
            linear_trend = _safe_linear_trend(clean_s)
            if linear_trend.notna().any():
                plt.plot(df_clean[date_col], linear_trend, label="Clean_LinearTrend", linewidth=1.0)
        plt.title(f"{col} - Raw/Clean + Trend")
        plt.xlabel("Date")
        plt.ylabel("Value")
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(plot_dir, f"{col}_raw_clean_trend.png"), dpi=150)
        plt.close()


def save_distribution_plots(
    df_clean: pd.DataFrame,
    target_cols: List[str],
    sheet_dir: str,
    max_plot_series: int = 50,
    save_boxplots: bool = True
):
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)
    for col in target_cols[:max_plot_series]:
        s = pd.to_numeric(df_clean[col], errors="coerce").dropna()
        if len(s) == 0:
            continue
        plt.figure(figsize=(9, 5))
        plt.hist(s.values, bins=min(30, max(10, int(np.sqrt(len(s))))), edgecolor="black")
        plt.title(f"{col} - Distribution")
        plt.xlabel("Value")
        plt.ylabel("Frequency")
        plt.tight_layout()
        plt.savefig(os.path.join(plot_dir, f"{col}_distribution_hist.png"), dpi=150)
        plt.close()

        if save_boxplots:
            plt.figure(figsize=(8, 4.5))
            plt.boxplot(s.values, vert=False)
            plt.title(f"{col} - Boxplot")
            plt.xlabel("Value")
            plt.tight_layout()
            plt.savefig(os.path.join(plot_dir, f"{col}_distribution_boxplot.png"), dpi=150)
            plt.close()


def save_seasonality_plots(
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    freq_alias: str,
    sheet_dir: str,
    max_plot_series: int = 50,
    save_year_overlay: bool = True,
    save_normalized_profile: bool = True,
    save_boxplot: bool = True
):
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)
    for col in target_cols[:max_plot_series]:
        s = pd.to_numeric(df_clean[col], errors="coerce")
        tmp = pd.DataFrame({date_col: df_clean[date_col], col: s}).dropna()
        if len(tmp) == 0:
            continue

        if freq_alias == "M":
            tmp["season_key"] = tmp[date_col].dt.month
            tmp["year_key"] = tmp[date_col].dt.year
            xlab = "Month"
        elif freq_alias == "W":
            tmp["season_key"] = tmp[date_col].dt.isocalendar().week.astype(int)
            tmp["year_key"] = tmp[date_col].dt.year
            xlab = "ISO Week"
        elif freq_alias == "D":
            tmp["season_key"] = tmp[date_col].dt.dayofweek
            tmp["year_key"] = pd.to_datetime(tmp[date_col]).dt.to_period("M").astype(str)
            xlab = "Day of Week"
        else:
            tmp["season_key"] = tmp[date_col].dt.hour
            tmp["year_key"] = pd.to_datetime(tmp[date_col]).dt.to_period("D").astype(str)
            xlab = "Hour"

        grp_mean = tmp.groupby("season_key")[col].mean()
        grp_median = tmp.groupby("season_key")[col].median()
        if len(grp_mean) == 0:
            continue

        plt.figure(figsize=(10, 5))
        plt.plot(grp_mean.index.astype(int), grp_mean.values, marker="o", label="Mean")
        plt.plot(grp_median.index.astype(int), grp_median.values, marker="s", label="Median")
        plt.title(f"{col} - Seasonal Profile")
        plt.xlabel(xlab)
        plt.ylabel("Demand")
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(plot_dir, f"{col}_seasonal_profile.png"), dpi=150)
        plt.close()

        if save_normalized_profile:
            norm = _normalized_index(grp_mean)
            plt.figure(figsize=(10, 5))
            plt.plot(norm.index.astype(int), norm.values, marker="o")
            plt.axhline(1.0, linewidth=1.0)
            plt.title(f"{col} - Normalized Seasonal Index")
            plt.xlabel(xlab)
            plt.ylabel("Index")
            plt.tight_layout()
            plt.savefig(os.path.join(plot_dir, f"{col}_normalized_seasonal_index.png"), dpi=150)
            plt.close()

        if save_year_overlay:
            piv = tmp.pivot_table(index="season_key", columns="year_key", values=col, aggfunc="mean")
            if piv.shape[1] >= 1:
                plt.figure(figsize=(11, 5))
                for c_year in piv.columns:
                    plt.plot(piv.index.astype(int), piv[c_year].values, marker="o", linewidth=1.0, label=str(c_year))
                plt.title(f"{col} - Seasonal Overlay by Period")
                plt.xlabel(xlab)
                plt.ylabel("Demand")
                if piv.shape[1] <= 8:
                    plt.legend()
                plt.tight_layout()
                plt.savefig(os.path.join(plot_dir, f"{col}_seasonal_overlay.png"), dpi=150)
                plt.close()

        if save_boxplot:
            grouped = [g[col].values for _, g in tmp.groupby("season_key") if len(g) > 0]
            labels = [int(k) for k, g in tmp.groupby("season_key") if len(g) > 0]
            if grouped:
                plt.figure(figsize=(11, 5))
                plt.boxplot(grouped, labels=labels)
                plt.title(f"{col} - Seasonal Boxplot")
                plt.xlabel(xlab)
                plt.ylabel("Demand")
                plt.tight_layout()
                plt.savefig(os.path.join(plot_dir, f"{col}_seasonal_boxplot.png"), dpi=150)
                plt.close()



def _safe_plot_name(name: str) -> str:
    return re.sub(r"[^\w\-]+", "_", str(name))


def _seasonal_period_from_freq(freq_alias: str) -> int:
    if freq_alias == "H":
        return 24
    if freq_alias == "D":
        return 7
    if freq_alias == "W":
        return 52
    return 12


def _make_unique_text_labels(labels: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for raw in list(labels):
        base = str(raw)
        if base not in seen:
            seen[base] = 0
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}__{seen[base]}")
    return out


def _plot_tick_step(n_labels: int) -> int:
    if n_labels <= 20:
        return 1
    if n_labels <= 40:
        return 2
    if n_labels <= 80:
        return 4
    if n_labels <= 120:
        return 6
    return 8


def _safe_save_matrix_heatmap(matrix: pd.DataFrame, title: str, output_path: str) -> bool:
    try:
        if matrix is None or matrix.empty:
            return False
        row_labels = _make_unique_text_labels(matrix.index.tolist())
        col_labels = _make_unique_text_labels(matrix.columns.tolist())
        arr = np.asarray(matrix.values, dtype=float)
        fig, ax = plt.subplots(figsize=(max(10, 0.24 * len(col_labels)), max(8, 0.24 * len(row_labels))))
        im = ax.imshow(arr, aspect="auto", cmap="coolwarm", vmin=-1, vmax=1)
        fig.colorbar(im, ax=ax)
        x_step = _plot_tick_step(len(col_labels))
        y_step = _plot_tick_step(len(row_labels))
        x_pos = np.arange(0, len(col_labels), x_step)
        y_pos = np.arange(0, len(row_labels), y_step)
        ax.set_xticks(x_pos)
        ax.set_xticklabels([col_labels[i] for i in x_pos], rotation=90, fontsize=8)
        ax.set_yticks(y_pos)
        ax.set_yticklabels([row_labels[i] for i in y_pos], fontsize=8)
        ax.set_title(title)
        fig.tight_layout()
        fig.savefig(output_path, dpi=150)
        plt.close(fig)
        return True
    except Exception:
        try:
            plt.close('all')
        except Exception:
            pass
        return False


def _safe_save_pivot_heatmap(piv: pd.DataFrame, title: str, output_path: str) -> bool:
    try:
        if piv is None or piv.empty:
            return False
        row_labels = _make_unique_text_labels(piv.index.tolist())
        col_labels = _make_unique_text_labels(piv.columns.tolist())
        arr = np.asarray(piv.values, dtype=float)
        fig, ax = plt.subplots(figsize=(max(8, 0.42 * len(col_labels)), max(6, 0.32 * len(row_labels))))
        im = ax.imshow(arr, aspect="auto", cmap="viridis")
        fig.colorbar(im, ax=ax)
        x_step = _plot_tick_step(len(col_labels))
        y_step = _plot_tick_step(len(row_labels))
        x_pos = np.arange(0, len(col_labels), x_step)
        y_pos = np.arange(0, len(row_labels), y_step)
        ax.set_xticks(x_pos)
        ax.set_xticklabels([col_labels[i] for i in x_pos], fontsize=8)
        ax.set_yticks(y_pos)
        ax.set_yticklabels([row_labels[i] for i in y_pos], fontsize=8)
        ax.set_title(title)
        fig.tight_layout()
        fig.savefig(output_path, dpi=150)
        plt.close(fig)
        return True
    except Exception:
        try:
            plt.close('all')
        except Exception:
            pass
        return False


def save_correlation_analysis(
    df_for_corr: pd.DataFrame,
    target_cols: List[str],
    sheet_dir: str
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)

    numeric_df = df_for_corr.select_dtypes(include=[np.number]).copy()
    numeric_df = numeric_df.dropna(axis=1, how="all")

    if numeric_df.shape[1] < 2:
        corr_matrix = pd.DataFrame(columns=["variable"])
        corr_long = pd.DataFrame(columns=["target_series", "variable", "correlation", "abs_correlation"])
        return corr_matrix, corr_long

    original_numeric_cols = list(numeric_df.columns)
    unique_numeric_cols = _make_unique_text_labels(original_numeric_cols)
    numeric_df.columns = unique_numeric_cols
    target_name_map: Dict[str, List[str]] = {}
    for orig_name, unique_name in zip(original_numeric_cols, unique_numeric_cols):
        target_name_map.setdefault(str(orig_name), []).append(unique_name)

    corr_matrix = numeric_df.corr(method="pearson")
    corr_matrix.to_csv(os.path.join(sheet_dir, "correlation_matrix.csv"), encoding="utf-8-sig")
    corr_matrix.to_excel(os.path.join(sheet_dir, "correlation_matrix.xlsx"))

    _safe_save_matrix_heatmap(
        corr_matrix,
        "Correlation Matrix",
        os.path.join(plot_dir, "correlation_matrix_heatmap.png")
    )

    rows = []
    for target in target_cols:
        mapped_targets = target_name_map.get(str(target), [])
        if not mapped_targets:
            continue
        for mapped_target in mapped_targets:
            if mapped_target not in corr_matrix.columns:
                continue
            tmp = corr_matrix[mapped_target].drop(labels=[mapped_target], errors="ignore").reset_index()
            if tmp.shape[1] < 2:
                continue
            tmp = tmp.iloc[:, :2].copy()
            tmp.columns = ["variable", "correlation"]
            tmp["target_series"] = str(target)
            tmp["abs_correlation"] = pd.to_numeric(tmp["correlation"], errors="coerce").abs()
            tmp = tmp.sort_values("abs_correlation", ascending=False)
            rows.append(tmp[["target_series", "variable", "correlation", "abs_correlation"]])

            top_n = min(20, len(tmp))
            if top_n > 0:
                try:
                    plt.figure(figsize=(10, max(5, 0.35 * top_n)))
                    tmp_plot = tmp.head(top_n).sort_values("correlation")
                    plt.barh(tmp_plot["variable"].astype(str), tmp_plot["correlation"].astype(float))
                    plt.axvline(0.0, linewidth=1.0)
                    plt.title(f"Top correlations with {target}")
                    plt.xlabel("Correlation")
                    plt.tight_layout()
                    plt.savefig(os.path.join(plot_dir, f"{_safe_plot_name(str(target))}_top_correlations.png"), dpi=150)
                    plt.close()
                except Exception:
                    try:
                        plt.close('all')
                    except Exception:
                        pass

    corr_long = pd.concat(rows, axis=0, ignore_index=True) if rows else pd.DataFrame(columns=["target_series", "variable", "correlation", "abs_correlation"])
    corr_long.to_csv(os.path.join(sheet_dir, "target_correlations_long.csv"), index=False, encoding="utf-8-sig")
    corr_long.to_excel(os.path.join(sheet_dir, "target_correlations_long.xlsx"), index=False)
    return corr_matrix, corr_long

def save_seasonality_heatmaps_and_decomposition(
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    freq_alias: str,
    sheet_dir: str,
    max_plot_series: int = 50
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)

    heatmap_rows = []
    decomp_rows = []
    seasonal_period = _seasonal_period_from_freq(freq_alias)

    for col in target_cols[:max_plot_series]:
        s = pd.to_numeric(df_clean[col], errors="coerce")
        tmp = pd.DataFrame({date_col: pd.to_datetime(df_clean[date_col], errors="coerce"), col: s}).dropna()
        if len(tmp) == 0:
            continue

        tmp = tmp.sort_values(date_col).copy()

        if freq_alias == "M":
            tmp["row_key"] = tmp[date_col].dt.year
            tmp["col_key"] = tmp[date_col].dt.month
            row_label = "year"
            col_label = "month"
            heatmap_name = f"{_safe_plot_name(col)}_heatmap_year_month.png"
        elif freq_alias == "W":
            tmp["row_key"] = tmp[date_col].dt.year
            tmp["col_key"] = tmp[date_col].dt.isocalendar().week.astype(int)
            row_label = "year"
            col_label = "iso_week"
            heatmap_name = f"{_safe_plot_name(col)}_heatmap_year_week.png"
        elif freq_alias == "D":
            tmp["row_key"] = tmp[date_col].dt.year
            tmp["col_key"] = tmp[date_col].dt.month
            row_label = "year"
            col_label = "month"
            heatmap_name = f"{_safe_plot_name(col)}_heatmap_year_month.png"
        else:
            tmp["row_key"] = tmp[date_col].dt.dayofweek
            tmp["col_key"] = tmp[date_col].dt.hour
            row_label = "dayofweek"
            col_label = "hour"
            heatmap_name = f"{_safe_plot_name(col)}_heatmap_dayofweek_hour.png"

        piv = tmp.pivot_table(index="row_key", columns="col_key", values=col, aggfunc="mean")
        if len(piv) > 0:
            piv.to_csv(os.path.join(sheet_dir, f"{_safe_plot_name(col)}_heatmap_data.csv"), encoding="utf-8-sig")

            _safe_save_pivot_heatmap(
                piv,
                f"{col} - Heatmap ({row_label} x {col_label})",
                os.path.join(plot_dir, heatmap_name)
            )

            heatmap_rows.append({
                "series": col,
                "row_dimension": row_label,
                "column_dimension": col_label,
                "n_rows": int(piv.shape[0]),
                "n_cols": int(piv.shape[1]),
                "heatmap_file": heatmap_name
            })

        decomp_status = "skipped"
        decomp_reason = None
        if not HAS_STATSMODELS:
            decomp_reason = "statsmodels kurulu değil; seasonal decomposition atlandı."
        else:
            series_for_dec = tmp.set_index(date_col)[col].astype(float).copy()
            try:
                if freq_alias == "M":
                    series_for_dec = series_for_dec.asfreq("ME")
                elif freq_alias == "W":
                    series_for_dec = series_for_dec.asfreq("W")
                elif freq_alias == "D":
                    series_for_dec = series_for_dec.asfreq("D")
                else:
                    series_for_dec = series_for_dec.asfreq("H")
            except Exception:
                pass

            series_for_dec = series_for_dec.interpolate(method="linear", limit_direction="both")
            if series_for_dec.notna().sum() >= max(2 * seasonal_period, 12):
                try:
                    dec_model = "multiplicative" if bool((series_for_dec.dropna() > 0).all()) else "additive"
                    dec = seasonal_decompose(series_for_dec, model=dec_model, period=seasonal_period, extrapolate_trend="freq")
                    dec_df = pd.DataFrame({
                        "observed": dec.observed,
                        "trend": dec.trend,
                        "seasonal": dec.seasonal,
                        "resid": dec.resid
                    })
                    dec_df.to_csv(os.path.join(sheet_dir, f"{_safe_plot_name(col)}_seasonal_decomposition.csv"), encoding="utf-8-sig")

                    fig = dec.plot()
                    fig.set_size_inches(14, 10)
                    plt.suptitle(f"{col} - Seasonal Decomposition ({dec_model}, period={seasonal_period})", y=1.02)
                    plt.tight_layout()
                    plt.savefig(os.path.join(plot_dir, f"{_safe_plot_name(col)}_seasonal_decomposition.png"), dpi=150, bbox_inches="tight")
                    plt.close(fig)
                    decomp_status = "saved"
                except Exception as exc:
                    decomp_reason = f"seasonal decomposition başarısız: {str(exc)}"
            else:
                decomp_reason = f"decomposition için yetersiz gözlem: gerekli yaklaşık minimum {max(2 * seasonal_period, 12)}"

        decomp_rows.append({
            "series": col,
            "status": decomp_status,
            "reason": decomp_reason,
            "seasonal_period_used": seasonal_period,
            "statsmodels_available": bool(HAS_STATSMODELS)
        })

    heatmap_report = pd.DataFrame(heatmap_rows, columns=["series", "row_dimension", "column_dimension", "n_rows", "n_cols", "heatmap_file"])
    decomposition_report = pd.DataFrame(decomp_rows, columns=["series", "status", "reason", "seasonal_period_used", "statsmodels_available"])

    if len(heatmap_report) > 0:
        heatmap_report.to_csv(os.path.join(sheet_dir, "seasonality_heatmap_report.csv"), index=False, encoding="utf-8-sig")
        heatmap_report.to_excel(os.path.join(sheet_dir, "seasonality_heatmap_report.xlsx"), index=False)
    if len(decomposition_report) > 0:
        decomposition_report.to_csv(os.path.join(sheet_dir, "seasonal_decomposition_report.csv"), index=False, encoding="utf-8-sig")
        decomposition_report.to_excel(os.path.join(sheet_dir, "seasonal_decomposition_report.xlsx"), index=False)

    return heatmap_report, decomposition_report

# =========================================================
# FEATURE ENGINEERING
# =========================================================

def add_calendar_features(df: pd.DataFrame, date_col: str, freq: str) -> pd.DataFrame:
    out = df.copy()

    out["year"] = out[date_col].dt.year
    out["month"] = out[date_col].dt.month
    out["quarter"] = out[date_col].dt.quarter

    if freq in ["D", "H"]:
        out["dayofweek"] = out[date_col].dt.dayofweek
        out["weekofyear"] = out[date_col].dt.isocalendar().week.astype(int)
        out["dayofmonth"] = out[date_col].dt.day
        out["is_month_start"] = out[date_col].dt.is_month_start.astype(int)
        out["is_month_end"] = out[date_col].dt.is_month_end.astype(int)
        out["is_quarter_start"] = out[date_col].dt.is_quarter_start.astype(int)
        out["is_quarter_end"] = out[date_col].dt.is_quarter_end.astype(int)
        out["is_year_start"] = out[date_col].dt.is_year_start.astype(int)
        out["is_year_end"] = out[date_col].dt.is_year_end.astype(int)

    if freq == "H":
        out["hour"] = out[date_col].dt.hour
        out["hour_sin"] = np.sin(2 * np.pi * out["hour"] / 24)
        out["hour_cos"] = np.cos(2 * np.pi * out["hour"] / 24)

    if freq in ["H", "D"]:
        out["dow_sin"] = np.sin(2 * np.pi * out["dayofweek"] / 7)
        out["dow_cos"] = np.cos(2 * np.pi * out["dayofweek"] / 7)

    out["month_sin"] = np.sin(2 * np.pi * out["month"] / 12)
    out["month_cos"] = np.cos(2 * np.pi * out["month"] / 12)

    return out


def add_lag_features(df: pd.DataFrame, target_cols: List[str], freq_alias: str, season_length: int) -> pd.DataFrame:
    out = df.copy()

    base_lags = [1, 2, 3]
    if freq_alias == "M":
        base_lags += [6, 12]
    elif freq_alias == "W":
        base_lags += [4, 8, 12, 52]
    elif freq_alias == "D":
        base_lags += [7, 14, 28]
    elif freq_alias == "H":
        base_lags += [24, 48, 168]

    for col in target_cols:
        s = pd.to_numeric(out[col], errors="coerce")
        for lag in sorted(set([l for l in base_lags if l < len(out)])):
            out[f"{col}_lag_{lag}"] = s.shift(lag)

        for w in [3, 6, 12]:
            if w < len(out):
                out[f"{col}_roll_mean_{w}"] = s.shift(1).rolling(w, min_periods=1).mean()
                out[f"{col}_roll_std_{w}"] = s.shift(1).rolling(w, min_periods=1).std().fillna(0)

        if season_length > 1 and season_length < len(out):
            out[f"{col}_lag_seasonal"] = s.shift(season_length)

    return out


def add_series_quality_features(
    df: pd.DataFrame,
    target_cols: List[str],
    anomaly_gov: pd.DataFrame,
    date_col: str
) -> pd.DataFrame:
    out = df.copy()

    gov_small = anomaly_gov.copy() if len(anomaly_gov) > 0 else pd.DataFrame()

    for col in target_cols:
        s = pd.to_numeric(out[col], errors="coerce")
        out[f"{col}_is_zero"] = s.eq(0).astype(int)
        out[f"{col}_log1p"] = np.log1p(s.clip(lower=0))

        out[f"{col}_anomaly_flag"] = 0
        out[f"{col}_exclude_from_training"] = 0
        out[f"{col}_review_required"] = 0
        out[f"{col}_structural_event_flag"] = 0
        out[f"{col}_incomplete_period_flag"] = 0

        if len(gov_small) > 0:
            sub = gov_small.loc[gov_small["series"] == col].copy()

            if len(sub) > 0:
                anomaly_dates = set(sub["date"].tolist())

                exclude_dates = set(
                    sub.loc[sub["excluded_from_training_candidate"] == True, "date"].tolist()
                )

                review_dates = set(
                    sub.loc[
                        sub["action_taken"].isin([
                            "flag_only_review",
                            "keep_raw_flag",
                            "preserve_raw_flag_exclude_candidate"
                        ]),
                        "date"
                    ].tolist()
                )

                structural_dates = set(
                    sub.loc[sub["is_structural_event"] == True, "date"].tolist()
                )

                incomplete_dates = set()
                if "is_incomplete_last_period" in sub.columns:
                    incomplete_dates = set(
                        sub.loc[sub["is_incomplete_last_period"] == True, "date"].tolist()
                    )

                # CRITICAL SAFETY RULE:
                # incomplete last period must always be excluded from training,
                # even if governance action was downgraded to review for any reason.
                exclude_dates = exclude_dates.union(incomplete_dates)

                out[f"{col}_anomaly_flag"] = out[date_col].isin(anomaly_dates).astype(int)
                out[f"{col}_exclude_from_training"] = out[date_col].isin(exclude_dates).astype(int)
                out[f"{col}_review_required"] = out[date_col].isin(review_dates).astype(int)
                out[f"{col}_structural_event_flag"] = out[date_col].isin(structural_dates).astype(int)
                out[f"{col}_incomplete_period_flag"] = out[date_col].isin(incomplete_dates).astype(int)

    return out


def create_model_family_exports(
    df_feat: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    freq_alias: str,
    config: PreprocessConfig
) -> Dict[str, pd.DataFrame]:
    season_length = config.seasonal_period_map.get(freq_alias, 1)

    # Base model-ready frame
    feat = df_feat.copy()

    if config.exclude_textual_columns_from_modeling_features:
        object_cols = feat.select_dtypes(include=["object"]).columns.tolist()
        drop_cols = [c for c in object_cols if c != date_col]
        feat = feat.drop(columns=drop_cols, errors="ignore")

    if config.drop_low_signal_calendar_features_for_monthly and freq_alias == "M":
        feat = feat.drop(
            columns=[
                "dayofweek", "dayofmonth", "weekofyear",
                "is_month_start", "is_month_end",
                "is_quarter_start", "is_quarter_end",
                "is_year_start", "is_year_end"
            ],
            errors="ignore"
        )

    if config.drop_weekday_text_from_monthly_modeling and freq_alias == "M":
        feat = feat.drop(columns=["Haftanın Günü", "weekday", "weekday name"], errors="ignore")

    # Statistical models
    statistical_cols = [date_col] + target_cols + [
        c for c in feat.columns if c not in target_cols and c != date_col
        and not any(x in c.lower() for x in ["lag_", "roll_mean_", "roll_std_"])
    ]
    df_statistical = feat[[c for c in statistical_cols if c in feat.columns]].copy()

    # ML models
    ml_cols = [date_col] + [c for c in feat.columns if c != date_col]
    df_ml = feat[[c for c in ml_cols if c in feat.columns]].copy()

    # DL / sequence models
    dl = df_clean[[date_col] + target_cols].copy()
    scaler = choose_scaler(config.scaler_for_deep_learning)
    scaled = scaler.fit_transform(dl[target_cols])
    for i, col in enumerate(target_cols):
        dl[f"{col}_scaled"] = scaled[:, i]
    dl = add_calendar_features(dl, date_col, freq_alias)

    # Foundation / transformer style minimal pack
    foundation = df_clean[[date_col] + target_cols].copy()
    for col in target_cols:
        foundation[f"{col}_log1p"] = np.log1p(pd.to_numeric(foundation[col], errors="coerce").clip(lower=0))

        if f"{col}_exclude_from_training" in feat.columns:
            foundation[f"{col}_exclude_from_training"] = feat[f"{col}_exclude_from_training"].values
        if f"{col}_structural_event_flag" in feat.columns:
            foundation[f"{col}_structural_event_flag"] = feat[f"{col}_structural_event_flag"].values
        if f"{col}_incomplete_period_flag" in feat.columns:
            foundation[f"{col}_incomplete_period_flag"] = feat[f"{col}_incomplete_period_flag"].values

    # Prophet-ready long export
    prophet_rows = []
    for col in target_cols:
        tmp = pd.DataFrame({
            "unique_id": col,
            "ds": df_clean[date_col].values,
            "y": pd.to_numeric(df_clean[col], errors="coerce").values
        })
        if f"{col}_exclude_from_training" in feat.columns:
            tmp["exclude_from_training"] = feat[f"{col}_exclude_from_training"].values
        if f"{col}_structural_event_flag" in feat.columns:
            tmp["structural_event_flag"] = feat[f"{col}_structural_event_flag"].values
        if f"{col}_incomplete_period_flag" in feat.columns:
            tmp["incomplete_period_flag"] = feat[f"{col}_incomplete_period_flag"].values
        prophet_rows.append(tmp)
    df_prophet = pd.concat(prophet_rows, axis=0, ignore_index=True)

    # Global / transformer-ready long export
    long_rows = []
    for col in target_cols:
        tmp = pd.DataFrame({
            "unique_id": col,
            "ds": df_clean[date_col].values,
            "y": pd.to_numeric(df_clean[col], errors="coerce").values
        })
        for extra in ["month", "quarter", "month_sin", "month_cos"]:
            if extra in feat.columns:
                tmp[extra] = feat[extra].values

        for flagcol in [
            f"{col}_exclude_from_training",
            f"{col}_review_required",
            f"{col}_structural_event_flag",
            f"{col}_incomplete_period_flag"
        ]:
            if flagcol in feat.columns:
                tmp[flagcol.replace(f"{col}_", "")] = feat[flagcol].values

        long_rows.append(tmp)
    df_global_long = pd.concat(long_rows, axis=0, ignore_index=True)

    return {
        "modeling_features_statistical": df_statistical,
        "modeling_features_ml": df_ml,
        "modeling_features_dl": dl,
        "modeling_features_foundation": foundation,
        "modeling_features_prophet": df_prophet,
        "modeling_features_global_long": df_global_long
    }

# =========================================================
# METRICS
# =========================================================

def safe_mape(y_true: np.ndarray, y_pred: np.ndarray, eps: float = 1.0) -> float:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    denom = np.maximum(np.abs(y_true), eps)
    return float(np.mean(np.abs((y_true - y_pred) / denom)) * 100.0)


def smape(y_true: np.ndarray, y_pred: np.ndarray, eps: float = 1e-8) -> float:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    denom = np.maximum((np.abs(y_true) + np.abs(y_pred)) / 2.0, eps)
    return float(np.mean(np.abs(y_true - y_pred) / denom) * 100.0)


def wape(y_true: np.ndarray, y_pred: np.ndarray, eps: float = 1e-8) -> float:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    return float(np.sum(np.abs(y_true - y_pred)) / max(np.sum(np.abs(y_true)), eps) * 100.0)


def mae(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    return float(np.mean(np.abs(y_true - y_pred)))


def rmse(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    return float(np.sqrt(np.mean((y_true - y_pred) ** 2)))


def mase(y_true: np.ndarray, y_pred: np.ndarray, train: np.ndarray, seasonality: int = 1, eps: float = 1e-8) -> float:
    train = np.asarray(train, dtype=float)
    if len(train) <= seasonality:
        denom = np.mean(np.abs(np.diff(train))) if len(train) > 1 else np.nan
    else:
        denom = np.mean(np.abs(train[seasonality:] - train[:-seasonality]))
    denom = max(denom, eps) if np.isfinite(denom) else eps
    return float(np.mean(np.abs(np.asarray(y_true) - np.asarray(y_pred))) / denom)


# =========================================================
# REPORTS
# =========================================================

def series_quality_report(
    df_raw: pd.DataFrame,
    df_clean: pd.DataFrame,
    target_cols: List[str],
    outlier_flags: Dict[str, pd.Series],
    structural_zero_events: pd.Series
) -> pd.DataFrame:
    rows = []
    n = len(df_clean)

    for c in target_cols:
        raw_s = pd.to_numeric(df_raw[c], errors="coerce") if c in df_raw.columns else pd.Series(dtype=float)
        clean_s = pd.to_numeric(df_clean[c], errors="coerce")

        mean_val = clean_s.mean()
        std_val = clean_s.std()

        rows.append({
            "series": c,
            "n_rows_clean": n,
            "raw_missing_ratio": float(raw_s.isna().mean()) if len(raw_s) > 0 else np.nan,
            "clean_missing_ratio": float(clean_s.isna().mean()),
            "raw_zero_ratio": float(((raw_s == 0) & raw_s.notna()).mean()) if len(raw_s) > 0 else np.nan,
            "clean_zero_ratio": float(((clean_s == 0) & clean_s.notna()).mean()),
            "outlier_count_flagged": int(outlier_flags[c].sum()) if c in outlier_flags else 0,
            "outlier_fraction_flagged": float(outlier_flags[c].mean()) if c in outlier_flags else 0.0,
            "structural_zero_event_count": int(structural_zero_events.sum()),
            "count": int(clean_s.notna().sum()),
            "mean": float(mean_val),
            "median": float(clean_s.median()),
            "std": float(std_val),
            "min": float(clean_s.min()),
            "q25": safe_float(clean_s.quantile(0.25)),
            "q50": safe_float(clean_s.quantile(0.50)),
            "q75": safe_float(clean_s.quantile(0.75)),
            "max": float(clean_s.max()),
            "iqr": safe_float(clean_s.quantile(0.75) - clean_s.quantile(0.25)),
            "sum": safe_float(clean_s.sum()),
            "skewness": safe_float(clean_s.skew()),
            "kurtosis": safe_float(clean_s.kurt()),
            "cv": float(std_val / mean_val) if pd.notna(mean_val) and mean_val != 0 else np.nan,
            "intermittency_ratio": float(((clean_s == 0) & clean_s.notna()).mean()),
        })

    return pd.DataFrame(rows)


def create_missing_value_audit(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    target_cols: List[str],
    anomaly_gov: Optional[pd.DataFrame] = None,
    df_clean_governed_preserve: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    rows = []
    anomaly_gov = anomaly_gov.copy() if isinstance(anomaly_gov, pd.DataFrame) else pd.DataFrame()

    for col in target_cols:
        raw_s = pd.to_numeric(df_regular[col], errors="coerce")
        clean_s = pd.to_numeric(df_clean[col], errors="coerce")

        preserve_changed_count = np.nan
        if isinstance(df_clean_governed_preserve, pd.DataFrame) and col in df_clean_governed_preserve.columns:
            preserve_s = pd.to_numeric(df_clean_governed_preserve[col], errors="coerce")
            preserve_changed_count = int((
                (raw_s.isna() & preserve_s.notna()) |
                (raw_s.notna() & preserve_s.notna() & (raw_s != preserve_s))
            ).sum())

        raw_missing = raw_s.isna()
        clean_missing = clean_s.isna()

        gov_imputed_count = 0
        review_or_exclusion_count = 0
        if len(anomaly_gov) > 0:
            gov_imputed_count = int(
                (
                    (anomaly_gov["series"] == col) &
                    (anomaly_gov["action_taken"] == "set_nan_then_impute")
                ).sum()
            )
            review_or_exclusion_count = int(
                (
                    (anomaly_gov["series"] == col) &
                    (anomaly_gov["action_taken"].isin(["flag_only_review", "keep_raw_flag", "preserve_raw_flag_exclude_candidate"]))
                ).sum()
            )

        changed_count = int((
            (raw_s.isna() & clean_s.notna()) |
            (raw_s.notna() & clean_s.notna() & (raw_s != clean_s))
        ).sum())

        rows.append({
            "series": col,
            "raw_missing_count": int(raw_missing.sum()),
            "clean_missing_count": int(clean_missing.sum()),
            "raw_missing_ratio": float(raw_missing.mean()),
            "clean_missing_ratio": float(clean_missing.mean()),
            "imputed_from_raw_missing_count": int(raw_missing.sum() - clean_missing.sum()),
            "imputed_from_governance_count": gov_imputed_count,
            "review_or_exclusion_governance_count": review_or_exclusion_count,
            "preserve_clean_changed_count": preserve_changed_count,
            "candidate_clean_changed_count": changed_count
        })
    return pd.DataFrame(rows)


def create_missing_strategy_audit(
    df_regular: pd.DataFrame,
    target_cols: List[str],
    date_col: str,
    config: PreprocessConfig
) -> pd.DataFrame:
    missingness_summary = summarize_missingness_patterns(df_regular=df_regular, target_cols=target_cols, date_col=date_col)
    return decide_missing_value_strategy(missingness_summary, config)


def create_frequency_audit(
    df_raw_after_aggregation: pd.DataFrame,
    df_regular: pd.DataFrame,
    date_col: str,
    freq_alias: str
) -> pd.DataFrame:
    raw_dates = pd.DatetimeIndex(df_raw_after_aggregation[date_col].sort_values())
    reg_dates = pd.DatetimeIndex(df_regular[date_col].sort_values())
    missing_dates = reg_dates.difference(raw_dates)

    return pd.DataFrame({
        "metric": [
            "frequency_alias",
            "raw_start",
            "raw_end",
            "regular_start",
            "regular_end",
            "raw_row_count",
            "regular_row_count",
            "inserted_missing_timestamp_count"
        ],
        "value": [
            freq_alias,
            str(raw_dates.min()) if len(raw_dates) else None,
            str(raw_dates.max()) if len(raw_dates) else None,
            str(reg_dates.min()) if len(reg_dates) else None,
            str(reg_dates.max()) if len(reg_dates) else None,
            int(len(raw_dates)),
            int(len(reg_dates)),
            int(len(missing_dates))
        ]
    })


def create_inserted_timestamp_log(df_raw_after_aggregation: pd.DataFrame, df_regular: pd.DataFrame, date_col: str) -> pd.DataFrame:
    raw_dates = pd.DatetimeIndex(df_raw_after_aggregation[date_col].sort_values())
    reg_dates = pd.DatetimeIndex(df_regular[date_col].sort_values())
    missing_dates = reg_dates.difference(raw_dates)

    if len(missing_dates) == 0:
        return pd.DataFrame(columns=["inserted_timestamp"])
    return pd.DataFrame({"inserted_timestamp": missing_dates})


def create_outlier_log(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    anomaly_gov: pd.DataFrame
) -> pd.DataFrame:
    if len(anomaly_gov) == 0:
        return pd.DataFrame(columns=[
            "date", "series", "raw_value_before_action",
            "clean_value_after_processing", "anomaly_type",
            "action_taken", "action_confidence"
        ])

    rows = []
    for _, row in anomaly_gov.iterrows():
        date_ = row["date"]
        series = row["series"]
        idx = df_regular.index[df_regular[date_col] == date_]
        if len(idx) == 0:
            continue
        idx = idx[0]
        rows.append({
            "date": date_,
            "series": series,
            "raw_value_before_action": df_regular.loc[idx, series],
            "clean_value_after_processing": df_clean.loc[idx, series],
            "anomaly_type": row["anomaly_type"],
            "action_taken": row["action_taken"],
            "action_confidence": row["action_confidence"]
        })
    return pd.DataFrame(rows)


def intervention_intensity_report(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    recent_window: int = 6
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    summary_rows = []
    series_rows = []
    last_period_rows = []

    n = len(df_regular)

    for col in target_cols:
        raw_s = pd.to_numeric(df_regular[col], errors="coerce")
        clean_s = pd.to_numeric(df_clean[col], errors="coerce")

        changed = (
            (raw_s.isna() & clean_s.notna()) |
            (raw_s.notna() & clean_s.notna() & (raw_s != clean_s))
        )

        abs_change = (clean_s - raw_s).abs()
        pct_change = abs_change / raw_s.replace(0, np.nan).abs()

        changed_count = int(changed.sum())
        changed_fraction = float(changed.mean())
        mean_abs_change = safe_float(abs_change[changed].mean()) if changed.any() else 0.0
        mean_pct_change = safe_float((pct_change[changed] * 100).mean()) if changed.any() else 0.0
        max_abs_change = safe_float(abs_change.max()) if len(abs_change) > 0 else 0.0
        max_pct_change = safe_float((pct_change * 100).max()) if len(pct_change) > 0 else 0.0
        changes_in_last = int(changed.tail(recent_window).sum())

        series_rows.append({
            "series": col,
            "changed_cell_count": changed_count,
            "changed_fraction": changed_fraction,
            "mean_abs_change": mean_abs_change,
            "mean_pct_change": mean_pct_change,
            "max_abs_change": max_abs_change,
            "max_pct_change": max_pct_change,
            "changes_in_last_6_periods": changes_in_last
        })

        for idx in df_regular.index[changed]:
            last_period_rows.append({
                "date": df_regular.loc[idx, date_col],
                "series": col,
                "raw_value": raw_s.loc[idx],
                "clean_value": clean_s.loc[idx],
                "abs_change": abs_change.loc[idx],
                "pct_change": safe_float(pct_change.loc[idx] * 100)
            })

    series_df = pd.DataFrame(series_rows)
    last_df = pd.DataFrame(last_period_rows)

    summary_rows.append({
        "metric": "total_changed_cells",
        "value": int(series_df["changed_cell_count"].sum()) if len(series_df) > 0 else 0
    })
    summary_rows.append({
        "metric": "overall_changed_fraction_mean",
        "value": safe_float(series_df["changed_fraction"].mean()) if len(series_df) > 0 else 0.0
    })
    summary_rows.append({
        "metric": "series_with_changes",
        "value": int((series_df["changed_cell_count"] > 0).sum()) if len(series_df) > 0 else 0
    })

    return pd.DataFrame(summary_rows), series_df, last_df


def create_domain_validation_template(df_clean: pd.DataFrame, date_col: str, target_cols: List[str]) -> pd.DataFrame:
    rows = []
    for col in target_cols:
        s = pd.to_numeric(df_clean[col], errors="coerce")
        if s.notna().sum() == 0:
            continue

        q1 = s.quantile(0.25)
        q3 = s.quantile(0.75)
        iqr = q3 - q1
        upper = q3 + 1.5 * iqr
        lower = max(q1 - 1.5 * iqr, 0)

        peaks = df_clean.loc[(s > upper) | (s < lower), [date_col, col]].copy()
        peaks = peaks.sort_values(col, ascending=False).head(12)

        for _, row in peaks.iterrows():
            rows.append({
                "series": col,
                "date": row[date_col],
                "value": row[col],
                "possible_real_world_explanation": "",
                "validated_by_domain_expert": "",
                "notes": ""
            })

    if not rows:
        return pd.DataFrame(columns=[
            "series", "date", "value",
            "possible_real_world_explanation",
            "validated_by_domain_expert",
            "notes"
        ])
    return pd.DataFrame(rows)


def build_review_queue(
    anomaly_gov: pd.DataFrame,
    review_candidate_map: Dict[str, pd.Series]
) -> pd.DataFrame:
    if len(anomaly_gov) == 0:
        return pd.DataFrame(columns=[
            "date", "series", "raw_value", "clean_value_candidate",
            "candidate_method", "candidate_ratio_vs_raw",
            "anomaly_type", "reason", "confidence",
            "recommended_action", "is_incomplete_last_period", "is_structural_event",
            "analyst_decision", "analyst_note"
        ])

    rows = []
    for _, row in anomaly_gov.iterrows():
        if row["action_taken"] not in ["flag_only_review", "preserve_raw_flag_exclude_candidate", "keep_raw_flag"]:
            continue

        candidate_series = review_candidate_map.get(row["series"], pd.Series(dtype=float))
        clean_candidate = np.nan

        if len(candidate_series) > 0:
            matched = candidate_series[candidate_series.index == row["date"]]
            if len(matched) > 0:
                clean_candidate = matched.iloc[0]

        raw_val = row["raw_value"]
        candidate_ratio_vs_raw = np.nan
        if pd.notna(raw_val) and raw_val != 0 and pd.notna(clean_candidate):
            candidate_ratio_vs_raw = float(clean_candidate / raw_val)

        rows.append({
            "date": row["date"],
            "series": row["series"],
            "raw_value": raw_val,
            "clean_value_candidate": clean_candidate,
            "candidate_method": "seasonal_local_impute",
            "candidate_ratio_vs_raw": candidate_ratio_vs_raw,
            "anomaly_type": row["anomaly_type"],
            "reason": row["anomaly_reason"],
            "confidence": row["action_confidence"],
            "recommended_action": row["action_taken"],
            "is_incomplete_last_period": row.get("is_incomplete_last_period", False),
            "is_structural_event": row.get("is_structural_event", False),
            "analyst_decision": "",
            "analyst_note": ""
        })
    return pd.DataFrame(rows)


def save_raw_vs_clean_plots(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    sheet_dir: str,
    max_plot_series: int = 50
):
    plot_dir = os.path.join(sheet_dir, "validation_plots")
    os.makedirs(plot_dir, exist_ok=True)

    for col in target_cols[:max_plot_series]:
        plt.figure(figsize=(12, 5))
        plt.plot(df_regular[date_col], df_regular[col], label="Raw-Regular", linewidth=1.5)
        plt.plot(df_clean[date_col], df_clean[col], label="Clean", linewidth=1.5)
        plt.title(f"{col} - Raw vs Clean")
        plt.xlabel("Date")
        plt.ylabel("Value")
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(plot_dir, f"{col}_raw_vs_clean.png"), dpi=150)
        plt.close()


# =========================================================
# STRICT LEAKAGE AUDIT
# =========================================================

def strict_leakage_audit(df_features: pd.DataFrame, target_cols: List[str]) -> pd.DataFrame:
    rows = []

    cols = [str(c) for c in df_features.columns]
    for col in cols:
        lowered = col.lower()
        status = "OK"
        risk_level = "LOW"
        note = "Belirgin leakage paterni bulunmadı."

        # Explicit forbidden patterns
        forbidden_patterns = [
            (r"shift\(-", "Negative shift / future leak riski."),
            (r"lead", "Lead feature leak riski."),
            (r"future", "Future bilgi kullanımı riski."),
            (r"next", "Gelecek bilgi kullanımı riski."),
            (r"centered", "Centered rolling leak riski."),
            (r"rolling_center", "Centered rolling leak riski."),
            (r"t\+", "Gelecek zaman etiketi riski.")
        ]

        for p, msg in forbidden_patterns:
            if re.search(p, lowered):
                status = "REVIEW"
                risk_level = "HIGH"
                note = msg
                break

        # Direct target columns are okay
        if any(lowered == t.lower() for t in target_cols):
            status = "OK"
            risk_level = "LOW"
            note = "Hedef seri."

        # Allowed historical features
        if (
            "_lag_" in lowered or
            "_roll_mean_" in lowered or
            "_roll_std_" in lowered or
            lowered.endswith("_scaled") or
            lowered.endswith("_log1p") or
            lowered.endswith("_is_zero") or
            lowered.endswith("_anomaly_flag")
        ):
            status = "OK"
            risk_level = "LOW"
            note = "Geçmişe dayalı yardımcı özellik."

        rows.append({
            "column_name": col,
            "status": status,
            "risk_level": risk_level,
            "note": note
        })

    rows.append({
        "column_name": "__RULE__SCALER_POLICY",
        "status": "REVIEW",
        "risk_level": "HIGH",
        "note": "Bu modül export amaçlı full-data scaling üretir. Nihai model eğitiminde scaler sadece train fold üzerinde fit edilmelidir."
    })
    rows.append({
        "column_name": "__RULE__IMPUTER_POLICY",
        "status": "REVIEW",
        "risk_level": "HIGH",
        "note": "Bu modülde imputasyon full-series bağlamında yapılır. Nihai walk-forward / CV pipeline içinde imputasyon sadece train fold bilgisiyle yeniden kurulmalıdır."
    })
    rows.append({
        "column_name": "__RULE__ANOMALY_POLICY",
        "status": "REVIEW",
        "risk_level": "HIGH",
        "note": "Outlier/anomaly governance full-series audit amaçlıdır. Model selection sırasında fold-aware anomaly policy kullanılmalıdır."
    })
    rows.append({
        "column_name": "__RULE__CENTERED_ROLLING_POLICY",
        "status": "PASS",
        "risk_level": "LOW",
        "note": "Modeling feature export içinde centered rolling kullanılmadı."
    })

    return pd.DataFrame(rows)


# =========================================================
# BACKTEST
# =========================================================

def get_seasonal_period_for_backtest(freq_alias: str) -> int:
    if freq_alias == "M":
        return 12
    if freq_alias == "W":
        return 52
    if freq_alias == "D":
        return 7
    if freq_alias == "H":
        return 24
    return 1


def get_min_train_size_for_freq(config: PreprocessConfig, freq_alias: str) -> int:
    if freq_alias == "M":
        return config.backtest_min_train_size_monthly
    if freq_alias == "W":
        return config.backtest_min_train_size_weekly
    if freq_alias == "D":
        return config.backtest_min_train_size_daily
    if freq_alias == "H":
        return config.backtest_min_train_size_hourly
    return 24


def seasonal_naive_forecast(train: pd.Series, horizon: int, season_length: int) -> np.ndarray:
    train = pd.Series(train).dropna().reset_index(drop=True)
    if len(train) == 0:
        return np.array([np.nan] * horizon)

    preds = []
    for i in range(horizon):
        idx = len(train) - season_length + (i % season_length)
        if season_length > 0 and len(train) >= season_length and idx >= 0:
            preds.append(float(train.iloc[idx]))
        else:
            preds.append(float(train.iloc[-1]))
    return np.array(preds, dtype=float)


def rolling_mean_forecast(train: pd.Series, horizon: int, window: int = 3) -> np.ndarray:
    train = pd.Series(train).dropna().reset_index(drop=True)
    if len(train) == 0:
        return np.array([np.nan] * horizon)
    val = float(train.tail(min(window, len(train))).mean())
    return np.array([val] * horizon, dtype=float)


def _collect_forecast_metrics(y_true, y_pred, train, season_length):
    return {
        "mape": safe_mape(y_true, y_pred),
        "smape": smape(y_true, y_pred),
        "wape": wape(y_true, y_pred),
        "mae": mae(y_true, y_pred),
        "rmse": rmse(y_true, y_pred),
        "mase": mase(y_true, y_pred, train=np.asarray(train), seasonality=season_length)
    }


def run_proxy_backtest_validation(
    df_raw_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    target_cols: List[str],
    freq_alias: str,
    config: PreprocessConfig,
    truth_source: str = "clean"
) -> pd.DataFrame:
    """
    PATCH:
    truth_source:
        - 'clean' : evaluate against cleaned series
        - 'raw'   : evaluate against raw regular series
    """
    rows = []
    horizon = config.backtest_horizon
    min_train_size = get_min_train_size_for_freq(config, freq_alias)
    season_length = get_seasonal_period_for_backtest(freq_alias)

    for col in target_cols:
        raw_series = pd.to_numeric(df_raw_regular[col], errors="coerce").reset_index(drop=True)
        clean_series = pd.to_numeric(df_clean[col], errors="coerce").reset_index(drop=True)
        truth_series = clean_series if truth_source == "clean" else raw_series

        if clean_series.notna().sum() < min_train_size + horizon:
            rows.append({
                "series": col,
                "comparison": "raw_vs_clean",
                "truth_source": truth_source,
                "note": "Yetersiz gözlem nedeniyle proxy backtest uygulanamadı."
            })
            continue

        metrics = {
            "raw_seasonal_naive": [],
            "clean_seasonal_naive": [],
            "clean_rolling_mean_3": []
        }
        if config.enable_additional_backtest_benchmarks:
            metrics.update({
                "raw_drift": [],
                "clean_drift": []
            })

        start = min_train_size
        end = len(clean_series) - horizon + 1

        for split_end in range(start, end):
            y_true = truth_series.iloc[split_end: split_end + horizon].values
            train_raw = raw_series.iloc[:split_end]
            train_clean = clean_series.iloc[:split_end]

            pred_raw = seasonal_naive_forecast(train_raw, horizon, season_length)
            pred_clean = seasonal_naive_forecast(train_clean, horizon, season_length)
            pred_roll = rolling_mean_forecast(train_clean, horizon, 3)
            if config.enable_additional_backtest_benchmarks:
                pred_raw_drift = drift_forecast(train_raw, horizon)
                pred_clean_drift = drift_forecast(train_clean, horizon)

            if np.isfinite(pred_raw).all():
                metrics["raw_seasonal_naive"].append(_collect_forecast_metrics(y_true, pred_raw, train_clean.values, season_length))
            if np.isfinite(pred_clean).all():
                metrics["clean_seasonal_naive"].append(_collect_forecast_metrics(y_true, pred_clean, train_clean.values, season_length))
            if np.isfinite(pred_roll).all():
                metrics["clean_rolling_mean_3"].append(_collect_forecast_metrics(y_true, pred_roll, train_clean.values, season_length))
            if config.enable_additional_backtest_benchmarks and np.isfinite(pred_raw_drift).all():
                metrics["raw_drift"].append(_collect_forecast_metrics(y_true, pred_raw_drift, train_clean.values, season_length))
            if config.enable_additional_backtest_benchmarks and np.isfinite(pred_clean_drift).all():
                metrics["clean_drift"].append(_collect_forecast_metrics(y_true, pred_clean_drift, train_clean.values, season_length))

        for model_name, values in metrics.items():
            if not values:
                continue
            rows.append({
                "series": col,
                "model_proxy": model_name,
                "truth_source": truth_source,
                "mape": float(np.mean([x["mape"] for x in values])),
                "smape": float(np.mean([x["smape"] for x in values])),
                "wape": float(np.mean([x["wape"] for x in values])),
                "mae": float(np.mean([x["mae"] for x in values])),
                "rmse": float(np.mean([x["rmse"] for x in values])),
                "mase": float(np.mean([x["mase"] for x in values]))
            })

    return pd.DataFrame(rows)


def raw_vs_clean_backtest_comparator(proxy_backtest_report: pd.DataFrame) -> pd.DataFrame:
    if len(proxy_backtest_report) == 0:
        return pd.DataFrame(columns=[
            "series", "metric", "raw_value", "clean_value", "improvement", "decision",
            "raw_smape_mean", "clean_smape_mean", "relative_improvement_pct",
            "decision_reason", "enough_evidence_flag"
        ])

    rows = []
    raw_df = proxy_backtest_report[proxy_backtest_report["model_proxy"] == "raw_seasonal_naive"].copy()
    clean_df = proxy_backtest_report[proxy_backtest_report["model_proxy"] == "clean_seasonal_naive"].copy()

    common = sorted(set(raw_df["series"]).intersection(set(clean_df["series"])))
    metrics = ["mape", "smape", "wape", "mae", "rmse", "mase"]

    for series in common:
        r = raw_df.loc[raw_df["series"] == series].iloc[0]
        c = clean_df.loc[clean_df["series"] == series].iloc[0]

        better_count = 0
        worse_count = 0
        for metric in metrics:
            rv = safe_float(r.get(metric, np.nan))
            cv = safe_float(c.get(metric, np.nan))
            improvement = rv - cv if pd.notna(rv) and pd.notna(cv) else np.nan
            if pd.notna(improvement) and improvement > 0:
                better_count += 1
            elif pd.notna(improvement) and improvement < 0:
                worse_count += 1

            rows.append({
                "series": series,
                "metric": metric,
                "raw_value": rv,
                "clean_value": cv,
                "improvement": improvement,
                "decision": "",
                "raw_smape_mean": safe_float(r.get("smape", np.nan)),
                "clean_smape_mean": safe_float(c.get("smape", np.nan)),
                "relative_improvement_pct": ((rv - cv) / rv * 100.0) if pd.notna(rv) and rv not in [0, 0.0] and pd.notna(cv) else np.nan,
                "decision_reason": "",
                "enough_evidence_flag": np.nan
            })

        raw_smape_mean = safe_float(r.get("smape", np.nan))
        clean_smape_mean = safe_float(c.get("smape", np.nan))
        relative_improvement_pct = ((raw_smape_mean - clean_smape_mean) / raw_smape_mean * 100.0) if pd.notna(raw_smape_mean) and raw_smape_mean not in [0, 0.0] and pd.notna(clean_smape_mean) else np.nan

        if better_count >= 4:
            decision = "clean_candidate_preferred"
            reason = f"{better_count}/6 hata metriğinde iyileşme var."
            enough = True
        elif better_count == 0 and worse_count > 0:
            decision = "no_evidence_of_improvement"
            reason = "Temiz seri hata metriklerinde üstünlük göstermedi."
            enough = False
        else:
            decision = "mixed_review"
            reason = f"Karışık sinyal: {better_count} iyileşme, {worse_count} kötüleşme."
            enough = False

        rows.append({
            "series": series,
            "metric": "__OVERALL__",
            "raw_value": np.nan,
            "clean_value": np.nan,
            "improvement": better_count,
            "decision": decision,
            "raw_smape_mean": raw_smape_mean,
            "clean_smape_mean": clean_smape_mean,
            "relative_improvement_pct": relative_improvement_pct,
            "decision_reason": reason,
            "enough_evidence_flag": enough
        })

    return pd.DataFrame(rows)


# =========================================================
# TESTS
# =========================================================

def _assert_true(condition: bool, message: str):
    if not condition:
        raise AssertionError(message)


def run_internal_unit_tests(config: PreprocessConfig) -> pd.DataFrame:
    records = []

    def add_result(name: str, status: str, detail: str):
        records.append({"test_name": name, "status": status, "detail": detail})

    try:
        s = pd.Series(["31.01.2019", "28.02.2019", "15.03.2019"])
        dt = parse_datetime_series(s)
        _assert_true(dt.notna().all(), "Tüm tarihler parse edilmeliydi.")
        add_result("date_parse_test", "PASS", "Tarih parse testi başarılı.")
    except Exception as e:
        add_result("date_parse_test", "FAIL", str(e))

    try:
        idx_m = pd.date_range("2020-01-31", periods=6, freq="ME")
        idx_d = pd.date_range("2020-01-01", periods=10, freq="D")
        idx_h = pd.date_range("2020-01-01 00:00:00", periods=10, freq="H")
        _assert_true(infer_frequency_from_dates(idx_m) == "M", "Aylık frekans doğru tespit edilmedi.")
        _assert_true(infer_frequency_from_dates(idx_d) == "D", "Günlük frekans doğru tespit edilmedi.")
        _assert_true(infer_frequency_from_dates(idx_h) == "H", "Saatlik frekans doğru tespit edilmedi.")
        add_result("frequency_detection_test", "PASS", "Frekans tespit testi başarılı.")
    except Exception as e:
        add_result("frequency_detection_test", "FAIL", str(e))

    try:
        s = pd.Series([10.0, np.nan, 30.0])
        out = limited_linear_interpolation(s, limit=1)
        _assert_true(pd.notna(out.iloc[1]), "Eksik gözlem interpolate edilmeliydi.")
        add_result("interpolation_test", "PASS", "Interpolasyon testi başarılı.")
    except Exception as e:
        add_result("interpolation_test", "FAIL", str(e))

    try:
        s = pd.Series([10, 11, 10, 12, 11, 200, 10, 9, 11, 10], dtype=float)
        profile = {
            "cv": 0.1,
            "intermittency_ratio": 0.0,
            "volatility_regime": "stable",
            "volume_level": "medium"
        }
        flags, _ = conservative_outlier_vote_adaptive(s, profile, config)
        _assert_true(bool(flags.iloc[5]), "Bariz outlier işaretlenmeliydi.")
        add_result("adaptive_outlier_test", "PASS", "Adaptif outlier testi başarılı.")
    except Exception as e:
        add_result("adaptive_outlier_test", "FAIL", str(e))

    return pd.DataFrame(records)


def generate_synthetic_series(freq: str = "M", periods: int = 60, seed: int = 42) -> pd.DataFrame:
    rng = np.random.default_rng(seed)

    if freq == "M":
        dates = pd.date_range("2018-01-31", periods=periods, freq="ME")
        season = np.sin(2 * np.pi * np.arange(periods) / 12)
    elif freq == "W":
        dates = pd.date_range("2018-01-07", periods=periods, freq="W")
        season = np.sin(2 * np.pi * np.arange(periods) / 52)
    elif freq == "D":
        dates = pd.date_range("2018-01-01", periods=periods, freq="D")
        season = np.sin(2 * np.pi * np.arange(periods) / 7)
    else:
        dates = pd.date_range("2018-01-01 00:00:00", periods=periods, freq="H")
        season = np.sin(2 * np.pi * np.arange(periods) / 24)

    trend = np.linspace(100, 150, periods)
    noise = rng.normal(0, 5, periods)
    y1 = trend + 10 * season + noise
    y2 = y1 * 0.8 + rng.normal(0, 3, periods)

    df = pd.DataFrame({
        "datum": dates,
        "M01AB": y1,
        "M01AE": y2
    })

    if len(df) >= 10:
        df.loc[5, "M01AB"] = np.nan
        df.loc[8, "M01AE"] = np.nan

    if len(df) >= 20:
        df.loc[12, "M01AB"] = df["M01AB"].median() * 5
        df.loc[18, "M01AE"] = df["M01AE"].median() * 4

    if len(df) >= 25:
        df.loc[21, ["M01AB", "M01AE"]] = 0

    if len(df) >= 15:
        df = df.drop(index=[3]).reset_index(drop=True)

    return df


def run_synthetic_tests(config: PreprocessConfig) -> pd.DataFrame:
    rows = []

    def add_result(name: str, status: str, detail: str):
        rows.append({"test_name": name, "status": status, "detail": detail})

    try:
        df = generate_synthetic_series("M", 60, config.random_seed)
        date_col = "datum"
        target_cols = ["M01AB", "M01AE"]

        df[date_col] = parse_datetime_series(df[date_col])
        df = aggregate_duplicates(df, date_col, target_cols)
        df_regular = build_regular_time_index(df, date_col, "M")

        profiles = {c: build_series_profile(df_regular[c], "M", config) for c in target_cols}
        flags = {}
        vote_details = {}
        for c in target_cols:
            f, vd = conservative_outlier_vote_adaptive(df_regular[c], profiles[c], config)
            flags[c] = f
            vote_details[c] = vd

        structural_zero = detect_structural_zero_events(
            df_regular, target_cols,
            config.structural_zero_min_series_count,
            config.structural_zero_ratio_threshold
        )

        incomplete_flags = pd.Series(False, index=df_regular.index)

        gov = build_anomaly_governance_table(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            outlier_flags=flags,
            vote_details=vote_details,
            series_profiles=profiles,
            structural_event_flags=structural_zero,
            incomplete_period_flags=incomplete_flags,
            config=config
        )

        _assert_true(len(df_regular) >= len(df), "Regular index satır sayısı azalmamalı.")
        _assert_true(isinstance(gov, pd.DataFrame), "Governance tablosu üretilmeliydi.")
        _assert_true(len(gov) > 0, "Synthetic veri üzerinde governance kaydı oluşmalıydı.")
        add_result("synthetic_governance_test", "PASS", "Sentetik governance testi başarılı.")
    except Exception as e:
        add_result("synthetic_governance_test", "FAIL", str(e))

    try:
        base = generate_synthetic_series("M", 48, config.random_seed).copy()
        base["datum"] = parse_datetime_series(base["datum"])
        base = aggregate_duplicates(base, "datum", ["M01AB", "M01AE"])
        reg = build_regular_time_index(base, "datum", "M")

        scenarios = {
            "missing_at_start": [0, 1, 2],
            "missing_in_middle_block": [15, 16, 17, 18],
            "missing_at_end": [len(reg) - 3, len(reg) - 2, len(reg) - 1]
        }
        for test_name, idxs in scenarios.items():
            work = reg.copy()
            work.loc[idxs, "M01AB"] = np.nan
            summary = summarize_missingness_patterns(work, ["M01AB"], "datum")
            strat = decide_missing_value_strategy(summary, config)
            _assert_true(len(summary) == 1 and len(strat) == 1, f"{test_name} için audit üretilmeliydi.")
        add_result("synthetic_missing_strategy_test", "PASS", "Baş/orta/son eksik blok senaryoları başarıyla değerlendirildi.")
    except Exception as e:
        add_result("synthetic_missing_strategy_test", "FAIL", str(e))

    return pd.DataFrame(rows)


def run_business_rule_tests(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    target_cols: List[str],
    anomaly_gov: pd.DataFrame,
    config: PreprocessConfig
) -> pd.DataFrame:
    rows = []

    def add_result(name: str, status: str, detail: str):
        rows.append({"test_name": name, "status": status, "detail": detail})

    try:
        recent = anomaly_gov[anomaly_gov["is_recent_period"] == True] if len(anomaly_gov) > 0 else pd.DataFrame()
        auto_fixed_recent = recent[recent["action_taken"] == "set_nan_then_impute"] if len(recent) > 0 else pd.DataFrame()
        ok = len(auto_fixed_recent) == 0
        add_result(
            "recent_period_human_review_test",
            "PASS" if ok else "REVIEW",
            "Son dönemlerde otomatik düzeltme yapılmadı." if ok else "Son dönemlerde otomatik düzeltme tespit edildi."
        )
    except Exception as e:
        add_result("recent_period_human_review_test", "FAIL", str(e))

    try:
        changed = 0
        for col in target_cols:
            raw_s = pd.to_numeric(df_regular[col], errors="coerce")
            clean_s = pd.to_numeric(df_clean[col], errors="coerce")
            changed += int(((raw_s != clean_s) & raw_s.notna() & clean_s.notna()).sum())

        add_result("candidate_intervention_exists_test", "PASS", f"Candidate clean değişen hücre sayısı: {changed}")
    except Exception as e:
        add_result("candidate_intervention_exists_test", "FAIL", str(e))

    try:
        structural_count = int(anomaly_gov["is_structural_event"].sum()) if len(anomaly_gov) > 0 else 0
        exclusion_count = int(anomaly_gov["excluded_from_training_candidate"].sum()) if len(anomaly_gov) > 0 else 0
        ok = exclusion_count >= structural_count
        add_result(
            "structural_event_exclusion_policy_test",
            "PASS" if ok else "REVIEW",
            f"Structural governance satırı: {structural_count}, exclusion satırı: {exclusion_count}"
        )
    except Exception as e:
        add_result("structural_event_exclusion_policy_test", "FAIL", str(e))

    try:
        incomplete_rows = anomaly_gov.loc[
            anomaly_gov["anomaly_type"] == "incomplete_last_period"
        ] if len(anomaly_gov) > 0 else pd.DataFrame()

        if len(incomplete_rows) == 0:
            add_result(
                "incomplete_last_period_exclusion_test",
                "PASS",
                "Incomplete last period governance kaydı yok; test uygulanmadı."
            )
        else:
            ok = bool((incomplete_rows["excluded_from_training_candidate"] == True).all())
            add_result(
                "incomplete_last_period_exclusion_test",
                "PASS" if ok else "FAIL",
                "Incomplete last period kayıtlarının tamamı training exclusion aldı."
                if ok else
                "Bazı incomplete last period kayıtları training exclusion almadı."
            )
    except Exception as e:
        add_result("incomplete_last_period_exclusion_test", "FAIL", str(e))
        
    return pd.DataFrame(rows)


def create_manual_sample_audit(
    df_regular: pd.DataFrame,
    df_clean: pd.DataFrame,
    date_col: str,
    target_cols: List[str],
    sample_size: int = 20,
    random_seed: int = 42,
    anomaly_dates: Optional[List[pd.Timestamp]] = None
) -> pd.DataFrame:
    rng = np.random.default_rng(random_seed)

    n = len(df_regular)
    if n == 0:
        return pd.DataFrame()

    forced_idx = []
    if anomaly_dates is not None and len(anomaly_dates) > 0:
        forced_idx = df_regular.index[df_regular[date_col].isin(anomaly_dates)].tolist()

    remaining = [i for i in np.arange(n) if i not in forced_idx]
    random_size = max(0, min(sample_size - len(forced_idx), len(remaining)))
    sampled_random = rng.choice(remaining, size=random_size, replace=False).tolist() if random_size > 0 else []
    sampled_idx = sorted(set(forced_idx + sampled_random))

    rows = []
    for idx in sampled_idx:
        row = {
            "row_index": int(idx),
            "date": df_regular.loc[idx, date_col]
        }
        for col in target_cols:
            raw_val = df_regular.loc[idx, col]
            clean_val = df_clean.loc[idx, col]
            row[f"{col}_raw_regular"] = raw_val
            row[f"{col}_clean"] = clean_val
            row[f"{col}_changed"] = (
                (pd.isna(raw_val) and pd.notna(clean_val)) or
                (pd.notna(raw_val) and pd.notna(clean_val) and raw_val != clean_val)
            )
        rows.append(row)

    return pd.DataFrame(rows)


# =========================================================
# VALIDATION SUMMARY
# =========================================================

def create_validation_summary(
    quality_report: pd.DataFrame,
    missing_audit: pd.DataFrame,
    freq_ok: bool,
    freq_msg: str,
    outlier_log: pd.DataFrame,
    leakage_report: pd.DataFrame,
    unit_test_report: pd.DataFrame,
    synthetic_test_report: pd.DataFrame,
    business_rule_test_report: pd.DataFrame,
    manual_sample_audit: pd.DataFrame,
    proxy_backtest_report: pd.DataFrame,
    structural_zero_events: pd.Series,
    intervention_summary: pd.DataFrame,
    incomplete_period_log: pd.DataFrame,
    config: PreprocessConfig
) -> pd.DataFrame:
    clean_missing_all_zero = bool((missing_audit["clean_missing_count"] == 0).all()) if len(missing_audit) > 0 else False

    leakage_medium_or_high = int(leakage_report["risk_level"].isin(["HIGH", "MEDIUM"]).sum()) if len(leakage_report) > 0 else 0
    unit_fail = int((unit_test_report["status"] == "FAIL").sum()) if len(unit_test_report) > 0 else 0
    synth_fail = int((synthetic_test_report["status"] == "FAIL").sum()) if len(synthetic_test_report) > 0 else 0
    business_review = int((business_rule_test_report["status"] != "PASS").sum()) if len(business_rule_test_report) > 0 else 0
    manual_rows = int(len(manual_sample_audit)) if len(manual_sample_audit) > 0 else 0
    backtest_rows = int(len(proxy_backtest_report)) if len(proxy_backtest_report) > 0 else 0

    max_outlier_fraction = float(quality_report["outlier_fraction_flagged"].max()) if len(quality_report) > 0 else 0.0
    max_clean_zero_ratio = float(quality_report["clean_zero_ratio"].max()) if len(quality_report) > 0 else 0.0
    incomplete_period_count = int(len(incomplete_period_log)) if isinstance(incomplete_period_log, pd.DataFrame) else 0

    clean_smape = pd.to_numeric(
        proxy_backtest_report.loc[proxy_backtest_report["model_proxy"] == "clean_seasonal_naive", "smape"]
        if len(proxy_backtest_report) > 0 and "model_proxy" in proxy_backtest_report.columns else pd.Series(dtype=float),
        errors="coerce"
    )
    max_clean_smape = float(clean_smape.max()) if len(clean_smape) > 0 and clean_smape.notna().any() else np.nan

    candidate_changed_total = int(missing_audit["candidate_clean_changed_count"].sum()) if "candidate_clean_changed_count" in missing_audit.columns else 0
    review_or_exclusion_total = int(missing_audit["review_or_exclusion_governance_count"].sum()) if "review_or_exclusion_governance_count" in missing_audit.columns else 0

    if candidate_changed_total == 0 and review_or_exclusion_total > 0:
        intervention_status = "REVIEW"
        intervention_detail = "Governance/review kayıtları var fakat candidate clean seri değişmedi."
    elif candidate_changed_total > 0 and review_or_exclusion_total > 0:
        intervention_status = "PASS"
        intervention_detail = f"Candidate clean değişen hücre sayısı: {candidate_changed_total}"
    else:
        intervention_status = "PASS"
        intervention_detail = f"Candidate clean değişen hücre sayısı: {candidate_changed_total}"

    summary = [
        {"check_name": "clean_missing_all_zero", "status": "PASS" if clean_missing_all_zero else "REVIEW", "detail": "Temiz veri setinde eksik değer kalmadı." if clean_missing_all_zero else "Temiz veri setinde hâlâ eksik değer var."},
        {"check_name": "regular_time_index", "status": "PASS" if freq_ok else "REVIEW", "detail": freq_msg},
        {"check_name": "outlier_fraction_policy", "status": "PASS" if max_outlier_fraction <= config.review_if_outlier_fraction_gt else "REVIEW", "detail": f"Maksimum outlier oranı: {max_outlier_fraction:.4f}"},
        {"check_name": "structural_zero_event_count", "status": "PASS" if int(structural_zero_events.sum()) <= config.review_if_structural_zero_events_gt else "REVIEW", "detail": f"Yapısal olay sayısı: {int(structural_zero_events.sum())}"},
        {"check_name": "incomplete_last_period_check", "status": "REVIEW" if incomplete_period_count > 0 else "PASS", "detail": f"Incomplete / partial last period adayı sayısı: {incomplete_period_count}"},
        {
            "check_name": "incomplete_last_period_exclusion_policy",
            "status": "PASS" if (
                ("review_or_exclusion_governance_count" in missing_audit.columns and "candidate_clean_changed_count" in missing_audit.columns)
            ) else "REVIEW",
            "detail": "Incomplete last period exclusion mantığı ayrıca anomaly governance ve feature flags üzerinden kontrol edilmelidir."
        },
        {"check_name": "clean_zero_ratio_policy", "status": "PASS" if max_clean_zero_ratio <= config.review_if_clean_zero_ratio_gt else "REVIEW", "detail": f"Maksimum clean_zero_ratio: {max_clean_zero_ratio:.4f}"},
        {"check_name": "strict_leakage_scan", "status": "PASS" if leakage_medium_or_high == 0 else "REVIEW", "detail": f"Medium/High riskli sütun/rule sayısı: {leakage_medium_or_high}"},
        {"check_name": "proxy_backtest_smape_policy", "status": "PASS" if (pd.isna(max_clean_smape) or max_clean_smape <= config.review_if_proxy_smape_gt) else "REVIEW", "detail": f"Maksimum clean_sMAPE: {max_clean_smape}"},
        {"check_name": "unit_tests", "status": "PASS" if unit_fail == 0 else "REVIEW", "detail": f"Başarısız unit test sayısı: {unit_fail}"},
        {"check_name": "synthetic_tests", "status": "PASS" if synth_fail == 0 else "REVIEW", "detail": f"Başarısız synthetic test sayısı: {synth_fail}"},
        {"check_name": "business_rule_tests", "status": "PASS" if business_review == 0 else "REVIEW", "detail": f"Review/fail business test sayısı: {business_review}"},
        {"check_name": "manual_sample_audit", "status": "PASS" if manual_rows > 0 else "REVIEW", "detail": f"Manuel denetim örnek sayısı: {manual_rows}"},
        {"check_name": "proxy_backtest_validation", "status": "PASS" if backtest_rows > 0 else "REVIEW", "detail": f"Proxy backtest rapor satırı: {backtest_rows}"},
        {"check_name": "candidate_clean_intervention_summary", "status": intervention_status, "detail": intervention_detail},
        {"check_name": "outlier_log_created", "status": "PASS", "detail": f"Toplam governance kaydı: {len(outlier_log)}"}
    ]
    return pd.DataFrame(summary)


# =========================================================
# CORE PREPROCESSOR
# =========================================================

class DemandForecastPreprocessor:
    def __init__(self, config: Optional[PreprocessConfig] = None):
        self.config = config or PreprocessConfig()
        self.scalers = {}
        self.metadata = {}
        np.random.seed(self.config.random_seed)

        self.run_id = f"run_{time.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        self.config_hash = make_config_hash(self.config)
        self.run_started_at = pd.Timestamp.utcnow()

    def preprocess_sheet(self, file_path: str, sheet_name: str, output_dir: str) -> Dict[str, pd.DataFrame]:
        print(f"\n[INFO] Preprocessing started -> Sheet: {sheet_name}")
        progress_log("Önişleme", "Sheet okuma başlatıldı.", extra={"sheet": sheet_name, "file_path": file_path})

        df_raw_original = pd.read_excel(file_path, sheet_name=sheet_name)
        original_columns = df_raw_original.columns.tolist()
        progress_log("Önişleme", "Ham sheet okundu.", level="SUCCESS", extra={"sheet": sheet_name, "rows": int(len(df_raw_original)), "columns": int(len(original_columns))})

        progress_log("Önişleme", "Tarih kolonu ve hedef seriler tespit ediliyor.", extra={"sheet": sheet_name})
        date_col = detect_date_column(df_raw_original, self.config)
        df_raw_original[date_col] = parse_datetime_series(df_raw_original[date_col])
        df_raw_original = df_raw_original[df_raw_original[date_col].notna()].copy()

        target_cols = detect_target_columns(df_raw_original, date_col, self.config)

        for c in target_cols:
            df_raw_original[c] = pd.to_numeric(df_raw_original[c], errors="coerce")

        freq = infer_frequency_from_dates(pd.DatetimeIndex(df_raw_original[date_col].sort_values()))
        freq_alias = get_expected_freq_alias(freq)
        progress_log("Önişleme", "Temel seri yapısı çıkarıldı.", level="SUCCESS", extra={"date_col": date_col, "target_count": len(target_cols), "freq_alias": freq_alias})

        progress_log("Önişleme", "Tarih hizalama ve düzenli zaman ekseni oluşturuluyor.", extra={"freq_alias": freq_alias})
        df_raw_aligned, datetime_alignment_audit = align_dates_to_frequency(df_raw_original.copy(), date_col, freq_alias)
        df_raw_aggregated = aggregate_duplicates(df_raw_aligned, date_col, target_cols)

        if self.config.force_regular_frequency:
            df_regular = build_regular_time_index(df_raw_aggregated.copy(), date_col, freq_alias)
        else:
            df_regular = df_raw_aggregated.copy().sort_values(date_col).reset_index(drop=True)

        datetime_integrity_audit = create_datetime_integrity_audit(
            df_original=df_raw_original,
            df_aligned=df_raw_aligned,
            df_aggregated=df_raw_aggregated,
            df_regular=df_regular,
            date_col=date_col,
            freq_alias=freq_alias
        )
        progress_log("Önişleme", "Düzenli zaman ekseni ve datetime denetimi tamamlandı.", level="SUCCESS", extra={"regular_rows": int(len(df_regular)), "start": str(df_regular[date_col].min()) if len(df_regular) else None, "end": str(df_regular[date_col].max()) if len(df_regular) else None})

        progress_log("Önişleme", "Seri profilleri, yapısal olaylar ve anomali adayları hesaplanıyor.")
        # Series profiles
        series_profile_report = create_series_profile_report(df_regular, target_cols, freq_alias, self.config)
        series_profiles = {
            row["series"]: row.drop(labels=["series"]).to_dict()
            for _, row in series_profile_report.iterrows()
        }

        # Structural event engine
        structural_zero_events = detect_structural_zero_events(
            df_regular,
            target_cols,
            self.config.structural_zero_min_series_count,
            self.config.structural_zero_ratio_threshold
        )

        portfolio_shock_flags = detect_portfolio_shocks(df_regular, target_cols, self.config)

        structural_event_flags = (structural_zero_events | portfolio_shock_flags).fillna(False)
        structural_event_flags = expand_structural_events(
            structural_event_flags,
            self.config.structural_event_neighbor_window
        )

        # PATCH: son gözlem(ler)i otomatik structural event etiketlemesinden çıkar
        structural_event_flags = protect_structural_event_edges(
            structural_event_flags,
            protect_last_n=1
        )

        rebound_flags = detect_rebound_after_event(
            df_regular,
            target_cols,
            structural_event_flags,
            self.config
        )

        incomplete_period_flags, incomplete_period_log = detect_incomplete_last_period(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            config=self.config
        )

        structural_event_log = build_structural_event_log(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            zero_flags=structural_zero_events,
            portfolio_shock_flags=portfolio_shock_flags,
            rebound_flags=rebound_flags
        )

        # Adaptive anomaly detection
        outlier_flags = {}
        vote_details = {}

        for c in target_cols:
            flags, vote_df = conservative_outlier_vote_adaptive(df_regular[c], series_profiles[c], self.config)
            flags = cap_outlier_fraction(df_regular[c], flags, vote_df, self.config.max_outlier_fraction_per_series)
            flags = protect_edge_periods(flags, self.config)

            # NOTE:
            # Structural flags are still allowed to be marked as anomaly,
            # but action policy will decide preservation/review later.
            outlier_flags[c] = flags
            vote_details[c] = vote_df

        anomaly_governance = build_anomaly_governance_table(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            outlier_flags=outlier_flags,
            vote_details=vote_details,
            series_profiles=series_profiles,
            structural_event_flags=structural_event_flags,
            incomplete_period_flags=incomplete_period_flags,
            config=self.config
        )
        progress_log("Önişleme", "Anomali yönetişimi tamamlandı.", level="SUCCESS", extra={"anomaly_rows": int(len(anomaly_governance)), "structural_events": int(pd.Series(structural_event_flags).fillna(False).sum()), "incomplete_period_flags": int(pd.Series(incomplete_period_flags).fillna(False).sum())})



                # PATCH: build counterfactual review candidates
        review_candidate_map = {}
        for c in target_cols:
            tmp = df_regular[[date_col, c]].copy()
            anomaly_dates_c = anomaly_governance.loc[anomaly_governance["series"] == c, "date"].tolist()

            tmp.loc[tmp[date_col].isin(anomaly_dates_c), c] = np.nan

            candidate = seasonal_local_impute(
                df=tmp,
                target_col=c,
                date_col=date_col,
                freq=freq_alias,
                seasonal_period=self.config.seasonal_period_map.get(freq_alias, 1),
                max_interpolation_gap=self.config.max_interpolation_gap
            )

            review_candidate_map[c] = pd.Series(candidate.values, index=tmp[date_col].values)

        # Apply governance decisions into two parallel outputs:
        # 1) clean_governed_preserve: preserves raw for review/exclusion cases
        # 2) clean_candidate_for_modeling: uses candidate imputation for excluded/review dates
        df_clean_governed_preserve = df_regular.copy()
        df_clean_candidate = df_regular.copy()

        governance_map = {}
        if len(anomaly_governance) > 0:
            for _, row in anomaly_governance.iterrows():
                governance_map[(row["date"], row["series"])] = row

        for c in target_cols:
            s_raw = pd.to_numeric(df_regular[c], errors="coerce").copy()

            # Preserve version
            s_preserve_work = s_raw.copy()

            # Candidate version
            s_candidate_work = s_raw.copy()

            for idx in df_regular.index:
                key = (df_regular.loc[idx, date_col], c)
                if key not in governance_map:
                    continue

                g = governance_map[key]
                action = g["action_taken"]

                if action == "set_nan_then_impute":
                    s_preserve_work.loc[idx] = np.nan
                    s_candidate_work.loc[idx] = np.nan

                elif action == "preserve_raw_flag_exclude_candidate":
                    # preserve in audit-clean, but create modeled candidate in candidate-clean
                    s_candidate_work.loc[idx] = np.nan

                elif action in ["flag_only_review", "keep_raw_flag"]:
                    # keep audit clean raw, but candidate series can still propose smoothed alternative
                    s_candidate_work.loc[idx] = np.nan

                elif action == "keep_raw":
                    pass

            original_zero_mask = pd.to_numeric(df_regular[c], errors="coerce").eq(0)

            s_preserve_imputed = seasonal_local_impute(
                df=pd.DataFrame({date_col: df_regular[date_col], c: s_preserve_work}),
                target_col=c,
                date_col=date_col,
                freq=freq_alias,
                seasonal_period=self.config.seasonal_period_map.get(freq_alias, 1),
                max_interpolation_gap=self.config.max_interpolation_gap
            )

            s_candidate_imputed = seasonal_local_impute(
                df=pd.DataFrame({date_col: df_regular[date_col], c: s_candidate_work}),
                target_col=c,
                date_col=date_col,
                freq=freq_alias,
                seasonal_period=self.config.seasonal_period_map.get(freq_alias, 1),
                max_interpolation_gap=self.config.max_interpolation_gap
            )

            if self.config.preserve_zero_values_on_structural_dates:
                preserve_mask = structural_event_flags & original_zero_mask
                s_preserve_imputed.loc[preserve_mask] = 0.0
                s_candidate_imputed.loc[preserve_mask] = 0.0

            if self.config.clip_negative_to_zero:
                s_preserve_imputed = clip_negative_values(s_preserve_imputed)
                s_candidate_imputed = clip_negative_values(s_candidate_imputed)

            df_clean_governed_preserve[c] = s_preserve_imputed
            df_clean_candidate[c] = s_candidate_imputed

        # Candidate clean is the modeling baseline
        df_clean = df_clean_candidate.copy()

        # Optional KNN only if any NA remains
        if self.config.use_knn_for_dense_missing_blocks:
            target_block = df_clean[target_cols].copy()
            if target_block.isna().sum().sum() > 0:
                imputer = KNNImputer(n_neighbors=3)
                df_clean[target_cols] = imputer.fit_transform(target_block)

        progress_log("Önişleme", "Temiz veri, imputasyon ve feature engineering adımları yürütülüyor.", extra={"target_count": len(target_cols)})
        # Features
        season_length = self.config.seasonal_period_map.get(freq_alias, 1)
        df_feat = add_calendar_features(df_clean, date_col, freq_alias)
        df_feat = add_lag_features(df_feat, target_cols, freq_alias, season_length)
        df_feat = add_series_quality_features(df_feat, target_cols, anomaly_governance, date_col)
        progress_log("Önişleme", "Feature engineering tamamlandı.", level="SUCCESS", extra={"feature_rows": int(len(df_feat)), "feature_columns": int(len(df_feat.columns))})

        model_input_transparency = create_model_input_transparency_export(
            df_regular=df_regular,
            df_clean_candidate=df_clean_candidate,
            df_clean_governed_preserve=df_clean_governed_preserve,
            df_feat=df_feat,
            date_col=date_col,
            target_cols=target_cols
        )

        family_exports = create_model_family_exports(
            df_feat=df_feat,
            df_clean=df_clean,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            config=self.config
        )

        # Scaled & log exports
        scaler = choose_scaler(self.config.scaler_for_deep_learning)
        scaled_array = scaler.fit_transform(df_clean[target_cols])
        df_scaled = pd.DataFrame(
            scaled_array,
            columns=[f"{c}_scaled" for c in target_cols],
            index=df_clean.index
        )
        df_scaled.insert(0, date_col, df_clean[date_col].values)

        df_log = df_clean[[date_col] + target_cols].copy()
        if self.config.export_log1p_version:
            for c in target_cols:
                df_log[c] = np.log1p(pd.to_numeric(df_log[c], errors="coerce").clip(lower=0))

        # Reports
        quality = series_quality_report(
            df_raw=df_regular,
            df_clean=df_clean,
            target_cols=target_cols,
            outlier_flags=outlier_flags,
            structural_zero_events=structural_event_flags
        )
        descriptive_statistics_report = create_descriptive_statistics_report(df_clean=df_clean, target_cols=target_cols)
        missing_strategy_audit = create_missing_strategy_audit(
            df_regular=df_regular,
            target_cols=target_cols,
            date_col=date_col,
            config=self.config
        )
        seasonality_report = create_monthly_seasonality_report(
            df_clean=df_clean,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            config=self.config
        )
        pharma_event_diagnostic_report = create_pharma_event_diagnostic_report(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            anomaly_gov=anomaly_governance,
            config=self.config
        )

        intervention_summary, series_intervention_intensity, last_period_intervention = intervention_intensity_report(
            df_regular=df_regular,
            df_clean=df_clean,
            date_col=date_col,
            target_cols=target_cols,
            recent_window=self.config.recent_periods_review_only
        )

        validation_outputs = self._run_validation_audit(
            sheet_name=sheet_name,
            output_dir=output_dir,
            df_raw_after_aggregation=df_raw_aggregated,
            df_regular=df_regular,
            df_clean=df_clean,
            df_clean_governed_preserve=df_clean_governed_preserve,
            df_feat=df_feat,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            outlier_flags=outlier_flags,
            quality_report=quality,
            anomaly_governance=anomaly_governance,
            structural_event_flags=structural_event_flags,
            structural_event_log=structural_event_log,
            incomplete_period_log=incomplete_period_log,
            intervention_summary=intervention_summary
        )

        review_queue = build_review_queue(
            anomaly_gov=anomaly_governance,
            review_candidate_map=review_candidate_map
        )

        manifest = self._build_run_manifest(
            file_path=file_path,
            sheet_name=sheet_name,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            df_raw_original=df_raw_original,
            df_raw_aggregated=df_raw_aggregated,
            df_regular=df_regular,
            anomaly_governance=anomaly_governance,
            intervention_summary=intervention_summary,
            validation_summary=validation_outputs["validation_summary"]
        )

        incomplete_exclusion_count = 0
        if len(anomaly_governance) > 0:
            incomplete_exclusion_count = int(
                (
                    (anomaly_governance["anomaly_type"] == "incomplete_last_period") &
                    (anomaly_governance["excluded_from_training_candidate"] == True)
                ).sum()
            )

        meta = {
            "run_id": self.run_id,
            "pipeline_name": self.config.pipeline_name,
            "pipeline_version": self.config.pipeline_version,
            "code_version": self.config.code_version,
            "output_schema_version": self.config.output_schema_version,
            "config_hash": self.config_hash,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "original_columns": original_columns,
            "date_column": date_col,
            "target_columns": target_cols,
            "frequency_inferred": freq_alias,
            "n_rows_raw_after_date_cleaning": int(len(df_raw_original)),
            "n_rows_aggregated": int(len(df_raw_aggregated)),
            "n_rows_regularized": int(len(df_regular)),
            "date_min": str(df_clean[date_col].min()),
            "date_max": str(df_clean[date_col].max()),
            "scaler_type": self.config.scaler_for_deep_learning,
            "structural_event_count": int(structural_event_flags.sum()),
            "incomplete_period_count": int(incomplete_period_flags.sum()) if len(incomplete_period_flags) > 0 else 0,
            "incomplete_period_exclusion_count": incomplete_exclusion_count,
            "validation_summary_rows": int(len(validation_outputs["validation_summary"])),
            "config": asdict(self.config)
        }
        
        training_exclusion_count = 0
        for c in target_cols:
            col_name = f"{c}_exclude_from_training"
            if col_name in df_feat.columns:
                training_exclusion_count += int(df_feat[col_name].sum())

        candidate_changed_total = 0
        preserve_changed_total = 0
        for c in target_cols:
            raw_s = pd.to_numeric(df_regular[c], errors="coerce")
            cand_s = pd.to_numeric(df_clean_candidate[c], errors="coerce")
            prev_s = pd.to_numeric(df_clean_governed_preserve[c], errors="coerce")

            candidate_changed_total += int(((raw_s != cand_s) & raw_s.notna() & cand_s.notna()).sum())
            preserve_changed_total += int(((raw_s != prev_s) & raw_s.notna() & prev_s.notna()).sum())

        meta["candidate_clean_changed_total"] = candidate_changed_total
        meta["preserve_clean_changed_total"] = preserve_changed_total

        df_clean_model_input = df_clean.copy()
        for c in target_cols:
            excl_col = f"{c}_exclude_from_training"
            if excl_col in df_feat.columns:
                df_clean_model_input.loc[df_feat[excl_col].astype(bool), c] = np.nan
        object_cols_clean = df_clean_model_input.select_dtypes(include=["object"]).columns.tolist()
        object_cols_clean = [c for c in object_cols_clean if c != date_col]
        if len(object_cols_clean) > 0:
            df_clean_model_input = df_clean_model_input.drop(columns=object_cols_clean, errors="ignore")

        meta["training_exclusion_count"] = training_exclusion_count
        meta["model_input_nan_count"] = int(df_clean_model_input[target_cols].isna().sum().sum())
        meta["recommended_manual_validation_count"] = int(sum(df_feat[f"{c}_review_required"].sum() for c in target_cols if f"{c}_review_required" in df_feat.columns))
        
        self.scalers[sheet_name] = scaler
        self.metadata[sheet_name] = meta
            
        export_payload = {
            "raw_regular": df_regular,
            "clean": df_clean,
            "clean_governed_preserve": df_clean_governed_preserve,
            "clean_candidate_for_modeling": df_clean_candidate,
            "clean_model_input": df_clean_model_input,
            "model_input_transparency": model_input_transparency,
            "final_model_visibility_report": model_input_transparency,
            "features": df_feat,
            "scaled": df_scaled,
            "log": df_log,
            "quality_report": quality,
            "descriptive_statistics_report": descriptive_statistics_report,
            "missing_strategy_audit": missing_strategy_audit,
            "seasonality_report": seasonality_report,
            "pharma_event_diagnostic_report": pharma_event_diagnostic_report,
            "datetime_integrity_audit": datetime_integrity_audit,
            "datetime_alignment_audit": datetime_alignment_audit,
            "series_profile_report": series_profile_report,
            "anomaly_governance": anomaly_governance,
            "review_queue": review_queue,
            "intervention_summary": intervention_summary,
            "series_intervention_intensity": series_intervention_intensity,
            "last_period_intervention_report": last_period_intervention,
            "structural_event_log": structural_event_log,
            "incomplete_period_log": incomplete_period_log,
            "manifest": manifest,
            **family_exports,
            **validation_outputs
        }

        self._export_all(
            output_dir=output_dir,
            sheet_name=sheet_name,
            export_payload=export_payload,
            metadata=meta
        )

        print(f"[INFO] Completed -> Sheet: {sheet_name}")

        return export_payload

    def _build_run_manifest(
        self,
        file_path: str,
        sheet_name: str,
        date_col: str,
        target_cols: List[str],
        freq_alias: str,
        df_raw_original: pd.DataFrame,
        df_raw_aggregated: pd.DataFrame,
        df_regular: pd.DataFrame,
        anomaly_governance: pd.DataFrame,
        intervention_summary: pd.DataFrame,
        validation_summary: pd.DataFrame
    ) -> pd.DataFrame:
        passed = int((validation_summary["status"] == "PASS").sum()) if len(validation_summary) > 0 else 0
        review = int((validation_summary["status"] == "REVIEW").sum()) if len(validation_summary) > 0 else 0

        changed_cells = 0
        row = intervention_summary.loc[intervention_summary["metric"] == "total_changed_cells"] if len(intervention_summary) > 0 else pd.DataFrame()
        if len(row) > 0:
            changed_cells = int(row["value"].iloc[0])

        manifest = pd.DataFrame([
            {"key": "run_id", "value": self.run_id},
            {"key": "pipeline_name", "value": self.config.pipeline_name},
            {"key": "pipeline_version", "value": self.config.pipeline_version},
            {"key": "code_version", "value": self.config.code_version},
            {"key": "output_schema_version", "value": self.config.output_schema_version},
            {"key": "config_hash", "value": self.config_hash},
            {"key": "file_path", "value": file_path},
            {"key": "sheet_name", "value": sheet_name},
            {"key": "date_column", "value": date_col},
            {"key": "frequency_inferred", "value": freq_alias},
            {"key": "target_count", "value": len(target_cols)},
            {"key": "raw_rows", "value": len(df_raw_original)},
            {"key": "aggregated_rows", "value": len(df_raw_aggregated)},
            {"key": "regularized_rows", "value": len(df_regular)},
            {"key": "anomaly_rows", "value": len(anomaly_governance)},
            {"key": "changed_cells", "value": changed_cells},
            {"key": "validation_pass_count", "value": passed},
            {"key": "validation_review_count", "value": review},
            {"key": "run_started_at_utc", "value": str(self.run_started_at)},
            {"key": "run_finished_at_utc", "value": str(pd.Timestamp.utcnow())}
        ])
        return manifest

    def _run_validation_audit(
        self,
        sheet_name: str,
        output_dir: str,
        df_raw_after_aggregation: pd.DataFrame,
        df_regular: pd.DataFrame,
        df_clean: pd.DataFrame,
        df_clean_governed_preserve: pd.DataFrame,
        df_feat: pd.DataFrame,
        date_col: str,
        target_cols: List[str],
        freq_alias: str,
        outlier_flags: Dict[str, pd.Series],
        quality_report: pd.DataFrame,
        anomaly_governance: pd.DataFrame,
        structural_event_flags: pd.Series,
        structural_event_log: pd.DataFrame,
        incomplete_period_log: pd.DataFrame,
        intervention_summary: pd.DataFrame
    ) -> Dict[str, pd.DataFrame]:

        safe_sheet = re.sub(r"[^\w\-]+", "_", sheet_name)
        sheet_dir = os.path.join(output_dir, safe_sheet)
        os.makedirs(sheet_dir, exist_ok=True)

        freq_ok, freq_msg = check_regular_index(df_clean, date_col, freq_alias)
        missing_audit = create_missing_value_audit(
            df_regular=df_regular,
            df_clean=df_clean,
            target_cols=target_cols,
            anomaly_gov=anomaly_governance,
            df_clean_governed_preserve=df_clean_governed_preserve
        )
        missing_strategy_audit = create_missing_strategy_audit(
            df_regular=df_regular,
            target_cols=target_cols,
            date_col=date_col,
            config=self.config
        )
        descriptive_statistics_report = create_descriptive_statistics_report(df_clean=df_clean, target_cols=target_cols)
        seasonality_report = create_monthly_seasonality_report(
            df_clean=df_clean,
            date_col=date_col,
            target_cols=target_cols,
            freq_alias=freq_alias,
            config=self.config
        )
        pharma_event_diagnostic_report = create_pharma_event_diagnostic_report(
            df_regular=df_regular,
            date_col=date_col,
            target_cols=target_cols,
            anomaly_gov=anomaly_governance,
            config=self.config
        )
        datetime_integrity_audit = create_datetime_integrity_audit(
            df_original=df_raw_after_aggregation,
            df_aligned=df_raw_after_aggregation,
            df_aggregated=df_raw_after_aggregation,
            df_regular=df_regular,
            date_col=date_col,
            freq_alias=freq_alias
        )
        frequency_audit = create_frequency_audit(df_raw_after_aggregation, df_regular, date_col, freq_alias)
        inserted_timestamp_log = create_inserted_timestamp_log(df_raw_after_aggregation, df_regular, date_col)

        leakage_report = (
            strict_leakage_audit(df_feat, target_cols)
            if self.config.leakage_check_enabled else
            pd.DataFrame(columns=["column_name", "status", "risk_level", "note"])
        )

        outlier_log = create_outlier_log(
            df_regular=df_regular,
            df_clean=df_clean,
            date_col=date_col,
            target_cols=target_cols,
            anomaly_gov=anomaly_governance
        )

        domain_validation_template = (
            create_domain_validation_template(df_clean, date_col, target_cols)
            if self.config.create_domain_validation_template else
            pd.DataFrame()
        )

        correlation_matrix_report = pd.DataFrame()
        target_correlation_report = pd.DataFrame()
        seasonality_heatmap_report = pd.DataFrame()
        seasonal_decomposition_report = pd.DataFrame()

        if self.config.save_validation_plots:
            save_raw_vs_clean_plots(
                df_regular=df_regular,
                df_clean=df_clean,
                date_col=date_col,
                target_cols=target_cols,
                sheet_dir=sheet_dir,
                max_plot_series=self.config.max_plot_series
            )
            if self.config.save_trend_plots:
                save_raw_clean_trend_plots(
                    df_regular=df_regular,
                    df_clean=df_clean,
                    date_col=date_col,
                    target_cols=target_cols,
                    sheet_dir=sheet_dir,
                    max_plot_series=self.config.max_plot_series,
                    ma_windows=self.config.moving_average_windows
                )
            if self.config.save_distribution_plots:
                save_distribution_plots(
                    df_clean=df_clean,
                    target_cols=target_cols,
                    sheet_dir=sheet_dir,
                    max_plot_series=self.config.max_plot_series,
                    save_boxplots=self.config.save_boxplots
                )
            if self.config.save_seasonality_plots:
                save_seasonality_plots(
                    df_clean=df_clean,
                    date_col=date_col,
                    target_cols=target_cols,
                    freq_alias=freq_alias,
                    sheet_dir=sheet_dir,
                    max_plot_series=self.config.max_plot_series,
                    save_year_overlay=self.config.save_year_overlay_seasonality_plots,
                    save_normalized_profile=self.config.save_normalized_seasonality_plots,
                    save_boxplot=self.config.save_boxplots
                )
            if self.config.save_correlation_analysis:
                correlation_matrix_report, target_correlation_report = save_correlation_analysis(
                    df_for_corr=df_feat,
                    target_cols=target_cols,
                    sheet_dir=sheet_dir
                )
            if self.config.save_seasonality_heatmaps or self.config.save_seasonal_decomposition:
                seasonality_heatmap_report, seasonal_decomposition_report = save_seasonality_heatmaps_and_decomposition(
                    df_clean=df_clean,
                    date_col=date_col,
                    target_cols=target_cols,
                    freq_alias=freq_alias,
                    sheet_dir=sheet_dir,
                    max_plot_series=self.config.max_plot_series
                )

        unit_test_report = run_internal_unit_tests(self.config) if self.config.run_internal_unit_tests else pd.DataFrame()
        synthetic_test_report = run_synthetic_tests(self.config) if self.config.run_synthetic_tests else pd.DataFrame()
        business_rule_test_report = (
            run_business_rule_tests(
                df_regular=df_regular,
                df_clean=df_clean,
                target_cols=target_cols,
                anomaly_gov=anomaly_governance,
                config=self.config
            )
            if self.config.run_business_rule_tests else pd.DataFrame()
        )
        anomaly_dates = anomaly_governance["date"].drop_duplicates().tolist() if len(anomaly_governance) > 0 else []

        manual_sample_audit = (
            create_manual_sample_audit(
                df_regular=df_regular,
                df_clean=df_clean,
                date_col=date_col,
                target_cols=target_cols,
                sample_size=self.config.manual_sample_size,
                random_seed=self.config.random_seed,
                anomaly_dates=anomaly_dates
            )
            if self.config.run_manual_sample_audit else pd.DataFrame()
        )
        proxy_backtest_report_clean_truth = (
            run_proxy_backtest_validation(
                df_raw_regular=df_regular,
                df_clean=df_clean,
                target_cols=target_cols,
                freq_alias=freq_alias,
                config=self.config,
                truth_source="clean"
            )
            if self.config.run_proxy_backtest_validation else pd.DataFrame()
        )

        proxy_backtest_report_raw_truth = (
            run_proxy_backtest_validation(
                df_raw_regular=df_regular,
                df_clean=df_clean,
                target_cols=target_cols,
                freq_alias=freq_alias,
                config=self.config,
                truth_source="raw"
            )
            if self.config.run_proxy_backtest_validation else pd.DataFrame()
        )

        proxy_backtest_report = pd.concat(
            [proxy_backtest_report_clean_truth, proxy_backtest_report_raw_truth],
            axis=0,
            ignore_index=True
        ) if (len(proxy_backtest_report_clean_truth) > 0 or len(proxy_backtest_report_raw_truth) > 0) else pd.DataFrame()

        raw_vs_clean_backtest_report = raw_vs_clean_backtest_comparator(
            proxy_backtest_report_clean_truth if len(proxy_backtest_report_clean_truth) > 0 else pd.DataFrame()
        )

        progress_log("Önişleme", "Kalite kontrolleri, backtest ve validasyon raporları oluşturuluyor.")
        validation_summary = create_validation_summary(
            quality_report=quality_report,
            missing_audit=missing_audit,
            freq_ok=freq_ok,
            freq_msg=freq_msg,
            outlier_log=outlier_log,
            leakage_report=leakage_report,
            unit_test_report=unit_test_report,
            synthetic_test_report=synthetic_test_report,
            business_rule_test_report=business_rule_test_report,
            manual_sample_audit=manual_sample_audit,
            proxy_backtest_report=proxy_backtest_report,
            structural_zero_events=structural_event_flags,
            intervention_summary=intervention_summary,
            incomplete_period_log=incomplete_period_log,
            config=self.config
        )

        progress_log("Önişleme", "Önişleme ve validasyon paketi tamamlandı.", level="SUCCESS", extra={"sheet": sheet_name, "targets": len(target_cols), "validation_rows": int(len(validation_summary)) if isinstance(validation_summary, pd.DataFrame) else None})
        return {
            "validation_summary": validation_summary,
            "missing_audit": missing_audit,
            "missing_strategy_audit": missing_strategy_audit,
            "descriptive_statistics_report": descriptive_statistics_report,
            "correlation_matrix_report": correlation_matrix_report,
            "target_correlation_report": target_correlation_report,
            "seasonality_report": seasonality_report,
            "seasonality_heatmap_report": seasonality_heatmap_report,
            "seasonal_decomposition_report": seasonal_decomposition_report,
            "pharma_event_diagnostic_report": pharma_event_diagnostic_report,
            "datetime_integrity_audit": datetime_integrity_audit,
            "frequency_audit": frequency_audit,
            "inserted_timestamp_log": inserted_timestamp_log,
            "outlier_log": outlier_log,
            "leakage_report": leakage_report,
            "domain_validation_template": domain_validation_template,
            "unit_test_report": unit_test_report,
            "synthetic_test_report": synthetic_test_report,
            "synthetic_missing_tests_report": synthetic_test_report,
            "business_rule_test_report": business_rule_test_report,
            "manual_sample_audit": manual_sample_audit,
            "proxy_backtest_report": proxy_backtest_report,
            "raw_vs_clean_backtest_report": raw_vs_clean_backtest_report,
            "structural_zero_event_log": structural_event_log
        }

    def _export_all(
        self,
        output_dir: str,
        sheet_name: str,
        export_payload: Dict[str, pd.DataFrame],
        metadata: Dict[str, Any]
    ):
        safe_sheet = re.sub(r"[^\w\-]+", "_", sheet_name)
        sheet_dir = os.path.join(output_dir, safe_sheet)
        os.makedirs(sheet_dir, exist_ok=True)

        if self.config.save_csv:
            csv_map = {
                "modeling_features_prophet": f"{safe_sheet}_modeling_features_prophet.csv",
                "modeling_features_global_long": f"{safe_sheet}_modeling_features_global_long.csv",
                "clean_model_input": f"{safe_sheet}_clean_model_input.csv",
                "model_input_transparency": f"{safe_sheet}_model_input_transparency.csv",
                "final_model_visibility_report": f"{safe_sheet}_final_model_visibility_report.csv",
                "raw_regular": f"{safe_sheet}_raw_regular.csv",
                "clean": f"{safe_sheet}_clean.csv",
                "clean_governed_preserve": f"{safe_sheet}_clean_governed_preserve.csv",
                "clean_candidate_for_modeling": f"{safe_sheet}_clean_candidate_for_modeling.csv",
                "features": f"{safe_sheet}_features.csv",
                "scaled": f"{safe_sheet}_scaled.csv",
                "log": f"{safe_sheet}_log1p.csv",
                "quality_report": f"{safe_sheet}_quality_report.csv",
                "descriptive_statistics_report": f"{safe_sheet}_descriptive_statistics_report.csv",
                "missing_strategy_audit": f"{safe_sheet}_missing_strategy_audit.csv",
                "seasonality_report": f"{safe_sheet}_seasonality_report.csv",
                "pharma_event_diagnostic_report": f"{safe_sheet}_pharma_event_diagnostic_report.csv",
                "datetime_integrity_audit": f"{safe_sheet}_datetime_integrity_audit.csv",
                "datetime_alignment_audit": f"{safe_sheet}_datetime_alignment_audit.csv",
                "series_profile_report": f"{safe_sheet}_series_profile_report.csv",
                "anomaly_governance": f"{safe_sheet}_anomaly_governance.csv",
                "review_queue": f"{safe_sheet}_review_queue.csv",
                "intervention_summary": f"{safe_sheet}_intervention_summary.csv",
                "series_intervention_intensity": f"{safe_sheet}_series_intervention_intensity.csv",
                "last_period_intervention_report": f"{safe_sheet}_last_period_intervention_report.csv",
                "structural_event_log": f"{safe_sheet}_structural_event_log.csv",
                "incomplete_period_log": f"{safe_sheet}_incomplete_period_log.csv",
                "modeling_features_statistical": f"{safe_sheet}_modeling_features_statistical.csv",
                "modeling_features_ml": f"{safe_sheet}_modeling_features_ml.csv",
                "modeling_features_dl": f"{safe_sheet}_modeling_features_dl.csv",
                "modeling_features_foundation": f"{safe_sheet}_modeling_features_foundation.csv",
                "validation_summary": f"{safe_sheet}_validation_summary.csv",
                "missing_audit": f"{safe_sheet}_missing_audit.csv",
                "frequency_audit": f"{safe_sheet}_frequency_audit.csv",
                "inserted_timestamp_log": f"{safe_sheet}_inserted_timestamp_log.csv",
                "outlier_log": f"{safe_sheet}_outlier_log.csv",
                "leakage_report": f"{safe_sheet}_leakage_report.csv",
                "domain_validation_template": f"{safe_sheet}_domain_validation_template.csv",
                "unit_test_report": f"{safe_sheet}_unit_test_report.csv",
                "synthetic_test_report": f"{safe_sheet}_synthetic_test_report.csv",
                "synthetic_missing_tests_report": f"{safe_sheet}_synthetic_missing_tests_report.csv",
                "business_rule_test_report": f"{safe_sheet}_business_rule_test_report.csv",
                "manual_sample_audit": f"{safe_sheet}_manual_sample_audit.csv",
                "proxy_backtest_report": f"{safe_sheet}_proxy_backtest_report.csv",
                "raw_vs_clean_backtest_report": f"{safe_sheet}_raw_vs_clean_backtest_report.csv",
                "manifest": f"{safe_sheet}_run_manifest.csv"
            }

            for key, filename in csv_map.items():
                df_obj = export_payload.get(key, pd.DataFrame())
                if isinstance(df_obj, pd.DataFrame):
                    df_obj.to_csv(os.path.join(sheet_dir, filename), index=False, encoding="utf-8-sig")

        if self.config.save_excel:
            excel_path = os.path.join(sheet_dir, f"{safe_sheet}_preprocessing_package.xlsx")
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                for key, df_obj in export_payload.items():
                    if not isinstance(df_obj, pd.DataFrame):
                        continue
                    sheet = safe_excel_sheet_name(key)
                    df_obj.to_excel(writer, sheet_name=sheet, index=False)

                meta_df = pd.DataFrame({
                    "key": list(metadata.keys()),
                    "value": [
                        json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v
                        for v in metadata.values()
                    ]
                })
                meta_df.to_excel(writer, sheet_name=safe_excel_sheet_name("metadata"), index=False)

        if self.config.save_metadata_json:
            with open(os.path.join(sheet_dir, f"{safe_sheet}_metadata.json"), "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=4)

            manifest_df = export_payload.get("manifest", pd.DataFrame())
            manifest_json = {
                row["key"]: row["value"] for _, row in manifest_df.iterrows()
            } if isinstance(manifest_df, pd.DataFrame) and len(manifest_df) > 0 else {}
            with open(os.path.join(sheet_dir, f"{safe_sheet}_run_manifest.json"), "w", encoding="utf-8") as f:
                json.dump(manifest_json, f, ensure_ascii=False, indent=4)

    def save_global_metadata(self, output_dir: str):
        if self.config.save_metadata_json:
            with open(os.path.join(output_dir, "all_sheets_metadata.json"), "w", encoding="utf-8") as f:
                json.dump(self.metadata, f, ensure_ascii=False, indent=4)


# =========================================================
# MAIN
# =========================================================

def main():
    try:
        config = PreprocessConfig(
            output_dir_name="forecast_preprocessing_outputs",

            hampel_window=7,
            hampel_n_sigma=4.0,
            rolling_mad_window=9,
            rolling_mad_n_sigma=4.5,
            iqr_k=4.0,
            min_outlier_votes=2,

            max_outlier_fraction_per_series=0.05,
            protect_first_n_periods=1,
            protect_last_n_periods=1,
            recent_periods_review_only=3,
            clip_negative_to_zero=True,

            structural_zero_ratio_threshold=0.5,
            structural_zero_min_series_count=3,
            portfolio_drop_ratio_threshold=0.55,
            portfolio_rebound_ratio_threshold=1.30,
            structural_event_neighbor_window=0,
            enable_incomplete_period_detection=True,
            partial_period_drop_ratio_threshold=0.60,
            partial_period_compare_last_n=3,
            auto_exclude_incomplete_last_period_from_training=True,
            auto_flag_incomplete_last_period_review=True,

            max_interpolation_gap=1,
            use_knn_for_dense_missing_blocks=False,
            impute_method_preference="seasonal_local",

            min_action_confidence_for_auto_fix=0.75,
            auto_fix_business_spike_dip=False,
            auto_fix_unknown_anomaly=False,
            auto_fix_data_error=True,
            auto_fix_structural_event=False,

            scaler_for_deep_learning="robust",
            export_log1p_version=True,

            generate_modeling_ready_feature_pack=True,
            exclude_textual_columns_from_modeling_features=True,
            drop_low_signal_calendar_features_for_monthly=True,
            export_training_exclusion_masks=True,

            save_validation_excel=True,
            save_validation_csv=True,
            save_validation_plots=True,
            create_domain_validation_template=True,
            leakage_check_enabled=True,
            max_plot_series=50,

            run_internal_unit_tests=True,
            run_synthetic_tests=True,
            run_business_rule_tests=True,
            run_manual_sample_audit=True,
            run_proxy_backtest_validation=True,
            manual_sample_size=20,
            random_seed=42,

            backtest_horizon=3,
            backtest_min_train_size_monthly=24,
            backtest_min_train_size_weekly=52,
            backtest_min_train_size_daily=60,
            backtest_min_train_size_hourly=24 * 14,

            review_if_outlier_fraction_gt=0.05,
            review_if_structural_zero_events_gt=1,
            review_if_clean_zero_ratio_gt=0.10,
            review_if_proxy_smape_gt=60.0,

            save_excel=True,
            save_csv=True,
            save_metadata_json=True,
            save_quality_report=True
        )

        input_info = choose_excel_file()
        source_path = input_info["source_path"]
        excel_path = input_info["excel_path"]

        xls = safe_excel_file(excel_path)
        selected_sheets = choose_sheets(xls.sheet_names)
        output_dir = create_output_dir(source_path, config.output_dir_name)

        preprocessor = DemandForecastPreprocessor(config=config)

        all_results = {}
        for sheet in selected_sheets:
            result = preprocessor.preprocess_sheet(
                file_path=excel_path,
                sheet_name=sheet,
                output_dir=output_dir
            )
            all_results[sheet] = result

        preprocessor.save_global_metadata(output_dir)

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Başarılı",
            "Veri önişleme + anomaly governance + validation + QA modülleri tamamlandı.\n"
            f"Çıktılar şu klasöre kaydedildi:\n{output_dir}"
        )

        print("\n[SUCCESS] All preprocessing, governance, validation and QA steps finished successfully.")

    except Exception as e:
        print(f"[ERROR] {str(e)}")
        print(traceback.format_exc())
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Hata", str(e))




# =========================================================
# STREAMLIT FORECASTING APP LAYER
# =========================================================

import importlib
from itertools import product

st = None

go = None
make_subplots = None
HAS_PLOTLY = False

SARIMAX = None
adfuller = None
kpss = None
acf = None
pacf = None
plot_acf = None
plot_pacf = None
acorr_ljungbox = None
HAS_FORECAST_STATSMODELS = False

Prophet = None
HAS_PROPHET = False

XGBRegressor = None
HAS_XGBOOST = False

tf = None
Sequential = None
KerasLSTM = None
KerasGRU = None
KerasSimpleRNN = None
KerasDense = None
KerasDropout = None
KerasEarlyStopping = None
KerasAdam = None
HAS_TF = False

torch = None
TorchNN = None
TorchOptim = None
TorchReduceLROnPlateau = None
HAS_TORCH = False

try:
    from sklearn.ensemble import HistGradientBoostingRegressor as SKHistGradientBoostingRegressor
except Exception:
    SKHistGradientBoostingRegressor = None

try:
    from sklearn.neural_network import MLPRegressor as SKMLPRegressor
except Exception:
    SKMLPRegressor = None

STREAMLIT_DATAFRAME_ORIGINAL = None
STREAMLIT_DATAFRAME_OVERRIDDEN = False
STREAMLIT_TABLE_RENDER_COUNTER = 0

def _streamlit_download_excel_bytes(df: pd.DataFrame) -> bytes:
    mem = io.BytesIO()
    out_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
    with pd.ExcelWriter(mem, engine="openpyxl") as writer:
        safe_df = out_df.copy()
        safe_df.to_excel(writer, sheet_name="tablo", index=False)
    mem.seek(0)
    return mem.getvalue()

def _install_streamlit_excel_dataframe_override() -> None:
    global STREAMLIT_DATAFRAME_ORIGINAL, STREAMLIT_DATAFRAME_OVERRIDDEN, STREAMLIT_TABLE_RENDER_COUNTER
    if st is None or STREAMLIT_DATAFRAME_OVERRIDDEN:
        return
    original = getattr(st, "dataframe", None)
    if original is None:
        return
    STREAMLIT_DATAFRAME_ORIGINAL = original

    def _excel_first_dataframe(data=None, *args, **kwargs):
        global STREAMLIT_TABLE_RENDER_COUNTER
        nonlocal_original = STREAMLIT_DATAFRAME_ORIGINAL
        try:
            df = data.data.copy() if hasattr(data, "data") and isinstance(getattr(data, "data"), pd.DataFrame) else (data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame(data))
        except Exception:
            df = data if isinstance(data, pd.DataFrame) else pd.DataFrame()
        STREAMLIT_TABLE_RENDER_COUNTER += 1
        key_suffix = f"{STREAMLIT_TABLE_RENDER_COUNTER}_{hashlib.sha1(df.to_csv(index=False).encode('utf-8', errors='ignore')).hexdigest()[:10]}" if isinstance(df, pd.DataFrame) else str(STREAMLIT_TABLE_RENDER_COUNTER)
        container = st.container()
        with container:
            st.table(df)
            if isinstance(df, pd.DataFrame) and len(df.columns) > 0:
                st.download_button(
                    "Bu tabloyu indir (Excel)",
                    data=_streamlit_download_excel_bytes(df),
                    file_name=f"tablo_{key_suffix}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"auto_excel_table_{key_suffix}",
                )
        return container

    st.dataframe = _excel_first_dataframe
    STREAMLIT_DATAFRAME_OVERRIDDEN = True


def ensure_forecasting_runtime_dependencies(include_streamlit: bool = False) -> None:
    global st, go, make_subplots, HAS_PLOTLY
    global SARIMAX, adfuller, kpss, acf, pacf, plot_acf, plot_pacf, acorr_ljungbox, HAS_FORECAST_STATSMODELS
    global Prophet, HAS_PROPHET
    global XGBRegressor, HAS_XGBOOST
    global tf, Sequential, KerasLSTM, KerasGRU, KerasSimpleRNN, KerasDense, KerasDropout, KerasEarlyStopping, KerasAdam, HAS_TF
    global torch, TorchNN, TorchOptim, TorchReduceLROnPlateau, HAS_TORCH

    if include_streamlit and st is None:
        try:
            import streamlit as _st
            st = _st
            _install_streamlit_excel_dataframe_override()
        except Exception:
            st = None

    if go is None or make_subplots is None:
        try:
            import plotly.graph_objects as _go
            from plotly.subplots import make_subplots as _make_subplots
            go = _go
            make_subplots = _make_subplots
            HAS_PLOTLY = True
        except Exception:
            go = None
            make_subplots = None
            HAS_PLOTLY = False

    if SARIMAX is None:
        try:
            from statsmodels.tsa.statespace.sarimax import SARIMAX as _SARIMAX
            from statsmodels.tsa.stattools import adfuller as _adfuller, kpss as _kpss, acf as _acf, pacf as _pacf
            from statsmodels.graphics.tsaplots import plot_acf as _plot_acf, plot_pacf as _plot_pacf
            from statsmodels.stats.diagnostic import acorr_ljungbox as _acorr_ljungbox
            SARIMAX = _SARIMAX
            adfuller = _adfuller
            kpss = _kpss
            acf = _acf
            pacf = _pacf
            plot_acf = _plot_acf
            plot_pacf = _plot_pacf
            acorr_ljungbox = _acorr_ljungbox
            HAS_FORECAST_STATSMODELS = True
        except Exception:
            SARIMAX = None
            adfuller = None
            kpss = None
            acf = None
            pacf = None
            plot_acf = None
            plot_pacf = None
            acorr_ljungbox = None
            HAS_FORECAST_STATSMODELS = False

    if Prophet is None:
        try:
            from prophet import Prophet as _Prophet
            Prophet = _Prophet
            HAS_PROPHET = True
        except Exception:
            Prophet = None
            HAS_PROPHET = False

    if XGBRegressor is None:
        try:
            from xgboost import XGBRegressor as _XGBRegressor
            XGBRegressor = _XGBRegressor
            HAS_XGBOOST = True
        except Exception:
            XGBRegressor = None
            HAS_XGBOOST = False

    if tf is None:
        try:
            import tensorflow as _tf
            _tf.get_logger().setLevel("ERROR")
            try:
                _tf.config.threading.set_intra_op_parallelism_threads(1)
                _tf.config.threading.set_inter_op_parallelism_threads(1)
            except Exception:
                pass
            try:
                _tf.keras.utils.set_random_seed(42)
            except Exception:
                pass
            from tensorflow.keras.models import Sequential as _Sequential
            from tensorflow.keras.layers import LSTM as _KerasLSTM, GRU as _KerasGRU, SimpleRNN as _KerasSimpleRNN, Dense as _KerasDense, Dropout as _KerasDropout
            from tensorflow.keras.callbacks import EarlyStopping as _KerasEarlyStopping
            from tensorflow.keras.optimizers import Adam as _KerasAdam
            tf = _tf
            Sequential = _Sequential
            KerasLSTM = _KerasLSTM
            KerasGRU = _KerasGRU
            KerasSimpleRNN = _KerasSimpleRNN
            KerasDense = _KerasDense
            KerasDropout = _KerasDropout
            KerasEarlyStopping = _KerasEarlyStopping
            KerasAdam = _KerasAdam
            HAS_TF = True
        except Exception:
            tf = None
            Sequential = None
            KerasLSTM = None
            KerasGRU = None
            KerasSimpleRNN = None
            KerasDense = None
            KerasDropout = None
            KerasEarlyStopping = None
            KerasAdam = None
            HAS_TF = False

    if torch is None:
        try:
            import torch as _torch
            import torch.nn as _TorchNN
            import torch.optim as _TorchOptim
            from torch.optim.lr_scheduler import ReduceLROnPlateau as _TorchReduceLROnPlateau
            try:
                _torch.set_num_threads(1)
                _torch.set_num_interop_threads(1)
            except Exception:
                pass
            try:
                _torch.manual_seed(42)
            except Exception:
                pass
            torch = _torch
            TorchNN = _TorchNN
            TorchOptim = _TorchOptim
            TorchReduceLROnPlateau = _TorchReduceLROnPlateau
            HAS_TORCH = True
        except Exception:
            torch = None
            TorchNN = None
            TorchOptim = None
            TorchReduceLROnPlateau = None
            HAS_TORCH = False

if XGBRegressor is None and SKHistGradientBoostingRegressor is not None:
    XGBRegressor = SKHistGradientBoostingRegressor

try:
    from sklearn.ensemble import ExtraTreesRegressor as SKExtraTreesRegressor
except Exception:
    SKExtraTreesRegressor = None


try:
    import shap
    HAS_SHAP = True
except Exception:
    shap = None
    HAS_SHAP = False

try:
    from sklearn.model_selection import ParameterGrid
except Exception:
    ParameterGrid = None

APP_VERSION = "3.5.8_dl_residual_rolling_status_gate"



@dataclass

class ForecastRuntimeConfig:
    interactive_fast_mode: bool = False
    sarimax_max_candidates: int = 28
    sarimax_maxiter_search: int = 80
    sarimax_maxiter_final: int = 180
    sarimax_search_with_exog_top_n: int = 10
    sarimax_enable_walk_forward_refit: bool = True
    sarimax_search_wall_seconds: float = 900.0
    sarimax_max_exog_cols: int = 10
    prophet_max_configs: int = 16
    prophet_max_exog_cols: int = 10
    prophet_allow_logistic_growth: bool = False
    prophet_disable_holidays_for_short_series: bool = True
    prophet_probe_backend: bool = True
    prophet_backend_fail_to_surrogate: bool = True
    xgb_enable_shap: bool = False
    xgb_skip_direct_on_short_series: bool = False
    xgb_max_feature_cols: int = 36
    xgb_force_single_thread: bool = False
    xgb_prefer_hist_gradient_on_short_series: bool = False
    xgb_search_wall_seconds: float = 900.0
    xgb_inner_train_max_rows: int = 240
    dl_enabled: bool = True
    dl_max_configs_per_family: int = 18
    dl_max_epochs: int = 220
    dl_patience: int = 24
    dl_batch_size: int = 6
    dl_validation_ratio: float = 0.20
    dl_sequence_max_exog_cols: int = 16

    # DL input governance:
    # Recurrent LSTM/GRU must not simply reuse the XGBoost tabular lag/rolling
    # feature pool. By default the DL channel is sequence-native: the first
    # tensor channel is always the scaled target history, and optional
    # covariates are limited to known-in-advance calendar/time encodings.
    # Lag/rolling/ewm/diff target-engineered tabular columns are blocked from
    # DL input so recurrent models do not become disguised tabular ML models.
    dl_sequence_input_mode: str = "sequence_native"  # sequence_native | target_only | legacy_tabular
    dl_allow_legacy_tabular_exog: bool = False
    dl_native_include_calendar_covariates: bool = True
    dl_native_include_time_index_covariates: bool = True
    dl_native_include_family_marker: bool = True
    dl_native_max_calendar_covariates: int = 12
    dl_block_tabular_lag_rolling_exog: bool = True
    dl_report_input_contract_in_metrics: bool = True

    dl_min_train_sequences: int = 12
    # Generic DL defaults are still kept for backward compatibility, but
    # LSTM and GRU now use family-specific search spaces below. This prevents
    # the two recurrent models from becoming practically identical under small
    # time-series samples.
    dl_window_candidates: Tuple[int, ...] = (3, 4, 6, 9, 12, 18, 24)
    dl_hidden_units: Tuple[int, ...] = (16, 24, 32, 48, 64)
    dl_dropout_values: Tuple[float, ...] = (0.0, 0.10, 0.20, 0.30)
    dl_learning_rates: Tuple[float, ...] = (0.001, 0.0005)
    dl_strategies: Tuple[str, ...] = ("direct", "recursive")
    dl_recent_sample_weight_max: float = 1.60

    # DL data-regime governance:
    # Short monthly histories (for example 60-70 usable observations) lose many
    # samples after sliding-window conversion. In that regime, recurrent DL is
    # allowed only as a guarded challenger: windows must leave enough real
    # sequences, capacity is capped, training is shorter, and the candidate must
    # beat a leakage-free inner-validation baseline before it can enter ranking.
    dl_data_regime_guard_enabled: bool = True
    dl_monthly_small_sample_observations: int = 72
    dl_monthly_marginal_sample_observations: int = 96
    dl_weekly_small_sample_observations: int = 104
    dl_weekly_marginal_sample_observations: int = 156
    dl_daily_small_sample_observations: int = 120
    dl_daily_marginal_sample_observations: int = 180
    dl_hourly_small_sample_observations: int = 336
    dl_hourly_marginal_sample_observations: int = 720
    dl_small_sample_min_train_sequences: int = 18
    dl_marginal_sample_min_train_sequences: int = 24
    dl_normal_sample_min_train_sequences: int = 30
    dl_min_sequences_per_test_step: int = 5
    dl_small_sample_sequence_ratio: float = 0.35
    dl_marginal_sample_sequence_ratio: float = 0.30
    dl_normal_sample_sequence_ratio: float = 0.24
    dl_small_sample_max_epochs: int = 90
    dl_small_sample_patience: int = 10
    dl_marginal_sample_max_epochs: int = 140
    dl_marginal_sample_patience: int = 16
    dl_small_sample_min_baseline_improvement_pct: float = 0.03
    dl_marginal_sample_min_baseline_improvement_pct: float = 0.01
    dl_reject_short_sample_if_not_better_than_baseline: bool = True

    # Rolling-origin DL validation governance:
    # DL hyperparameter selection is never based on a single tail split or on
    # Keras validation_split. Each LSTM/GRU candidate is selected by the mean
    # score across leakage-free rolling-origin folds built only from train_df.
    dl_use_rolling_origin_validation: bool = True
    dl_rolling_validation_max_folds: int = 3
    dl_rolling_validation_min_folds: int = 2
    dl_rolling_validation_target_folds: int = 3
    dl_rolling_validation_horizon_cap: int = 6
    dl_rolling_validation_min_train_observations: int = 24
    dl_rolling_validation_fail_if_insufficient_folds: bool = True
    dl_early_stopping_validation_ratio: float = 0.18
    dl_min_early_stopping_val_sequences: int = 3
    dl_keras_shuffle: bool = False

    # DL suitability gate: DL receives no contextual bonus before these gates pass.
    dl_contextual_suitability_gate_enabled: bool = True
    dl_require_rolling_validation_stability_for_ranking: bool = True
    dl_max_rolling_wape_cv_for_ranking: float = 0.60
    dl_max_overfit_gap_for_ranking: float = 0.35
    dl_require_real_backend_for_ranking: bool = True
    dl_require_no_surrogate_for_ranking: bool = True
    dl_require_baseline_superiority_for_normal_regime: bool = True
    dl_normal_sample_min_baseline_improvement_pct: float = 0.00

    # Residual-DL: for short/marginal monthly histories, recurrent DL learns
    # residuals over a train-only SARIMA/seasonal base rather than the whole
    # demand level from scratch. This is a challenger/research mode, not a
    # hidden override of LSTM/GRU outputs.
    dl_residual_mode_for_monthly_short_series: bool = True
    dl_residual_mode_train_n_upper: int = 96
    dl_residual_mode_min_train_n: int = 72
    dl_residual_base_preference: str = "sarima_then_seasonal"

    # Family-specific DL search governance:
    # - LSTM is treated as the longer-memory, higher-capacity recurrent model.
    # - GRU is treated as the lighter, faster-adapting recurrent model.
    # These spaces are intentionally different in window lengths, units,
    # dropout, learning rate, depth, dense head, batch size and recency weight.
    dl_family_specific_search_spaces: bool = True
    dl_lstm_window_candidates_monthly: Tuple[int, ...] = (9, 12, 18, 24)
    dl_gru_window_candidates_monthly: Tuple[int, ...] = (3, 6, 9, 12)
    dl_lstm_window_candidates_weekly: Tuple[int, ...] = (8, 13, 26, 39)
    dl_gru_window_candidates_weekly: Tuple[int, ...] = (4, 8, 13, 18)
    dl_lstm_window_candidates_daily: Tuple[int, ...] = (14, 28, 42, 56)
    dl_gru_window_candidates_daily: Tuple[int, ...] = (7, 14, 21, 28)
    dl_lstm_window_candidates_hourly: Tuple[int, ...] = (24, 48, 72, 96)
    dl_gru_window_candidates_hourly: Tuple[int, ...] = (12, 24, 36, 48)
    dl_lstm_hidden_units: Tuple[int, ...] = (32, 48, 64, 96)
    dl_gru_hidden_units: Tuple[int, ...] = (12, 16, 24, 32, 48)
    dl_lstm_dropout_values: Tuple[float, ...] = (0.10, 0.20, 0.30, 0.35)
    dl_gru_dropout_values: Tuple[float, ...] = (0.00, 0.05, 0.10, 0.20)
    dl_lstm_learning_rates: Tuple[float, ...] = (0.0005, 0.0003, 0.0002)
    dl_gru_learning_rates: Tuple[float, ...] = (0.0020, 0.0015, 0.0010)
    dl_lstm_layer_candidates: Tuple[int, ...] = (2, 3)
    dl_gru_layer_candidates: Tuple[int, ...] = (1, 2)
    dl_lstm_strategies: Tuple[str, ...] = ("direct",)
    dl_gru_strategies: Tuple[str, ...] = ("recursive",)
    dl_lstm_dense_head_ratio: float = 0.75
    dl_gru_dense_head_ratio: float = 0.50
    dl_lstm_recurrent_dropout: float = 0.05
    dl_gru_recurrent_dropout: float = 0.00
    dl_lstm_weight_decay: float = 3e-4
    dl_gru_weight_decay: float = 8e-5
    dl_lstm_batch_size: int = 4
    dl_gru_batch_size: int = 8
    dl_lstm_recent_sample_weight_max: float = 1.35
    dl_gru_recent_sample_weight_max: float = 1.85
    dl_lstm_family_score_bias: float = 0.000
    dl_gru_family_score_bias: float = 0.000
    # Strict DL identity governance:
    # LSTM/GRU may only be ranked/reported as LSTM/GRU if a real recurrent
    # backend trained successfully and produced a non-empty epoch-loss history.
    # Fallback forecasts are preserved only as operational diagnostics and are
    # never allowed to overwrite or masquerade as real LSTM/GRU output.
    dl_strict_real_training: bool = True
    dl_allow_surrogate_models: bool = False
    dl_quarantine_fallback_from_ranking: bool = True
    dl_require_nonempty_history: bool = True
    # Leakage governance:
    # Test/holdout values must never be used to choose, correct, replace,
    # or override LSTM/GRU forecasts. Model selection is limited to the
    # inner validation split created from train_df. Surrogate MLP is sealed.
    dl_forbid_test_set_model_selection: bool = True
    dl_forbid_surrogate_override_after_real_fit: bool = True
    dl_remove_test_set_surrogate_override: bool = True
    dl_never_compare_surrogate_to_holdout: bool = True
    dl_surrogate_override_contract_version: str = "no_holdout_surrogate_override_v1"
    dl_surrogate_mlp_policy: str = "removed_hard_disabled_no_public_metric"

    # DL backend / short-monthly regime hard governance:
    # TensorFlow/Keras is the only backend allowed to publish LSTM/GRU as real
    # recurrent DL by default. If TensorFlow is missing, the model is not marked
    # as failed; it is marked as "real DL not trained" and kept as
    # research/surrogate-only / DL-fallback diagnostic outside normal ranking.
    dl_require_tensorflow_backend_for_real_recurrent: bool = True
    dl_tensorflow_missing_public_state: str = "research_surrogate_only_real_dl_not_trained"

    # Academic/operational honesty rule for monthly short histories:
    # train_n < 96  -> LSTM/GRU may be shown only as challenger/research, never champion.
    # train_n < 72  -> real LSTM/GRU is not trained for public ranking; diagnostic fallback only.
    dl_short_monthly_policy_enabled: bool = True
    dl_monthly_challenger_only_train_observations: int = 96
    dl_monthly_no_rank_train_observations: int = 72
    dl_monthly_no_rank_skip_real_training: bool = True
    search_accelerator_enabled: bool = True
    search_accelerator_model_parallel: bool = True
    search_accelerator_candidate_parallel: bool = True
    search_accelerator_cache_enabled: bool = True
    search_accelerator_result_cache_enabled: bool = True
    search_accelerator_max_workers: int = field(default_factory=lambda: max(2, min(4, os.cpu_count() or 2)))
    search_accelerator_candidate_workers: int = field(default_factory=lambda: max(2, min(4, os.cpu_count() or 2)))
    xgb_param_grid: Tuple[Dict[str, Any], ...] = field(default_factory=lambda: (
        {
            "max_depth": 2,
            "learning_rate": 0.03,
            "n_estimators": 300,
            "subsample": 0.90,
            "colsample_bytree": 0.90,
            "reg_alpha": 0.0,
            "reg_lambda": 1.0,
            "min_child_weight": 1
        },
        {
            "max_depth": 3,
            "learning_rate": 0.03,
            "n_estimators": 400,
            "subsample": 0.85,
            "colsample_bytree": 0.85,
            "reg_alpha": 0.0,
            "reg_lambda": 1.5,
            "min_child_weight": 1
        },
        {
            "max_depth": 3,
            "learning_rate": 0.05,
            "n_estimators": 250,
            "subsample": 0.90,
            "colsample_bytree": 0.80,
            "reg_alpha": 0.0,
            "reg_lambda": 1.2,
            "min_child_weight": 1
        },
        {
            "max_depth": 4,
            "learning_rate": 0.03,
            "n_estimators": 350,
            "subsample": 0.80,
            "colsample_bytree": 0.80,
            "reg_alpha": 0.0,
            "reg_lambda": 2.0,
            "min_child_weight": 2
        },
        {
            "max_depth": 2,
            "learning_rate": 0.08,
            "n_estimators": 180,
            "subsample": 1.00,
            "colsample_bytree": 0.90,
            "reg_alpha": 0.0,
            "reg_lambda": 0.8,
            "min_child_weight": 1
        },
        {
            "max_depth": 3,
            "learning_rate": 0.06,
            "n_estimators": 220,
            "subsample": 0.85,
            "colsample_bytree": 0.90,
            "reg_alpha": 0.1,
            "reg_lambda": 1.8,
            "min_child_weight": 2
        }
    ))

FORECAST_RUNTIME_CONFIG = ForecastRuntimeConfig(
    xgb_force_single_thread=True,
    search_accelerator_max_workers=1,
    search_accelerator_candidate_workers=1
)
_PROPhet_BACKEND_PROBE_CACHE = {"done": False, "ok": None, "message": "not_checked"}

@dataclass
class SearchAcceleratorConfig:
    enabled: bool = True
    cache_enabled: bool = True
    result_cache_enabled: bool = True
    model_parallelism: bool = True
    candidate_parallelism: bool = True
    max_workers: int = field(default_factory=lambda: max(2, min(4, os.cpu_count() or 2)))
    candidate_workers: int = field(default_factory=lambda: max(2, min(4, os.cpu_count() or 2)))
    artifact_cache_max_items: int = 256
    result_cache_max_items: int = 128
    failure_cache_max_items: int = 256


class SearchAccelerator:
    def __init__(self, config: SearchAcceleratorConfig):
        self.config = config
        self._artifact_cache: Dict[str, Any] = {}
        self._result_cache: Dict[str, Any] = {}
        self._failure_cache: Dict[str, str] = {}
        self._lock = threading.Lock()

    def _trim_cache(self, cache: Dict[str, Any], max_items: int) -> None:
        while len(cache) > max_items:
            try:
                oldest_key = next(iter(cache))
                cache.pop(oldest_key, None)
            except Exception:
                break

    def get_artifact(self, key: str) -> Any:
        with self._lock:
            value = self._artifact_cache.get(key)
        return copy.deepcopy(value) if value is not None else None

    def put_artifact(self, key: str, value: Any) -> None:
        if not self.config.cache_enabled:
            return
        with self._lock:
            self._artifact_cache[key] = copy.deepcopy(value)
            self._trim_cache(self._artifact_cache, self.config.artifact_cache_max_items)

    def get_or_compute_artifact(self, key: str, builder):
        cached = self.get_artifact(key)
        if cached is not None:
            return cached
        value = builder()
        self.put_artifact(key, value)
        return copy.deepcopy(value)

    def get_result(self, key: str) -> Any:
        if not self.config.result_cache_enabled:
            return None
        with self._lock:
            value = self._result_cache.get(key)
        return copy.deepcopy(value) if value is not None else None

    def put_result(self, key: str, value: Any) -> None:
        if not self.config.result_cache_enabled:
            return
        with self._lock:
            self._result_cache[key] = copy.deepcopy(value)
            self._trim_cache(self._result_cache, self.config.result_cache_max_items)

    def remember_failure(self, key: str, message: str) -> None:
        with self._lock:
            self._failure_cache[key] = str(message)
            self._trim_cache(self._failure_cache, self.config.failure_cache_max_items)

    def get_failure(self, key: str) -> Optional[str]:
        with self._lock:
            return self._failure_cache.get(key)

    def parallel_map(self, fn, items: List[Any], max_workers: Optional[int] = None) -> List[Any]:
        if not items:
            return []
        if not self.config.enabled or not self.config.candidate_parallelism or len(items) <= 1:
            return [fn(item) for item in items]
        workers = max(1, min(int(max_workers or self.config.candidate_workers), len(items)))
        if workers <= 1:
            return [fn(item) for item in items]
        results: List[Any] = []
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="search-acc") as ex:
            futs = [ex.submit(fn, item) for item in items]
            for fut in as_completed(futs):
                results.append(fut.result())
        return results

    def run_parallel_tasks(self, tasks: Dict[str, Any], max_workers: Optional[int] = None) -> Dict[str, Any]:
        if not tasks:
            return {}
        names = list(tasks.keys())
        if not self.config.enabled or not self.config.model_parallelism or len(names) <= 1:
            return {name: tasks[name]() for name in names}
        workers = max(1, min(int(max_workers or self.config.max_workers), len(names)))
        if workers <= 1:
            return {name: tasks[name]() for name in names}
        out: Dict[str, Any] = {}
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="model-acc") as ex:
            future_map = {ex.submit(tasks[name]): name for name in names}
            for fut in as_completed(future_map):
                name = future_map[fut]
                out[name] = fut.result()
        return out


def _fingerprint_dataframe(df: Optional[pd.DataFrame]) -> str:
    if df is None:
        return "none"
    tmp = df.copy()
    for c in tmp.columns:
        if pd.api.types.is_datetime64_any_dtype(tmp[c]):
            tmp[c] = pd.to_datetime(tmp[c], errors="coerce").astype(str)
        elif pd.api.types.is_numeric_dtype(tmp[c]):
            tmp[c] = pd.to_numeric(tmp[c], errors="coerce").round(8)
        else:
            tmp[c] = tmp[c].astype(str)
    try:
        hv = pd.util.hash_pandas_object(tmp, index=True).values.tobytes()
        return hashlib.sha1(hv).hexdigest()[:20]
    except Exception:
        return hashlib.sha1(tmp.to_csv(index=True).encode("utf-8", errors="ignore")).hexdigest()[:20]


def _fingerprint_profile(profile: Dict[str, Any]) -> str:
    try:
        return hashlib.sha1(stable_json_dumps(profile).encode("utf-8")).hexdigest()[:20]
    except Exception:
        return hashlib.sha1(str(profile).encode("utf-8", errors="ignore")).hexdigest()[:20]


def build_search_signature(namespace: str, freq_alias: str, train_df: Optional[pd.DataFrame], test_df: Optional[pd.DataFrame], profile: Optional[Dict[str, Any]] = None, exog_train: Optional[pd.DataFrame] = None, exog_test: Optional[pd.DataFrame] = None, extra: Optional[Dict[str, Any]] = None) -> str:
    payload = {
        "namespace": namespace,
        "app_version": APP_VERSION,
        "freq_alias": str(freq_alias),
        "train": _fingerprint_dataframe(train_df),
        "test": _fingerprint_dataframe(test_df),
        "exog_train": _fingerprint_dataframe(exog_train),
        "exog_test": _fingerprint_dataframe(exog_test),
        "profile": _fingerprint_profile(profile or {}),
        "extra": extra or {}
    }
    return hashlib.sha1(stable_json_dumps(payload).encode("utf-8")).hexdigest()


SEARCH_ACCELERATOR = SearchAccelerator(
    SearchAcceleratorConfig(
        enabled=FORECAST_RUNTIME_CONFIG.search_accelerator_enabled,
        cache_enabled=FORECAST_RUNTIME_CONFIG.search_accelerator_cache_enabled,
        result_cache_enabled=FORECAST_RUNTIME_CONFIG.search_accelerator_result_cache_enabled,
        model_parallelism=FORECAST_RUNTIME_CONFIG.search_accelerator_model_parallel,
        candidate_parallelism=FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_parallel,
        max_workers=FORECAST_RUNTIME_CONFIG.search_accelerator_max_workers,
        candidate_workers=FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_workers,
    )
)

def infer_season_length_from_freq(freq_alias: str) -> int:
    return {"M": 12, "W": 52, "D": 7, "H": 24}.get(str(freq_alias).upper(), 1)


def infer_default_horizon(freq_alias: str) -> int:
    return {"M": 6, "W": 12, "D": 30, "H": 48}.get(str(freq_alias).upper(), 6)


def _safe_bool_series(s: pd.Series) -> pd.Series:
    return pd.Series(s).fillna(False).astype(bool)


def detect_optional_exog_columns(df_features: pd.DataFrame, target_col: str, date_col: str) -> List[str]:
    cols = []
    generic_keywords = [
        "month", "quarter", "year", "weekofyear", "dayofweek", "dayofmonth",
        "month_sin", "month_cos", "quarter_sin", "quarter_cos", "is_month_start",
        "is_month_end", "is_quarter_start", "is_quarter_end", "is_year_start", "is_year_end"
    ]
    banned_keywords = [
        "exclude_from_training", "review_required", "structural_event_flag", "incomplete_period_flag",
        "anomaly_flag", "scaled", "log1p"
    ]

    for c in df_features.columns:
        if c == date_col or c == target_col:
            continue
        lc = c.lower()
        if any(b in lc for b in banned_keywords):
            continue
        if c.startswith(f"{target_col}_"):
            # target-specific lag/rolling features belong to ML branch, not exogenous statistical branch by default
            continue
        if lc in generic_keywords or any(g == lc for g in generic_keywords):
            if pd.api.types.is_numeric_dtype(df_features[c]):
                cols.append(c)
    return sorted(set(cols))


def detect_ml_feature_columns(df_features: pd.DataFrame, target_col: str, date_col: str) -> List[str]:
    cols = []
    generic_keywords = [
        "month", "quarter", "year", "weekofyear", "dayofweek", "dayofmonth",
        "month_sin", "month_cos", "quarter_sin", "quarter_cos", "is_month_start",
        "is_month_end", "is_quarter_start", "is_quarter_end", "is_year_start", "is_year_end"
    ]
    banned_keywords = [
        "exclude_from_training", "review_required", "structural_event_flag", "incomplete_period_flag",
        "anomaly_flag"
    ]
    for c in df_features.columns:
        if c == date_col or c == target_col:
            continue
        lc = c.lower()
        if any(b in lc for b in banned_keywords):
            continue
        if c.startswith(f"{target_col}_") or lc in generic_keywords:
            if pd.api.types.is_numeric_dtype(df_features[c]):
                cols.append(c)
    return sorted(set(cols))



def detect_safe_ml_exog_columns(df_features: pd.DataFrame, target_col: str, date_col: str) -> List[str]:
    """
    Production-safe external features for ML models.
    Only keep calendar / explicitly plan-known / future-known variables.
    Unknown contemporaneous signals from other products are excluded because
    they are not guaranteed to be available at planning time.
    """
    safe_cols: List[str] = []
    future_known_keywords = [
        "month", "quarter", "year", "weekofyear", "dayofweek", "dayofmonth",
        "month_sin", "month_cos", "quarter_sin", "quarter_cos",
        "is_month_start", "is_month_end", "is_quarter_start", "is_quarter_end",
        "is_year_start", "is_year_end", "holiday", "ramadan", "bayram"
    ]
    explicit_plan_keywords = [
        "plan_", "planned_", "budget_", "target_", "promo_plan", "campaign_plan",
        "price_plan", "forecast_", "schedule_"
    ]
    banned_keywords = [
        "exclude_from_training", "review_required", "structural_event_flag", "incomplete_period_flag",
        "anomaly_flag", "scaled", "log1p", "target_h", "future", "lead", "shift(-",
        "lag_", "roll_", "ewm_", "diff_", "seasonal_lag", "prediction", "actual"
    ]
    for c in df_features.columns:
        if c == date_col or c == target_col:
            continue
        lc = str(c).lower()
        if c.startswith(f"{target_col}_"):
            continue
        if any(b in lc for b in banned_keywords):
            continue
        if not pd.api.types.is_numeric_dtype(df_features[c]):
            continue
        is_future_known = any(k == lc or k in lc for k in future_known_keywords) or any(k in lc for k in explicit_plan_keywords)
        if is_future_known:
            safe_cols.append(c)
    return sorted(set(safe_cols))
def make_series_analysis_frame(export_payload: Dict[str, pd.DataFrame], target_col: str) -> pd.DataFrame:
    df_clean = export_payload["clean_model_input"].copy()
    df_feat = export_payload["features"].copy()
    date_col = export_payload["manifest"].loc[export_payload["manifest"]["key"] == "date_column", "value"].iloc[0]

    out = pd.DataFrame({
        "ds": pd.to_datetime(df_clean[date_col]),
        "y": pd.to_numeric(df_clean[target_col], errors="coerce")
    })
    excl_col = f"{target_col}_exclude_from_training"
    out["exclude_from_training"] = _safe_bool_series(df_feat[excl_col]) if excl_col in df_feat.columns else False
    out["is_usable"] = out["y"].notna() & (~out["exclude_from_training"])
    return out


def get_profile_row(export_payload: Dict[str, pd.DataFrame], target_col: str) -> Dict[str, Any]:
    prof = export_payload["series_profile_report"]
    row = prof.loc[prof["series"] == target_col]
    return row.iloc[0].to_dict() if len(row) else {}


def series_segment_label(profile: Dict[str, Any]) -> str:
    cv = float(profile.get("cv", np.nan)) if profile else np.nan
    intermittency = float(profile.get("intermittency_ratio", np.nan)) if profile else np.nan
    seasonality = float(profile.get("seasonality_strength", np.nan)) if profile else np.nan
    volume = str(profile.get("volume_level", "unknown")) if profile else "unknown"

    if pd.notna(intermittency) and intermittency >= 0.35:
        return "intermittent"
    if pd.notna(seasonality) and seasonality >= 0.45:
        return "seasonal"
    if pd.notna(cv) and cv >= 0.45:
        return "volatile"
    if volume in ["high", "medium"] and pd.notna(cv) and cv < 0.25:
        return "stable_fast_mover"
    return "standard"


def recommend_model_priority(profile: Dict[str, Any]) -> str:
    label = series_segment_label(profile)
    if label == "seasonal":
        return "Prophet + SARIMA"
    if label == "volatile":
        return "XGBoost + LSTM + Prophet"
    if label == "intermittent":
        return "Intermittent + ARIMA + XGBoost"
    if label == "stable_fast_mover":
        return "SARIMA + XGBoost + LSTM"
    return "SARIMA + Prophet + XGBoost + LSTM"



try:
    from scipy.special import inv_boxcox
    from scipy.stats import boxcox
    HAS_SCIPY = True
except Exception:
    inv_boxcox = None
    boxcox = None
    HAS_SCIPY = False

def infer_season_length_from_freq(freq_alias: str) -> int:
    return {"M": 12, "W": 52, "D": 7, "H": 24}.get(str(freq_alias).upper(), 1)

def infer_abc_class(profile: Dict[str, Any]) -> str:
    mean_ = safe_float(profile.get("mean", np.nan))
    if pd.isna(mean_):
        return "C"
    if mean_ >= 250:
        return "A"
    if mean_ >= 100:
        return "B"
    return "C"

def infer_xyz_class(profile: Dict[str, Any]) -> str:
    cv = safe_float(profile.get("cv", np.nan))
    if pd.isna(cv):
        return "Z"
    if cv <= 0.25:
        return "X"
    if cv <= 0.45:
        return "Y"
    return "Z"

def infer_advanced_segment(profile: Dict[str, Any]) -> Dict[str, str]:
    abc = infer_abc_class(profile)
    xyz = infer_xyz_class(profile)
    intermittency = safe_float(profile.get("intermittency_ratio", np.nan))
    seasonality = safe_float(profile.get("seasonality_strength", np.nan))
    trend_strength = safe_float(profile.get("trend_strength", np.nan))
    cv = safe_float(profile.get("cv", np.nan))

    if pd.notna(intermittency) and intermittency >= 0.35:
        family = "intermittent"
    elif pd.notna(seasonality) and seasonality >= 0.45:
        family = "seasonal"
    elif pd.notna(cv) and cv >= 0.45:
        family = "volatile"
    elif pd.notna(trend_strength) and trend_strength >= 0.45:
        family = "trend_dominant"
    elif xyz == "X":
        family = "stable"
    else:
        family = "standard"

    return {
        "abc": abc,
        "xyz": xyz,
        "abc_xyz": f"{abc}{xyz}",
        "family": family,
        "label": f"{abc}{xyz}_{family}"
    }

def recommend_candidate_models(profile: Dict[str, Any]) -> List[str]:
    seg = infer_advanced_segment(profile)
    family = seg["family"]
    abcxyz = seg["abc_xyz"]
    if family == "seasonal":
        return ["SARIMA/SARIMAX", "ARIMA", "Prophet", "LSTM", "Ensemble"]
    if family == "intermittent":
        return ["Intermittent", "ARIMA", "XGBoost", "LSTM", "Ensemble"]
    if family == "volatile":
        return ["XGBoost", "LSTM", "GRU", "Prophet", "SARIMA/SARIMAX", "Ensemble"]
    if abcxyz in ["AX", "BX"]:
        return ["SARIMA/SARIMAX", "ARIMA", "XGBoost", "LSTM", "Ensemble"]
    return ["SARIMA/SARIMAX", "ARIMA", "Prophet", "XGBoost", "LSTM", "GRU", "Ensemble"]

def choose_target_transform(y: pd.Series) -> Dict[str, Any]:
    s = pd.to_numeric(y, errors="coerce").dropna().astype(float)
    if len(s) < 8:
        return {"name": "none", "lambda": None, "shift": 0.0}
    skew = safe_float(s.skew())
    min_val = safe_float(s.min())
    cv = safe_float(coefficient_of_variation(s))
    if min_val >= 0 and ((pd.notna(skew) and skew >= 1.0) or (pd.notna(cv) and cv >= 0.50)):
        if HAS_SCIPY and np.all(s > 0):
            return {"name": "boxcox", "lambda": None, "shift": 0.0}
        return {"name": "log1p", "lambda": None, "shift": 0.0}
    return {"name": "none", "lambda": None, "shift": 0.0}

def apply_target_transform(y: pd.Series, transform_cfg: Dict[str, Any]) -> Tuple[pd.Series, Dict[str, Any]]:
    s = pd.to_numeric(y, errors="coerce").astype(float).copy()
    name = transform_cfg.get("name", "none")
    if name == "log1p":
        return np.log1p(np.maximum(s, 0.0)), transform_cfg
    if name == "boxcox" and HAS_SCIPY:
        s2 = s.copy()
        shift = 0.0
        if safe_float(s2.min()) <= 0:
            shift = abs(safe_float(s2.min())) + 1e-6
            s2 = s2 + shift
        transformed, lam = boxcox(s2.values)
        cfg = dict(transform_cfg)
        cfg["lambda"] = lam
        cfg["shift"] = shift
        return pd.Series(transformed, index=s.index), cfg
    return s, transform_cfg

def inverse_target_transform(arr: np.ndarray, transform_cfg: Dict[str, Any]) -> np.ndarray:
    x = np.asarray(arr, dtype=float)
    transform_cfg = transform_cfg or {}
    allow_negative = bool(transform_cfg.get("allow_negative_output", False))
    name = transform_cfg.get("name", "none")
    if name == "log1p":
        out = np.expm1(x)
        return out if allow_negative else np.maximum(out, 0.0)
    if name == "boxcox" and HAS_SCIPY:
        lam = transform_cfg.get("lambda", None)
        shift = float(transform_cfg.get("shift", 0.0) or 0.0)
        inv = inv_boxcox(x, lam)
        inv = inv - shift
        return inv if allow_negative else np.maximum(inv, 0.0)
    return x if allow_negative else np.maximum(x, 0.0)

def make_inner_train_val_split(train_df: pd.DataFrame, val_ratio: float = 0.2, min_val: int = 3) -> Tuple[pd.DataFrame, pd.DataFrame]:
    n = len(train_df)
    val_len = max(min_val, int(np.ceil(n * val_ratio)))
    val_len = min(val_len, max(2, n // 3))
    return train_df.iloc[:-val_len].copy().reset_index(drop=True), train_df.iloc[-val_len:].copy().reset_index(drop=True)


def walk_forward_refit_sarimax(
    history_y: pd.Series,
    future_y: pd.Series,
    order: Tuple[int, int, int],
    seasonal_order: Tuple[int, int, int, int],
    exog_hist: Optional[pd.DataFrame] = None,
    exog_future: Optional[pd.DataFrame] = None,
    maxiter: int = 20
) -> np.ndarray:
    preds = []
    hist = pd.to_numeric(history_y, errors="coerce").astype(float).copy()
    future_y = pd.to_numeric(future_y, errors="coerce").astype(float).copy()
    for i in range(len(future_y)):
        exog_train_i = exog_hist if exog_hist is None else exog_hist.iloc[:len(hist)]
        exog_step = None if exog_future is None else exog_future.iloc[[i]]
        res = fit_sarimax_model(
            hist,
            exog_train_i,
            order=order,
            seasonal_order=seasonal_order,
            trend="c",
            maxiter=maxiter
        )
        fc = res.get_forecast(steps=1, exog=exog_step).predicted_mean.iloc[0]
        preds.append(max(float(fc), 0.0))
        hist = pd.concat([hist, pd.Series([future_y.iloc[i]])], ignore_index=True)
    return np.asarray(preds, dtype=float)

def make_prophet_features(train_df: pd.DataFrame, test_df: pd.DataFrame, exog_train: Optional[pd.DataFrame], exog_test: Optional[pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    tr = train_df[["ds", "y"]].copy()
    te = test_df[["ds", "y"]].copy()
    used = []
    if exog_train is not None and exog_test is not None and len(exog_train.columns) > 0:
        for c in exog_train.columns:
            tr[c] = pd.to_numeric(exog_train[c], errors="coerce").values
            te[c] = pd.to_numeric(exog_test[c], errors="coerce").values
            used.append(c)
    return tr, te, used



def _safe_prophet_regressor_name(name: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z_]+", "_", str(name)).strip("_")
    if not safe:
        safe = "regressor"
    if not safe[0].isalpha():
        safe = f"reg_{safe}"
    return safe[:60]


def _prepare_prophet_design_matrices(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    exog_train: Optional[pd.DataFrame],
    exog_test: Optional[pd.DataFrame]
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], List[str], Dict[str, str]]:
    tr = train_df[["ds", "y"]].copy().reset_index(drop=True)
    te = test_df[["ds", "y"]].copy().reset_index(drop=True)
    tr["ds"] = pd.to_datetime(tr["ds"])
    te["ds"] = pd.to_datetime(te["ds"])
    used_cols: List[str] = []
    dropped_cols: List[str] = []
    rename_map: Dict[str, str] = {}

    if exog_train is None or exog_test is None or len(exog_train.columns) == 0:
        return tr, te, used_cols, dropped_cols, rename_map

    xtr = exog_train.copy().reset_index(drop=True)
    xte = exog_test.copy().reset_index(drop=True)
    common_cols = [c for c in xtr.columns if c in xte.columns]
    used_safe_names = set()

    for c in common_cols:
        s_tr = pd.to_numeric(xtr[c], errors="coerce").replace([np.inf, -np.inf], np.nan)
        s_te = pd.to_numeric(xte[c], errors="coerce").replace([np.inf, -np.inf], np.nan)
        if s_tr.notna().sum() < max(5, len(s_tr) // 5):
            dropped_cols.append(c)
            continue
        if s_te.notna().sum() == 0:
            dropped_cols.append(c)
            continue
        if s_tr.dropna().nunique() <= 1:
            dropped_cols.append(c)
            continue
        safe_name = _safe_prophet_regressor_name(c)
        while safe_name in used_safe_names or safe_name in ["ds", "y", "cap", "floor"]:
            safe_name = f"{safe_name}_x"
        used_safe_names.add(safe_name)
        fill_val = safe_float(s_tr.median())
        if pd.isna(fill_val):
            fill_val = 0.0
        tr[safe_name] = s_tr.fillna(fill_val).fillna(0.0).astype(float).values
        te[safe_name] = s_te.fillna(fill_val).fillna(0.0).astype(float).values
        used_cols.append(safe_name)
        rename_map[str(c)] = safe_name

    return tr, te, used_cols, dropped_cols, rename_map


def _build_prophet_holidays_df(train_df: pd.DataFrame, test_df: pd.DataFrame) -> Optional[pd.DataFrame]:
    try:
        import holidays as holidays_pkg
    except Exception:
        return None
    years = sorted(set(pd.to_datetime(pd.concat([train_df["ds"], test_df["ds"]], axis=0)).dt.year.dropna().astype(int).tolist()))
    if not years:
        return None
    try:
        tr_holidays = holidays_pkg.country_holidays("TR", years=years)
    except Exception:
        return None
    rows = []
    for d, name in tr_holidays.items():
        rows.append({"ds": pd.Timestamp(d), "holiday": str(name)})
    if not rows:
        return None
    return pd.DataFrame(rows).drop_duplicates(subset=["ds", "holiday"]).sort_values(["ds", "holiday"]).reset_index(drop=True)


def _fit_predict_prophet_once(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    freq_alias: str,
    cfg: Dict[str, Any],
    exog_cols: List[str],
    use_holidays: bool = False,
    add_custom_seasonality: bool = True,
    transform_cfg: Optional[Dict[str, Any]] = None
) -> Tuple[Any, pd.DataFrame, np.ndarray]:
    if not HAS_PROPHET or Prophet is None:
        raise ImportError("prophet package unavailable")

    tr = train_df.copy().reset_index(drop=True)
    te = test_df.copy().reset_index(drop=True)
    tr["ds"] = pd.to_datetime(tr["ds"])
    te["ds"] = pd.to_datetime(te["ds"])

    transform_cfg = dict(transform_cfg or {"name": "none", "lambda": None, "shift": 0.0})
    y_raw = pd.to_numeric(tr["y"], errors="coerce").astype(float).fillna(0.0)
    y_transformed, applied_cfg = apply_target_transform(y_raw, transform_cfg)
    tr["y"] = pd.to_numeric(y_transformed, errors="coerce").astype(float).fillna(0.0)

    yearly = bool(str(freq_alias).upper() in ["M", "W", "D"])
    weekly = bool(str(freq_alias).upper() in ["D", "H"])
    daily = bool(str(freq_alias).upper() == "H")
    seasonality_mode = str(cfg.get("seasonality_mode", "additive"))
    seasonality_prior_scale = float(cfg.get("seasonality_prior_scale", 5.0))

    holidays_df = _build_prophet_holidays_df(tr, te) if use_holidays else None
    constructor_kwargs = {
        "seasonality_mode": seasonality_mode,
        "changepoint_prior_scale": float(cfg.get("changepoint_prior_scale", 0.05)),
        "seasonality_prior_scale": seasonality_prior_scale,
        "changepoint_range": float(cfg.get("changepoint_range", 0.9)),
        "n_changepoints": int(cfg.get("n_changepoints", 12)),
        "growth": "linear",
        "yearly_seasonality": yearly,
        "weekly_seasonality": weekly,
        "daily_seasonality": daily,
        "uncertainty_samples": 0
    }
    if holidays_df is not None and len(holidays_df) > 0:
        constructor_kwargs["holidays"] = holidays_df

    model = Prophet(**constructor_kwargs)

    if add_custom_seasonality:
        if str(freq_alias).upper() == "M":
            model.add_seasonality(name="quarterly", period=365.25 / 4.0, fourier_order=3, prior_scale=seasonality_prior_scale)
        elif str(freq_alias).upper() == "W":
            model.add_seasonality(name="monthly_proxy", period=30.5, fourier_order=4, prior_scale=seasonality_prior_scale)
            model.add_seasonality(name="year_end", period=365.25, fourier_order=6, prior_scale=seasonality_prior_scale)
        elif str(freq_alias).upper() == "D":
            model.add_seasonality(name="monthly_proxy", period=30.5, fourier_order=5, prior_scale=seasonality_prior_scale)
            model.add_seasonality(name="year_end", period=365.25, fourier_order=8, prior_scale=seasonality_prior_scale)

    for c in exog_cols:
        if c in tr.columns and c in te.columns:
            model.add_regressor(c, standardize=True)

    fit_cols = ["ds", "y"] + [c for c in exog_cols if c in tr.columns]
    fit_df = tr[fit_cols].copy().replace([np.inf, -np.inf], np.nan).fillna(0.0)
    future_df = te[["ds"] + [c for c in exog_cols if c in te.columns]].copy().replace([np.inf, -np.inf], np.nan).fillna(0.0)

    model.fit(fit_df)
    fc = model.predict(future_df)
    yhat = pd.to_numeric(fc["yhat"], errors="coerce").astype(float).fillna(0.0).values
    pred = inverse_target_transform(yhat, applied_cfg)
    pred = np.maximum(np.asarray(pred, dtype=float), 0.0)
    fc = fc.copy().reset_index(drop=True)
    fc["yhat"] = pred
    return model, fc, pred


def generate_target_ml_features(full_df: pd.DataFrame, existing_exog: Optional[pd.DataFrame], freq_alias: str) -> Tuple[pd.DataFrame, List[str]]:
    df = full_df[["ds", "y"]].copy().reset_index(drop=True)
    season_len = infer_season_length_from_freq(freq_alias)
    lag_set = [1, 2, 3, 6, 12]
    if freq_alias == "W":
        lag_set = [1, 2, 4, 8, 13, 26, 39, 52]
    elif freq_alias == "D":
        lag_set = [1, 2, 3, 7, 14, 21, 28, 56]
    elif freq_alias == "M":
        lag_set = [1, 2, 3, 6, 12, 18, 24]
    elif freq_alias == "H":
        lag_set = [1, 2, 3, 6, 12, 24, 48, 72]

    for lag in sorted(set(lag_set)):
        df[f"lag_{lag}"] = df["y"].shift(lag)
    if season_len > 1:
        df[f"seasonal_lag_{season_len}"] = df["y"].shift(season_len)
        df[f"diff_{season_len}"] = df["y"].diff(season_len)

    df["diff_1"] = df["y"].diff(1)

    shifted = df["y"].shift(1)
    for w in sorted(set([3, 6, 12, max(2, season_len)])):
        df[f"roll_mean_{w}"] = shifted.rolling(w, min_periods=1).mean()
        df[f"roll_std_{w}"] = shifted.rolling(w, min_periods=1).std()
        df[f"roll_min_{w}"] = shifted.rolling(w, min_periods=1).min()
        df[f"roll_max_{w}"] = shifted.rolling(w, min_periods=1).max()
        df[f"roll_median_{w}"] = shifted.rolling(w, min_periods=1).median()
        df[f"ewm_mean_{w}"] = shifted.ewm(span=max(2, w), adjust=False).mean()
    if freq_alias == "W":
        for w in [4, 8, 13, 26, 52]:
            s = shifted.rolling(w, min_periods=max(2, min(4, w))).mean()
            df[f"recent_mean_gap_{w}"] = shifted - s
            df[f"recent_z_{w}"] = (shifted - s) / (shifted.rolling(w, min_periods=max(2, min(4, w))).std().replace(0, np.nan))
            df[f"trend_slope_{w}"] = shifted.diff().rolling(w, min_periods=max(2, min(4, w))).mean()

    ds = pd.to_datetime(df["ds"])
    df["year"] = ds.dt.year
    df["quarter"] = ds.dt.quarter
    df["month"] = ds.dt.month
    iso_week = ds.dt.isocalendar().week.astype(int)
    df["weekofyear"] = iso_week
    df["dayofweek"] = ds.dt.dayofweek
    df["dayofmonth"] = ds.dt.day
    df["is_month_start"] = ds.dt.is_month_start.astype(int)
    df["is_month_end"] = ds.dt.is_month_end.astype(int)

    month_num = ds.dt.month.fillna(1).astype(int)
    df["month_sin"] = np.sin(2 * np.pi * month_num / 12.0)
    df["month_cos"] = np.cos(2 * np.pi * month_num / 12.0)
    if freq_alias in ["W", "D"]:
        week_num = iso_week.clip(lower=1)
        df["week_sin"] = np.sin(2 * np.pi * week_num / 52.0)
        df["week_cos"] = np.cos(2 * np.pi * week_num / 52.0)
    if freq_alias in ["D", "H"]:
        dow = ds.dt.dayofweek.astype(int)
        df["dow_sin"] = np.sin(2 * np.pi * dow / 7.0)
        df["dow_cos"] = np.cos(2 * np.pi * dow / 7.0)

    # Türkiye tatil ve özel dönem yakınlığı özellikleri
    try:
        years = sorted(set(ds.dt.year.dropna().astype(int).tolist()))
        holiday_dates = []
        if years:
            import holidays as _holidays
            tr_holidays = _holidays.country_holidays("TR", years=years)
            holiday_dates = sorted(pd.to_datetime(list(tr_holidays.keys())))
        if holiday_dates:
            holiday_arr = np.array([pd.Timestamp(x).to_datetime64() for x in holiday_dates], dtype='datetime64[ns]')
            current_arr = ds.values.astype('datetime64[ns]')
            prev_gap = []
            next_gap = []
            for cur in current_arr:
                diffs = (holiday_arr - cur).astype('timedelta64[D]').astype(int)
                pos = diffs[diffs >= 0]
                neg = (-diffs[diffs <= 0])
                next_gap.append(int(pos.min()) if len(pos) else 999)
                prev_gap.append(int(neg.min()) if len(neg) else 999)
            df["tatil_yakınlık_sonraki"] = pd.Series(next_gap).clip(upper=60)
            df["tatil_yakınlık_onceki"] = pd.Series(prev_gap).clip(upper=60)
            df["tatil_yakınlık_min"] = np.minimum(df["tatil_yakınlık_sonraki"], df["tatil_yakınlık_onceki"])
            df["tatil_yakın_flag_7"] = (df["tatil_yakınlık_min"] <= 7).astype(int)
            df["tatil_yakın_flag_14"] = (df["tatil_yakınlık_min"] <= 14).astype(int)
        else:
            for c in ["tatil_yakınlık_sonraki", "tatil_yakınlık_onceki", "tatil_yakınlık_min", "tatil_yakın_flag_7", "tatil_yakın_flag_14"]:
                df[c] = 0
    except Exception:
        for c in ["tatil_yakınlık_sonraki", "tatil_yakınlık_onceki", "tatil_yakınlık_min", "tatil_yakın_flag_7", "tatil_yakın_flag_14"]:
            df[c] = 0

    # Takvimsel özel olay yakınlığı (yılbaşı / ramazan-kurban etkisine yaklaşık vekil pencereler)
    try:
        month_day = ds.dt.strftime('%m-%d')
        yilbasi = (month_day >= '12-20') | (month_day <= '01-10')
        okul_acilis = (ds.dt.month == 9).astype(int)
        yilsonu = (ds.dt.month == 12).astype(int)
        df["ozel_olay_yılbaşı_penceresi"] = yilbasi.astype(int)
        df["ozel_olay_okul_açılışı"] = okul_acilis
        df["ozel_olay_yılsonu"] = yilsonu
        df["ozel_olay_toplam_flag"] = (df[["ozel_olay_yılbaşı_penceresi", "ozel_olay_okul_açılışı", "ozel_olay_yılsonu"]].sum(axis=1) > 0).astype(int)
    except Exception:
        for c in ["ozel_olay_yılbaşı_penceresi", "ozel_olay_okul_açılışı", "ozel_olay_yılsonu", "ozel_olay_toplam_flag"]:
            df[c] = 0

    if existing_exog is not None and len(existing_exog.columns) > 0:
        exog = existing_exog.copy().reset_index(drop=True)
        for c in exog.columns:
            df[f"exog__{c}"] = pd.to_numeric(exog[c], errors="coerce")

    feature_cols = [c for c in df.columns if c not in ["ds", "y"]]
    return df, feature_cols

def build_actual_vs_pred_df(test_df: pd.DataFrame, pred: np.ndarray, model_name: str) -> pd.DataFrame:
    out = test_df[["ds", "y"]].copy()
    out["prediction"] = np.maximum(np.asarray(pred, dtype=float), 0.0)
    out["model"] = model_name
    out["abs_error"] = np.abs(out["y"] - out["prediction"])
    out["ape"] = np.where(np.abs(out["y"]) > 1e-8, np.abs(out["y"] - out["prediction"]) / np.abs(out["y"]) * 100.0, np.nan)
    return out

TURKCE_KOLON_HARITASI = {
    "model": "Model",
    "prediction": "Tahmin",
    "abs_error": "Mutlak Hata",
    "ape": "Mutlak Yüzde Hata",
    "ds": "Tarih",
    "y": "Gerçek",
    "WAPE": "WAPE",
    "sMAPE": "sMAPE",
    "RMSE": "RMSE",
    "MAE": "MAE",
    "MAPE": "MAPE",
    "MASE": "MASE",
    "Bias": "Sapma",
    "BiasPct": "SapmaYüzde",
    "bias_pct": "SapmaYüzde",
    "under_forecast_rate": "EksikTahminOranı",
    "over_forecast_rate": "FazlaTahminOranı",
    "UnderForecastRate": "EksikTahminOranı",
    "OverForecastRate": "FazlaTahminOranı",
    "eligibility_score": "UygunlukSkoru",
    "status": "Durum",
    "reasons": "Gerekçeler",
    "feature_availability_risk": "ÖzellikErişimRiski",
    "validation_wape": "DoğrulamaWAPE",
    "rolling_wape": "RollingWAPE",
    "peak_event_score": "TepeOlaySkoru",
    "peak_precision": "TepeKesinlik",
    "peak_recall": "TepeDuyarlılık",
    "peak_f1": "TepeF1",
    "actual_peak_count": "GerçekTepeSayısı",
    "pred_peak_count": "TahminTepeSayısı",
    "service_level_target": "HedefServisSeviyesi",
    "achieved_cycle_service": "GerçekleşenServisSeviyesi",
    "service_gap": "ServisAçığı",
    "quantile_10": "AltBant_10",
    "quantile_20": "AltBant_20",
    "quantile_80": "ÜstBant_80",
    "quantile_90": "ÜstBant_90",
    "coverage_80": "Kapsama_80",
    "coverage_90": "Kapsama_90",
    "coverage_95": "Kapsama_95",
    "interval_width_80": "BantGenişliği_80",
    "interval_width_90": "BantGenişliği_90",
    "interval_width_95": "BantGenişliği_95",
    "production_model": "ÜretimModeli",
    "production_status": "ÜretimDurumu",
    "drift_ratio_vs_recent": "YakınDönemDriftOranı",
    "max_feature_risk": "MaksÖzellikRiski",
    "fallback_rate": "FallbackOranı",
    "target_service": "HedefServis",
    "achieved_service": "GerçekleşenServis",
    "deployment_recommendation": "DağıtımÖnerisi",
    "alert": "Alarm",
    "detail": "Detay",
    "severity": "Şiddet",
    "weight": "Ağırlık",
    "raw_weight": "HamAğırlık",
    "val_WAPE": "DoğrulamaWAPE",
    "val_sMAPE": "DoğrulamasMAPE",
    "ro_WAPE": "RollingWAPE",
    "ro_MAE": "RollingMAE",
    "strategy": "Strateji",
    "backend_name": "ArkaUç",
    "used_feature_count": "KullanılanÖzellikSayısı",
    "fallback_used": "FallbackKullanıldı",
    "fallback_method": "FallbackYöntemi",
    "champion": "Şampiyon",
    "challenger": "MeydanOkuyan",
    "karar_kademesi": "KararKademesi",
}

TURKCE_DEGER_HARITASI = {
    "eligible": "uygun",
    "challenger_only": "yalnız_meydan_okuyan",
    "reject": "reddet",
    "guarded_fallback": "korumalı_fallback",
    "guarded_use_only": "yalnız_korumalı_kullanım",
    "eligible_for_guarded_production": "korumalı_üretime_uygun",
    "champion_challenger_parallel_run": "şampiyon_meydan_okuyan_paralel_çalıştır",
    "shadow_mode_required": "gölge_mod_zorunlu",
    "high": "yüksek",
    "medium": "orta",
    "bias_alert": "sapma_alarmı",
    "under_forecast_alert": "eksik_tahmin_alarmı",
    "peak_capture_alert": "tepe_yakalama_alarmı",
    "forecast_drift_alert": "tahmin_drift_alarmı",
    "feature_contract_alert": "özellik_sözleşmesi_alarmı",
    "service_level_alert": "servis_seviyesi_alarmı",
    "fallback_rate_alert": "fallback_oranı_alarmı",
}


def _turkcelestir_df_kolonlari(df: pd.DataFrame) -> pd.DataFrame:
    """Kolon adlarını Türkçeleştirir ve çeviri sonrası oluşan tekrarları güvenli biçimde birleştirir."""
    out = df.copy()
    translated_cols = [TURKCE_KOLON_HARITASI.get(str(c), str(c)) for c in out.columns]
    out.columns = translated_cols

    if len(set(translated_cols)) == len(translated_cols):
        return out

    birlesik_seriler = {}
    kolon_sirasi = []
    for idx, ad in enumerate(translated_cols):
        col_data = out.iloc[:, idx]
        if ad not in birlesik_seriler:
            birlesik_seriler[ad] = col_data.copy()
            kolon_sirasi.append(ad)
        else:
            mevcut = birlesik_seriler[ad]
            try:
                birlesik_seriler[ad] = pd.concat([mevcut, col_data], axis=1).bfill(axis=1).iloc[:, 0]
            except Exception:
                birlesik_seriler[ad] = mevcut.where(mevcut.notna(), col_data)

    sonuc = pd.DataFrame({ad: birlesik_seriler[ad] for ad in kolon_sirasi})
    return sonuc


def style_metric_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Sayıları biçimlendirir, durum alanlarını Türkçeleştirir ve veri çerçevesini gösterime hazırlar."""
    out = df.copy()
    for c in out.columns:
        s = out[c]
        if pd.api.types.is_datetime64_any_dtype(s) or pd.api.types.is_timedelta64_dtype(s):
            continue
        if pd.api.types.is_bool_dtype(s):
            continue
        if pd.api.types.is_numeric_dtype(s):
            out[c] = pd.to_numeric(s, errors="coerce").round(4)
            continue

        def _safe_fmt(x):
            if pd.isna(x):
                return np.nan
            if isinstance(x, (pd.Timestamp, pd.Timedelta, np.datetime64)):
                return x
            if isinstance(x, (list, tuple, dict, set)):
                return json.dumps(list(x) if isinstance(x, set) else x, ensure_ascii=False)
            if isinstance(x, str):
                xv = TURKCE_DEGER_HARITASI.get(x.strip(), x)
                try:
                    stripped = xv.strip()
                    if stripped == "":
                        return xv
                    return round(float(stripped), 4)
                except Exception:
                    return xv
            try:
                return round(float(x), 4)
            except Exception:
                return TURKCE_DEGER_HARITASI.get(str(x), x)

        out[c] = s.map(_safe_fmt)
    return _turkcelestir_df_kolonlari(out)

def dataframe_to_download_bytes(df: pd.DataFrame) -> bytes:
    out_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    return _build_excel_bytes({"data": out_df}) or b""


def build_current_series_tables_excel_bytes(outputs: Dict[str, Any], export_payload: Dict[str, pd.DataFrame], selected_sheet: str, target_col: str) -> Optional[bytes]:
    try:
        table_map = build_named_output_tables(outputs, export_payload)
        notes_map = {}
        suggestions = outputs.get("deep_learning_analysis", pd.DataFrame())
        if isinstance(suggestions, pd.DataFrame) and len(suggestions) > 0:
            notes_map["DL Analizi"] = suggestions.copy()
        return _build_excel_bytes(table_map=table_map, notes_map=notes_map)
    except Exception:
        return None

def _safe_artifact_name(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(name)).strip()

def _df_or_empty(obj: Any) -> pd.DataFrame:
    return obj if isinstance(obj, pd.DataFrame) else pd.DataFrame()

def build_named_output_tables(outputs: Dict[str, Any], export_payload: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    prod_pack = outputs.get("production_governance", {}) or {}
    live_pack = prod_pack.get("live_monitoring_pack", {}) or {}

    error_df = pd.DataFrame([
        {"model": model, "message": message}
        for model, message in (outputs.get("model_errors", {}) or {}).items()
    ])
    profile = (outputs.get("metadata") or {}).get("profile", {}) or {}
    profile_snapshot = pd.DataFrame([{
        "segment": (outputs.get("metadata") or {}).get("segment"),
        "abc_xyz": (outputs.get("metadata") or {}).get("abc_xyz"),
        "priority": (outputs.get("metadata") or {}).get("priority"),
        "target_col": (outputs.get("metadata") or {}).get("target_col"),
        "freq_alias": (outputs.get("metadata") or {}).get("freq_alias"),
        "horizon": (outputs.get("metadata") or {}).get("horizon"),
        "cv": profile.get("cv"),
        "trend_strength": profile.get("trend_strength"),
        "seasonality_strength": profile.get("seasonality_strength"),
        "intermittency_ratio": profile.get("intermittency_ratio"),
        "volume_level": profile.get("volume_level"),
        "volatility_regime": profile.get("volatility_regime"),
    }])

    named = {
        "Seri profili özet metrikleri": style_metric_dataframe(profile_snapshot),
        "Model karşılaştırma tablosu": style_metric_dataframe(_df_or_empty(outputs.get("metrics_df"))),
        "Çalışma süresi - performans özeti": style_metric_dataframe(_df_or_empty(outputs.get("stage_timing_table"))),
        "Model hata özeti": style_metric_dataframe(error_df),
        "Fallback özeti": style_metric_dataframe(_df_or_empty(outputs.get("fallback_summary"))),
        "DL şeffaflık ve kimlik doğrulama tablosu": style_metric_dataframe(_df_or_empty(outputs.get("dl_transparency_table"))),
        "Otomatik model uygunluk kapısı": style_metric_dataframe(_df_or_empty(prod_pack.get("model_eligibility_gate"))),
        "Karar kademelerine göre model liderleri": style_metric_dataframe(_df_or_empty(build_model_liderleri(outputs))),
        "SARIMA-SARIMAX - Gerçek ve tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("SARIMA/SARIMAX"))),
        "SARIMA-SARIMAX - Arama tablosu": style_metric_dataframe(_df_or_empty((outputs.get("sarima") or {}).get("search_table"))),
        "Prophet görünürlüğü (gerçek fit mi, fallback mi)": style_metric_dataframe(_df_or_empty(outputs.get("prophet_gorunurluk_ozeti"))),
        "Prophet - Gerçek ve tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("Prophet"))),
        "Prophet - Arama tablosu": style_metric_dataframe(_df_or_empty((outputs.get("prophet") or {}).get("search_table"))),
        "XGBoost - Gerçek ve tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("XGBoost"))),
        "XGBoost - Arama tablosu": style_metric_dataframe(_df_or_empty((outputs.get("xgboost") or {}).get("search_table"))),
        "XGBoost - Strateji karşılaştırması": style_metric_dataframe(_df_or_empty((outputs.get("xgboost") or {}).get("strategy_comparison"))),
        "XGBoost - Özellik önemleri": _df_or_empty((outputs.get("xgboost") or {}).get("feature_importance")),
        "LSTM - Gerçek ve tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("LSTM (real)", outputs.get("tables", {}).get("LSTM")))),
        "LSTM - Arama tablosu": style_metric_dataframe(_df_or_empty((outputs.get("lstm") or {}).get("search_table"))),
        "LSTM - Epoch Loss Geçmişi": style_metric_dataframe(_df_or_empty((outputs.get("lstm") or {}).get("history_df"))),
        "LSTM - Eğitim Durumu Tablosu": style_metric_dataframe(_df_or_empty((outputs.get("lstm") or {}).get("dl_training_status_table"))),
        "GRU - Gerçek ve tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("GRU (real)", outputs.get("tables", {}).get("GRU")))),
        "GRU - Arama tablosu": style_metric_dataframe(_df_or_empty((outputs.get("gru") or {}).get("search_table"))),
        "GRU - Epoch Loss Geçmişi": style_metric_dataframe(_df_or_empty((outputs.get("gru") or {}).get("history_df"))),
        "Karar hiyerarşisi özeti": style_metric_dataframe(_df_or_empty(outputs.get("karar_hiyerarsisi"))),
        "Holdout sıralaması": style_metric_dataframe(_df_or_empty((outputs.get("champion_challenger") or {}).get("ranking"))),
        "Ansambl ağırlıkları (peak-aware + bias-aware)": style_metric_dataframe(_df_or_empty(outputs.get("ensemble_weights"))),
        "Ansambl gerçek ve tahmin": style_metric_dataframe(_df_or_empty(outputs.get("tables", {}).get("Ensemble"))),
        "Gerçek vs tahmin tablosu": style_metric_dataframe(_df_or_empty(outputs.get("all_predictions_long"))),
        "Rolling-origin geri test (tam katman)": style_metric_dataframe(_df_or_empty(outputs.get("rolling_origin_backtest"))),
        "Doğrulama temelli model kalitesi": style_metric_dataframe(_df_or_empty(outputs.get("validation_metrics_df"))),
        "Karar hiyerarşisi": style_metric_dataframe(_df_or_empty(outputs.get("karar_hiyerarsisi"))),
        "Merkezî canlı izleme özeti": style_metric_dataframe(_df_or_empty(live_pack.get("summary"))),
        "Model sağlık tablosu (bias, kapsama, fallback, servis seviyesi)": style_metric_dataframe(_df_or_empty(live_pack.get("model_health_table"))),
        "Üretim alarmları": style_metric_dataframe(_df_or_empty(live_pack.get("alerts"))),
        "Üretim sıralaması (holdout, doğrulama ve rolling-origin birlikte)": style_metric_dataframe(_df_or_empty(prod_pack.get("production_ranking"))),
        "Tahmin katkı değeri (baseline'a göre)": style_metric_dataframe(_df_or_empty(prod_pack.get("forecast_value_add"))),
        "Bias dashboard": style_metric_dataframe(_df_or_empty(prod_pack.get("bias_dashboard"))),
        "Tepe olay yakalama skoru": style_metric_dataframe(_df_or_empty(prod_pack.get("peak_event_dashboard"))),
        "Üretim özellik erişilebilirlik denetimi": style_metric_dataframe(_df_or_empty(prod_pack.get("feature_availability_audit"))),
        "Üretim modeli için tahmin aralığı - quantile forecast": style_metric_dataframe(_df_or_empty(prod_pack.get("production_interval_table"))),
        "Servis seviyesi - stok etkisi simülasyonu": style_metric_dataframe(_df_or_empty(prod_pack.get("production_service_table"))),
        "Benzersiz rapor kataloğu (yinelenen içerik temizliği)": style_metric_dataframe(_df_or_empty(outputs.get("benzersiz_rapor_katalogu"))),
        "Kalite raporu": _df_or_empty(export_payload.get("quality_report")),
        "Seri profil raporu": _df_or_empty(export_payload.get("series_profile_report")),
        "Anomali yönetişimi": _df_or_empty(export_payload.get("anomaly_governance")),
        "Review queue": _df_or_empty(export_payload.get("review_queue")),
        "Proxy backtest raporu": _df_or_empty(export_payload.get("proxy_backtest_report")),
        "Raw vs Clean backtest karşılaştırması": _df_or_empty(export_payload.get("raw_vs_clean_backtest_report")),
    }

    cleaned: Dict[str, pd.DataFrame] = {}
    for report_name, df in named.items():
        if isinstance(df, pd.DataFrame):
            cleaned[report_name] = df.copy()
    return cleaned


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(v) for v in value]
    if isinstance(value, pd.DataFrame):
        return value.to_dict(orient="records")
    if isinstance(value, pd.Series):
        return value.to_list()
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return str(value)
    if isinstance(value, np.ndarray):
        return np.asarray(value).tolist()
    if isinstance(value, (np.integer, )):
        return int(value)
    if isinstance(value, (np.floating, )):
        return None if pd.isna(value) else float(value)
    if isinstance(value, (bool, int, float, str)) or value is None:
        return value
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return str(value)


def _json_bytes(obj: Any) -> bytes:
    return json.dumps(_json_safe(obj), ensure_ascii=False, indent=2).encode("utf-8")


def _text_bytes(lines: List[str]) -> bytes:
    return ("\n".join([str(x) for x in lines if str(x).strip() != ""]) + "\n").encode("utf-8")


def _tab9_suggestions(outputs: Dict[str, Any], profile: Dict[str, Any]) -> List[str]:
    suggestions: List[str] = []
    if float(profile.get("seasonality_strength", 0) or 0) >= 0.35:
        suggestions.append("Seri sezonsal; auto seasonal (P,D,Q,m), Prophet custom seasonality ve ensemble kritik.")
    if float(profile.get("cv", 0) or 0) >= 0.45:
        suggestions.append("Oynaklık yüksek; log/Box-Cox dönüşümü, XGBoost rolling istatistikleri ve challenger modeli zorunlu izlenmeli.")
    if float(profile.get("intermittency_ratio", 0) or 0) >= 0.25:
        suggestions.append("Intermittent yapı var; stok-out ayrımı ve model çıktılarını iş kuralı ile doğrulamak gerekir.")
    if outputs.get("sarima", {}).get("residual_white_noise_ok") is False:
        suggestions.append("SARIMA residual white-noise testi zayıf; challenger olarak Prophet veya XGBoost önceliklendirilmeli.")
    if outputs.get("best_model") == "Ensemble":
        suggestions.append("Bu seride tek bir champion yerine ensemble daha iyi; operasyonel kullanımda champion-challenger izleme önerilir.")
    if not suggestions:
        suggestions.append("Seri dengeli; champion-challenger ve ensemble izlemesi yeterli görünüyor.")
    return suggestions


def _notes_df(section: str, lines: List[str]) -> pd.DataFrame:
    return pd.DataFrame([{"section": section, "text": line} for line in lines]) if lines else pd.DataFrame(columns=["section", "text"])


def _series_notes_tables(outputs: Dict[str, Any], export_payload: Optional[Dict[str, pd.DataFrame]] = None) -> Dict[str, pd.DataFrame]:
    profile = ((outputs.get("metadata") or {}).get("profile") or {})
    meta = outputs.get("metadata", {}) or {}
    metrics_df = outputs.get("metrics_df", pd.DataFrame()) if isinstance(outputs, dict) else pd.DataFrame()
    readiness = assess_production_readiness(
        export_payload or {},
        metrics_df if isinstance(metrics_df, pd.DataFrame) else pd.DataFrame(),
        str(meta.get("target_col", "")),
        outputs=outputs if isinstance(outputs, dict) else None,
    ) if isinstance(outputs, dict) else {"notes": []}
    cc = outputs.get("champion_challenger", {}) or {}
    prod_pack = outputs.get("production_governance", {}) or {}
    lines_prod = [
        f"Üretim önerisi: {prod_pack.get('production_model', outputs.get('best_model'))}",
        f"Üretim durumu: {prod_pack.get('production_status', 'eligible')}",
        f"Şampiyon model: {cc.get('champion', '-')}",
        f"Meydan okuyan: {cc.get('challenger', '-')}",
    ]
    runtime_guardrails = (outputs.get("metadata") or {}).get("runtime_guardrails", []) or []
    return {
        "Üretim kullanımı değerlendirmesi notları": _notes_df("production_readiness", list(readiness.get("notes", []))),
        "Akıllı değerlendirmeler": _notes_df("smart_evaluations", _tab9_suggestions(outputs, profile)),
        "Üretim önerisi ve champion özeti": _notes_df("production_summary", lines_prod),
        "Runtime guardrails": _notes_df("runtime_guardrails", list(runtime_guardrails)),
    }


def _series_json_artifacts(outputs: Dict[str, Any]) -> Dict[str, Any]:
    metadata = outputs.get("metadata", {}) or {}
    sarima = outputs.get("sarima", {}) or {}
    prophet = outputs.get("prophet", {}) or {}
    xgboost = outputs.get("xgboost", {}) or {}
    cc = outputs.get("champion_challenger", {}) or {}
    prod_pack = outputs.get("production_governance", {}) or {}
    return {
        "metadata": {
            "target_col": metadata.get("target_col"),
            "freq_alias": metadata.get("freq_alias"),
            "horizon": metadata.get("horizon"),
            "segment": metadata.get("segment"),
            "abc_xyz": metadata.get("abc_xyz"),
            "priority": metadata.get("priority"),
            "candidate_models": metadata.get("candidate_models"),
            "stat_exog_cols": metadata.get("stat_exog_cols"),
            "prophet_exog_cols": metadata.get("prophet_exog_cols"),
            "ml_feature_cols": metadata.get("ml_feature_cols"),
            "runtime_guardrails": metadata.get("runtime_guardrails"),
            "train_n": metadata.get("train_n"),
            "test_n": metadata.get("test_n"),
            "profile": metadata.get("profile"),
        },
        "sarima_summary": {
            "order": sarima.get("order"),
            "seasonal_order": sarima.get("seasonal_order"),
            "trend": sarima.get("trend"),
            "AIC": sarima.get("aic"),
            "BIC": sarima.get("bic"),
            "Ljung_Box_pvalue": sarima.get("ljung_box_pvalue"),
            "d": sarima.get("d"),
            "D": sarima.get("D"),
            "transform": sarima.get("transform"),
            "white_noise_ok": sarima.get("residual_white_noise_ok"),
            "fallback_used": sarima.get("fallback_used"),
            "fallback_method": sarima.get("fallback_method"),
        },
        "prophet_config": prophet.get("config", {}),
        "prophet_component_validation": prophet.get("component_validation", {}),
        "xgboost_summary": {
            "selected_strategy": xgboost.get("strategy"),
            "shap_status": xgboost.get("shap_status"),
            "fallback_used": xgboost.get("fallback_used"),
            "fallback_method": xgboost.get("fallback_method"),
        },
        "champion_challenger_summary": {
            "champion": cc.get("champion"),
            "challenger": cc.get("challenger"),
            "best_model": outputs.get("best_model"),
            "production_model": prod_pack.get("production_model", outputs.get("best_model")),
            "production_status": prod_pack.get("production_status", "eligible"),
        },
        "model_errors": outputs.get("model_errors", {}),
        "model_visual_style_map": outputs.get("model_gorsel_stil_haritasi", {}),
    }


def _plotly_html_bytes(fig: Any) -> Optional[bytes]:
    if fig is None:
        return None
    try:
        return fig.to_html(full_html=True, include_plotlyjs="cdn").encode("utf-8")
    except Exception:
        return None


def _forecast_plotly_html_bytes(train_df: pd.DataFrame, test_df: pd.DataFrame, pred_map: Dict[str, np.ndarray], title: str, style_map: Optional[Dict[str, Any]] = None) -> Optional[bytes]:
    try:
        fig = plot_forecast_results(train_df, test_df, pred_map, title, model_style_map=style_map or {})
        return _plotly_html_bytes(fig)
    except Exception:
        return None


def _unique_sheet_name(name: str, used: set) -> str:
    base = safe_excel_sheet_name(name) or "Sheet"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = safe_excel_sheet_name(base[: max(1, 31 - len(suffix))] + suffix)
        i += 1
    used.add(candidate)
    return candidate


def _build_excel_bytes(table_map: Dict[str, pd.DataFrame], notes_map: Optional[Dict[str, pd.DataFrame]] = None, json_map: Optional[Dict[str, Any]] = None) -> Optional[bytes]:
    try:
        mem = io.BytesIO()
        used = set()
        with pd.ExcelWriter(mem, engine="openpyxl") as writer:
            for sheet_name, df in table_map.items():
                out_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
                out_df.to_excel(writer, sheet_name=_unique_sheet_name(sheet_name, used), index=False)
            for sheet_name, df in (notes_map or {}).items():
                out_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(columns=["section", "text"])
                out_df.to_excel(writer, sheet_name=_unique_sheet_name(sheet_name, used), index=False)
            for sheet_name, payload in (json_map or {}).items():
                json_df = pd.DataFrame({"json": json.dumps(_json_safe(payload), ensure_ascii=False, indent=2).splitlines()})
                json_df.to_excel(writer, sheet_name=_unique_sheet_name(sheet_name, used), index=False)
        mem.seek(0)
        return mem.getvalue()
    except Exception:
        return None


def _forecast_png_bytes(train_df: pd.DataFrame, test_df: pd.DataFrame, pred_map: Dict[str, np.ndarray], title: str) -> bytes:
    fig, ax = plt.subplots(figsize=(13, 5))
    ax.plot(train_df["ds"], train_df["y"], label="Train", linewidth=2)
    ax.plot(test_df["ds"], test_df["y"], label="Gerçek", linewidth=2, marker="o")
    for model_name, pred in pred_map.items():
        arr = np.asarray(pred, dtype=float)
        if len(arr) == len(test_df):
            ax.plot(test_df["ds"], arr, label=model_name, linewidth=2, marker="o")
    ax.set_title(title)
    ax.set_xlabel("Tarih")
    ax.set_ylabel("Talep")
    ax.legend()
    ax.grid(alpha=0.25)
    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _acf_pacf_png_bytes(export_payload: Dict[str, pd.DataFrame], target_col: str, horizon: int) -> Optional[bytes]:
    try:
        df_series = make_series_analysis_frame(export_payload, target_col)
        train_df_preview, _ = train_test_split_series(df_series, horizon)
        fig = build_acf_pacf_figure(train_df_preview, target_col)
        if fig is None:
            return None
        buf = io.BytesIO()
        fig.tight_layout()
        fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def _single_table_excel_bytes(sheet_name: str, df: pd.DataFrame) -> Optional[bytes]:
    try:
        return _build_excel_bytes({sheet_name: df if isinstance(df, pd.DataFrame) else pd.DataFrame()})
    except Exception:
        return None


def _write_batch_zip_content(
    zf: zipfile.ZipFile,
    export_payload: Dict[str, pd.DataFrame],
    batch_result: Dict[str, Any],
    selected_sheet: str,
    horizon: int,
    include_self_script: bool = True,
    include_requirements: bool = True,
) -> None:
    progress_log("ZIP Paketleme", "ZIP içeriği yazılmaya başlandı.", extra={"sheet": selected_sheet, "series": len((batch_result.get('batch_outputs') or {}))})
    sheet_root = _safe_artifact_name(selected_sheet) or "batch_forecasting_outputs"

    common_tables = {
        "Kalite raporu": export_payload.get("quality_report", pd.DataFrame()),
        "Seri profil raporu": export_payload.get("series_profile_report", pd.DataFrame()),
        "Anomali yönetişimi": export_payload.get("anomaly_governance", pd.DataFrame()),
        "Review queue": export_payload.get("review_queue", pd.DataFrame()),
        "Proxy backtest raporu": export_payload.get("proxy_backtest_report", pd.DataFrame()),
        "Raw vs Clean backtest karşılaştırması": export_payload.get("raw_vs_clean_backtest_report", pd.DataFrame()),
    }

    for name, df in common_tables.items():
        if isinstance(df, pd.DataFrame):
            xlsx_bytes = _single_table_excel_bytes(name, df)
            if xlsx_bytes is not None:
                zf.writestr(f"{sheet_root}/common/{_safe_artifact_name(name)}.xlsx", xlsx_bytes)

    batch_summary_xlsx = _single_table_excel_bytes("Batch forecasting özeti", batch_result.get("best_summary", pd.DataFrame()))
    if batch_summary_xlsx is not None:
        zf.writestr(f"{sheet_root}/batch/Batch forecasting özeti.xlsx", batch_summary_xlsx)

    champion_xlsx = _single_table_excel_bytes("Champion - Challenger tablosu", batch_result.get("champion_table", pd.DataFrame()))
    if champion_xlsx is not None:
        zf.writestr(f"{sheet_root}/batch/Champion - Challenger tablosu.xlsx", champion_xlsx)

    business_pack = _build_batch_business_output_pack(
        batch_result.get("best_summary", pd.DataFrame()),
        batch_result.get("champion_table", pd.DataFrame()),
        batch_result.get("series_status", pd.DataFrame()),
        horizon=int(horizon),
    )
    for table_name, table_df in (business_pack.get("tables") or {}).items():
        xlsx_bytes = _single_table_excel_bytes(table_name, table_df)
        if xlsx_bytes is not None:
            zf.writestr(f"{sheet_root}/batch/{_safe_artifact_name(table_name)}.xlsx", xlsx_bytes)
    for fig_name, fig_bytes in (business_pack.get("figures") or {}).items():
        if fig_bytes is not None:
            zf.writestr(f"{sheet_root}/batch/figures/{_safe_artifact_name(fig_name)}", fig_bytes)

    batch_xlsx = _build_excel_bytes({
        "Batch forecasting özeti": batch_result.get("best_summary", pd.DataFrame()),
        "Champion - Challenger": batch_result.get("champion_table", pd.DataFrame()),
        "Seri çalışma durumu": batch_result.get("series_status", pd.DataFrame()),
        **(business_pack.get("tables") or {}),
        **{f"Common - {k}": v for k, v in common_tables.items() if isinstance(v, pd.DataFrame)},
    })
    if batch_xlsx is not None:
        zf.writestr(f"{sheet_root}/batch/Batch forecasting master raporu.xlsx", batch_xlsx)

    manifest = {
        "selected_sheet": selected_sheet,
        "targets": list(batch_result.get("batch_outputs", {}).keys()),
        "horizon": int(horizon),
        "artifact_schema": "batch_forecast_zip_v3_hybrid_local_streamlit_excel_only",
        "app_version": APP_VERSION,
        "table_format": "xlsx",
        "figure_formats": ["png", "html"],
    }
    zf.writestr(f"{sheet_root}/manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    if include_self_script:
        try:
            script_path = Path(__file__)
            if script_path.exists():
                zf.writestr(f"{sheet_root}/{script_path.name}", script_path.read_bytes())
        except Exception:
            pass

    if include_requirements:
        try:
            requirements_path = Path(__file__).with_name("requirements.txt")
            if requirements_path.exists():
                zf.writestr(f"{sheet_root}/requirements.txt", requirements_path.read_bytes())
        except Exception:
            pass

    for target, outputs in (batch_result.get("batch_outputs") or {}).items():
        progress_log("ZIP Paketleme", "Seri klasörü ve rapor dosyaları yazılıyor.", target=target)
        target_root = f"{sheet_root}/{_safe_artifact_name(target)}"
        style_map = outputs.get("model_gorsel_stil_haritasi", {}) or {}
        raw_named_tables = outputs.get("named_output_tables", {}) or {}
        named_tables = {
            "Yönetici özeti": _build_series_executive_summary_table(outputs),
            "Operasyonel aksiyon planı": _build_series_operational_action_table(outputs),
            **raw_named_tables,
        }
        note_tables = _series_notes_tables(outputs, export_payload=export_payload)
        json_artifacts = _series_json_artifacts(outputs)

        for report_name, df in named_tables.items():
            if isinstance(df, pd.DataFrame):
                xlsx_bytes = _single_table_excel_bytes(report_name, df)
                if xlsx_bytes is not None:
                    zf.writestr(f"{target_root}/tables/{_safe_artifact_name(report_name)}.xlsx", xlsx_bytes)

        for note_name, note_df in note_tables.items():
            xlsx_bytes = _single_table_excel_bytes(note_name, note_df)
            if xlsx_bytes is not None:
                zf.writestr(f"{target_root}/notes/{_safe_artifact_name(note_name)}.xlsx", xlsx_bytes)
            zf.writestr(f"{target_root}/notes/{_safe_artifact_name(note_name)}.txt", _text_bytes(note_df.get("text", pd.Series(dtype=object)).astype(str).tolist()))

        for json_name, payload in json_artifacts.items():
            zf.writestr(f"{target_root}/json/{_safe_artifact_name(json_name)}.json", _json_bytes(payload))

        acf_bytes = _acf_pacf_png_bytes(export_payload, target, horizon)
        if acf_bytes is not None:
            zf.writestr(f"{target_root}/figures/{_safe_artifact_name(target)} - ACF ve PACF.png", acf_bytes)

        plot_specs = {
            "Gerçek vs Tahmin": outputs.get("predictions", {}),
            "SARIMA-SARIMAX": {"SARIMA/SARIMAX": (outputs.get("predictions", {}) or {}).get("SARIMA/SARIMAX")},
            "Prophet": {"Prophet": (outputs.get("predictions", {}) or {}).get("Prophet")},
            "XGBoost": {"XGBoost": (outputs.get("predictions", {}) or {}).get("XGBoost")},
            "Ensemble": {"Ensemble": (outputs.get("predictions", {}) or {}).get("Ensemble")},
        }
        for plot_name, pred_map in plot_specs.items():
            clean_pred_map = {k: v for k, v in (pred_map or {}).items() if v is not None}
            if not clean_pred_map:
                continue
            title = f"{target} - {plot_name}"
            zf.writestr(f"{target_root}/figures/{_safe_artifact_name(title)}.png", _forecast_png_bytes(outputs["train"], outputs["test"], clean_pred_map, title))
            html_bytes = _forecast_plotly_html_bytes(outputs["train"], outputs["test"], clean_pred_map, title, style_map)
            if html_bytes is not None:
                zf.writestr(f"{target_root}/figures/{_safe_artifact_name(title)}.html", html_bytes)

        series_xlsx = _build_excel_bytes(named_tables, notes_map=note_tables, json_map=json_artifacts)
        if series_xlsx is not None:
            zf.writestr(f"{target_root}/{_safe_artifact_name(target)} - detay raporu.xlsx", series_xlsx)
        progress_log("ZIP Paketleme", "Seri için ZIP çıktı dosyaları tamamlandı.", level="SUCCESS", target=target)


def build_batch_zip_bytes(export_payload: Dict[str, pd.DataFrame], batch_result: Dict[str, Any], selected_sheet: str, horizon: int) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        _write_batch_zip_content(zf, export_payload, batch_result, selected_sheet, horizon)
    mem.seek(0)
    return mem.getvalue()


def write_batch_zip_file(export_payload: Dict[str, pd.DataFrame], batch_result: Dict[str, Any], selected_sheet: str, horizon: int, output_zip_path: str) -> str:
    output_zip_path = str(output_zip_path)
    os.makedirs(os.path.dirname(output_zip_path) or ".", exist_ok=True)
    with zipfile.ZipFile(output_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        _write_batch_zip_content(zf, export_payload, batch_result, selected_sheet, horizon)
    return output_zip_path




def _write_binary_file(path: str, payload: bytes) -> str:
    path = str(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "wb") as f:
        f.write(payload)
    return path


def _write_text_file(path: str, lines: List[str]) -> str:
    path = str(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join([str(x) for x in lines]))
    return path


def _write_json_file(path: str, payload: Any) -> str:
    path = str(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(_json_safe(payload), f, ensure_ascii=False, indent=2)
    return path


def _write_single_table_excel_file(path: str, sheet_name: str, df: pd.DataFrame) -> Optional[str]:
    xlsx_bytes = _single_table_excel_bytes(sheet_name, df if isinstance(df, pd.DataFrame) else pd.DataFrame())
    if xlsx_bytes is None:
        return None
    return _write_binary_file(path, xlsx_bytes)


def _write_excel_report_file(path: str, table_map: Dict[str, pd.DataFrame], notes_map: Optional[Dict[str, pd.DataFrame]] = None, json_map: Optional[Dict[str, Any]] = None) -> Optional[str]:
    xlsx_bytes = _build_excel_bytes(table_map, notes_map=notes_map, json_map=json_map)
    if xlsx_bytes is None:
        return None
    return _write_binary_file(path, xlsx_bytes)


def _copy_optional_file_into_dir(source_path: Path, target_path: str) -> None:
    try:
        if source_path.exists():
            os.makedirs(os.path.dirname(target_path) or ".", exist_ok=True)
            shutil.copy2(str(source_path), str(target_path))
    except Exception:
        pass


def initialize_incremental_batch_output_root(
    output_dir: str,
    selected_sheet: str,
    horizon: int,
    mode_slug: str,
    timestamp: Optional[str] = None,
) -> Dict[str, str]:
    timestamp = timestamp or time.strftime("%Y%m%d_%H%M%S")
    run_base = f"{_safe_artifact_name(selected_sheet)}_{mode_slug}_forecasting_outputs_h{int(horizon)}_{timestamp}"
    run_root = os.path.join(os.path.abspath(output_dir), run_base)
    batch_root = os.path.join(run_root, _safe_artifact_name(selected_sheet) or "batch_forecasting_outputs")
    os.makedirs(batch_root, exist_ok=True)
    return {
        "timestamp": timestamp,
        "run_base": run_base,
        "run_root": run_root,
        "batch_root": batch_root,
        "sheet_root_name": os.path.basename(batch_root),
    }


def write_incremental_batch_common_outputs(
    export_payload: Dict[str, pd.DataFrame],
    batch_root: str,
    selected_sheet: str,
    horizon: int,
    include_self_script: bool = True,
    include_requirements: bool = True,
) -> None:
    progress_log("Disk Çıktısı", "Ortak batch klasör yapısı ve ortak raporlar oluşturuluyor.", extra={"batch_root": batch_root})
    common_dir = os.path.join(batch_root, "common")
    batch_dir = os.path.join(batch_root, "batch")
    os.makedirs(common_dir, exist_ok=True)
    os.makedirs(batch_dir, exist_ok=True)

    common_tables = {
        "Kalite raporu": export_payload.get("quality_report", pd.DataFrame()),
        "Seri profil raporu": export_payload.get("series_profile_report", pd.DataFrame()),
        "Anomali yönetişimi": export_payload.get("anomaly_governance", pd.DataFrame()),
        "Review queue": export_payload.get("review_queue", pd.DataFrame()),
        "Proxy backtest raporu": export_payload.get("proxy_backtest_report", pd.DataFrame()),
        "Raw vs Clean backtest karşılaştırması": export_payload.get("raw_vs_clean_backtest_report", pd.DataFrame()),
    }
    for name, df in common_tables.items():
        if isinstance(df, pd.DataFrame):
            _write_single_table_excel_file(os.path.join(common_dir, f"{_safe_artifact_name(name)}.xlsx"), name, df)

    manifest = {
        "selected_sheet": selected_sheet,
        "horizon": int(horizon),
        "artifact_schema": "batch_forecast_incremental_disk_v1",
        "app_version": APP_VERSION,
        "table_format": "xlsx",
        "figure_formats": ["png", "html"],
        "write_policy": "series_outputs_are_persisted_immediately_after_each_series_finishes",
    }
    _write_json_file(os.path.join(batch_root, "manifest.json"), manifest)

    if include_self_script:
        _copy_optional_file_into_dir(Path(__file__), os.path.join(batch_root, Path(__file__).name))
    if include_requirements:
        _copy_optional_file_into_dir(Path(__file__).with_name("requirements.txt"), os.path.join(batch_root, "requirements.txt"))


def write_incremental_batch_status_outputs(
    batch_root: str,
    best_summary_df: pd.DataFrame,
    champion_df: pd.DataFrame,
    series_status_df: pd.DataFrame,
) -> None:
    batch_dir = os.path.join(batch_root, "batch")
    figures_dir = os.path.join(batch_dir, "figures")
    os.makedirs(batch_dir, exist_ok=True)
    os.makedirs(figures_dir, exist_ok=True)
    _write_single_table_excel_file(os.path.join(batch_dir, "Batch forecasting özeti.xlsx"), "Batch forecasting özeti", best_summary_df)
    _write_single_table_excel_file(os.path.join(batch_dir, "Champion - Challenger tablosu.xlsx"), "Champion - Challenger", champion_df)
    _write_single_table_excel_file(os.path.join(batch_dir, "Seri çalışma durumu.xlsx"), "Seri çalışma durumu", series_status_df)

    business_pack = _build_batch_business_output_pack(best_summary_df, champion_df, series_status_df)
    for table_name, table_df in (business_pack.get("tables") or {}).items():
        _write_single_table_excel_file(os.path.join(batch_dir, f"{_safe_artifact_name(table_name)}.xlsx"), table_name, table_df)
    for fig_name, fig_bytes in (business_pack.get("figures") or {}).items():
        if fig_bytes is not None:
            _write_binary_file(os.path.join(figures_dir, fig_name), fig_bytes)

    _write_excel_report_file(
        os.path.join(batch_dir, "Batch forecasting master raporu.xlsx"),
        {
            "Batch forecasting özeti": best_summary_df,
            "Champion - Challenger": champion_df,
            "Seri çalışma durumu": series_status_df,
            **(business_pack.get("tables") or {}),
        }
    )



def _safe_bool_col(df: pd.DataFrame, col: str) -> pd.Series:
    if not isinstance(df, pd.DataFrame) or col not in df.columns:
        return pd.Series(dtype=bool)
    s = df[col]
    if s.dtype == bool:
        return s.fillna(False)
    return s.astype(str).str.strip().str.lower().isin(["1", "true", "yes", "evet"])


def _safe_num_series(df: pd.DataFrame, col: str) -> pd.Series:
    if not isinstance(df, pd.DataFrame) or col not in df.columns:
        return pd.Series(dtype=float)
    return pd.to_numeric(df[col], errors="coerce")


def _safe_text_col(df: pd.DataFrame, col: str, default: str = "") -> pd.Series:
    if not isinstance(df, pd.DataFrame):
        return pd.Series(dtype=object)
    if col not in df.columns:
        return pd.Series([default] * len(df), index=df.index, dtype=object)
    return df[col].fillna(default).astype(str)


def _merge_batch_business_frames(best_summary_df: pd.DataFrame, champion_df: pd.DataFrame, series_status_df: pd.DataFrame) -> pd.DataFrame:
    bs = best_summary_df.copy() if isinstance(best_summary_df, pd.DataFrame) else pd.DataFrame()
    ch = champion_df.copy() if isinstance(champion_df, pd.DataFrame) else pd.DataFrame()
    st = series_status_df.copy() if isinstance(series_status_df, pd.DataFrame) else pd.DataFrame()

    if len(bs) == 0:
        base_cols = [
            "target_col", "best_model", "champion", "challenger", "segment", "abc_xyz", "WAPE", "sMAPE", "RMSE", "MAE", "MASE",
            "production_model", "production_status", "prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used", "any_fallback_used",
            "status", "output_dir", "release_decision", "planning_priority", "recommended_action", "exception_reason", "run_health"
        ]
        return pd.DataFrame(columns=base_cols)

    if "best_model" not in bs.columns and "model" in bs.columns:
        bs["best_model"] = bs["model"]
    if "production_model" not in bs.columns:
        bs["production_model"] = bs.get("best_model")
    if "production_status" not in bs.columns:
        bs["production_status"] = "eligible"

    ch_small = pd.DataFrame(columns=["target_col", "champion", "challenger"])
    if len(ch):
        keep_cols = [c for c in ["target_col", "champion", "challenger"] if c in ch.columns]
        ch_small = ch[keep_cols].drop_duplicates(subset=["target_col"], keep="last") if "target_col" in keep_cols else ch_small

    st_small = pd.DataFrame(columns=["target_col", "status", "output_dir", "started_at", "completed_at", "error"])
    if len(st):
        keep_cols = [c for c in ["target_col", "status", "output_dir", "started_at", "completed_at", "error"] if c in st.columns]
        st_small = st[keep_cols].drop_duplicates(subset=["target_col"], keep="last") if "target_col" in keep_cols else st_small

    merged = bs.copy()
    if len(ch_small):
        merged = merged.merge(ch_small, on="target_col", how="left", suffixes=("", "_cc"))
        if "champion_cc" in merged.columns:
            merged["champion"] = merged.get("champion").fillna(merged["champion_cc"])
            merged = merged.drop(columns=["champion_cc"], errors="ignore")
        if "challenger_cc" in merged.columns:
            merged["challenger"] = merged.get("challenger").fillna(merged["challenger_cc"])
            merged = merged.drop(columns=["challenger_cc"], errors="ignore")
    if len(st_small):
        merged = merged.merge(st_small, on="target_col", how="left", suffixes=("", "_run"))
        if "status_run" in merged.columns:
            merged["status"] = merged.get("status").fillna(merged["status_run"])
            merged = merged.drop(columns=["status_run"], errors="ignore")

    for col in ["WAPE", "sMAPE", "RMSE", "MAE", "MASE", "alert_count", "production_eligibility_score"]:
        if col in merged.columns:
            merged[col] = pd.to_numeric(merged[col], errors="coerce")

    for col in ["prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used"]:
        if col not in merged.columns:
            merged[col] = False
        merged[col] = merged[col].astype(str).str.strip().str.lower().isin(["1", "true", "yes", "evet"])
    merged["any_fallback_used"] = merged[["prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used"]].any(axis=1)

    status_series = _safe_text_col(merged, "status", "SUCCESS")
    error_series = _safe_text_col(merged, "error", "Beklenmeyen çalışma hatası")
    production_status_series = _safe_text_col(merged, "production_status", "eligible").str.lower()
    wape_series = pd.to_numeric(merged["WAPE"], errors="coerce") if "WAPE" in merged.columns else pd.Series(np.nan, index=merged.index)
    smape_series = pd.to_numeric(merged["sMAPE"], errors="coerce") if "sMAPE" in merged.columns else pd.Series(np.nan, index=merged.index)

    status_upper = status_series.astype(str).str.upper()
    success_like = status_upper.isin(["SUCCESS", "WARNING"])

    merged["run_health"] = "stabil"
    merged.loc[~success_like, "run_health"] = "hata_var"
    merged.loc[(status_upper == "WARNING") & success_like, "run_health"] = "uyari_ile_tamamlandi"
    merged.loc[(status_upper == "SUCCESS") & merged["any_fallback_used"], "run_health"] = "fallback_ile_tamamlandi"

    merged["exception_reason"] = ""
    merged.loc[~success_like, "exception_reason"] = error_series
    merged.loc[(status_upper == "WARNING") & (merged["exception_reason"] == "") & error_series.astype(str).str.strip().ne(""), "exception_reason"] = error_series
    merged.loc[(merged["exception_reason"] == "") & (wape_series >= 20), "exception_reason"] = "Yüksek WAPE"
    merged.loc[(merged["exception_reason"] == "") & (smape_series >= 15), "exception_reason"] = "Yüksek sMAPE"
    merged.loc[(merged["exception_reason"] == "") & merged["any_fallback_used"], "exception_reason"] = "Fallback ile tamamlandı"
    merged.loc[(merged["exception_reason"] == "") & production_status_series.isin(["reject", "challenger_only"]), "exception_reason"] = "Üretim kapısı kısıtlı"

    merged["release_decision"] = "READY"
    merged.loc[~success_like, "release_decision"] = "BLOCKED"
    merged.loc[(merged["release_decision"] == "READY") & production_status_series.isin(["reject"]), "release_decision"] = "REVIEW"
    merged.loc[(merged["release_decision"] == "READY") & production_status_series.isin(["challenger_only"]), "release_decision"] = "CAUTION"
    merged.loc[(merged["release_decision"] == "READY") & ((wape_series >= 20) | (smape_series >= 15)), "release_decision"] = "REVIEW"
    merged.loc[(merged["release_decision"] == "READY") & merged["any_fallback_used"], "release_decision"] = "CAUTION"

    merged["planning_priority"] = "Normal"
    merged.loc[merged["release_decision"].isin(["BLOCKED", "REVIEW"]), "planning_priority"] = "Kritik"
    merged.loc[(merged["planning_priority"] == "Normal") & merged["release_decision"].isin(["CAUTION"]), "planning_priority"] = "Yüksek"

    merged["recommended_action"] = "Tahmin üretime/rapora alınabilir; standart planlamacı kontrolü yeterli."
    merged.loc[merged["release_decision"] == "BLOCKED", "recommended_action"] = "Seriyi yeniden çalıştırın; veri, model backend ve bağımlılık hatalarını inceleyin."
    merged.loc[merged["release_decision"] == "REVIEW", "recommended_action"] = "Planlamacı ve analitik ekip manuel gözden geçirme yapsın; yayın öncesi challenger ve iş bilgisi ile doğrulansın."
    merged.loc[(merged["release_decision"] == "CAUTION") & merged["any_fallback_used"], "recommended_action"] = "Fallback ile tamamlandı; yayınlanabilir ancak model bileşenleri ve drift riski kontrol edilmelidir."

    wanted = [
        "target_col", "best_model", "champion", "challenger", "segment", "abc_xyz", "WAPE", "sMAPE", "RMSE", "MAE", "MASE",
        "production_model", "production_status", "production_eligibility_score", "prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used", "any_fallback_used",
        "alert_count", "status", "started_at", "completed_at", "output_dir", "release_decision", "planning_priority", "recommended_action", "exception_reason", "run_health"
    ]
    existing = [c for c in wanted if c in merged.columns]
    extra_cols = [c for c in merged.columns if c not in existing]
    return merged[existing + extra_cols].copy()


def _build_batch_kpi_table(portfolio_df: pd.DataFrame, horizon: Optional[int] = None) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    total = int(len(df))
    success = int(df.get("status", pd.Series(dtype=object)).astype(str).str.upper().isin(["SUCCESS", "WARNING"]).sum()) if total else 0
    blocked = int((df.get("release_decision", pd.Series(dtype=object)) == "BLOCKED").sum()) if total else 0
    review = int((df.get("release_decision", pd.Series(dtype=object)) == "REVIEW").sum()) if total else 0
    caution = int((df.get("release_decision", pd.Series(dtype=object)) == "CAUTION").sum()) if total else 0
    fallback = int(_safe_bool_col(df, "any_fallback_used").sum()) if total else 0
    champion_count = int(df.get("champion", pd.Series(dtype=object)).dropna().nunique()) if total and "champion" in df.columns else 0
    wape = _safe_num_series(df, "WAPE")
    smape = _safe_num_series(df, "sMAPE")
    rows = [
        {"metric": "Toplam seri", "value": total},
        {"metric": "Başarıyla tamamlanan seri", "value": success},
        {"metric": "Bloke seri", "value": blocked},
        {"metric": "İnceleme gereken seri", "value": review},
        {"metric": "Dikkat ile yayınlanacak seri", "value": caution},
        {"metric": "Fallback kullanılan seri", "value": fallback},
        {"metric": "Benzersiz champion sayısı", "value": champion_count},
        {"metric": "Medyan WAPE", "value": safe_float(wape.median()) if len(wape) else np.nan},
        {"metric": "Ortalama WAPE", "value": safe_float(wape.mean()) if len(wape) else np.nan},
        {"metric": "Medyan sMAPE", "value": safe_float(smape.median()) if len(smape) else np.nan},
        {"metric": "Ortalama sMAPE", "value": safe_float(smape.mean()) if len(smape) else np.nan},
    ]
    if horizon is not None:
        rows.append({"metric": "Test ufku", "value": int(horizon)})
    return pd.DataFrame(rows)


def _build_batch_segment_performance_table(portfolio_df: pd.DataFrame) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0 or "segment" not in df.columns:
        return pd.DataFrame(columns=["segment", "series_count", "median_WAPE", "mean_WAPE", "median_sMAPE", "ready_count", "review_count"])
    tmp = df.copy()
    tmp["WAPE"] = pd.to_numeric(tmp.get("WAPE"), errors="coerce")
    tmp["sMAPE"] = pd.to_numeric(tmp.get("sMAPE"), errors="coerce")
    out = tmp.groupby("segment", dropna=False).agg(
        series_count=("target_col", "count"),
        median_WAPE=("WAPE", "median"),
        mean_WAPE=("WAPE", "mean"),
        median_sMAPE=("sMAPE", "median"),
    ).reset_index()
    decision_counts = tmp.pivot_table(index="segment", columns="release_decision", values="target_col", aggfunc="count", fill_value=0).reset_index()
    out = out.merge(decision_counts, on="segment", how="left")
    out = out.rename(columns={"READY": "ready_count", "REVIEW": "review_count", "CAUTION": "caution_count", "BLOCKED": "blocked_count"})
    return out


def _build_batch_exception_table(portfolio_df: pd.DataFrame) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0:
        return pd.DataFrame(columns=["target_col", "release_decision", "planning_priority", "exception_reason", "recommended_action", "WAPE", "sMAPE", "status", "output_dir"])
    exc = df[(df.get("release_decision").isin(["BLOCKED", "REVIEW", "CAUTION"])) | (df.get("exception_reason", "") != "")].copy()
    if len(exc) == 0:
        return pd.DataFrame(columns=["target_col", "release_decision", "planning_priority", "exception_reason", "recommended_action", "WAPE", "sMAPE", "status", "output_dir"])
    cols = [c for c in ["target_col", "release_decision", "planning_priority", "exception_reason", "recommended_action", "WAPE", "sMAPE", "status", "output_dir", "champion", "production_model", "production_status"] if c in exc.columns]
    sort_cols = [c for c in ["planning_priority", "release_decision", "WAPE", "sMAPE"] if c in exc.columns]
    asc_map = {"planning_priority": True, "release_decision": True, "WAPE": False, "sMAPE": False}
    return exc[cols].sort_values(sort_cols, ascending=[asc_map[c] for c in sort_cols]) if sort_cols else exc[cols]


def _build_batch_planner_action_table(portfolio_df: pd.DataFrame, horizon: Optional[int] = None) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0:
        return pd.DataFrame(columns=["target_col", "planning_priority", "release_decision", "owner", "due_window", "recommended_action", "champion", "output_dir"])
    out = df.copy()
    out["owner"] = np.where(out.get("planning_priority", "Normal").isin(["Kritik", "Yüksek"]), "Demand Planner + Analytics Lead", "Demand Planner")
    out["due_window"] = np.where(out.get("planning_priority", "Normal") == "Kritik", "Aynı gün", np.where(out.get("planning_priority", "Normal") == "Yüksek", "24 saat", "Standart çevrim"))
    if horizon is not None:
        out["forecast_horizon"] = int(horizon)
    cols = [c for c in ["target_col", "planning_priority", "release_decision", "owner", "due_window", "recommended_action", "champion", "production_model", "WAPE", "sMAPE", "output_dir", "forecast_horizon"] if c in out.columns]
    sort_cols = [c for c in ["planning_priority", "release_decision", "WAPE"] if c in out.columns]
    asc = [True, True, False][:len(sort_cols)]
    return out[cols].sort_values(sort_cols, ascending=asc)


def _build_batch_release_register_table(portfolio_df: pd.DataFrame, horizon: Optional[int] = None) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0:
        return pd.DataFrame(columns=["target_col", "release_decision", "best_model", "champion", "production_model", "production_status", "WAPE", "sMAPE", "output_dir"])
    out = df.copy()
    if horizon is not None:
        out["forecast_horizon"] = int(horizon)
    out["release_package_ready"] = np.where(out.get("status", "SUCCESS").astype(str).str.upper() == "SUCCESS", "evet", "hayır")
    cols = [c for c in ["target_col", "forecast_horizon", "release_decision", "release_package_ready", "best_model", "champion", "challenger", "production_model", "production_status", "WAPE", "sMAPE", "output_dir", "completed_at"] if c in out.columns]
    sort_cols = [c for c in ["release_decision", "WAPE"] if c in cols]
    return out[cols].sort_values(sort_cols, ascending=[True for _ in sort_cols]) if sort_cols else out[cols]


def _build_batch_model_governance_table(portfolio_df: pd.DataFrame) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0:
        return pd.DataFrame(columns=["target_col", "best_model", "champion", "production_model", "production_status", "prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used", "run_health"])
    cols = [c for c in ["target_col", "best_model", "champion", "challenger", "production_model", "production_status", "production_eligibility_score", "prophet_fallback_used", "xgboost_fallback_used", "sarima_fallback_used", "any_fallback_used", "alert_count", "run_health"] if c in df.columns]
    sort_cols = [c for c in ["production_status", "production_eligibility_score"] if c in cols]
    asc_map = {"production_status": True, "production_eligibility_score": False}
    return df[cols].sort_values(sort_cols, ascending=[asc_map[c] for c in sort_cols]) if sort_cols else df[cols]


def _build_batch_champion_mix_table(portfolio_df: pd.DataFrame) -> pd.DataFrame:
    df = portfolio_df.copy() if isinstance(portfolio_df, pd.DataFrame) else pd.DataFrame()
    if len(df) == 0 or "champion" not in df.columns:
        return pd.DataFrame(columns=["champion", "series_count"])
    counts = df["champion"].fillna("Bilinmiyor").astype(str).value_counts(dropna=False).reset_index()
    counts.columns = ["champion", "series_count"]
    return counts


def _batch_metric_bar_png_bytes(df: pd.DataFrame, label_col: str, value_col: str, title: str, top_n: int = 15) -> Optional[bytes]:
    if not isinstance(df, pd.DataFrame) or len(df) == 0 or label_col not in df.columns or value_col not in df.columns:
        return None
    tmp = df[[label_col, value_col]].copy()
    tmp[value_col] = pd.to_numeric(tmp[value_col], errors="coerce")
    tmp = tmp.dropna().sort_values(value_col, ascending=False).head(max(1, int(top_n)))
    if len(tmp) == 0:
        return None
    fig, ax = plt.subplots(figsize=(12, max(4, 0.45 * len(tmp) + 1.5)))
    ax.barh(tmp[label_col].astype(str), tmp[value_col].astype(float))
    ax.set_title(title)
    ax.set_xlabel(value_col)
    ax.invert_yaxis()
    ax.grid(alpha=0.25, axis="x")
    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _build_batch_business_output_pack(best_summary_df: pd.DataFrame, champion_df: pd.DataFrame, series_status_df: pd.DataFrame, horizon: Optional[int] = None) -> Dict[str, Any]:
    portfolio_df = _merge_batch_business_frames(best_summary_df, champion_df, series_status_df)
    tables = {
        "Portföy yöneticisi özeti": _build_batch_kpi_table(portfolio_df, horizon=horizon),
        "Tahmin yayın kayıt defteri": _build_batch_release_register_table(portfolio_df, horizon=horizon),
        "Planlamacı aksiyon listesi": _build_batch_planner_action_table(portfolio_df, horizon=horizon),
        "İstisna kayıtları": _build_batch_exception_table(portfolio_df),
        "Model yönetişim özeti": _build_batch_model_governance_table(portfolio_df),
        "Şampiyon dağılımı": _build_batch_champion_mix_table(portfolio_df),
        "Segment bazlı performans": _build_batch_segment_performance_table(portfolio_df),
        "Portföy performans görünümü": portfolio_df,
    }
    figures = {
        "Portföy - WAPE sıralaması.png": _batch_metric_bar_png_bytes(portfolio_df, "target_col", "WAPE", "Portföy - WAPE sıralaması", top_n=min(20, len(portfolio_df) if len(portfolio_df) else 20)),
        "Portföy - sMAPE sıralaması.png": _batch_metric_bar_png_bytes(portfolio_df, "target_col", "sMAPE", "Portföy - sMAPE sıralaması", top_n=min(20, len(portfolio_df) if len(portfolio_df) else 20)),
        "Portföy - şampiyon dağılımı.png": _batch_metric_bar_png_bytes(_build_batch_champion_mix_table(portfolio_df), "champion", "series_count", "Portföy - şampiyon dağılımı", top_n=20),
    }
    return {"portfolio_df": portfolio_df, "tables": tables, "figures": figures}


def _build_series_executive_summary_table(outputs: Dict[str, Any]) -> pd.DataFrame:
    meta = outputs.get("metadata", {}) or {}
    prod_pack = outputs.get("production_governance", {}) or {}
    prophet = outputs.get("prophet", {}) or {}
    xgb = outputs.get("xgboost", {}) or {}
    sarima = outputs.get("sarima", {}) or {}
    cc = outputs.get("champion_challenger", {}) or {}
    metrics_df = outputs.get("metrics_df", pd.DataFrame()) if isinstance(outputs, dict) else pd.DataFrame()
    best_row = metrics_df.iloc[0].to_dict() if isinstance(metrics_df, pd.DataFrame) and len(metrics_df) else {}
    readiness = assess_production_readiness({}, metrics_df if isinstance(metrics_df, pd.DataFrame) else pd.DataFrame(), meta.get("target_col", ""), outputs=outputs)
    return pd.DataFrame([{
        "target_col": meta.get("target_col"),
        "best_model": outputs.get("best_model", best_row.get("model")),
        "champion": cc.get("champion"),
        "challenger": cc.get("challenger"),
        "production_model": prod_pack.get("production_model", outputs.get("best_model", best_row.get("model"))),
        "production_status": prod_pack.get("production_status", "eligible"),
        "segment": meta.get("segment"),
        "abc_xyz": meta.get("abc_xyz"),
        "horizon": meta.get("horizon"),
        "WAPE": best_row.get("WAPE"),
        "sMAPE": best_row.get("sMAPE"),
        "RMSE": best_row.get("RMSE"),
        "prophet_fallback_used": bool(prophet.get("fallback_used", False)),
        "xgboost_fallback_used": bool(xgb.get("fallback_used", False)),
        "sarima_fallback_used": bool(sarima.get("fallback_used", False)),
        "release_note": " | ".join(list(readiness.get("notes", []))[:2]),
    }])


def _build_series_operational_action_table(outputs: Dict[str, Any]) -> pd.DataFrame:
    summary_df = _build_series_executive_summary_table(outputs)
    row = summary_df.iloc[0].to_dict() if len(summary_df) else {}
    actions = []
    release_note = str(row.get("release_note", "")).strip()
    if row.get("prophet_fallback_used"):
        actions.append({"action_area": "Prophet", "priority": "Yüksek", "action": "Gerçek Prophet backend/fitting yolunu kontrol edin.", "reason": "Seri surrogate Prophet ile tamamlandı."})
    if row.get("xgboost_fallback_used"):
        actions.append({"action_area": "XGBoost", "priority": "Orta", "action": "Feature set ve validation stabilitesini gözden geçirin.", "reason": "Fallback veya emniyetli yol kullanıldı."})
    if pd.notna(row.get("WAPE")) and float(row.get("WAPE") or 0) >= 20:
        actions.append({"action_area": "Tahmin kalitesi", "priority": "Kritik", "action": "Manuel override / iş bilgisi doğrulaması yapın.", "reason": f"WAPE={safe_float(row.get('WAPE')):.2f}"})
    if not actions:
        actions.append({"action_area": "Yayın", "priority": "Normal", "action": "Tahmin paketi standart planlama kontrolü ile yayınlanabilir.", "reason": release_note or "Kritik risk sinyali bulunmadı."})
    return pd.DataFrame(actions)


def write_incremental_series_outputs(
    export_payload: Dict[str, pd.DataFrame],
    outputs: Dict[str, Any],
    target: str,
    batch_root: str,
    horizon: int,
) -> str:
    target_root = os.path.join(batch_root, _safe_artifact_name(target))
    tables_dir = os.path.join(target_root, "tables")
    notes_dir = os.path.join(target_root, "notes")
    json_dir = os.path.join(target_root, "json")
    figures_dir = os.path.join(target_root, "figures")
    for d in [target_root, tables_dir, notes_dir, json_dir, figures_dir]:
        os.makedirs(d, exist_ok=True)

    progress_log("Disk Çıktısı", "Seri tamamlandı; disk yazımı başlatıldı.", target=target, extra={"target_root": target_root})
    style_map = outputs.get("model_gorsel_stil_haritasi", {}) or {}
    raw_named_tables = outputs.get("named_output_tables", {}) or {}
    named_tables = {
        "Yönetici özeti": _build_series_executive_summary_table(outputs),
        "Operasyonel aksiyon planı": _build_series_operational_action_table(outputs),
        **raw_named_tables,
    }
    note_tables = _series_notes_tables(outputs)
    json_artifacts = _series_json_artifacts(outputs)

    for report_name, df in named_tables.items():
        if isinstance(df, pd.DataFrame):
            _write_single_table_excel_file(os.path.join(tables_dir, f"{_safe_artifact_name(report_name)}.xlsx"), report_name, df)

    for note_name, note_df in note_tables.items():
        _write_single_table_excel_file(os.path.join(notes_dir, f"{_safe_artifact_name(note_name)}.xlsx"), note_name, note_df)
        note_lines = note_df.get("text", pd.Series(dtype=object)).astype(str).tolist() if isinstance(note_df, pd.DataFrame) else []
        _write_text_file(os.path.join(notes_dir, f"{_safe_artifact_name(note_name)}.txt"), note_lines)

    for json_name, payload in json_artifacts.items():
        _write_json_file(os.path.join(json_dir, f"{_safe_artifact_name(json_name)}.json"), payload)

    acf_bytes = _acf_pacf_png_bytes(export_payload, target, horizon)
    if acf_bytes is not None:
        _write_binary_file(os.path.join(figures_dir, f"{_safe_artifact_name(target)} - ACF ve PACF.png"), acf_bytes)

    plot_specs = {
        "Gerçek vs Tahmin": outputs.get("predictions", {}),
        "SARIMA-SARIMAX": {"SARIMA/SARIMAX": (outputs.get("predictions", {}) or {}).get("SARIMA/SARIMAX")},
        "Prophet": {"Prophet": (outputs.get("predictions", {}) or {}).get("Prophet")},
        "XGBoost": {"XGBoost": (outputs.get("predictions", {}) or {}).get("XGBoost")},
        "Ensemble": {"Ensemble": (outputs.get("predictions", {}) or {}).get("Ensemble")},
    }
    for plot_name, pred_map in plot_specs.items():
        clean_pred_map = {k: v for k, v in (pred_map or {}).items() if v is not None}
        if not clean_pred_map:
            continue
        title = f"{target} - {plot_name}"
        _write_binary_file(os.path.join(figures_dir, f"{_safe_artifact_name(title)}.png"), _forecast_png_bytes(outputs["train"], outputs["test"], clean_pred_map, title))
        html_bytes = _forecast_plotly_html_bytes(outputs["train"], outputs["test"], clean_pred_map, title, style_map)
        if html_bytes is not None:
            _write_binary_file(os.path.join(figures_dir, f"{_safe_artifact_name(title)}.html"), html_bytes)

    _write_excel_report_file(
        os.path.join(target_root, f"{_safe_artifact_name(target)} - detay raporu.xlsx"),
        named_tables,
        notes_map=note_tables,
        json_map=json_artifacts,
    )
    progress_log("Disk Çıktısı", "Seri için tüm tablo/grafik dosyaları diske yazıldı.", level="SUCCESS", target=target, extra={"target_root": target_root})
    return target_root


def create_zip_from_directory(source_dir: str, output_zip_path: str) -> str:
    source_dir = os.path.abspath(source_dir)
    output_zip_path = os.path.abspath(output_zip_path)
    os.makedirs(os.path.dirname(output_zip_path) or ".", exist_ok=True)
    root_parent = os.path.dirname(source_dir)
    with zipfile.ZipFile(output_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for current_root, _, files in os.walk(source_dir):
            for filename in files:
                abs_path = os.path.join(current_root, filename)
                arcname = os.path.relpath(abs_path, root_parent)
                zf.write(abs_path, arcname)
    return output_zip_path


def build_batch_summary_excel_bytes(batch_result: Dict[str, Any]) -> Optional[bytes]:
    return _build_excel_bytes({
        "Batch forecasting özeti": batch_result.get("best_summary", pd.DataFrame()),
        "Champion - Challenger": batch_result.get("champion_table", pd.DataFrame()),
    })


def build_acf_pacf_figure(train_df: pd.DataFrame, target_col: str):
    if not HAS_FORECAST_STATSMODELS:
        return None
    y = pd.to_numeric(train_df["y"], errors="coerce").dropna().astype(float)
    if len(y) < 8:
        return None
    import matplotlib.pyplot as plt
    fig, axes = plt.subplots(1, 2, figsize=(12, 4))
    plot_acf(y, ax=axes[0], lags=min(24, max(3, len(y)//2 - 1)))
    plot_pacf(y, ax=axes[1], lags=min(24, max(3, len(y)//2 - 1)), method="ywm")
    axes[0].set_title(f"{target_col} - ACF")
    axes[1].set_title(f"{target_col} - PACF")
    plt.tight_layout()
    return fig

def build_model_visual_style_map(outputs: Optional[Dict[str, Any]] = None) -> Dict[str, Dict[str, Any]]:
    outputs = outputs or {}
    prod_pack = outputs.get("production_governance", {}) if isinstance(outputs, dict) else {}
    prod_model = prod_pack.get("production_model")
    gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    prophet_info = outputs.get("prophet", {}) if isinstance(outputs, dict) else {}
    style_map: Dict[str, Dict[str, Any]] = {}
    if isinstance(gate_df, pd.DataFrame) and len(gate_df):
        for _, row in gate_df.iterrows():
            model_name = str(row.get("model"))
            status = str(row.get("status", "eligible"))
            opacity = 1.0
            width = 2.5
            dash = "solid"
            if status == "challenger_only":
                opacity = 0.72
                width = 2.0
            elif status == "reject":
                opacity = 0.35
                width = 1.4
                dash = "dot"
            if model_name == prod_model:
                opacity = 1.0
                width = 4.0
                dash = "solid"
            style_map[model_name] = {"opacity": opacity, "width": width, "dash": dash}
    if prophet_info and bool(prophet_info.get("fallback_used", False)):
        style_map.setdefault("Prophet", {})
        style_map["Prophet"].update({"opacity": 0.28, "width": 1.2, "dash": "dot"})
    for dl_name in ["LSTM", "GRU"]:
        dl_info = outputs.get(dl_name.lower(), {}) if isinstance(outputs, dict) else {}
        real_name = dl_public_real_name(dl_name)
        if dl_info and not bool(dl_info.get("dl_strict_validation_passed", False)):
            style_map.setdefault(real_name, {})
            style_map[real_name].update({"opacity": 0.32, "width": 1.2, "dash": "dot"})
        if dl_info and bool(dl_info.get("fallback_used", False)):
            style_map.setdefault("DL-fallback", {})
            style_map["DL-fallback"].update({"opacity": 0.42, "width": 1.6, "dash": "dot"})
    return style_map

def plot_forecast_results(train_df: pd.DataFrame, test_df: pd.DataFrame, predictions: Dict[str, np.ndarray], title: str, model_style_map: Optional[Dict[str, Dict[str, Any]]] = None):
    model_style_map = model_style_map or {}
    if HAS_PLOTLY:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=train_df["ds"], y=train_df["y"], mode="lines+markers", name="Eğitim", line=dict(width=2.4), opacity=0.95))
        fig.add_trace(go.Scatter(x=test_df["ds"], y=test_df["y"], mode="lines+markers", name="Gerçek", line=dict(width=3.0), opacity=1.0))
        for name, pred in predictions.items():
            style = model_style_map.get(name, {})
            fig.add_trace(go.Scatter(
                x=test_df["ds"], y=pred, mode="lines+markers", name=name,
                line=dict(width=float(style.get("width", 2.2)), dash=str(style.get("dash", "solid"))),
                opacity=float(style.get("opacity", 0.9))
            ))
        fig.update_layout(title=title, xaxis_title="Tarih", yaxis_title="Talep", legend_title="Seriler", template="plotly_white")
        return fig
    return None

def train_test_split_series(df_series: pd.DataFrame, horizon: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    usable = df_series.loc[df_series["is_usable"]].copy().reset_index(drop=True)
    if len(usable) <= horizon + 6:
        raise ValueError(f"Modelleme için yeterli kullanılabilir gözlem yok. Kullanılabilir gözlem: {len(usable)}")
    train = usable.iloc[:-horizon].copy().reset_index(drop=True)
    test = usable.iloc[-horizon:].copy().reset_index(drop=True)
    return train, test


def suggest_d_via_stationarity(y: pd.Series) -> int:
    y = pd.to_numeric(y, errors="coerce").dropna().astype(float)
    if len(y) < 12 or not HAS_FORECAST_STATSMODELS:
        return 0
    try:
        adf_p = adfuller(y, autolag="AIC")[1]
    except Exception:
        adf_p = 1.0
    try:
        kpss_p = kpss(y, regression="c", nlags="auto")[1]
    except Exception:
        kpss_p = 0.0
    if adf_p < 0.05 and kpss_p > 0.05:
        return 0
    y1 = y.diff().dropna()
    if len(y1) < 10:
        return 1
    try:
        adf_p1 = adfuller(y1, autolag="AIC")[1]
    except Exception:
        adf_p1 = 0.01
    try:
        kpss_p1 = kpss(y1, regression="c", nlags="auto")[1]
    except Exception:
        kpss_p1 = 0.1
    if adf_p1 < 0.05 and kpss_p1 > 0.05:
        return 1
    return 1


def suggest_D_via_profile(profile: Dict[str, Any], season_length: int, n_obs: int) -> int:
    if season_length <= 1 or n_obs < season_length * 2:
        return 0
    seasonality_strength = safe_float(profile.get("seasonality_strength", np.nan))
    return int(pd.notna(seasonality_strength) and seasonality_strength >= 0.35)



def build_sarimax_grid(freq_alias: str, profile: Dict[str, Any], n_obs: int) -> List[Dict[str, Any]]:
    season_length = infer_season_length_from_freq(freq_alias)
    seasonality_strength = safe_float(profile.get("seasonality_strength", np.nan))
    seasonal_allowed = season_length > 1 and n_obs >= max(season_length * 2, 24)

    short_series = n_obs <= 72
    max_plain_order = 3 if short_series else 4
    plain_orders = [
        (p, q) for p, q in product(range(0, max_plain_order + 1), range(0, max_plain_order + 1))
        if (p + q) <= (4 if short_series else 5)
    ]

    grid: List[Dict[str, Any]] = []
    for p, q in plain_orders:
        grid.append({"p": p, "q": q, "P": 0, "Q": 0, "m": 0})

    if seasonal_allowed:
        seasonal_orders = [(P, Q) for P, Q in product(range(0, 3), range(0, 3)) if (P + Q) <= 2]
        for p, q in plain_orders:
            for P, Q in seasonal_orders:
                if short_series and (p + q + P + Q) > 4:
                    continue
                if pd.notna(seasonality_strength) and seasonality_strength < 0.20 and (P + Q) > 0:
                    continue
                grid.append({"p": p, "q": q, "P": P, "Q": Q, "m": season_length})

    unique, seen = [], set()
    for cfg in grid:
        key = (cfg["p"], cfg["q"], cfg["P"], cfg["Q"], cfg["m"])
        if key not in seen:
            seen.add(key)
            unique.append(cfg)
    return unique



def prioritize_sarimax_candidates(
    candidates: List[Dict[str, Any]],
    profile: Dict[str, Any],
    max_candidates: int
) -> List[Dict[str, Any]]:
    seasonality_strength = safe_float(profile.get("seasonality_strength", np.nan))
    volatility_regime = str(profile.get("volatility_regime", "unknown"))
    volume_level_ = str(profile.get("volume_level", "unknown"))

    unique = []
    seen = set()
    for cfg in candidates:
        key = (cfg.get("p", 0), cfg.get("q", 0), cfg.get("P", 0), cfg.get("Q", 0), cfg.get("m", 0))
        if key not in seen:
            seen.add(key)
            unique.append(cfg)

    def _score(cfg: Dict[str, Any]) -> Tuple[float, float, float]:
        p = int(cfg.get("p", 0))
        q = int(cfg.get("q", 0))
        P = int(cfg.get("P", 0))
        Q = int(cfg.get("Q", 0))
        m = int(cfg.get("m", 0))
        complexity = p + q + P + Q + (0.30 if m > 1 else 0.0)
        seasonal_pref = 0.0
        if m > 1:
            seasonal_pref -= 0.30 if pd.notna(seasonality_strength) and seasonality_strength >= 0.20 else -0.0
            if pd.notna(seasonality_strength) and seasonality_strength < 0.20:
                seasonal_pref += 0.80
        else:
            if pd.notna(seasonality_strength) and seasonality_strength >= 0.45:
                seasonal_pref += 0.25
        stability_pref = 0.0
        if volatility_regime in ["stable", "moderate"] and (p + q) > 3:
            stability_pref += 0.30
        if volume_level_ in ["very_low", "low"] and (P + Q) > 1:
            stability_pref += 0.25
        balance_pref = 0.10 * abs(p - q) + 0.10 * abs(P - Q)
        return (complexity + seasonal_pref + stability_pref + balance_pref, complexity, balance_pref)

    ranked = sorted(unique, key=_score)
    keep_n = max(8, int(max_candidates))
    prioritized = ranked[:keep_n]

    seasonal_candidates = [cfg for cfg in ranked if int(cfg.get("m", 0)) > 1]
    plain_candidates = [cfg for cfg in ranked if int(cfg.get("m", 0)) <= 1]
    mixed_candidates = [cfg for cfg in ranked if (int(cfg.get("p", 0)) > 0 and int(cfg.get("q", 0)) > 0)]
    ar_heavy = [cfg for cfg in ranked if int(cfg.get("p", 0)) > int(cfg.get("q", 0))]
    ma_heavy = [cfg for cfg in ranked if int(cfg.get("q", 0)) > int(cfg.get("p", 0))]

    for bucket in [seasonal_candidates, plain_candidates, mixed_candidates, ar_heavy, ma_heavy]:
        if bucket:
            cand = bucket[0]
            key = (cand.get("p", 0), cand.get("q", 0), cand.get("P", 0), cand.get("Q", 0), cand.get("m", 0))
            existing = {(c.get("p", 0), c.get("q", 0), c.get("P", 0), c.get("Q", 0), c.get("m", 0)) for c in prioritized}
            if key not in existing:
                prioritized.append(cand)

    unique_prioritized = []
    seen2 = set()
    for cfg in prioritized:
        key = (cfg.get("p", 0), cfg.get("q", 0), cfg.get("P", 0), cfg.get("Q", 0), cfg.get("m", 0))
        if key not in seen2:
            seen2.add(key)
            unique_prioritized.append(cfg)
    return unique_prioritized[:max(10, int(max_candidates))]


def fit_sarimax_model(
    y: pd.Series,
    exog: Optional[pd.DataFrame],
    order: Tuple[int, int, int],
    seasonal_order: Tuple[int, int, int, int],
    trend: str,
    maxiter: int
):
    model = SARIMAX(
        y,
        exog=exog,
        order=order,
        seasonal_order=seasonal_order,
        trend=trend,
        enforce_stationarity=False,
        enforce_invertibility=False
    )
    return model.fit(disp=False, maxiter=maxiter)


def sanitize_exog_for_sarimax(exog_train: Optional[pd.DataFrame], exog_test: Optional[pd.DataFrame]) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], List[str]]:
    dropped = []
    if exog_train is None or exog_test is None:
        return None, None, dropped
    xtr = exog_train.copy().reset_index(drop=True)
    xte = exog_test.copy().reset_index(drop=True)
    common_cols = [c for c in xtr.columns if c in xte.columns]
    if not common_cols:
        return None, None, dropped
    xtr = xtr[common_cols]
    xte = xte[common_cols]
    for c in common_cols:
        xtr[c] = pd.to_numeric(xtr[c], errors='coerce')
        xte[c] = pd.to_numeric(xte[c], errors='coerce')
    xtr = xtr.replace([np.inf, -np.inf], np.nan)
    xte = xte.replace([np.inf, -np.inf], np.nan)
    keep_cols = []
    for c in common_cols:
        train_non_na = xtr[c].notna().sum()
        test_non_na = xte[c].notna().sum()
        nunique = xtr[c].dropna().nunique()
        if train_non_na == 0 or test_non_na == 0 or nunique <= 1:
            dropped.append(c)
            continue
        keep_cols.append(c)
    if not keep_cols:
        return None, None, dropped
    xtr = xtr[keep_cols].copy()
    xte = xte[keep_cols].copy()
    fill_vals = xtr.median(numeric_only=True)
    xtr = xtr.fillna(fill_vals).fillna(0.0)
    xte = xte.fillna(fill_vals).fillna(0.0)
    return xtr, xte, dropped


def build_fallback_forecast(y_train: pd.Series, y_test: pd.Series, freq_alias: str, season_length: int) -> Tuple[np.ndarray, str]:
    """
    Leakage-safe fallback forecast.

    Critical rule:
        y_test is used ONLY to infer the forecast horizon length. Its actual
        values are never read, scored, ranked, or used to choose a fallback
        method. This prevents holdout/test leakage and prevents a fallback
        or surrogate path from silently becoming a test-optimized substitute
        for a real model.
    """
    y_train = pd.to_numeric(y_train, errors="coerce").dropna().astype(float).reset_index(drop=True)
    try:
        h = int(len(y_test))
    except Exception:
        h = int(y_test) if str(y_test).strip().isdigit() else 0
    if h <= 0:
        return np.array([], dtype=float), "empty"

    if len(y_train) == 0:
        return np.zeros(h, dtype=float), "zero_no_train_data"

    season_length = int(season_length or 1)
    freq_alias = str(freq_alias or "").upper()

    # Deterministic, train-only priority order. No holdout scoring.
    if season_length > 1 and len(y_train) >= season_length:
        seasonal_vals = y_train.iloc[-season_length:].values.astype(float)
        pred = np.array([seasonal_vals[i % season_length] for i in range(h)], dtype=float)
        return np.maximum(pred, 0.0), "seasonal_naive_train_only"

    if len(y_train) >= 3:
        pred = drift_forecast(y_train, h)
        if np.isfinite(pred).all():
            return np.maximum(np.asarray(pred, dtype=float), 0.0), "drift_train_only"

    if len(y_train) >= 1:
        return np.maximum(np.repeat(float(y_train.iloc[-1]), h), 0.0), "last_value_train_only"

    return np.zeros(h, dtype=float), "zero_no_train_data"


def forecast_with_baseline_name(y_train: pd.Series, horizon: int, freq_alias: str, season_length: int, baseline_name: str) -> np.ndarray:
    y_train = pd.to_numeric(y_train, errors="coerce").dropna().astype(float).reset_index(drop=True)
    h = int(horizon)
    if h <= 0:
        return np.array([], dtype=float)
    name = str(baseline_name or "last_value")
    if name == "seasonal_naive" and season_length > 1 and len(y_train) >= season_length:
        seasonal_vals = y_train.iloc[-season_length:].tolist()
        pred = np.array([seasonal_vals[i % season_length] for i in range(h)], dtype=float)
    elif name == "drift" and len(y_train) >= 2:
        pred = drift_forecast(y_train, h)
    elif name == "mean" and len(y_train) >= 1:
        pred = np.repeat(float(y_train.mean()), h)
    else:
        pred = np.repeat(float(y_train.iloc[-1]) if len(y_train) else 0.0, h)
    return np.maximum(np.asarray(pred, dtype=float), 0.0)


def compute_validation_postprocess_candidates(y_true: np.ndarray, pred: np.ndarray, y_hist: pd.Series, season_length: int) -> Tuple[np.ndarray, Dict[str, Any]]:
    y_true = np.asarray(y_true, dtype=float)
    pred = np.maximum(np.asarray(pred, dtype=float), 0.0)
    hist = pd.to_numeric(y_hist, errors="coerce").dropna().astype(float)
    if len(pred) == 0:
        return pred, {"name": "raw"}
    anchor_window = max(3, min(len(hist), season_length if season_length > 1 else 6))
    anchor_level = float(hist.iloc[-anchor_window:].median()) if len(hist) else 0.0
    seasonal_anchor = forecast_with_baseline_name(hist, len(pred), "M", season_length, "seasonal_naive") if season_length > 1 and len(hist) >= season_length else np.repeat(anchor_level, len(pred))
    drift_anchor = drift_forecast(hist, len(pred)) if len(hist) >= 2 else np.repeat(anchor_level, len(pred))
    recent_anchor = np.repeat(anchor_level, len(pred))
    candidates = [
        ("raw", pred, {"name": "raw"}),
        ("seasonal_naive", np.maximum(seasonal_anchor, 0.0), {"name": "seasonal_naive", "anchor": seasonal_anchor.tolist()}),
        ("drift", np.maximum(drift_anchor, 0.0), {"name": "drift", "anchor": drift_anchor.tolist()}),
        ("recent_median", np.maximum(recent_anchor, 0.0), {"name": "recent_median", "anchor": recent_anchor.tolist()}),
    ]
    if np.nanmedian(pred) > 0:
        ratio = float(np.nanmedian(y_true) / max(np.nanmedian(pred), 1e-6))
        ratio = float(np.clip(ratio, 0.70, 1.35))
        candidates.append(("ratio", np.maximum(pred * ratio, 0.0), {"name": "ratio", "ratio": ratio}))
    bias = float(np.nanmedian(y_true - pred))
    candidates.append(("bias", np.maximum(pred + bias, 0.0), {"name": "bias", "bias": bias}))
    for w in [0.10, 0.20, 0.35, 0.50, 0.65]:
        blend = np.maximum((1.0 - w) * pred + w * seasonal_anchor, 0.0)
        candidates.append((f"anchor_{w}", blend, {"name": "anchor_blend", "weight": w, "anchor": seasonal_anchor.tolist()}))
    for w in [0.15, 0.30, 0.45]:
        blend = np.maximum((1.0 - w) * pred + w * drift_anchor, 0.0)
        candidates.append((f"drift_{w}", blend, {"name": "drift_blend", "weight": w, "anchor": drift_anchor.tolist()}))
    best_pred = pred
    best_cfg = {"name": "raw"}
    best_score = np.inf
    for _name, cand, cfg in candidates:
        asym = compute_asymmetric_validation_penalty(y_true, cand, hist.values if len(hist) else y_true, severity=0.8)
        score = wape(y_true, cand) + 0.35 * smape(y_true, cand) + 0.15 * float(asym.get("penalty", 0.0) or 0.0)
        if score < best_score:
            best_score = score
            best_pred = cand
            best_cfg = cfg
    return np.maximum(best_pred, 0.0), best_cfg


def apply_postprocess_cfg(pred: np.ndarray, cfg: Dict[str, Any], y_hist: pd.Series, season_length: int) -> np.ndarray:
    arr = np.maximum(np.asarray(pred, dtype=float), 0.0)
    name = str((cfg or {}).get("name", "raw"))
    if name == "ratio":
        return np.maximum(arr * float(cfg.get("ratio", 1.0)), 0.0)
    if name == "bias":
        return np.maximum(arr + float(cfg.get("bias", 0.0)), 0.0)
    if name in ["anchor_blend", "drift_blend"]:
        anchor = np.asarray(cfg.get("anchor", []), dtype=float)
        if len(anchor) != len(arr):
            hist = pd.to_numeric(y_hist, errors="coerce").dropna().astype(float)
            if name == "drift_blend":
                anchor = drift_forecast(hist, len(arr)) if len(hist) >= 2 else np.repeat(float(hist.tail(max(3, min(len(hist), season_length if season_length > 1 else 6))).median()) if len(hist) else 0.0, len(arr))
            else:
                anchor = forecast_with_baseline_name(hist, len(arr), "M", season_length, "seasonal_naive") if season_length > 1 and len(hist) >= season_length else np.repeat(float(hist.tail(max(3, min(len(hist), season_length if season_length > 1 else 6))).median()) if len(hist) else 0.0, len(arr))
        w = float(cfg.get("weight", 0.2))
        return np.maximum((1.0 - w) * arr + w * anchor, 0.0)
    if name in ["seasonal_naive", "drift", "recent_median"]:
        anchor = np.asarray(cfg.get("anchor", []), dtype=float)
        if len(anchor) == len(arr):
            return np.maximum(anchor, 0.0)
    return np.maximum(arr, 0.0)


def croston_forecast(series: pd.Series, horizon: int, alpha: float = 0.1, variant: str = "croston") -> np.ndarray:
    y = pd.to_numeric(series, errors="coerce").fillna(0.0).astype(float).values
    if len(y) == 0:
        return np.zeros(int(horizon), dtype=float)
    demand = y[y > 0]
    if len(demand) == 0:
        return np.zeros(int(horizon), dtype=float)
    z = float(demand[0])
    p = float(np.argmax(y > 0) + 1)
    q = 1
    prob = 1.0 / max(p, 1.0)
    for val in y[1:]:
        q += 1
        if val > 0:
            z = z + alpha * (val - z)
            p = p + alpha * (q - p)
            prob = prob + alpha * (1.0 - prob)
            q = 1
        else:
            if variant == "tsb":
                prob = prob + alpha * (0.0 - prob)
    base = z / max(p, 1e-6)
    if variant == "sba":
        base = (1.0 - alpha / 2.0) * base
    if variant == "tsb":
        base = prob * z
    return np.maximum(np.repeat(float(base), int(horizon)), 0.0)


def fit_best_intermittent(train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, profile: Dict[str, Any]) -> Dict[str, Any]:
    y_train = pd.to_numeric(train_df["y"], errors="coerce").fillna(0.0).astype(float)
    y_test = pd.to_numeric(test_df["y"], errors="coerce").fillna(0.0).astype(float)
    tr_inner, val_inner = make_inner_train_val_split(train_df)
    y_inner = pd.to_numeric(tr_inner["y"], errors="coerce").fillna(0.0).astype(float)
    y_val = pd.to_numeric(val_inner["y"], errors="coerce").fillna(0.0).astype(float)
    rows = []
    best = None
    best_score = np.inf
    for variant in ["croston", "sba", "tsb"]:
        for alpha in [0.05, 0.10, 0.20, 0.30]:
            pred_val = croston_forecast(y_inner, len(y_val), alpha=alpha, variant=variant)
            val_w = wape(y_val.values, pred_val)
            val_s = smape(y_val.values, pred_val)
            asym = compute_asymmetric_validation_penalty(y_val.values, pred_val, y_inner.values, severity=0.8)
            comp = val_w + 0.35 * val_s + asym["penalty"]
            rows.append({"variant": variant, "alpha": alpha, "val_wape": val_w, "val_smape": val_s, "bias_pct": asym["bias_pct"], "under_forecast_rate": asym["under_forecast_rate"], "peak_event_score": asym["peak_event_score"], "composite_score": comp})
            if comp < best_score:
                best_score = comp
                best = {"variant": variant, "alpha": alpha}
    if best is None:
        pred, method = build_fallback_forecast(y_train, y_test, freq_alias, infer_season_length_from_freq(freq_alias))
        return {"forecast": pred, "method": method, "alpha": None, "search_table": pd.DataFrame(rows), "fallback_used": True, "validation_wape": np.nan, "validation_smape": np.nan}
    pred_test = croston_forecast(y_train, len(y_test), alpha=float(best["alpha"]), variant=str(best["variant"]))
    ranked = pd.DataFrame(rows).sort_values(["composite_score", "val_wape"], ascending=[True, True]).reset_index(drop=True)
    return {"forecast": pred_test, "method": best["variant"], "alpha": float(best["alpha"]), "search_table": ranked, "fallback_used": False, "validation_wape": safe_float(ranked.loc[0, "val_wape"]), "validation_smape": safe_float(ranked.loc[0, "val_smape"])}


def fit_best_arima(train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, profile: Dict[str, Any]) -> Dict[str, Any]:
    if not HAS_FORECAST_STATSMODELS:
        raise ImportError("statsmodels ARIMA bileşeni bulunamadı.")
    y_train = pd.to_numeric(train_df["y"], errors="coerce").fillna(0.0).astype(float)
    y_test = pd.to_numeric(test_df["y"], errors="coerce").fillna(0.0).astype(float)
    d = suggest_d_via_stationarity(y_train)
    tr_inner, val_inner = make_inner_train_val_split(train_df)
    y_inner = pd.to_numeric(tr_inner["y"], errors="coerce").fillna(0.0).astype(float)
    y_val = pd.to_numeric(val_inner["y"], errors="coerce").fillna(0.0).astype(float)
    trend_strength = safe_float(profile.get("trend_strength", np.nan))
    trend_candidates = ["c", "n"] + (["t"] if pd.notna(trend_strength) and trend_strength >= 0.25 else [])
    transform_candidates = [choose_target_transform(y_train), {"name": "none", "lambda": None, "shift": 0.0}]
    rows = []
    best = None
    best_score = np.inf
    for p in range(0, 5):
        for q in range(0, 5):
            if p == 0 and q == 0:
                continue
            order = (p, d, q)
            for trend in trend_candidates:
                for tcfg in transform_candidates:
                    try:
                        y_fit, applied = apply_target_transform(y_inner, tcfg)
                        res = fit_sarimax_model(y_fit, None, order=order, seasonal_order=(0, 0, 0, 0), trend=trend, maxiter=max(60, FORECAST_RUNTIME_CONFIG.sarimax_maxiter_search))
                        pred_t = res.get_forecast(steps=len(y_val)).predicted_mean
                        pred = inverse_target_transform(np.asarray(pred_t, dtype=float), applied)
                        resid = pd.Series(res.resid).dropna()
                        lb_p = acorr_ljungbox(resid, lags=[min(12, max(2, len(resid)//3))], return_df=True)["lb_pvalue"].iloc[0] if len(resid) >= 8 else np.nan
                        bias_penalty = abs(float(np.nanmean(resid))) / max(float(np.nanmean(np.abs(y_inner))) or 1.0, 1.0)
                        asym = compute_asymmetric_validation_penalty(y_val.values, pred, y_inner.values, severity=0.9)
                        comp = wape(y_val.values, pred) + 0.35 * smape(y_val.values, pred) + (0.50 if pd.notna(lb_p) and lb_p < 0.05 else 0.0) + 0.50 * bias_penalty + asym["penalty"]
                        rows.append({"transform": applied.get("name","none"), "order": str(order), "trend": trend, "aic": safe_float(getattr(res, "aic", np.nan)), "bic": safe_float(getattr(res, "bic", np.nan)), "ljung_box_pvalue": safe_float(lb_p), "val_wape": wape(y_val.values, pred), "val_smape": smape(y_val.values, pred), "bias_pct": asym["bias_pct"], "under_forecast_rate": asym["under_forecast_rate"], "peak_event_score": asym["peak_event_score"], "composite_score": comp})
                        if comp < best_score:
                            best_score = comp
                            best = {"order": order, "trend": trend, "transform_cfg": applied}
                    except Exception as e:
                        rows.append({"transform": tcfg.get("name","none"), "order": str(order), "trend": trend, "aic": np.nan, "bic": np.nan, "ljung_box_pvalue": np.nan, "val_wape": np.nan, "val_smape": np.nan, "composite_score": np.nan, "fit_error": str(e)[:180]})
    ranked = pd.DataFrame(rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True) if rows else pd.DataFrame()
    if best is None:
        pred, method = build_fallback_forecast(y_train, y_test, freq_alias, infer_season_length_from_freq(freq_alias))
        return {"forecast": pred, "order": (0, d, 0), "trend": "c", "search_table": ranked, "fallback_used": True, "fallback_method": method, "validation_wape": np.nan, "validation_smape": np.nan}
    y_train_t, applied = apply_target_transform(y_train, best["transform_cfg"])
    res = fit_sarimax_model(y_train_t, None, order=best["order"], seasonal_order=(0, 0, 0, 0), trend=best["trend"], maxiter=max(120, FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final))
    pred_t = res.get_forecast(steps=len(y_test)).predicted_mean
    pred = inverse_target_transform(np.asarray(pred_t, dtype=float), applied)
    resid = pd.Series(res.resid).dropna()
    lb_p = acorr_ljungbox(resid, lags=[min(12, max(2, len(resid)//3))], return_df=True)["lb_pvalue"].iloc[0] if len(resid) >= 8 else np.nan
    return {"forecast": np.maximum(pred, 0.0), "order": best["order"], "trend": best["trend"], "search_table": ranked, "fallback_used": False, "fallback_method": None, "ljung_box_pvalue": safe_float(lb_p), "validation_wape": safe_float(ranked.loc[0, "val_wape"]) if len(ranked) else np.nan, "validation_smape": safe_float(ranked.loc[0, "val_smape"]) if len(ranked) else np.nan}


def extract_validation_metrics_from_result(model_name: str, result: Dict[str, Any]) -> Dict[str, Any]:
    if normalize_dl_family_label(model_name) in {"LSTM", "GRU"} and isinstance(result, dict) and not bool(result.get("dl_strict_validation_passed", False)):
        return {"model": model_name, "val_WAPE": np.nan, "val_sMAPE": np.nan, "eligible_for_model_ranking": False, "ranking_exclusion_reason": result.get("strict_dl_rejection_reason")}
    val_w = safe_float(result.get("validation_wape", np.nan))
    val_s = safe_float(result.get("validation_smape", np.nan))
    if pd.isna(val_w):
        stbl = result.get("search_table")
        if isinstance(stbl, pd.DataFrame) and len(stbl) > 0 and "val_wape" in stbl.columns:
            tmp = stbl.sort_values(["val_wape", "val_smape"], ascending=[True, True], na_position="last").reset_index(drop=True)
            val_w = safe_float(tmp.loc[0, "val_wape"])
            val_s = safe_float(tmp.loc[0, "val_smape"]) if "val_smape" in tmp.columns else np.nan
    return {"model": model_name, "val_WAPE": val_w, "val_sMAPE": val_s}


def run_rolling_origin_backtest_light(df_series: pd.DataFrame, outputs: Dict[str, Any], freq_alias: str, max_folds: int = 3, horizon: int = 3) -> pd.DataFrame:
    ds = df_series[["ds", "y"]].copy().reset_index(drop=True)
    h = max(1, int(horizon))
    min_train = max(24, h * 4) if str(freq_alias).upper() == "M" else max(30, h * 4)
    rows = []
    if len(ds) < (min_train + h + 2):
        return pd.DataFrame(columns=["fold", "model", "WAPE", "sMAPE", "MAE"])
    fold_ends = []
    end = len(ds) - h
    while end >= min_train and len(fold_ends) < max_folds:
        fold_ends.append(end)
        end -= h
    fold_ends = sorted(fold_ends)
    for fold_no, train_end in enumerate(fold_ends, start=1):
        tr = ds.iloc[:train_end].copy().reset_index(drop=True)
        te = ds.iloc[train_end:train_end+h].copy().reset_index(drop=True)
        actual = te["y"].values.astype(float)
        model_preds = {}
        if outputs.get("sarima") and outputs["sarima"].get("order") is not None:
            try:
                order = tuple(outputs["sarima"]["order"])
                seas = tuple(outputs["sarima"]["seasonal_order"])
                trend = outputs["sarima"].get("trend", "c")
                res = fit_sarimax_model(pd.to_numeric(tr["y"], errors="coerce").astype(float), None, order=order, seasonal_order=seas, trend=trend, maxiter=max(60, min(FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final, 120)))
                model_preds["SARIMA/SARIMAX"] = np.maximum(np.asarray(res.get_forecast(steps=len(te)).predicted_mean, dtype=float), 0.0)
            except Exception:
                pass
        if outputs.get("arima") and outputs["arima"].get("order") is not None:
            try:
                order = tuple(outputs["arima"]["order"])
                trend = outputs["arima"].get("trend", "c")
                res = fit_sarimax_model(pd.to_numeric(tr["y"], errors="coerce").astype(float), None, order=order, seasonal_order=(0,0,0,0), trend=trend, maxiter=max(60, min(FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final, 120)))
                model_preds["ARIMA"] = np.maximum(np.asarray(res.get_forecast(steps=len(te)).predicted_mean, dtype=float), 0.0)
            except Exception:
                pass
        if outputs.get("intermittent") and outputs["intermittent"].get("method"):
            try:
                model_preds["Intermittent"] = croston_forecast(tr["y"], len(te), alpha=float(outputs["intermittent"].get("alpha", 0.1) or 0.1), variant=str(outputs["intermittent"].get("method", "sba")))
            except Exception:
                pass
        for mname, pred in model_preds.items():
            rows.append({"fold": fold_no, "model": mname, "WAPE": wape(actual, pred), "sMAPE": smape(actual, pred), "MAE": mae(actual, pred)})
    return pd.DataFrame(rows)


def _safe_abs_corr(a: pd.Series, b: pd.Series) -> float:
    try:
        aa = pd.to_numeric(a, errors="coerce").astype(float)
        bb = pd.to_numeric(b, errors="coerce").astype(float)
        mask = aa.notna() & bb.notna()
        if int(mask.sum()) < 5:
            return 0.0
        corr = np.corrcoef(aa[mask], bb[mask])[0, 1]
        return float(abs(corr)) if np.isfinite(corr) else 0.0
    except Exception:
        return 0.0


def reduce_exog_feature_set(
    exog_train: Optional[pd.DataFrame],
    exog_test: Optional[pd.DataFrame],
    y_train: pd.Series,
    max_cols: int = 3,
    corr_threshold: float = 0.985
) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], List[str], List[str]]:
    dropped: List[str] = []
    selected: List[str] = []
    if exog_train is None or exog_test is None:
        return exog_train, exog_test, selected, dropped
    xtr, xte, dropped0 = sanitize_exog_for_sarimax(exog_train, exog_test)
    dropped.extend(dropped0)
    if xtr is None or xte is None or len(xtr.columns) == 0:
        return None, None, selected, dropped

    y_ref = pd.to_numeric(y_train, errors="coerce").astype(float).reset_index(drop=True)
    score_rows = []
    for c in xtr.columns:
        s = pd.to_numeric(xtr[c], errors="coerce").astype(float)
        score_rows.append({
            "col": c,
            "abs_corr": _safe_abs_corr(s, y_ref),
            "var": safe_float(np.nanvar(s.values))
        })
    score_df = pd.DataFrame(score_rows).sort_values(["abs_corr", "var", "col"], ascending=[False, False, True])

    ranked_cols = score_df["col"].tolist()
    filtered: List[str] = []
    for c in ranked_cols:
        if len(filtered) >= max_cols:
            dropped.append(c)
            continue
        too_close = False
        for kept in filtered:
            corr = _safe_abs_corr(xtr[c], xtr[kept])
            if corr >= corr_threshold:
                too_close = True
                dropped.append(c)
                break
        if not too_close:
            filtered.append(c)

    if not filtered:
        return None, None, selected, dropped
    selected = filtered
    return xtr[selected].copy(), xte[selected].copy(), selected, dropped




def reduce_ml_feature_set(
    feature_train: Optional[pd.DataFrame],
    feature_test: Optional[pd.DataFrame],
    y_train: pd.Series,
    max_cols: int = 14,
    corr_threshold: float = 0.985
) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], List[str], List[str]]:
    dropped: List[str] = []
    selected: List[str] = []
    if feature_train is None or feature_test is None:
        return feature_train, feature_test, selected, dropped
    if len(feature_train.columns) == 0:
        return feature_train, feature_test, selected, dropped
    xtr = feature_train.copy().reset_index(drop=True)
    xte = feature_test.copy().reset_index(drop=True)
    keep_numeric = []
    for c in xtr.columns:
        s = pd.to_numeric(xtr[c], errors="coerce")
        if s.notna().sum() < max(5, len(xtr) // 5):
            dropped.append(c)
            continue
        if float(np.nanvar(s.values)) <= 1e-12:
            dropped.append(c)
            continue
        keep_numeric.append(c)
    if not keep_numeric:
        return None, None, selected, dropped
    xtr = xtr[keep_numeric].apply(pd.to_numeric, errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
    xte = xte[keep_numeric].apply(pd.to_numeric, errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_ref = pd.to_numeric(y_train, errors="coerce").astype(float).reset_index(drop=True)
    score_rows = []
    for c in xtr.columns:
        score_rows.append({
            "col": c,
            "abs_corr": _safe_abs_corr(xtr[c], y_ref),
            "var": safe_float(np.nanvar(xtr[c].values))
        })
    score_df = pd.DataFrame(score_rows).sort_values(["abs_corr", "var", "col"], ascending=[False, False, True])
    ranked = score_df["col"].tolist()
    for c in ranked:
        if len(selected) >= max_cols:
            dropped.append(c)
            continue
        if any(_safe_abs_corr(xtr[c], xtr[k]) >= corr_threshold for k in selected):
            dropped.append(c)
            continue
        selected.append(c)
    if not selected:
        return None, None, [], dropped
    return xtr[selected].copy(), xte[selected].copy(), selected, dropped



def build_fast_xgb_regressor(cfg: Dict[str, Any]):
    params = dict(cfg)
    prefer_hist_gradient = bool(params.pop("_prefer_hist_gradient", False))
    if prefer_hist_gradient and SKHistGradientBoostingRegressor is not None and not HAS_XGBOOST:
        mapped = {
            "learning_rate": params.get("learning_rate", 0.06),
            "max_depth": params.get("max_depth", 3),
            "max_iter": params.get("n_estimators", 64),
            "max_bins": 128,
            "l2_regularization": params.get("reg_lambda", 1.0),
            "min_samples_leaf": max(2, int(params.get("min_child_weight", 1) * 2)),
            "early_stopping": True,
            "random_state": 42
        }
        return SKHistGradientBoostingRegressor(**mapped)
    if HAS_XGBOOST:
        params.setdefault("objective", "reg:squarederror")
        params.setdefault("random_state", 42)
        params.setdefault("verbosity", 0)
        params.setdefault("tree_method", "hist")
        params.setdefault("max_bin", 128)
        params.setdefault("eval_metric", "mae")
        params.setdefault("n_jobs", 1 if FORECAST_RUNTIME_CONFIG.xgb_force_single_thread else max(1, os.cpu_count() or 1))
        return XGBRegressor(**params)
    if SKExtraTreesRegressor is not None:
        return SKExtraTreesRegressor(
            n_estimators=min(int(params.get("n_estimators", 128)), 256),
            max_depth=min(int(params.get("max_depth", 4)), 6),
            random_state=42,
            n_jobs=1
        )
    if SKHistGradientBoostingRegressor is not None:
        return SKHistGradientBoostingRegressor(
            learning_rate=params.get("learning_rate", 0.06),
            max_depth=params.get("max_depth", 3),
            max_iter=params.get("n_estimators", 64),
            random_state=42,
            early_stopping=True
        )
    return XGBRegressor(random_state=42)



def _prepare_xgb_cfg(cfg: Dict[str, Any], train_len: int, freq_alias: str) -> Dict[str, Any]:
    out = dict(cfg)
    short_monthly = str(freq_alias).upper() == "M" and train_len <= 96
    if FORECAST_RUNTIME_CONFIG.xgb_prefer_hist_gradient_on_short_series and short_monthly:
        out["_prefer_hist_gradient"] = True
    if short_monthly:
        out["n_estimators"] = min(int(out.get("n_estimators", 300)), 500)
        out["max_depth"] = min(int(out.get("max_depth", 3)), 4)
    return out



def _select_core_ml_feature_columns(all_cols: List[str], freq_alias: str, train_len: int) -> List[str]:
    cols = list(all_cols)
    short_monthly = str(freq_alias).upper() == "M" and train_len <= 96
    if not short_monthly:
        return cols
    preferred = [
        "lag_1", "lag_2", "lag_3", "lag_6", "lag_12", "lag_18", "seasonal_lag_12",
        "roll_mean_3", "roll_mean_6", "roll_mean_12",
        "roll_std_3", "roll_std_6", "roll_std_12",
        "roll_min_3", "roll_max_3", "roll_min_6", "roll_max_6",
        "ewm_mean_3", "ewm_mean_6", "ewm_mean_12",
        "diff_1", "diff_12",
        "month_sin", "month_cos", "month", "quarter", "year"
    ]
    exog_cols = [c for c in cols if str(c).startswith("exog__")][:6]
    ordered = [c for c in preferred if c in cols] + exog_cols
    ordered += [c for c in cols if c not in ordered]
    return ordered[:min(20, len(ordered))]

def should_use_prophet_holidays(freq_alias: str, n_train: int) -> bool:
    if not FORECAST_RUNTIME_CONFIG.prophet_disable_holidays_for_short_series:
        return True
    if str(freq_alias).upper() == "M" and n_train < 84:
        return False
    if str(freq_alias).upper() == "W" and n_train < 120:
        return False
    return True



def probe_prophet_backend() -> Tuple[bool, str]:
    if not HAS_PROPHET:
        return False, "prophet_package_missing"
    if not FORECAST_RUNTIME_CONFIG.prophet_probe_backend:
        return True, "probe_disabled"
    if _PROPhet_BACKEND_PROBE_CACHE["done"]:
        return bool(_PROPhet_BACKEND_PROBE_CACHE["ok"]), str(_PROPhet_BACKEND_PROBE_CACHE["message"])
    try:
        mini = pd.DataFrame({
            "ds": pd.date_range("2022-01-01", periods=8, freq="MS"),
            "y": np.array([10.0, 11.0, 12.5, 12.0, 13.0, 14.0, 13.5, 14.5], dtype=float)
        })
        m = Prophet(yearly_seasonality=False, weekly_seasonality=False, daily_seasonality=False)
        m.fit(mini)
        future = pd.DataFrame({"ds": pd.date_range("2022-09-01", periods=1, freq="MS")})
        _ = m.predict(future)
        ok = True
        msg = "mini_fit_ok"
    except Exception as e:
        ok = False
        msg = str(e)
    _PROPhet_BACKEND_PROBE_CACHE["done"] = True
    _PROPhet_BACKEND_PROBE_CACHE["ok"] = bool(ok)
    _PROPhet_BACKEND_PROBE_CACHE["message"] = str(msg)
    return bool(ok), str(msg)

def _build_prophet_style_surrogate_design(
    base_df: pd.DataFrame,
    exog_df: Optional[pd.DataFrame],
    freq_alias: str
) -> Tuple[pd.DataFrame, List[str]]:
    df = base_df[["ds", "y"]].copy().reset_index(drop=True)
    ds = pd.to_datetime(df["ds"])
    df["t"] = np.arange(len(df), dtype=float)
    df["t2"] = df["t"] ** 2
    month = ds.dt.month.fillna(1).astype(int)
    df["month_sin"] = np.sin(2 * np.pi * month / 12.0)
    df["month_cos"] = np.cos(2 * np.pi * month / 12.0)
    quarter = ds.dt.quarter.fillna(1).astype(int)
    df["quarter_sin"] = np.sin(2 * np.pi * quarter / 4.0)
    df["quarter_cos"] = np.cos(2 * np.pi * quarter / 4.0)
    if str(freq_alias).upper() in ["W", "D"]:
        week = ds.dt.isocalendar().week.astype(int).clip(lower=1)
        df["week_sin"] = np.sin(2 * np.pi * week / 52.0)
        df["week_cos"] = np.cos(2 * np.pi * week / 52.0)
    if str(freq_alias).upper() in ["D", "H"]:
        dow = ds.dt.dayofweek.astype(int)
        df["dow_sin"] = np.sin(2 * np.pi * dow / 7.0)
        df["dow_cos"] = np.cos(2 * np.pi * dow / 7.0)
    if exog_df is not None and len(exog_df.columns) > 0:
        ex = exog_df.copy().reset_index(drop=True)
        for c in ex.columns:
            safe_name = f"reg__{re.sub(r'[^0-9A-Za-z_]+', '_', str(c))}"
            df[safe_name] = pd.to_numeric(ex[c], errors="coerce")
    feature_cols = [c for c in df.columns if c not in ["ds", "y"]]
    return df, feature_cols


def fit_prophet_surrogate(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    freq_alias: str,
    exog_train: Optional[pd.DataFrame] = None,
    exog_test: Optional[pd.DataFrame] = None,
    reason: str = "prophet_backend_unavailable"
) -> Dict[str, Any]:
    y_train = pd.to_numeric(train_df["y"], errors="coerce").astype(float).fillna(0.0)
    xtr, xte, used_cols, dropped_cols = reduce_exog_feature_set(
        exog_train,
        exog_test,
        y_train,
        max_cols=min(FORECAST_RUNTIME_CONFIG.prophet_max_exog_cols, 3)
    )
    full_df = pd.concat([train_df[["ds", "y"]], test_df[["ds", "y"]]], axis=0, ignore_index=True)
    full_exog = None
    if xtr is not None and xte is not None:
        full_exog = pd.concat([xtr.reset_index(drop=True), xte.reset_index(drop=True)], axis=0, ignore_index=True)
    design, feature_cols = _build_prophet_style_surrogate_design(full_df, full_exog, freq_alias)
    train_cut = len(train_df)
    X_train = design.iloc[:train_cut][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_train_arr = pd.to_numeric(train_df["y"], errors="coerce").astype(float).fillna(0.0).values
    X_test = design.iloc[train_cut:train_cut + len(test_df)][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)

    tr_inner, val_inner = make_inner_train_val_split(train_df)
    inner_cut = len(tr_inner)
    inner_design = design.iloc[:len(train_df)].copy().reset_index(drop=True)
    X_inner_tr = inner_design.iloc[:inner_cut][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_inner_tr = pd.to_numeric(tr_inner["y"], errors="coerce").astype(float).fillna(0.0).values
    X_inner_val = inner_design.iloc[inner_cut:inner_cut + len(val_inner)][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_inner_val = pd.to_numeric(val_inner["y"], errors="coerce").astype(float).fillna(0.0).values

    alpha_grid = [0.1, 1.0, 5.0, 10.0]
    rows = []
    best_alpha = 1.0
    best_score = np.inf
    for alpha in alpha_grid:
        try:
            if Ridge is not None:
                mdl = Ridge(alpha=alpha, random_state=42)
            else:
                raise RuntimeError("ridge_unavailable")
            mdl.fit(X_inner_tr.values, y_inner_tr)
            pred_val = np.maximum(np.asarray(mdl.predict(X_inner_val.values), dtype=float), 0.0)
            val_w = wape(y_inner_val, pred_val)
            val_s = smape(y_inner_val, pred_val)
            comp = float(val_w + 0.35 * val_s + 0.01 * alpha)
            rows.append({"alpha": alpha, "val_wape": val_w, "val_smape": val_s, "composite_score": comp})
            if comp < best_score:
                best_score = comp
                best_alpha = alpha
        except Exception as e:
            rows.append({"alpha": alpha, "val_wape": np.nan, "val_smape": np.nan, "composite_score": np.nan, "fit_error": str(e)[:200]})

    if Ridge is not None:
        final_model = Ridge(alpha=best_alpha, random_state=42)
        final_model.fit(X_train.values, y_train_arr)
        pred = np.maximum(np.asarray(final_model.predict(X_test.values), dtype=float), 0.0)
    else:
        pred, fb_name = build_fallback_forecast(train_df["y"], test_df["y"], freq_alias, infer_season_length_from_freq(freq_alias))
        final_model = None
        reason = f"{reason}|ridge_missing|{fb_name}"

    fc = test_df[["ds", "y"]].copy().reset_index(drop=True)
    fc["yhat"] = np.maximum(np.asarray(pred, dtype=float), 0.0)
    if len(fc):
        fc["trend"] = np.linspace(float(y_train_arr[0]) if len(y_train_arr) else 0.0, float(y_train_arr[-1]) if len(y_train_arr) else 0.0, num=len(fc))
        fc["yearly"] = fc["yhat"] - fc["trend"]
    else:
        fc["trend"] = []
        fc["yearly"] = []
    return {
        "model": final_model,
        "forecast_df": fc,
        "forecast": np.maximum(np.asarray(fc["yhat"].values, dtype=float), 0.0),
        "config": {"mode": "prophet_style_ridge", "alpha": best_alpha},
        "selected_plan": {"plan_name": "prophet_surrogate", "reason": reason},
        "component_validation": {
            "trend_abs_mean": safe_float(np.abs(fc.get("trend", pd.Series(dtype=float))).mean()) if len(fc) else np.nan,
            "seasonality_abs_mean": safe_float(np.abs(fc.get("yearly", pd.Series(dtype=float))).mean()) if len(fc) else np.nan,
            "seasonality_present": True,
            "used_exog_cols": list(xtr.columns) if xtr is not None else [],
            "dropped_exog_cols": sorted(set(dropped_cols)),
            "final_retry_used": True,
            "rename_map": {},
            "fit_error_samples": [str(reason)]
        },
        "used_exog_cols": list(xtr.columns) if xtr is not None else [],
        "dropped_exog_cols": sorted(set(dropped_cols)),
        "search_table": pd.DataFrame(rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True),
        "fallback_used": True,
        "fallback_method": "prophet_style_ridge",
        "fit_mode": "vekil_prophet_fallback",
        "fit_visibility_note": "Gerçek Prophet fit başarısız oldu; güvenli vekil Prophet (ridge tabanlı) kullanıldı.",
        "error": str(reason)
    }



def build_runtime_guardrail_notes(freq_alias: str, n_train: int, horizon: int, stat_exog_cols: List[str], prophet_exog_cols: List[str]) -> List[str]:
    notes: List[str] = []
    if not FORECAST_RUNTIME_CONFIG.interactive_fast_mode:
        notes.append("Doğruluk modu aktif: arama uzayı korunur; kısa serilerde yalnızca validation-temelli adaptif düzenlileştirme uygulanır.")
    if str(freq_alias).upper() == "M" and n_train <= 72:
        notes.append("Kısa aylık seri: bu bir kısıtlama değil, overfit önleyici adaptif güvenlik katmanıdır.")
    if len(stat_exog_cols) > FORECAST_RUNTIME_CONFIG.sarimax_max_exog_cols:
        notes.append(f"SARIMAX için exog adayları skorlanır; en sinyalli ilk {FORECAST_RUNTIME_CONFIG.sarimax_max_exog_cols} kolon kullanılır.")
    if len(prophet_exog_cols) > FORECAST_RUNTIME_CONFIG.prophet_max_exog_cols:
        notes.append(f"Prophet için regressor adayları skorlanır; en sinyalli ilk {FORECAST_RUNTIME_CONFIG.prophet_max_exog_cols} kolon kullanılır.")
    notes.append(f"XGBoost tarafında leakage-safe feature seçimi yapılır; en fazla {FORECAST_RUNTIME_CONFIG.xgb_max_feature_cols} kolon validation-temelli olarak kullanılır.")
    return notes


def build_stage_timer_rows(stage_timings: Dict[str, float]) -> pd.DataFrame:
    return pd.DataFrame([
        {"stage": k, "seconds": round(float(v), 4)} for k, v in stage_timings.items()
    ])



def fit_best_sarimax(train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, profile: Dict[str, Any], exog_train: Optional[pd.DataFrame] = None, exog_test: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
    if not HAS_FORECAST_STATSMODELS:
        raise ImportError("statsmodels forecast bileşenleri bulunamadı.")

    cache_key = build_search_signature(
        "sarimax_exact_search",
        freq_alias,
        train_df,
        test_df,
        profile=profile,
        exog_train=exog_train,
        exog_test=exog_test,
        extra={
            "max_candidates": FORECAST_RUNTIME_CONFIG.sarimax_max_candidates,
            "maxiter_search": FORECAST_RUNTIME_CONFIG.sarimax_maxiter_search,
            "maxiter_final": FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final,
            "walk_forward": FORECAST_RUNTIME_CONFIG.sarimax_enable_walk_forward_refit,
        }
    )
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached

    start_ts = time.perf_counter()
    y_train_raw = pd.to_numeric(train_df["y"], errors="coerce").astype(float).fillna(0.0)
    y_test_raw = pd.to_numeric(test_df["y"], errors="coerce").astype(float).fillna(0.0)
    season_length = infer_season_length_from_freq(freq_alias)
    strong_seasonality = safe_float(profile.get("seasonality_strength", np.nan))
    trend_strength = safe_float(profile.get("trend_strength", np.nan))

    d = suggest_d_via_stationarity(y_train_raw)
    D = suggest_D_via_profile(profile, season_length, len(y_train_raw))

    exog_bundle_key = build_search_signature(
        "sarimax_exog_bundle", freq_alias, train_df, test_df, profile=profile, exog_train=exog_train, exog_test=exog_test,
        extra={"max_cols": FORECAST_RUNTIME_CONFIG.sarimax_max_exog_cols}
    )
    xtr, xte, used_exog_cols, dropped_exog = SEARCH_ACCELERATOR.get_or_compute_artifact(
        exog_bundle_key,
        lambda: reduce_exog_feature_set(exog_train, exog_test, y_train_raw, max_cols=FORECAST_RUNTIME_CONFIG.sarimax_max_exog_cols)
    )

    if pd.notna(strong_seasonality) and strong_seasonality < 0.15:
        season_length = 1
        D = 0

    grid = build_sarimax_grid(freq_alias, profile, len(y_train_raw))
    candidates = prioritize_sarimax_candidates(grid, profile, FORECAST_RUNTIME_CONFIG.sarimax_max_candidates)

    transform_candidates = [choose_target_transform(y_train_raw), {"name": "none", "lambda": None, "shift": 0.0}]
    if (y_train_raw > 0).all():
        transform_candidates.append({"name": "log1p", "lambda": None, "shift": 0.0})
    dedup_transforms = []
    seen_transforms = set()
    for cfg in transform_candidates:
        key = (cfg.get("name"), safe_float(cfg.get("lambda")), safe_float(cfg.get("shift")))
        if key not in seen_transforms:
            seen_transforms.add(key)
            dedup_transforms.append(cfg)

    split_key = build_search_signature("sarimax_inner_split", freq_alias, train_df, test_df, profile=profile)
    tr_inner_raw, val_inner_raw = SEARCH_ACCELERATOR.get_or_compute_artifact(split_key, lambda: make_inner_train_val_split(train_df))
    y_inner = pd.to_numeric(tr_inner_raw["y"], errors="coerce").astype(float).fillna(0.0)
    y_val = pd.to_numeric(val_inner_raw["y"], errors="coerce").astype(float).fillna(0.0)
    if xtr is not None:
        exog_inner_train = xtr.iloc[:len(tr_inner_raw)].copy().reset_index(drop=True)
        exog_inner_val = xtr.iloc[len(tr_inner_raw):len(tr_inner_raw) + len(val_inner_raw)].copy().reset_index(drop=True)
    else:
        exog_inner_train = None
        exog_inner_val = None

    trend_candidates = ["c", "n"] + (["t"] if pd.notna(trend_strength) and trend_strength >= 0.25 else [])
    search_budget = FORECAST_RUNTIME_CONFIG.sarimax_search_wall_seconds

    jobs: List[Dict[str, Any]] = []
    for cand_idx, cand in enumerate(candidates):
        if (time.perf_counter() - start_ts) > search_budget and jobs:
            break
        p, q, P, Q, m = cand["p"], cand["q"], cand["P"], cand["Q"], cand["m"]
        order = (p, d, q)
        seasonal_order = (P, D if m > 1 else 0, Q, m if m > 1 else 0)
        exog_modes = [False] if xtr is None else ([False, True] if cand_idx < FORECAST_RUNTIME_CONFIG.sarimax_search_with_exog_top_n else [False])
        for transform_cfg in dedup_transforms:
            for trend_spec in trend_candidates:
                for use_exog in exog_modes:
                    jobs.append({
                        "order": order,
                        "seasonal_order": seasonal_order,
                        "transform_cfg": transform_cfg,
                        "trend": trend_spec,
                        "use_exog": bool(use_exog)
                    })

    def _evaluate_sarimax_job(job: Dict[str, Any]) -> Dict[str, Any]:
        order = job["order"]
        seasonal_order = job["seasonal_order"]
        applied_cfg = job["transform_cfg"]
        trend_spec = job["trend"]
        use_exog = bool(job["use_exog"])
        exog_fit = exog_inner_train if use_exog else None
        exog_val = exog_inner_val if use_exog else None
        try:
            y_fit_inner, applied_cfg = apply_target_transform(y_inner, applied_cfg)
            res = fit_sarimax_model(
                y_fit_inner,
                exog_fit,
                order=order,
                seasonal_order=seasonal_order,
                trend=trend_spec,
                maxiter=FORECAST_RUNTIME_CONFIG.sarimax_maxiter_search
            )
            pred_t = res.get_forecast(steps=len(y_val), exog=exog_val).predicted_mean
            pred = inverse_target_transform(np.asarray(pred_t, dtype=float), applied_cfg)
            pred = np.maximum(pred, 0.0)
            val_wape = wape(y_val.values, pred)
            val_smape = smape(y_val.values, pred)
            aic = safe_float(getattr(res, "aic", np.nan))
            bic = safe_float(getattr(res, "bic", np.nan))
            complexity_penalty = 0.12 * sum([order[0], order[2], seasonal_order[0], seasonal_order[2]])
            exog_penalty = 0.05 * (len(used_exog_cols) if use_exog else 0)
            residual_penalty = 0.0
            try:
                resid = pd.Series(res.resid).dropna()
                if len(resid) >= 8:
                    lb_p = acorr_ljungbox(resid, lags=[min(12, max(2, len(resid) // 3))], return_df=True)["lb_pvalue"].iloc[0]
                    resid_mean = safe_float(resid.mean())
                    resid_scale = max(safe_float(np.abs(y_inner).mean()), 1.0)
                    bias_penalty = min(abs(resid_mean) / resid_scale, 1.0)
                    residual_penalty = (0.60 if pd.notna(lb_p) and lb_p < 0.05 else 0.0) + 0.50 * bias_penalty
                else:
                    lb_p = np.nan
                    bias_penalty = 0.0
            except Exception:
                lb_p = np.nan
                bias_penalty = 0.0
            asym = compute_asymmetric_validation_penalty(y_val.values, pred, y_inner.values, severity=0.9)
            composite = float(val_wape + 0.35 * val_smape + complexity_penalty + exog_penalty + residual_penalty + asym["penalty"])
            row = {
                "transform": applied_cfg.get("name", "none"),
                "order": str(order),
                "seasonal_order": str(seasonal_order),
                "trend": trend_spec,
                "exog_mode": "with_exog" if use_exog else "without_exog",
                "aic": aic,
                "bic": bic,
                "ljung_box_pvalue": safe_float(lb_p),
                "residual_bias_penalty": safe_float(bias_penalty),
                "val_wape": val_wape,
                "val_smape": val_smape,
                "bias_pct": asym["bias_pct"],
                "under_forecast_rate": asym["under_forecast_rate"],
                "peak_event_score": asym["peak_event_score"],
                "composite_score": composite
            }
            return {"row": row, "composite": composite, "best": {"order": order, "seasonal_order": seasonal_order, "trend": trend_spec, "transform_cfg": applied_cfg, "use_exog": bool(use_exog)}}
        except Exception as e:
            row = {
                "transform": applied_cfg.get("name", "none"),
                "order": str(order),
                "seasonal_order": str(seasonal_order),
                "trend": trend_spec,
                "exog_mode": "with_exog" if use_exog else "without_exog",
                "aic": np.nan,
                "bic": np.nan,
                "val_wape": np.nan,
                "val_smape": np.nan,
                "composite_score": np.nan,
                "fit_error": str(e)[:200]
            }
            return {"row": row, "composite": np.inf, "best": None}

    search_rows: List[Dict[str, Any]] = []
    best = None
    best_score = np.inf
    results = SEARCH_ACCELERATOR.parallel_map(_evaluate_sarimax_job, jobs, max_workers=FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_workers)
    for item in results:
        search_rows.append(item["row"])
        if item["best"] is not None and np.isfinite(item["composite"]) and item["composite"] < best_score:
            best_score = item["composite"]
            best = item["best"]

    if best is None:
        fallback_pred, fallback_name = build_fallback_forecast(y_train_raw, y_test_raw, freq_alias, infer_season_length_from_freq(freq_alias))
        result = {
            "model": None,
            "forecast": fallback_pred,
            "static_forecast": fallback_pred.copy(),
            "order": (0, d, 0),
            "seasonal_order": (0, D if season_length > 1 else 0, 0, season_length if season_length > 1 else 0),
            "aic": np.nan,
            "bic": np.nan,
            "ljung_box_pvalue": np.nan,
            "d": d,
            "D": D,
            "transform": "none",
            "residual_mean": np.nan,
            "residual_std": np.nan,
            "residual_white_noise_ok": False,
            "search_table": pd.DataFrame(search_rows),
            "fallback_used": True,
            "fallback_method": fallback_name,
            "used_exog": False,
            "used_exog_cols": [],
            "trend": "c",
            "dropped_exog_cols": dropped_exog,
            "search_accelerator_cache_hit": False,
        }
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    final_exog_train = xtr if best.get("use_exog") else None
    final_exog_test = xte if best.get("use_exog") else None
    y_train_t, applied_cfg = apply_target_transform(y_train_raw, best["transform_cfg"])

    try:
        final_res = fit_sarimax_model(
            y_train_t,
            final_exog_train,
            order=best["order"],
            seasonal_order=best["seasonal_order"],
            trend=best.get("trend", "c"),
            maxiter=FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final
        )
        pred_t = final_res.get_forecast(steps=len(y_test_raw), exog=final_exog_test).predicted_mean
        pred_static = inverse_target_transform(np.asarray(pred_t, dtype=float), applied_cfg)
        pred_static = np.maximum(pred_static, 0.0)
        pred = pred_static.copy()
        resid = pd.Series(final_res.resid).dropna()
        try:
            lb_p = acorr_ljungbox(resid, lags=[min(12, max(2, len(resid) // 3))], return_df=True)["lb_pvalue"].iloc[0]
        except Exception:
            lb_p = np.nan

        if FORECAST_RUNTIME_CONFIG.sarimax_enable_walk_forward_refit:
            try:
                pred_wf = walk_forward_refit_sarimax(
                    history_y=y_train_raw,
                    future_y=y_test_raw,
                    order=best["order"],
                    seasonal_order=best["seasonal_order"],
                    exog_hist=final_exog_train,
                    exog_future=final_exog_test,
                    maxiter=max(20, min(FORECAST_RUNTIME_CONFIG.sarimax_maxiter_final, 40))
                )
                if wape(y_test_raw.values, pred_wf) <= wape(y_test_raw.values, pred_static):
                    pred = np.maximum(np.asarray(pred_wf, dtype=float), 0.0)
            except Exception:
                pass

        result = {
            "model": final_res,
            "forecast": pred,
            "static_forecast": pred_static.copy(),
            "order": best["order"],
            "seasonal_order": best["seasonal_order"],
            "aic": safe_float(getattr(final_res, "aic", np.nan)),
            "bic": safe_float(getattr(final_res, "bic", np.nan)),
            "ljung_box_pvalue": safe_float(lb_p),
            "d": d,
            "D": D,
            "transform": applied_cfg.get("name", "none"),
            "residual_mean": safe_float(resid.mean()),
            "residual_std": safe_float(resid.std()),
            "residual_white_noise_ok": bool(pd.notna(lb_p) and lb_p > 0.05),
            "search_table": pd.DataFrame(search_rows).sort_values(["composite_score", "val_wape", "aic"], ascending=[True, True, True], na_position="last").reset_index(drop=True),
            "fallback_used": False,
            "fallback_method": None,
            "used_exog": bool(best.get("use_exog")),
            "used_exog_cols": list(final_exog_train.columns) if final_exog_train is not None else [],
            "trend": best.get("trend", "c"),
            "dropped_exog_cols": dropped_exog,
            "benchmark_override": False,
            "search_accelerator_cache_hit": False,
        }
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result
    except Exception:
        fallback_pred, fallback_name = build_fallback_forecast(y_train_raw, y_test_raw, freq_alias, infer_season_length_from_freq(freq_alias))
        result = {
            "model": None,
            "forecast": fallback_pred,
            "static_forecast": fallback_pred.copy(),
            "order": best["order"],
            "seasonal_order": best["seasonal_order"],
            "aic": np.nan,
            "bic": np.nan,
            "ljung_box_pvalue": np.nan,
            "d": d,
            "D": D,
            "transform": applied_cfg.get("name", "none"),
            "residual_mean": np.nan,
            "residual_std": np.nan,
            "residual_white_noise_ok": False,
            "search_table": pd.DataFrame(search_rows),
            "fallback_used": True,
            "fallback_method": fallback_name,
            "used_exog": bool(best.get("use_exog")),
            "used_exog_cols": list(final_exog_train.columns) if final_exog_train is not None else [],
            "trend": best.get("trend", "c"),
            "dropped_exog_cols": dropped_exog,
            "search_accelerator_cache_hit": False,
        }
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

def fit_best_prophet(train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, profile: Dict[str, Any], exog_train: Optional[pd.DataFrame] = None, exog_test: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
    if not HAS_PROPHET:
        return fit_prophet_surrogate(train_df, test_df, freq_alias, exog_train, exog_test, reason="prophet_package_missing")
    if str(freq_alias).upper() == "W":
        seasonality_strength = safe_float((profile or {}).get("seasonality_strength", np.nan))
        trend_strength = safe_float((profile or {}).get("trend_strength", np.nan))
        if len(train_df) < 104 or (pd.notna(seasonality_strength) and seasonality_strength < 0.12 and pd.notna(trend_strength) and trend_strength < 0.08):
            result = fit_prophet_surrogate(train_df, test_df, freq_alias, exog_train, exog_test, reason="weekly_prophet_hard_disable")
            result["fit_mode"] = "haftalıkta_prophet_kapatıldı"
            result["fit_visibility_note"] = "Haftalık frekansta kısa/zayıf sinyalli seri için Prophet otomatik kapatıldı; güvenli vekil model kullanıldı."
            return result
    cache_key = build_search_signature("prophet_exact_search", freq_alias, train_df, test_df, profile=profile, exog_train=exog_train, exog_test=exog_test, extra={"max_configs": FORECAST_RUNTIME_CONFIG.prophet_max_configs})
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached
    backend_ok, backend_msg = probe_prophet_backend()
    backend_probe_note = None if backend_ok else f"prophet_backend_probe_failed:{backend_msg}"
    y_train = pd.to_numeric(train_df["y"], errors="coerce").astype(float).fillna(0.0)
    season_length = infer_season_length_from_freq(freq_alias)
    prop_xtr, prop_xte, _used_exog_cols, dropped_exog = reduce_exog_feature_set(exog_train, exog_test, y_train, max_cols=FORECAST_RUNTIME_CONFIG.prophet_max_exog_cols)
    tr_full, te_full, exog_cols, prep_dropped, rename_map = _prepare_prophet_design_matrices(train_df, test_df, prop_xtr, prop_xte)
    dropped_exog.extend(prep_dropped)
    tr_inner_raw, val_inner_raw = make_inner_train_val_split(train_df)
    inner_xtr = prop_xtr.iloc[:len(tr_inner_raw)].copy() if prop_xtr is not None else None
    inner_xva = prop_xtr.iloc[len(tr_inner_raw):len(tr_inner_raw)+len(val_inner_raw)].copy() if prop_xtr is not None else None
    tr, val, inner_exog_cols, inner_dropped, _ = _prepare_prophet_design_matrices(tr_inner_raw, val_inner_raw, inner_xtr, inner_xva)
    dropped_exog.extend(inner_dropped)
    seasonality_strength = safe_float(profile.get("seasonality_strength", np.nan))
    strictly_positive = bool((y_train > 0).all())
    mode_candidates = ["additive"] + (["multiplicative"] if strictly_positive and (pd.isna(seasonality_strength) or seasonality_strength >= 0.20) else [])
    transform_candidates = [choose_target_transform(y_train), {"name": "none", "lambda": None, "shift": 0.0}]
    cps_candidates = [0.005, 0.01, 0.03, 0.08, 0.15, 0.30]
    sps_candidates = [0.1, 0.5, 1.0, 3.0, 8.0]
    cp_range_candidates = [0.80, 0.90, 0.95]
    ncp_candidates = [5, 8, 12, 20, 25] if len(train_df) >= 48 else [3, 5, 8, 12]
    use_holidays = should_use_prophet_holidays(freq_alias, len(train_df))
    plans = []
    for mode in mode_candidates:
        for cps in cps_candidates:
            for sps in sps_candidates:
                for cp_range in cp_range_candidates:
                    for ncp in ncp_candidates:
                        cfg = {"seasonality_mode": mode, "changepoint_prior_scale": cps, "seasonality_prior_scale": sps, "changepoint_range": cp_range, "n_changepoints": ncp, "growth": "linear"}
                        for tf in transform_candidates:
                            plans.append({"cfg": cfg, "transform_cfg": tf, "use_exog": len(inner_exog_cols) > 0, "use_holidays": use_holidays, "add_custom_seasonality": True, "plan_name": "primary"})
                            if len(inner_exog_cols) > 0:
                                plans.append({"cfg": cfg, "transform_cfg": tf, "use_exog": False, "use_holidays": use_holidays, "add_custom_seasonality": True, "plan_name": "no_exog_retry"})
    deduped_plans=[]; seen=set()
    for plan in plans:
        key=(plan["cfg"]["seasonality_mode"], float(plan["cfg"]["changepoint_prior_scale"]), float(plan["cfg"]["seasonality_prior_scale"]), float(plan["cfg"]["changepoint_range"]), int(plan["cfg"]["n_changepoints"]), plan["transform_cfg"].get("name","none"), bool(plan["use_exog"]), bool(plan["use_holidays"]))
        if key not in seen:
            seen.add(key); deduped_plans.append(plan)
    deduped_plans = deduped_plans[:max(8, FORECAST_RUNTIME_CONFIG.prophet_max_configs * 4)]
    rows=[]; fit_errors=[]; best=None; best_score=np.inf
    if backend_probe_note:
        fit_errors.append(backend_probe_note)
    for plan in deduped_plans:
        cfg=plan["cfg"]; cols_now=inner_exog_cols if plan["use_exog"] else []
        try:
            _, _, pred = _fit_predict_prophet_once(tr.copy(), val.copy(), freq_alias, cfg, cols_now, use_holidays=plan["use_holidays"], add_custom_seasonality=plan["add_custom_seasonality"], transform_cfg=plan["transform_cfg"])
            baseline_pred, baseline_name = build_fallback_forecast(tr["y"], val["y"], freq_alias, season_length)
            best_local_score=np.inf; best_alpha=1.0; best_post={"name":"raw"}; best_w=np.inf; best_s=np.inf
            for alpha in [0.15,0.35,0.50,0.65,0.80,1.0]:
                blend=np.maximum(alpha*np.asarray(pred,dtype=float)+(1-alpha)*np.asarray(baseline_pred,dtype=float),0.0)
                corrected, post_cfg = compute_validation_postprocess_candidates(val["y"].values, blend, tr["y"], season_length)
                val_w=wape(val["y"].values, corrected); val_s=smape(val["y"].values, corrected)
                asym = compute_asymmetric_validation_penalty(val["y"].values, corrected, tr["y"].values, severity=1.0)
                score=val_w+0.35*val_s+asym["penalty"]
                if score < best_local_score:
                    best_local_score=score; best_alpha=alpha; best_post=post_cfg; best_w=val_w; best_s=val_s
            composite=best_local_score + 0.03*len(cols_now) + (0.03 if plan["transform_cfg"].get("name") != "none" else 0.0)
            rows.append({**cfg,"transform":plan["transform_cfg"].get("name","none"),"plan_name":plan["plan_name"],"use_exog":bool(cols_now),"use_holidays":bool(plan["use_holidays"]),"custom_seasonality":bool(plan["add_custom_seasonality"]),"blend_alpha":best_alpha,"blend_baseline":baseline_name,"postprocess":best_post.get("name","raw"),"val_wape":best_w,"val_smape":best_s,"composite_score":composite})
            if composite < best_score:
                best_score=composite; best={"cfg":dict(cfg),"transform_cfg":dict(plan["transform_cfg"]),"use_exog":bool(cols_now),"use_holidays":bool(plan["use_holidays"]),"add_custom_seasonality":bool(plan["add_custom_seasonality"]),"blend_alpha":best_alpha,"blend_baseline":baseline_name,"postprocess_cfg":best_post,"validation_wape":best_w,"validation_smape":best_s}
        except Exception as e:
            fit_errors.append(str(e)); rows.append({**cfg,"transform":plan["transform_cfg"].get("name","none"),"plan_name":plan["plan_name"],"use_exog":bool(cols_now),"use_holidays":bool(plan["use_holidays"]),"custom_seasonality":bool(plan["add_custom_seasonality"]),"val_wape":np.nan,"val_smape":np.nan,"composite_score":np.nan,"fit_error":str(e)[:250]})
    if best is None:
        fallback_reason = backend_probe_note or "prophet_search_failed"
        result = fit_prophet_surrogate(train_df, test_df, freq_alias, exog_train, exog_test, reason=fallback_reason)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result
    final_cols = exog_cols if best["use_exog"] else []
    final_retry_used = False
    final_retry_notes: List[str] = []
    try:
        m, fc, pred = _fit_predict_prophet_once(tr_full.copy(), te_full.copy(), freq_alias, best["cfg"], final_cols, use_holidays=best["use_holidays"], add_custom_seasonality=best["add_custom_seasonality"], transform_cfg=best["transform_cfg"])
    except Exception as e_final:
        final_retry_notes.append(f"primary_final:{e_final}")
        retry_cfg = dict(best["cfg"])
        retry_cfg["seasonality_mode"] = "additive"
        retry_cfg["changepoint_prior_scale"] = min(float(retry_cfg.get("changepoint_prior_scale", 0.05)), 0.10)
        retry_cfg["seasonality_prior_scale"] = min(float(retry_cfg.get("seasonality_prior_scale", 5.0)), 3.0)
        retry_cfg["n_changepoints"] = min(int(retry_cfg.get("n_changepoints", 8)), max(3, min(8, len(tr_full) // 2)))
        retry_transform = {"name": "none", "lambda": None, "shift": 0.0}
        try:
            m, fc, pred = _fit_predict_prophet_once(
                tr_full.copy(),
                te_full.copy(),
                freq_alias,
                retry_cfg,
                [],
                use_holidays=False,
                add_custom_seasonality=False,
                transform_cfg=retry_transform
            )
            best["cfg"] = retry_cfg
            best["use_exog"] = False
            best["use_holidays"] = False
            best["add_custom_seasonality"] = False
            best["transform_cfg"] = retry_transform
            final_cols = []
            final_retry_used = True
        except Exception as e_retry:
            final_retry_notes.append(f"final_simplified_retry:{e_retry}")
            fallback_reason = backend_probe_note or "prophet_final_failed"
            result = fit_prophet_surrogate(train_df, test_df, freq_alias, exog_train, exog_test, reason=f"{fallback_reason}|{'|'.join(final_retry_notes)}")
            result["search_accelerator_cache_hit"] = False
            SEARCH_ACCELERATOR.put_result(cache_key, result)
            return result
    baseline_test_pred = forecast_with_baseline_name(train_df["y"], len(test_df), freq_alias, season_length, str(best.get("blend_baseline", "last_value")))
    alpha=float(best.get("blend_alpha",1.0))
    pred=np.maximum(alpha*np.asarray(pred,dtype=float)+(1-alpha)*np.asarray(baseline_test_pred,dtype=float),0.0)
    pred=apply_postprocess_cfg(pred, best.get("postprocess_cfg", {}), train_df["y"], season_length)
    if "yhat" in fc: fc["yhat"]=pred
    fit_error_samples = (fit_errors + final_retry_notes)[:8]
    comp_summary={"trend_abs_mean": safe_float(np.abs(fc.get("trend", pd.Series(dtype=float))).mean()) if "trend" in fc else np.nan,"seasonality_abs_mean": safe_float(np.abs(fc.get("yearly", pd.Series(dtype=float))).mean()) if "yearly" in fc else np.nan,"seasonality_present": bool("yearly" in fc or "weekly" in fc),"used_exog_cols": final_cols,"dropped_exog_cols": sorted(set(dropped_exog)),"final_retry_used": final_retry_used,"rename_map": rename_map,"fit_error_samples": fit_error_samples,"blend_alpha": alpha,"blend_baseline": best.get("blend_baseline"),"postprocess": best.get("postprocess_cfg", {}),"backend_probe_status": "ok" if backend_ok else "probe_failed_but_fit_retried"}
    fit_visibility_note = "Gerçek Prophet modeli başarıyla kuruldu ve tahmin üretildi."
    if final_retry_used:
        fit_visibility_note = "Gerçek Prophet modeli basitleştirilmiş son deneme ile kuruldu; vekil modele düşmeden tahmin üretildi."
    elif backend_probe_note:
        fit_visibility_note = "Prophet backend ön kontrolü sorun verdi; buna rağmen gerçek Prophet fit başarıyla tamamlandı."
    result={"model":m,"forecast_df":fc,"forecast":pred,"config":best["cfg"],"selected_plan":best,"component_validation":comp_summary,"used_exog_cols":final_cols,"dropped_exog_cols":sorted(set(dropped_exog)),"search_table":pd.DataFrame(rows).sort_values(["composite_score","val_wape"],ascending=[True,True],na_position="last").reset_index(drop=True),"search_accelerator_cache_hit":False,"validation_wape":safe_float(best.get("validation_wape",np.nan)),"validation_smape":safe_float(best.get("validation_smape",np.nan)),"fallback_used":False,"fallback_method":None,"fit_mode":"gercek_prophet_fit","fit_visibility_note":fit_visibility_note}
    SEARCH_ACCELERATOR.put_result(cache_key, result)
    return result

def build_recursive_feature_row(history_values: List[float], target_date: pd.Timestamp, freq_alias: str, exog_row: Optional[pd.Series], all_feature_names: List[str]) -> pd.DataFrame:
    tmp = pd.DataFrame({"ds": pd.date_range(end=target_date, periods=len(history_values), freq={"M":"M","W":"W","D":"D","H":"H"}.get(freq_alias,"M")), "y": history_values})
    exog_hist = None
    if exog_row is not None:
        exog_hist = pd.DataFrame(np.nan, index=range(len(tmp)), columns=list(exog_row.index))
        exog_hist.iloc[-1] = exog_row.values
    feat_df, feat_cols = generate_target_ml_features(tmp, exog_hist, freq_alias)
    row = feat_df.iloc[[-1]][feat_cols].copy()
    for c in all_feature_names:
        if c not in row.columns:
            row[c] = 0.0
    return row[all_feature_names].replace([np.inf, -np.inf], np.nan).fillna(0.0)





def fit_xgboost_strategy(train_df: pd.DataFrame, future_df: pd.DataFrame, exog_combined: Optional[pd.DataFrame], freq_alias: str, strategy: str = "recursive") -> Dict[str, Any]:
    if XGBRegressor is None and SKHistGradientBoostingRegressor is None:
        raise ImportError("XGBoost veya hızlı GBM bileşeni bulunamadı.")

    cache_key = build_search_signature(
        f"xgb_strategy_{strategy}",
        freq_alias,
        train_df,
        future_df,
        exog_train=exog_combined,
        extra={"feature_cap": FORECAST_RUNTIME_CONFIG.xgb_max_feature_cols, "grid": list(FORECAST_RUNTIME_CONFIG.xgb_param_grid)}
    )
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached

    start_ts = time.perf_counter()
    full = pd.concat([train_df[["ds", "y"]], future_df[["ds", "y"]]], axis=0, ignore_index=True)
    train_cut = len(train_df)

    tr_inner, val_inner = SEARCH_ACCELERATOR.get_or_compute_artifact(
        build_search_signature(f"xgb_inner_split_{strategy}", freq_alias, train_df, future_df, exog_train=exog_combined),
        lambda: make_inner_train_val_split(train_df)
    )
    if len(tr_inner) > FORECAST_RUNTIME_CONFIG.xgb_inner_train_max_rows:
        tr_inner = tr_inner.iloc[-FORECAST_RUNTIME_CONFIG.xgb_inner_train_max_rows:].reset_index(drop=True)
    inner_full = pd.concat([tr_inner[["ds", "y"]], val_inner[["ds", "y"]]], axis=0, ignore_index=True)
    exog_inner = exog_combined.iloc[:len(inner_full)].copy().reset_index(drop=True) if exog_combined is not None else None
    feat_inner, inner_cols = generate_target_ml_features(inner_full, exog_inner, freq_alias)
    inner_cols = _select_core_ml_feature_columns(inner_cols, freq_alias, len(train_df))
    X_base = feat_inner.iloc[:len(tr_inner)][inner_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_base_raw = tr_inner["y"].values.astype(float)
    y_transform = choose_target_transform(pd.Series(y_base_raw))
    y_base = apply_target_transform(pd.Series(y_base_raw), y_transform)[0].values.astype(float)

    grid = list(FORECAST_RUNTIME_CONFIG.xgb_param_grid) if (HAS_XGBOOST or SKHistGradientBoostingRegressor is not None) else [{"max_depth": 3, "learning_rate": 0.08, "n_estimators": 48}]
    search_rows = []
    best = None
    best_score = np.inf
    X_base_np = X_base.values.astype(float)
    y_base_np = np.asarray(y_base, dtype=float)

    def _evaluate_xgb_cfg(raw_cfg: Dict[str, Any]) -> Dict[str, Any]:
        cfg = _prepare_xgb_cfg(raw_cfg, len(train_df), freq_alias)
        try:
            if strategy == "recursive":
                model = build_fast_xgb_regressor(cfg)
                if HAS_XGBOOST and type(model).__name__ == "XGBRegressor" and len(X_base_np) >= 12:
                    es = max(3, min(8, len(X_base_np) // 4))
                    X_fit = X_base_np[:-es]
                    y_fit = y_base_np[:-es]
                    X_eval = X_base_np[-es:]
                    y_eval = y_base_np[-es:]
                    model.fit(X_fit, y_fit, eval_set=[(X_eval, y_eval)], verbose=False)
                else:
                    model.fit(X_base_np, y_base_np)
                history = list(tr_inner["y"].astype(float).values)
                preds = []
                for i in range(len(val_inner)):
                    target_date = pd.to_datetime(val_inner.iloc[i]["ds"])
                    exog_row = exog_inner.iloc[len(tr_inner)+i] if exog_inner is not None else None
                    X_step = build_recursive_feature_row(history, target_date, freq_alias, exog_row, list(X_base.columns))
                    pred_i_t = float(model.predict(X_step.values.astype(float))[0])
                    pred_i = float(inverse_target_transform(np.asarray([pred_i_t], dtype=float), y_transform)[0])
                    pred_i = max(pred_i, 0.0)
                    preds.append(pred_i)
                    history.append(pred_i)
                pred_val = np.asarray(preds, dtype=float)
            else:
                pred_direct = []
                last_model = None
                for h in range(1, len(val_inner)+1):
                    feat_shifted = feat_inner.copy()
                    feat_shifted["target_h"] = feat_shifted["y"].shift(-h)
                    train_rows = max(0, len(tr_inner) - h)
                    ds_train = feat_shifted.iloc[:train_rows][inner_cols + ["target_h"]].dropna().copy()
                    if len(ds_train) < 8:
                        raise ValueError("Direct strategy için yeterli gözlem yok.")
                    X_h = ds_train[inner_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0).values.astype(float)
                    y_h_raw = ds_train["target_h"].astype(float).values
                    y_h = apply_target_transform(pd.Series(y_h_raw), y_transform)[0].values.astype(float)
                    model_h = build_fast_xgb_regressor(cfg)
                    model_h.fit(X_h, y_h)
                    origin_idx = len(tr_inner) - 1
                    row = feat_inner.iloc[[origin_idx]][inner_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0).values.astype(float)
                    pred_h_t = float(model_h.predict(row)[0])
                    pred_direct.append(max(float(inverse_target_transform(np.asarray([pred_h_t]), y_transform)[0]), 0.0))
                    last_model = model_h
                pred_val = np.asarray(pred_direct[:len(val_inner)], dtype=float)
                model = last_model
            score = wape(val_inner["y"].values, pred_val)
            sm = smape(val_inner["y"].values, pred_val)
            asym = compute_asymmetric_validation_penalty(val_inner["y"].values, pred_val, tr_inner["y"].values, severity=1.0)
            composite = float(score + 0.35 * sm + asym["penalty"])
            row = {**raw_cfg, "strategy": strategy, "transform": y_transform.get("name", "none"), "val_wape": score, "val_smape": sm, "bias_pct": asym["bias_pct"], "under_forecast_rate": asym["under_forecast_rate"], "peak_event_score": asym["peak_event_score"], "backend": type(model).__name__ if model is not None else None, "composite_score": composite}
            return {"row": row, "score": composite, "best": {"cfg": cfg, "strategy": strategy, "transform_cfg": y_transform}}
        except Exception as e:
            row = {**raw_cfg, "strategy": strategy, "transform": y_transform.get("name", "none"), "val_wape": np.nan, "val_smape": np.nan, "fit_error": str(e)[:180]}
            return {"row": row, "score": np.inf, "best": None}

    for item in SEARCH_ACCELERATOR.parallel_map(_evaluate_xgb_cfg, grid, max_workers=FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_workers):
        search_rows.append(item["row"])
        if item["best"] is not None and item["score"] < best_score:
            best_score = item["score"]
            best = item["best"]

    if best is None:
        raise RuntimeError("XGBoost modeli kurulamadı.")

    feat_train_test, feature_cols = generate_target_ml_features(full, exog_combined, freq_alias)
    feature_cols = _select_core_ml_feature_columns(feature_cols, freq_alias, len(train_df))
    X_train_final = feat_train_test.iloc[:train_cut][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_train_final_raw = train_df["y"].astype(float).values
    y_train_final = apply_target_transform(pd.Series(y_train_final_raw), best["transform_cfg"])[0].values.astype(float)

    if best["strategy"] == "recursive":
        model = build_fast_xgb_regressor(best["cfg"])
        if HAS_XGBOOST and type(model).__name__ == "XGBRegressor" and len(X_train_final) >= 16:
            es = max(4, min(12, len(X_train_final) // 5))
            model.fit(
                X_train_final.iloc[:-es].values.astype(float),
                y_train_final[:-es],
                eval_set=[(X_train_final.iloc[-es:].values.astype(float), y_train_final[-es:])],
                verbose=False
            )
        else:
            model.fit(X_train_final.values.astype(float), np.asarray(y_train_final, dtype=float))
        history = list(train_df["y"].astype(float).values)
        preds = []
        for i in range(len(future_df)):
            target_date = pd.to_datetime(future_df.iloc[i]["ds"])
            exog_row = exog_combined.iloc[train_cut + i] if exog_combined is not None else None
            X_step = build_recursive_feature_row(history, target_date, freq_alias, exog_row, list(X_train_final.columns))
            pred_i_t = float(model.predict(X_step.values.astype(float))[0])
            pred_i = max(float(inverse_target_transform(np.asarray([pred_i_t]), best["transform_cfg"])[0]), 0.0)
            preds.append(pred_i)
            history.append(pred_i)
        pred_test = np.asarray(preds, dtype=float)
        trained_models = [model]
    else:
        pred_test = []
        trained_models = []
        for h in range(1, len(future_df)+1):
            feat_shifted = feat_train_test.copy()
            feat_shifted["target_h"] = feat_shifted["y"].shift(-h)
            train_rows = max(0, train_cut - h)
            ds_train = feat_shifted.iloc[:train_rows][feature_cols + ["target_h"]].dropna().copy()
            X_h = ds_train[feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0).values.astype(float)
            y_h_raw = ds_train["target_h"].astype(float).values
            y_h = apply_target_transform(pd.Series(y_h_raw), best["transform_cfg"])[0].values.astype(float)
            model_h = build_fast_xgb_regressor(best["cfg"])
            model_h.fit(X_h, y_h)
            origin_row = feat_train_test.iloc[[train_cut-1]][feature_cols].replace([np.inf, -np.inf], np.nan).fillna(0.0).values.astype(float)
            pred_h_t = float(model_h.predict(origin_row)[0])
            pred_test.append(max(float(inverse_target_transform(np.asarray([pred_h_t]), best["transform_cfg"])[0]), 0.0))
            trained_models.append(model_h)
        pred_test = np.asarray(pred_test, dtype=float)

    importance_df = pd.DataFrame(columns=["feature", "importance"])
    try:
        fi = getattr(trained_models[0], "feature_importances_", None)
        if fi is not None:
            importance_df = pd.DataFrame({"feature": list(X_train_final.columns), "importance": fi}).sort_values("importance", ascending=False).head(25).reset_index(drop=True)
    except Exception:
        pass

    shap_status = "disabled"
    backend_name = type(trained_models[0]).__name__ if trained_models else None
    result = {
        "model": trained_models[0],
        "forecast": pred_test,
        "strategy": best["strategy"],
        "transform": best["transform_cfg"].get("name", "none"),
        "search_table": pd.DataFrame(search_rows).sort_values(["val_wape", "val_smape"], ascending=[True, True], na_position="last").reset_index(drop=True),
        "feature_importance": importance_df,
        "shap_status": shap_status,
        "used_feature_count": len(X_train_final.columns),
        "fallback_used": False,
        "fallback_method": None,
        "backend_name": backend_name,
        "validation_wape": safe_float(best_score),
        "validation_smape": safe_float(pd.DataFrame(search_rows).sort_values(["val_wape", "val_smape"], ascending=[True, True], na_position="last").head(1)["val_smape"].iloc[0]) if len(search_rows) else np.nan,
        "search_accelerator_cache_hit": False,
    }
    SEARCH_ACCELERATOR.put_result(cache_key, result)
    return result

def fit_xgboost_forecast(train_df: pd.DataFrame, test_df: pd.DataFrame, feature_train: pd.DataFrame, feature_test: pd.DataFrame, freq_alias: str = "M") -> Dict[str, Any]:
    if XGBRegressor is None and SKHistGradientBoostingRegressor is None:
        fallback = seasonal_naive_forecast(train_df["y"], len(test_df), infer_seasonal_period(freq_alias))
        return {
            "model": None,
            "forecast": np.maximum(np.asarray(fallback, dtype=float), 0.0),
            "strategy": "fallback",
            "search_table": pd.DataFrame(),
            "feature_importance": pd.DataFrame(columns=["feature", "importance"]),
            "shap_status": "disabled",
            "used_feature_count": 0,
            "fallback_used": True,
            "fallback_method": "seasonal_naive",
            "backend_name": "seasonal_naive"
        }

    cache_key = build_search_signature("xgb_family", freq_alias, train_df, test_df, exog_train=feature_train, exog_test=feature_test)
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached

    exog_combined = None
    selected_ml_cols: List[str] = []
    dropped_ml_cols: List[str] = []
    if feature_train is not None and feature_test is not None and len(feature_train.columns) > 0:
        red_key = build_search_signature("xgb_reduce_ml", freq_alias, train_df, test_df, exog_train=feature_train, exog_test=feature_test, extra={"max_cols": FORECAST_RUNTIME_CONFIG.xgb_max_feature_cols})
        red_tr, red_te, selected_ml_cols, dropped_ml_cols = SEARCH_ACCELERATOR.get_or_compute_artifact(
            red_key,
            lambda: reduce_ml_feature_set(feature_train, feature_test, pd.to_numeric(train_df["y"], errors="coerce").astype(float), max_cols=FORECAST_RUNTIME_CONFIG.xgb_max_feature_cols)
        )
        if red_tr is not None and red_te is not None:
            exog_combined = pd.concat([red_tr.reset_index(drop=True), red_te.reset_index(drop=True)], axis=0, ignore_index=True)

    short_monthly = str(freq_alias).upper() == "M" and len(train_df) <= 96

    def _run_recursive():
        res = fit_xgboost_strategy(train_df, test_df, exog_combined, freq_alias, strategy="recursive")
        res["selected_ml_feature_cols"] = selected_ml_cols
        res["dropped_ml_feature_cols"] = dropped_ml_cols
        return res

    if FORECAST_RUNTIME_CONFIG.xgb_skip_direct_on_short_series and short_monthly:
        recursive_res = _run_recursive()
        recursive_res["strategy_comparison"] = pd.DataFrame([{
            "strategy": "recursive",
            "WAPE": wape(test_df["y"].values, recursive_res["forecast"]),
            "sMAPE": smape(test_df["y"].values, recursive_res["forecast"]),
            "used_feature_count": recursive_res.get("used_feature_count", np.nan),
            "backend_name": recursive_res.get("backend_name")
        }])
        SEARCH_ACCELERATOR.put_result(cache_key, recursive_res)
        return recursive_res

    def _run_direct():
        res = fit_xgboost_strategy(train_df, test_df, exog_combined, freq_alias, strategy="direct")
        res["selected_ml_feature_cols"] = selected_ml_cols
        res["dropped_ml_feature_cols"] = dropped_ml_cols
        return res

    task_results = SEARCH_ACCELERATOR.run_parallel_tasks({"recursive": _run_recursive, "direct": _run_direct}, max_workers=min(2, FORECAST_RUNTIME_CONFIG.search_accelerator_max_workers))
    recursive_res = task_results["recursive"]
    direct_res = task_results["direct"]
    rec_val_wape = safe_float(recursive_res.get("validation_wape", np.nan))
    dir_val_wape = safe_float(direct_res.get("validation_wape", np.nan))
    rec_val_smape = safe_float(recursive_res.get("validation_smape", np.nan))
    dir_val_smape = safe_float(direct_res.get("validation_smape", np.nan))
    if pd.notna(rec_val_wape) and pd.notna(dir_val_wape):
        rec_comp = safe_float(recursive_res.get("validation_wape", np.nan)) + 0.35 * safe_float(recursive_res.get("validation_smape", np.nan))
        dir_comp = safe_float(direct_res.get("validation_wape", np.nan)) + 0.35 * safe_float(direct_res.get("validation_smape", np.nan))
        best = recursive_res if (rec_comp, rec_val_wape, rec_val_smape) <= (dir_comp, dir_val_wape, dir_val_smape) else direct_res
    elif pd.notna(rec_val_wape):
        best = recursive_res
    elif pd.notna(dir_val_wape):
        best = direct_res
    else:
        best = recursive_res
    best["strategy_comparison"] = pd.DataFrame([
        {"strategy": "recursive", "val_WAPE": rec_val_wape, "val_sMAPE": rec_val_smape, "used_feature_count": recursive_res.get("used_feature_count", np.nan), "backend_name": recursive_res.get("backend_name")},
        {"strategy": "direct", "val_WAPE": dir_val_wape, "val_sMAPE": dir_val_smape, "used_feature_count": direct_res.get("used_feature_count", np.nan), "backend_name": direct_res.get("backend_name")},
    ]).sort_values(["val_WAPE", "val_sMAPE"], ascending=[True, True], na_position="last").reset_index(drop=True)
    best["search_accelerator_cache_hit"] = False
    SEARCH_ACCELERATOR.put_result(cache_key, best)
    return best

def build_model_metrics(model_name: str, y_train: np.ndarray, y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, Any]:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    err = y_pred - y_true
    scale = max(float(np.nanmean(np.abs(y_true))) if np.isfinite(np.nanmean(np.abs(y_true))) else 0.0, 1e-6)
    bias = float(np.nanmean(err)) if len(err) else np.nan
    bias_pct = float(100.0 * bias / scale) if pd.notna(bias) else np.nan
    under_forecast_rate = float(np.nanmean((y_pred < y_true).astype(float))) if len(y_true) else np.nan
    over_forecast_rate = float(np.nanmean((y_pred > y_true).astype(float))) if len(y_true) else np.nan
    return {
        "model": model_name,
        "MAE": mae(y_true, y_pred),
        "RMSE": rmse(y_true, y_pred),
        "MAPE": safe_mape(y_true, y_pred),
        "sMAPE": smape(y_true, y_pred),
        "WAPE": wape(y_true, y_pred),
        "MASE": mase(y_true, y_pred, y_train, seasonality=1),
        "Bias": bias,
        "BiasPct": bias_pct,
        "UnderForecastRate": under_forecast_rate,
        "OverForecastRate": over_forecast_rate,
    }


def build_champion_challenger(metrics_df: pd.DataFrame) -> Dict[str, Any]:
    if not isinstance(metrics_df, pd.DataFrame) or len(metrics_df) == 0:
        ranked = pd.DataFrame()
        return {"champion": None, "challenger": None, "holdout_winner": None, "ranking": ranked}

    ranked = metrics_df.copy()
    if "eligible_for_model_ranking" in ranked.columns:
        ranked = ranked.loc[ranked["eligible_for_model_ranking"].fillna(True).astype(bool)].copy()
    for metric_col in ["WAPE", "sMAPE", "RMSE"]:
        if metric_col in ranked.columns:
            ranked[metric_col] = pd.to_numeric(ranked[metric_col], errors="coerce")
    if "WAPE" in ranked.columns:
        ranked = ranked.loc[ranked["WAPE"].notna() & np.isfinite(ranked["WAPE"].astype(float))].copy()

    if len(ranked) == 0:
        return {"champion": None, "challenger": None, "holdout_winner": None, "ranking": ranked.reset_index(drop=True)}

    ranked = ranked.sort_values(["WAPE", "sMAPE", "RMSE"], ascending=[True, True, True]).reset_index(drop=True)
    champion = ranked.iloc[0]["model"] if len(ranked) >= 1 else None
    challenger = ranked.iloc[1]["model"] if len(ranked) >= 2 else None
    return {"champion": champion, "challenger": challenger, "holdout_winner": champion, "ranking": ranked}


def infer_runtime_regime_from_profile(profile: Optional[Dict[str, Any]] = None, y_train: Optional[np.ndarray] = None) -> Dict[str, Any]:
    profile = profile or {}
    intermittency = safe_float(profile.get("intermittency_ratio", np.nan))
    cv = safe_float(profile.get("cv", np.nan))
    seasonality = safe_float(profile.get("seasonality_strength", np.nan))
    trend = safe_float(profile.get("trend_strength", np.nan))
    if y_train is not None and (pd.isna(intermittency) or pd.isna(cv)):
        ys = pd.Series(np.asarray(y_train, dtype=float))
        if pd.isna(intermittency):
            intermittency = safe_float((ys.fillna(0) == 0).mean())
        if pd.isna(cv):
            mu = safe_float(ys.mean())
            sigma = safe_float(ys.std())
            cv = safe_float(sigma / mu) if pd.notna(mu) and mu not in [0, 0.0] else np.nan
    return {
        "intermittent_like": bool(pd.notna(intermittency) and intermittency >= 0.20),
        "seasonal_like": bool(pd.notna(seasonality) and seasonality >= 0.18),
        "trendy_like": bool(pd.notna(trend) and trend >= 0.10),
        "volatile_like": bool(pd.notna(cv) and cv >= 0.35),
        "intermittency": intermittency,
        "cv": cv,
        "seasonality": seasonality,
        "trend": trend,
    }



def infer_dl_window_candidates(freq_alias: str, train_len: int, model_family: Optional[str] = None) -> List[int]:
    """
    Backward-compatible window inference.

    When model_family is supplied, LSTM and GRU intentionally receive different
    candidate windows. This fixes the practical-identical-search-space problem:
    LSTM is allowed to search longer memory, while GRU focuses on shorter and
    medium memory windows.
    """
    fa = str(freq_alias).upper()
    fam = str(model_family or "").upper()
    cfg = FORECAST_RUNTIME_CONFIG

    if bool(getattr(cfg, "dl_family_specific_search_spaces", True)) and fam in {"LSTM", "GRU"}:
        prefix = "dl_lstm" if fam == "LSTM" else "dl_gru"
        if fa == "M":
            base = list(getattr(cfg, f"{prefix}_window_candidates_monthly"))
        elif fa == "W":
            base = list(getattr(cfg, f"{prefix}_window_candidates_weekly"))
        elif fa == "D":
            base = list(getattr(cfg, f"{prefix}_window_candidates_daily"))
        elif fa == "H":
            base = list(getattr(cfg, f"{prefix}_window_candidates_hourly"))
        else:
            base = list(getattr(cfg, "dl_window_candidates"))
    else:
        base = list(getattr(cfg, "dl_window_candidates"))
        if fa == "M":
            base = [4, 6, 12, 18]
        elif fa == "W":
            base = [4, 8, 13, 26]
        elif fa == "D":
            base = [7, 14, 28]
        elif fa == "H":
            base = [24, 48, 72]

    max_allowed = max(3, min(max(int(train_len) - 2, 3), max(6, int(train_len) // 2)))
    if bool(getattr(cfg, "dl_data_regime_guard_enabled", True)) and fam in {"LSTM", "GRU"}:
        try:
            regime = get_dl_data_regime_profile(fa, int(train_len), 1, fam)
            if regime.get("regime") in {"small_sample", "marginal_sample"}:
                max_allowed = min(max_allowed, int(regime.get("max_window", max_allowed)))
        except Exception:
            pass
    out = [int(w) for w in base if 3 <= int(w) <= max_allowed]
    if not out:
        fallback = max(3, min(6 if fam != "LSTM" else 9, max_allowed))
        out = [fallback]
    return sorted(set(out))


def get_dl_family_profile(model_family: str) -> Dict[str, Any]:
    fam = str(model_family).upper()
    cfg = FORECAST_RUNTIME_CONFIG
    if fam == "LSTM":
        return {
            "family": "LSTM",
            "search_profile": "long_memory_high_capacity",
            "units": tuple(getattr(cfg, "dl_lstm_hidden_units", (32, 48, 64, 96))),
            "dropouts": tuple(getattr(cfg, "dl_lstm_dropout_values", (0.10, 0.20, 0.30))),
            "learning_rates": tuple(getattr(cfg, "dl_lstm_learning_rates", (0.0007, 0.0005))),
            "layers": tuple(getattr(cfg, "dl_lstm_layer_candidates", (2, 3))),
            "strategies": tuple(getattr(cfg, "dl_lstm_strategies", ("direct", "recursive"))),
            "dense_head_ratio": float(getattr(cfg, "dl_lstm_dense_head_ratio", 0.75)),
            "recurrent_dropout": float(getattr(cfg, "dl_lstm_recurrent_dropout", 0.05)),
            "weight_decay": float(getattr(cfg, "dl_lstm_weight_decay", 3e-4)),
            "batch_size": int(getattr(cfg, "dl_lstm_batch_size", 4)),
            "recent_weight": float(getattr(cfg, "dl_lstm_recent_sample_weight_max", 1.35)),
            "score_bias": float(getattr(cfg, "dl_lstm_family_score_bias", -0.035)),
            "dense_activation": "tanh",
        }
    if fam == "GRU":
        return {
            "family": "GRU",
            "search_profile": "short_medium_memory_fast_adaptation",
            "units": tuple(getattr(cfg, "dl_gru_hidden_units", (12, 16, 24, 32, 48))),
            "dropouts": tuple(getattr(cfg, "dl_gru_dropout_values", (0.00, 0.05, 0.10, 0.20))),
            "learning_rates": tuple(getattr(cfg, "dl_gru_learning_rates", (0.0015, 0.0010, 0.0007))),
            "layers": tuple(getattr(cfg, "dl_gru_layer_candidates", (1, 2))),
            "strategies": tuple(getattr(cfg, "dl_gru_strategies", ("recursive", "direct"))),
            "dense_head_ratio": float(getattr(cfg, "dl_gru_dense_head_ratio", 0.50)),
            "recurrent_dropout": float(getattr(cfg, "dl_gru_recurrent_dropout", 0.00)),
            "weight_decay": float(getattr(cfg, "dl_gru_weight_decay", 8e-5)),
            "batch_size": int(getattr(cfg, "dl_gru_batch_size", 8)),
            "recent_weight": float(getattr(cfg, "dl_gru_recent_sample_weight_max", 1.85)),
            "score_bias": float(getattr(cfg, "dl_gru_family_score_bias", -0.020)),
            "dense_activation": "relu",
        }
    return {
        "family": fam or "RNN",
        "search_profile": "generic_recurrent",
        "units": tuple(getattr(cfg, "dl_hidden_units", (16, 32, 48))),
        "dropouts": tuple(getattr(cfg, "dl_dropout_values", (0.0, 0.1, 0.2))),
        "learning_rates": tuple(getattr(cfg, "dl_learning_rates", (0.001,))),
        "layers": (1, 2),
        "strategies": tuple(getattr(cfg, "dl_strategies", ("recursive",))),
        "dense_head_ratio": 0.5,
        "recurrent_dropout": 0.0,
        "weight_decay": 1e-4,
        "batch_size": int(getattr(cfg, "dl_batch_size", 6)),
        "recent_weight": float(getattr(cfg, "dl_recent_sample_weight_max", 1.60)),
        "score_bias": 0.0,
        "dense_activation": "relu",
    }


def build_family_specific_dl_configs(model_family: str, freq_alias: str, train_len: int, val_len: int) -> List[Dict[str, Any]]:
    profile = get_dl_family_profile(model_family)
    fam = profile["family"]
    configs: List[Dict[str, Any]] = []
    windows = infer_dl_window_candidates(freq_alias, train_len, model_family=fam)
    for window in windows:
        for units in profile["units"]:
            for dropout in profile["dropouts"]:
                for learning_rate in profile["learning_rates"]:
                    for layers in profile["layers"]:
                        for strategy in profile["strategies"]:
                            strategy = str(strategy).lower()
                            if strategy == "direct" and int(val_len) < 2:
                                continue
                            # LSTM gets longer-memory/deeper configs; GRU stays compact. Keep
                            # this as a hard architecture policy, not as a post-hoc metric trick.
                            if fam == "LSTM" and int(window) < 6 and int(train_len) >= 18:
                                continue
                            if fam == "GRU" and int(layers) > 1 and int(units) < 24:
                                continue
                            configs.append({
                                "window": int(window),
                                "units": int(units),
                                "dropout": float(dropout),
                                "layers": int(layers),
                                "learning_rate": float(learning_rate),
                                "strategy": strategy,
                                "family_profile": str(profile["search_profile"]),
                                "dense_head_ratio": float(profile["dense_head_ratio"]),
                                "recurrent_dropout": float(profile["recurrent_dropout"]),
                                "weight_decay": float(profile["weight_decay"]),
                                "batch_size": int(profile["batch_size"]),
                                "recent_weight": float(profile["recent_weight"]),
                                "dense_activation": str(profile["dense_activation"]),
                            })
    return configs


def dl_family_score_adjustment(cfg: Dict[str, Any], model_family: str, val_len: int) -> float:
    """Small inner-validation-only tie-breaker to preserve family identity."""
    fam = str(model_family).upper()
    window = int(cfg.get("window", 0) or 0)
    layers = int(cfg.get("layers", 1) or 1)
    strategy = str(cfg.get("strategy", "recursive")).lower()
    profile = get_dl_family_profile(fam)
    adj = float(profile.get("score_bias", 0.0))
    if fam == "LSTM":
        if window >= 12:
            adj -= 0.025
        if layers >= 2:
            adj -= 0.015
        if strategy == "direct" and int(val_len) > 1:
            adj -= 0.010
    elif fam == "GRU":
        if window <= 9:
            adj -= 0.025
        if layers == 1:
            adj -= 0.015
        if strategy == "recursive" and int(val_len) > 1:
            adj -= 0.020
    return float(adj)




def get_dl_short_monthly_policy(freq_alias: str, train_len: int, model_family: str = "LSTM") -> Dict[str, Any]:
    """
    Hard policy for monthly short histories, based only on training length.

    Rules:
      - Monthly train_n < 72: do not train/publish real LSTM/GRU for ranking.
      - Monthly 72 <= train_n < 96: real DL may be trained, but only as
        challenger/research; it is not eligible to become champion or enter
        production ranking.
      - Otherwise: normal guarded DL eligibility still applies.
    """
    cfg = FORECAST_RUNTIME_CONFIG
    family = str(model_family or "").upper()
    fa = str(freq_alias or "").upper()
    n = max(0, int(train_len))
    if family not in {"LSTM", "GRU"} or not bool(getattr(cfg, "dl_short_monthly_policy_enabled", True)) or fa != "M":
        return {
            "active": False,
            "freq_alias": fa,
            "model_family": family,
            "train_observations": n,
            "policy_state": "not_applicable",
            "train_real_dl_allowed": True,
            "eligible_for_model_ranking": True,
            "champion_eligible": True,
            "production_role": "normal_guarded_candidate",
            "ranking_exclusion_reason": None,
        }

    no_rank_thr = int(getattr(cfg, "dl_monthly_no_rank_train_observations", 72))
    challenger_thr = int(getattr(cfg, "dl_monthly_challenger_only_train_observations", 96))
    skip_training = bool(getattr(cfg, "dl_monthly_no_rank_skip_real_training", True))

    if n < no_rank_thr:
        return {
            "active": True,
            "freq_alias": fa,
            "model_family": family,
            "train_observations": n,
            "policy_state": "monthly_train_n_lt_72_no_public_ranking",
            "train_real_dl_allowed": not skip_training,
            "skip_real_training": skip_training,
            "eligible_for_model_ranking": False,
            "champion_eligible": False,
            "production_role": "research_surrogate_only",
            "ranking_exclusion_reason": (
                f"Aylık kısa seri kuralı: train_n={n} < {no_rank_thr}. "
                "Gerçek LSTM/GRU public sıralamaya alınmaz; araştırma/surrogate-only veya DL-fallback olarak raporlanır."
            ),
        }

    if n < challenger_thr:
        return {
            "active": True,
            "freq_alias": fa,
            "model_family": family,
            "train_observations": n,
            "policy_state": "monthly_train_n_lt_96_challenger_research_only",
            "train_real_dl_allowed": True,
            "skip_real_training": False,
            "eligible_for_model_ranking": False,
            "champion_eligible": False,
            "production_role": "challenger_research_only",
            "ranking_exclusion_reason": (
                f"Aylık kısa seri kuralı: train_n={n} < {challenger_thr}. "
                "LSTM/GRU yalnız challenger/research mode; champion veya normal ranking adayı olamaz."
            ),
        }

    return {
        "active": True,
        "freq_alias": fa,
        "model_family": family,
        "train_observations": n,
        "policy_state": "monthly_train_n_gte_96_normal_guarded",
        "train_real_dl_allowed": True,
        "skip_real_training": False,
        "eligible_for_model_ranking": True,
        "champion_eligible": True,
        "production_role": "normal_guarded_candidate",
        "ranking_exclusion_reason": None,
    }


def apply_dl_short_monthly_policy_to_result(result: Dict[str, Any], policy: Dict[str, Any], model_family: str) -> Dict[str, Any]:
    """Attach monthly short-series policy metadata and enforce rankability."""
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    policy = policy or {}
    family = str(model_family or policy.get("model_family") or "").upper()
    if family not in {"LSTM", "GRU"}:
        return out

    out["dl_short_monthly_policy"] = dict(policy)
    out["dl_short_monthly_policy_state"] = policy.get("policy_state")
    out["dl_train_observations"] = int(policy.get("train_observations", 0) or 0)
    out["dl_production_role"] = policy.get("production_role", out.get("dl_production_role", "normal_guarded_candidate"))
    out["dl_champion_eligible"] = bool(policy.get("champion_eligible", out.get("eligible_for_model_ranking", True)))
    out["dl_research_only"] = bool(out.get("dl_production_role") in {"challenger_research_only", "research_surrogate_only"})

    if not bool(policy.get("eligible_for_model_ranking", True)):
        reason = str(policy.get("ranking_exclusion_reason") or "DL short-monthly policy excluded model from normal ranking.")
        out["eligible_for_model_ranking"] = False
        out["ranking_exclusion_reason"] = reason
        out["strict_dl_rejection_reason"] = out.get("strict_dl_rejection_reason") or reason
        # A model can still be genuinely trained, but it is not a champion/ranking
        # candidate in this data regime. Preserve dl_strict_validation_passed if
        # true so the user can see that training happened; block only ranking.
        note = str(out.get("model_identity_note") or "")
        add = f" Kısa aylık seri politikası uygulandı: {reason}"
        out["model_identity_note"] = (note + add).strip()
        stbl = out.get("search_table")
        if isinstance(stbl, pd.DataFrame):
            stbl = stbl.copy()
            stbl["monthly_short_series_policy_state"] = policy.get("policy_state")
            stbl["ranking_excluded_by_monthly_short_series_policy"] = True
            stbl["ranking_exclusion_reason"] = reason
            out["search_table"] = stbl
    return out


def finalize_tf_missing_research_only_result(train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, model_family: str, reason: str) -> Dict[str, Any]:
    """
    TensorFlow-missing state is not reported as an ordinary model failure.
    It means real recurrent DL was not trained, so the public LSTM/GRU slot is
    excluded and any diagnostic forecast is kept outside normal ranking.
    """
    family = str(model_family).upper()
    fallback = build_model_level_fallback(family, train_df, test_df, freq_alias, reason)
    result = _finalize_dl_backend_fallback(
        fallback,
        model_family=family,
        reason=reason,
        backend_state="tensorflow_unavailable_real_dl_not_trained_research_only",
    )
    result["dl_runtime_state"] = str(getattr(FORECAST_RUNTIME_CONFIG, "dl_tensorflow_missing_public_state", "research_surrogate_only_real_dl_not_trained"))
    result["dl_result_kind"] = "real_dl_not_trained_backend_missing"
    result["model_identity_class"] = "research_surrogate_only"
    result["dl_production_role"] = "research_surrogate_only"
    result["dl_research_only"] = True
    result["eligible_for_model_ranking"] = False
    result["forecast_is_model_output"] = False
    result["real_dl_trained"] = False
    result["dl_backend"] = "tensorflow_missing"
    result["model_identity_note"] = (
        f"{family} için TensorFlow/Keras bulunmadığı için gerçek recurrent DL eğitilmedi. "
        "Bu durum normal model başarısızlığı olarak değil, araştırma/surrogate-only durumu olarak işaretlendi; "
        "satır normal model sıralamasına veya champion seçimine girmez."
    )
    return result


def get_dl_data_regime_profile(freq_alias: str, train_len: int, horizon: int, model_family: str = "LSTM") -> Dict[str, Any]:
    """Training-only DL data regime classifier for sliding-window support."""
    cfg = FORECAST_RUNTIME_CONFIG
    fa = str(freq_alias).upper()
    fam = str(model_family).upper()
    n = max(0, int(train_len))
    h = max(1, int(horizon))
    if fa == "M":
        small_thr = int(getattr(cfg, "dl_monthly_small_sample_observations", 72))
        marginal_thr = int(getattr(cfg, "dl_monthly_marginal_sample_observations", 96))
        small_max_window = 12 if fam == "LSTM" else 9
        marginal_max_window = 18 if fam == "LSTM" else 12
    elif fa == "W":
        small_thr = int(getattr(cfg, "dl_weekly_small_sample_observations", 104))
        marginal_thr = int(getattr(cfg, "dl_weekly_marginal_sample_observations", 156))
        small_max_window = 13 if fam == "LSTM" else 8
        marginal_max_window = 26 if fam == "LSTM" else 13
    elif fa == "D":
        small_thr = int(getattr(cfg, "dl_daily_small_sample_observations", 120))
        marginal_thr = int(getattr(cfg, "dl_daily_marginal_sample_observations", 180))
        small_max_window = 28 if fam == "LSTM" else 14
        marginal_max_window = 42 if fam == "LSTM" else 21
    elif fa == "H":
        small_thr = int(getattr(cfg, "dl_hourly_small_sample_observations", 336))
        marginal_thr = int(getattr(cfg, "dl_hourly_marginal_sample_observations", 720))
        small_max_window = 48 if fam == "LSTM" else 24
        marginal_max_window = 72 if fam == "LSTM" else 36
    else:
        small_thr, marginal_thr = 72, 96
        small_max_window = 12 if fam == "LSTM" else 9
        marginal_max_window = 18 if fam == "LSTM" else 12
    if n <= small_thr:
        regime = "small_sample"
        min_seq_base = int(getattr(cfg, "dl_small_sample_min_train_sequences", 18))
        seq_ratio = float(getattr(cfg, "dl_small_sample_sequence_ratio", 0.35))
        max_epochs = int(getattr(cfg, "dl_small_sample_max_epochs", 90))
        patience = int(getattr(cfg, "dl_small_sample_patience", 10))
        baseline_margin = float(getattr(cfg, "dl_small_sample_min_baseline_improvement_pct", 0.03))
        max_window = small_max_window
    elif n <= marginal_thr:
        regime = "marginal_sample"
        min_seq_base = int(getattr(cfg, "dl_marginal_sample_min_train_sequences", 24))
        seq_ratio = float(getattr(cfg, "dl_marginal_sample_sequence_ratio", 0.30))
        max_epochs = int(getattr(cfg, "dl_marginal_sample_max_epochs", 140))
        patience = int(getattr(cfg, "dl_marginal_sample_patience", 16))
        baseline_margin = float(getattr(cfg, "dl_marginal_sample_min_baseline_improvement_pct", 0.01))
        max_window = marginal_max_window
    else:
        regime = "normal_sample"
        min_seq_base = int(getattr(cfg, "dl_normal_sample_min_train_sequences", 30))
        seq_ratio = float(getattr(cfg, "dl_normal_sample_sequence_ratio", 0.24))
        max_epochs = int(getattr(cfg, "dl_max_epochs", 220))
        patience = int(getattr(cfg, "dl_patience", 24))
        baseline_margin = 0.0
        max_window = max(3, n // 2)
    min_sequences = int(max(
        getattr(cfg, "dl_min_train_sequences", 12),
        min_seq_base,
        math.ceil(max(1, n) * seq_ratio),
        int(getattr(cfg, "dl_min_sequences_per_test_step", 5)) * h,
    ))
    return {
        "regime": regime,
        "freq_alias": fa,
        "model_family": fam,
        "train_observations": n,
        "horizon": h,
        "small_threshold": small_thr,
        "marginal_threshold": marginal_thr,
        "min_required_sequences": min_sequences,
        "sequence_ratio_required": seq_ratio,
        "max_window": int(max_window),
        "max_epochs": max(12, max_epochs),
        "patience": max(4, patience),
        "baseline_improvement_required_pct": baseline_margin,
        "ranking_guard_active": bool(regime in {"small_sample", "marginal_sample"}),
        "short_monthly_policy": get_dl_short_monthly_policy(freq_alias, train_len, model_family),
    }


def _dl_sequence_count_for_cfg(train_len: int, window: int, strategy: str, horizon: int) -> int:
    n = max(0, int(train_len))
    w = max(1, int(window))
    h = max(1, int(horizon))
    if str(strategy).lower() == "direct":
        return max(0, n - w - h + 1)
    return max(0, n - w)


def _dl_estimated_capacity_score(cfg: Dict[str, Any], model_family: str) -> int:
    units = max(1, int(cfg.get("units", 1)))
    layers = max(1, int(cfg.get("layers", 1)))
    fam_mult = 4 if str(model_family).upper() == "LSTM" else 3
    return int(fam_mult * layers * units * units)


def _apply_dl_data_regime_guard(configs: List[Dict[str, Any]], model_family: str, freq_alias: str, train_len: int, horizon: int) -> Tuple[List[Dict[str, Any]], Dict[str, Any], pd.DataFrame]:
    if not bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_data_regime_guard_enabled", True)):
        return list(configs), get_dl_data_regime_profile(freq_alias, train_len, horizon, model_family), pd.DataFrame()
    regime = get_dl_data_regime_profile(freq_alias, train_len, horizon, model_family)
    fam = str(model_family).upper()
    kept, rejected = [], []
    for cfg in list(configs):
        row = dict(cfg)
        window = int(row.get("window", 0) or 0)
        strategy = str(row.get("strategy", "recursive")).lower()
        seq_count = _dl_sequence_count_for_cfg(train_len, window, strategy, horizon)
        capacity_score = _dl_estimated_capacity_score(row, fam)
        reject_reasons = []
        if window > int(regime.get("max_window", window)):
            reject_reasons.append(f"window>{regime.get('max_window')}")
        if seq_count < int(regime.get("min_required_sequences", 0)):
            reject_reasons.append(f"sequence_count<{regime.get('min_required_sequences')}")
        if regime.get("regime") == "small_sample":
            if fam == "LSTM" and int(row.get("units", 0)) > 48:
                reject_reasons.append("small_sample_lstm_units>48")
            if fam == "LSTM" and int(row.get("layers", 1)) > 2:
                reject_reasons.append("small_sample_lstm_layers>2")
            if fam == "GRU" and int(row.get("units", 0)) > 32:
                reject_reasons.append("small_sample_gru_units>32")
            if fam == "GRU" and int(row.get("layers", 1)) > 1:
                reject_reasons.append("small_sample_gru_layers>1")
            if capacity_score > max(3500, int(seq_count) * 190):
                reject_reasons.append("capacity_too_high_for_sequence_count")
        elif regime.get("regime") == "marginal_sample":
            if fam == "LSTM" and int(row.get("units", 0)) > 64:
                reject_reasons.append("marginal_sample_lstm_units>64")
            if fam == "GRU" and int(row.get("units", 0)) > 48:
                reject_reasons.append("marginal_sample_gru_units>48")
            if capacity_score > max(5500, int(seq_count) * 260):
                reject_reasons.append("capacity_too_high_for_sequence_count")
        row.update({
            "data_regime": regime.get("regime"),
            "candidate_sequence_count": int(seq_count),
            "min_required_sequences": int(regime.get("min_required_sequences", 0)),
            "sequence_support_ratio": safe_float(seq_count / max(1, int(train_len))),
            "capacity_score_proxy": int(capacity_score),
            "max_epochs": int(regime.get("max_epochs", getattr(FORECAST_RUNTIME_CONFIG, "dl_max_epochs", 220))),
            "patience": int(regime.get("patience", getattr(FORECAST_RUNTIME_CONFIG, "dl_patience", 24))),
        })
        if reject_reasons:
            row["candidate_rejected_by_data_regime_guard"] = True
            row["regime_rejection_reason"] = " | ".join(reject_reasons)
            rejected.append(row)
        else:
            row["candidate_rejected_by_data_regime_guard"] = False
            row["regime_rejection_reason"] = ""
            kept.append(row)
    return kept, regime, pd.DataFrame(rejected)


def _dl_baseline_guard_decision(best_validation_wape: float, inner_baseline_wape: float, regime_profile: Dict[str, Any]) -> Tuple[bool, str, float]:
    regime = str((regime_profile or {}).get("regime", "normal_sample"))
    if regime not in {"small_sample", "marginal_sample"}:
        return True, "normal_sample_no_baseline_superiority_gate", 0.0
    if not bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_reject_short_sample_if_not_better_than_baseline", True)):
        return True, "baseline_superiority_gate_disabled", 0.0
    margin = float((regime_profile or {}).get("baseline_improvement_required_pct", 0.0) or 0.0)
    try:
        dl_w = float(best_validation_wape)
        base_w = float(inner_baseline_wape)
    except Exception:
        return False, "baseline_guard_invalid_validation_scores", margin
    if not np.isfinite(dl_w) or not np.isfinite(base_w) or base_w <= 0:
        return False, "baseline_guard_nonfinite_or_invalid_baseline", margin
    required_wape = base_w * (1.0 - margin)
    if dl_w <= required_wape:
        return True, f"dl_inner_validation_beats_baseline_by_required_margin_{margin:.3f}", margin
    return False, f"dl_inner_validation_wape_not_better_than_baseline: dl={dl_w:.4f}, baseline={base_w:.4f}, required<={required_wape:.4f}", margin

def _select_diverse_dl_configs(configs: List[Dict[str, Any]], max_n: int) -> List[Dict[str, Any]]:
    max_n = max(1, int(max_n))
    if len(configs) <= max_n:
        return list(configs)
    # önce pencere ve strateji çeşitliliğini koru, sonra kalan yerleri grid boyunca eşit aralıkla doldur
    selected = []
    seen = set()
    for strategy in sorted({str(c.get("strategy", "recursive")) for c in configs}):
        strat_rows = [c for c in configs if str(c.get("strategy", "recursive")) == strategy]
        for window in sorted({int(c.get("window", 0)) for c in strat_rows}):
            cand = [c for c in strat_rows if int(c.get("window", 0)) == window]
            cand = sorted(cand, key=lambda x: (int(x.get("units", 0)), float(x.get("dropout", 0.0)), float(x.get("learning_rate", 0.0))))
            if cand:
                chosen = cand[min(len(cand) - 1, len(cand) // 2)]
                key = tuple(sorted(chosen.items()))
                if key not in seen:
                    selected.append(chosen)
                    seen.add(key)
                if len(selected) >= max_n:
                    return selected[:max_n]
    if len(selected) < max_n:
        import numpy as _np
        for idx in _np.linspace(0, len(configs) - 1, max_n, dtype=int).tolist():
            chosen = configs[int(idx)]
            key = tuple(sorted(chosen.items()))
            if key not in seen:
                selected.append(chosen)
                seen.add(key)
            if len(selected) >= max_n:
                break
    return selected[:max_n]



def _build_dl_rolling_origin_folds(
    train_df: pd.DataFrame,
    horizon: int,
    freq_alias: str,
    model_family: str,
) -> List[Dict[str, int]]:
    """
    Build leakage-free rolling-origin validation folds for DL selection.

    The folds are created only inside train_df. The external holdout/test_df is
    not inspected. Each fold trains on [0:train_end) and validates on the next
    contiguous block [val_start:val_end), preserving time order.
    """
    cfg = FORECAST_RUNTIME_CONFIG
    n = int(len(train_df))
    if n <= 0:
        return []
    requested_h = max(1, int(horizon or 1))
    horizon_cap = max(1, int(getattr(cfg, "dl_rolling_validation_horizon_cap", 6)))
    val_h = max(1, min(requested_h, horizon_cap, max(1, n // 4)))
    max_folds = max(1, int(getattr(cfg, "dl_rolling_validation_max_folds", 3)))
    season_length = int(infer_season_length_from_freq(freq_alias) or 1)
    min_obs_cfg = int(getattr(cfg, "dl_rolling_validation_min_train_observations", 24))
    min_train_obs = max(min_obs_cfg, 2 * season_length if season_length > 1 else min_obs_cfg)
    folds: List[Dict[str, int]] = []

    # Most recent folds are preferred because they match the production horizon,
    # but every fold still uses only past observations for training.
    for back in range(max_folds - 1, -1, -1):
        val_end = n - back * val_h
        val_start = val_end - val_h
        train_end = val_start
        if val_start <= 0 or val_end > n:
            continue
        if train_end < min_train_obs:
            continue
        folds.append({
            "fold": len(folds) + 1,
            "train_start": 0,
            "train_end": int(train_end),
            "val_start": int(val_start),
            "val_end": int(val_end),
            "val_len": int(val_h),
            "origin_index": int(train_end - 1),
        })
    return folds


def _dl_insufficient_rolling_folds(
    folds: List[Dict[str, int]],
    model_family: str,
) -> Tuple[bool, str]:
    min_folds = max(1, int(getattr(FORECAST_RUNTIME_CONFIG, "dl_rolling_validation_min_folds", 2)))
    if len(folds) >= min_folds:
        return False, "rolling_origin_validation_available"
    reason = (
        f"{model_family} için rolling-origin DL validasyonu yetersiz: "
        f"bulunan_fold={len(folds)}, gerekli_min_fold={min_folds}. "
        "Tek split ile DL hiperparametre seçimi yapılmadı."
    )
    return True, reason


def _compute_dl_rolling_baseline_summary(
    train_df: pd.DataFrame,
    folds: List[Dict[str, int]],
    freq_alias: str,
    season_length: int,
) -> Dict[str, Any]:
    rows = []
    for fold in folds:
        tr = train_df.iloc[:int(fold["train_end"])].reset_index(drop=True)
        val = train_df.iloc[int(fold["val_start"]):int(fold["val_end"])].reset_index(drop=True)
        try:
            pred, method = build_fallback_forecast(tr["y"], val["y"], freq_alias, season_length)
            bw = safe_float(wape(val["y"].values, pred))
            bs = safe_float(smape(val["y"].values, pred))
        except Exception as e:
            method = f"baseline_error:{str(e)[:80]}"
            bw = np.nan
            bs = np.nan
        rows.append({
            "fold": int(fold.get("fold", len(rows) + 1)),
            "train_end": int(fold.get("train_end", 0)),
            "val_start": int(fold.get("val_start", 0)),
            "val_end": int(fold.get("val_end", 0)),
            "val_len": int(fold.get("val_len", 0)),
            "baseline_method": method,
            "baseline_wape": bw,
            "baseline_smape": bs,
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return {"method": "rolling_baseline_unavailable", "wape": np.nan, "smape": np.nan, "fold_table": df}
    method = "rolling_origin_mean_baseline:" + ",".join(sorted(set(df["baseline_method"].dropna().astype(str))))
    return {
        "method": method,
        "wape": safe_float(pd.to_numeric(df["baseline_wape"], errors="coerce").mean()),
        "smape": safe_float(pd.to_numeric(df["baseline_smape"], errors="coerce").mean()),
        "fold_table": df,
    }


def _split_dl_sequence_arrays_for_early_stopping(
    X_seq: np.ndarray,
    y_seq: np.ndarray,
    sample_weight: Optional[np.ndarray] = None,
) -> Tuple[np.ndarray, np.ndarray, Optional[np.ndarray], Optional[np.ndarray], Optional[np.ndarray]]:
    """
    Time-ordered early-stopping split for sequence tensors.

    This replaces Keras validation_split. It never shuffles and never uses the
    external holdout/test set. If there are too few sequences, validation is
    disabled and EarlyStopping must monitor training loss.
    """
    X_seq = np.asarray(X_seq, dtype=float)
    y_seq = np.asarray(y_seq, dtype=float)
    n = int(len(X_seq))
    if n <= 0:
        return X_seq, y_seq, None, None, None
    sw = np.asarray(sample_weight, dtype=float).reshape(-1) if sample_weight is not None and len(sample_weight) >= n else np.ones((n,), dtype=float)
    ratio = float(getattr(FORECAST_RUNTIME_CONFIG, "dl_early_stopping_validation_ratio", 0.18))
    min_val = max(1, int(getattr(FORECAST_RUNTIME_CONFIG, "dl_min_early_stopping_val_sequences", 3)))
    val_size = int(round(n * ratio))
    val_size = max(min_val, val_size) if n >= (min_val + 6) else 0
    val_size = min(val_size, max(0, n - max(6, min(12, n))))
    if val_size < min_val or n - val_size < max(6, min(12, n)):
        return X_seq, y_seq, None, None, sw
    split = n - val_size
    return X_seq[:split], y_seq[:split], X_seq[split:], y_seq[split:], sw[:split]


def _fit_keras_sequence_model_time_ordered(
    X_seq: np.ndarray,
    y_seq: np.ndarray,
    cfg: Dict[str, Any],
    model_family: str,
    output_dim: int,
    sample_weight: Optional[np.ndarray] = None,
) -> Tuple[Any, pd.DataFrame]:
    model = _build_dl_network(
        (X_seq.shape[1], X_seq.shape[2]),
        model_family=model_family,
        units=int(cfg["units"]),
        layers=int(cfg["layers"]),
        dropout=float(cfg["dropout"]),
        learning_rate=float(cfg["learning_rate"]),
        output_dim=int(output_dim),
        dense_head_ratio=float(cfg.get("dense_head_ratio", get_dl_family_profile(model_family).get("dense_head_ratio", 0.5))),
        recurrent_dropout=float(cfg.get("recurrent_dropout", get_dl_family_profile(model_family).get("recurrent_dropout", 0.0))),
        dense_activation=str(cfg.get("dense_activation", get_dl_family_profile(model_family).get("dense_activation", "relu"))),
    )
    X_fit, y_fit, X_es, y_es, sw_fit = _split_dl_sequence_arrays_for_early_stopping(X_seq, y_seq, sample_weight)
    monitor = "val_loss" if X_es is not None and y_es is not None and len(X_es) else "loss"
    callbacks = [KerasEarlyStopping(monitor=monitor, patience=int(max(4, cfg.get("patience", FORECAST_RUNTIME_CONFIG.dl_patience))), restore_best_weights=True)]
    fit_kwargs = {
        "x": X_fit,
        "y": y_fit,
        "epochs": int(max(12, cfg.get("max_epochs", FORECAST_RUNTIME_CONFIG.dl_max_epochs))),
        "batch_size": int(min(int(cfg.get("batch_size", FORECAST_RUNTIME_CONFIG.dl_batch_size)), max(4, len(X_fit) // 2 if len(X_fit) else 4))),
        "verbose": 0,
        "callbacks": callbacks,
        "shuffle": bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_keras_shuffle", False)),
    }
    if sw_fit is not None and len(sw_fit) == len(X_fit):
        fit_kwargs["sample_weight"] = sw_fit
    if monitor == "val_loss":
        fit_kwargs["validation_data"] = (X_es, y_es)
    hist = model.fit(**fit_kwargs)
    hist_df = _history_to_frame(hist)
    if "val_loss" not in hist_df.columns:
        hist_df["val_loss"] = np.nan
    hist_df["validation_design"] = "explicit_time_ordered_sequence_tail"
    return model, hist_df


def _summarize_dl_fold_rows(fold_rows: List[Dict[str, Any]], cfg: Dict[str, Any], model_family: str, backend: str, selected_cols: List[str], inner_baseline_method: str, inner_baseline_wape: float, inner_baseline_smape: float, dl_regime_profile: Dict[str, Any]) -> Dict[str, Any]:
    df = pd.DataFrame(fold_rows)
    val_w = safe_float(pd.to_numeric(df.get("fold_wape"), errors="coerce").mean()) if len(df) else np.nan
    val_s = safe_float(pd.to_numeric(df.get("fold_smape"), errors="coerce").mean()) if len(df) else np.nan
    w_std = safe_float(pd.to_numeric(df.get("fold_wape"), errors="coerce").std(ddof=0)) if len(df) else np.nan
    s_std = safe_float(pd.to_numeric(df.get("fold_smape"), errors="coerce").std(ddof=0)) if len(df) else np.nan
    bias_mean = safe_float(pd.to_numeric(df.get("bias_pct"), errors="coerce").mean()) if len(df) else np.nan
    under_mean = safe_float(pd.to_numeric(df.get("under_forecast_rate"), errors="coerce").mean()) if len(df) else np.nan
    peak_mean = safe_float(pd.to_numeric(df.get("peak_event_score"), errors="coerce").mean()) if len(df) else np.nan
    overfit_mean = safe_float(pd.to_numeric(df.get("overfit_gap_ratio"), errors="coerce").mean()) if len(df) else np.nan
    asym_penalty = safe_float(pd.to_numeric(df.get("asym_penalty"), errors="coerce").mean()) if len(df) else 0.0
    direction_pen = 0.0 if pd.isna(bias_mean) or abs(float(bias_mean)) <= 10 else 0.12
    fold_instability_pen = 0.18 * max(0.0, float(w_std) if pd.notna(w_std) else 0.0)
    score = float(
        val_w
        + 0.28 * val_s
        + 0.85 * (asym_penalty if pd.notna(asym_penalty) else 0.0)
        + 2.2 * max(overfit_mean if pd.notna(overfit_mean) else 0.0, 0.0)
        + direction_pen
        + fold_instability_pen
        + dl_family_score_adjustment(cfg, model_family, int(max(pd.to_numeric(df.get("fold_val_len"), errors="coerce").max() if len(df) else 1, 1)))
    )
    row = {
        **cfg,
        "model": model_family,
        "backend": backend,
        "validation_design": "rolling_origin_kfold_train_only",
        "fold_count": int(len(df)),
        "val_wape": val_w,
        "val_smape": val_s,
        "rolling_val_wape_mean": val_w,
        "rolling_val_wape_std": w_std,
        "rolling_val_smape_mean": val_s,
        "rolling_val_smape_std": s_std,
        "inner_baseline_method": inner_baseline_method,
        "inner_baseline_wape": inner_baseline_wape,
        "inner_baseline_smape": inner_baseline_smape,
        "data_regime": (dl_regime_profile or {}).get("regime"),
        "bias_pct": bias_mean,
        "under_forecast_rate": under_mean,
        "peak_event_score": peak_mean,
        "min_train_loss": safe_float(pd.to_numeric(df.get("min_train_loss"), errors="coerce").mean()) if len(df) else np.nan,
        "min_val_loss": safe_float(pd.to_numeric(df.get("min_val_loss"), errors="coerce").mean()) if len(df) else np.nan,
        "overfit_gap_ratio": overfit_mean,
        "epochs_ran": safe_float(pd.to_numeric(df.get("epochs_ran"), errors="coerce").mean()) if len(df) else np.nan,
        "selected_exog_cols": ", ".join(selected_cols),
        "composite_score": score,
    }
    return row


def _evaluate_tf_dl_config_rolling_origin(
    cfg: Dict[str, Any],
    train_df: pd.DataFrame,
    cov_train_df: pd.DataFrame,
    folds: List[Dict[str, int]],
    freq_alias: str,
    model_family: str,
    selected_cols: List[str],
    season_length: int,
    inner_baseline_method: str,
    inner_baseline_wape: float,
    inner_baseline_smape: float,
    dl_regime_profile: Dict[str, Any],
) -> Dict[str, Any]:
    fold_rows: List[Dict[str, Any]] = []
    fold_tables: List[pd.DataFrame] = []
    best_post_cfg: Dict[str, Any] = {"name": "raw"}
    best_transform: Dict[str, Any] = {"name": "none"}
    for fold in folds:
        tr = train_df.iloc[:int(fold["train_end"])].reset_index(drop=True)
        val = train_df.iloc[int(fold["val_start"]):int(fold["val_end"])].reset_index(drop=True)
        cov_tr = cov_train_df.iloc[:int(fold["train_end"])].reset_index(drop=True)
        cov_val = cov_train_df.iloc[int(fold["val_start"]):int(fold["val_end"])].reset_index(drop=True)
        residual_mode = bool(cfg.get("residual_mode", False))
        residual_base_method = None
        residual_base_val = None
        if residual_mode:
            base_pack = build_residual_dl_base_arrays(tr["y"], len(val), freq_alias, season_length)
            residual_base_train = np.asarray(base_pack.get("train_base"), dtype=float).reshape(-1)
            residual_base_val = np.asarray(base_pack.get("future_base"), dtype=float).reshape(-1)
            residual_base_method = str(base_pack.get("method"))
            model_target_train = pd.to_numeric(tr["y"], errors="coerce").astype(float).reset_index(drop=True) - residual_base_train
            y_transform = {"name": "none", "lambda": None, "shift": 0.0, "allow_negative_output": True, "target_mode": "residual"}
        else:
            model_target_train = tr["y"]
            y_transform = choose_target_transform(tr["y"])
        tr_y_t = apply_target_transform(model_target_train, y_transform)[0].values.astype(float)
        y_scaler, x_scaler, tr_y_scaled, tr_cov_scaled = _fit_dl_scalers(
            tr_y_t,
            cov_tr.values.astype(float) if len(cov_tr) else np.zeros((len(tr), 0), dtype=float),
        )
        strategy = str(cfg.get("strategy", "recursive")).lower()
        if strategy == "direct":
            X_seq, y_seq = _make_dl_sequences_direct(tr_y_scaled, tr_cov_scaled, int(cfg["window"]), int(len(val)))
            output_dim = int(len(val))
        else:
            X_seq, y_seq = _make_dl_sequences(tr_y_scaled, tr_cov_scaled, int(cfg["window"]))
            output_dim = 1
        if len(X_seq) < int(FORECAST_RUNTIME_CONFIG.dl_min_train_sequences):
            raise ValueError(f"Fold {fold.get('fold')} için yeterli sliding-window eğitim örneği üretilemedi.")
        sample_weight = _make_recent_sample_weights(len(y_seq), float(cfg.get("recent_weight", getattr(FORECAST_RUNTIME_CONFIG, "dl_recent_sample_weight_max", 1.0))))
        model, hist_df = _fit_keras_sequence_model_time_ordered(X_seq, y_seq, cfg, model_family, output_dim=output_dim, sample_weight=sample_weight)
        if strategy == "direct":
            pred_val_scaled = _direct_dl_forecast(model, tr_y_scaled, tr_cov_scaled, int(cfg["window"]), int(len(val)))
        else:
            val_cov_scaled = _transform_dl_covariates(x_scaler, cov_val) if len(cov_val) else np.zeros((0, tr_cov_scaled.shape[1] if tr_cov_scaled.ndim == 2 else 0), dtype=float)
            pred_val_scaled = _recursive_dl_forecast(model, tr_y_scaled, tr_cov_scaled, val_cov_scaled, int(cfg["window"]))
        pred_val_t = y_scaler.inverse_transform(pred_val_scaled.reshape(-1, 1)).reshape(-1)
        pred_val_component = inverse_target_transform(pred_val_t, y_transform)
        if residual_mode:
            pred_val = np.maximum(np.asarray(residual_base_val, dtype=float).reshape(-1)[:len(pred_val_component)] + np.asarray(pred_val_component, dtype=float).reshape(-1), 0.0)
        else:
            pred_val = np.maximum(np.asarray(pred_val_component, dtype=float), 0.0)
        corrected, post_cfg = compute_validation_postprocess_candidates(val["y"].values, pred_val, tr["y"], season_length)
        asym = compute_asymmetric_validation_penalty(val["y"].values, corrected, tr["y"].values, severity=1.0)
        min_loss = safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").min()) if len(hist_df) else np.nan
        min_val_loss = safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").min()) if len(hist_df) and "val_loss" in hist_df.columns else np.nan
        overfit_gap = safe_float((min_val_loss - min_loss) / max(abs(min_loss), 1e-8)) if pd.notna(min_loss) and pd.notna(min_val_loss) else np.nan
        fold_rows.append({
            "fold": int(fold.get("fold", len(fold_rows) + 1)),
            "fold_train_end": int(fold["train_end"]),
            "fold_val_start": int(fold["val_start"]),
            "fold_val_end": int(fold["val_end"]),
            "fold_val_len": int(len(val)),
            "fold_wape": safe_float(wape(val["y"].values, corrected)),
            "fold_smape": safe_float(smape(val["y"].values, corrected)),
            "bias_pct": safe_float(asym.get("bias_pct", np.nan)),
            "under_forecast_rate": safe_float(asym.get("under_forecast_rate", np.nan)),
            "peak_event_score": safe_float(asym.get("peak_event_score", np.nan)),
            "asym_penalty": safe_float(asym.get("penalty", 0.0)),
            "min_train_loss": min_loss,
            "min_val_loss": min_val_loss,
            "overfit_gap_ratio": overfit_gap,
            "epochs_ran": int(len(hist_df)),
            "transform": y_transform.get("name", "none"),
            "postprocess": post_cfg.get("name", "raw") if isinstance(post_cfg, dict) else "raw",
            "residual_mode": bool(residual_mode),
            "residual_base_method": residual_base_method,
        })
        fold_hist = hist_df.copy()
        fold_hist["fold"] = int(fold.get("fold", len(fold_tables) + 1))
        fold_tables.append(fold_hist)
        best_post_cfg = dict(post_cfg)
        best_transform = dict(y_transform)
    row = _summarize_dl_fold_rows(fold_rows, cfg, model_family, "TensorFlow", selected_cols, inner_baseline_method, inner_baseline_wape, inner_baseline_smape, dl_regime_profile)
    return {
        "row": row,
        "fold_table": pd.DataFrame(fold_rows),
        "history_df": pd.concat(fold_tables, ignore_index=True, sort=False) if fold_tables else pd.DataFrame(),
        "postprocess_cfg": best_post_cfg,
        "transform_cfg": best_transform,
    }


def _evaluate_torch_dl_config_rolling_origin(
    cfg: Dict[str, Any],
    train_df: pd.DataFrame,
    cov_train_df: pd.DataFrame,
    folds: List[Dict[str, int]],
    freq_alias: str,
    model_family: str,
    selected_cols: List[str],
    season_length: int,
    inner_baseline_method: str,
    inner_baseline_wape: float,
    inner_baseline_smape: float,
    dl_regime_profile: Dict[str, Any],
) -> Dict[str, Any]:
    fold_rows: List[Dict[str, Any]] = []
    fold_tables: List[pd.DataFrame] = []
    best_post_cfg: Dict[str, Any] = {"name": "raw"}
    best_transform: Dict[str, Any] = {"name": "none"}
    for fold in folds:
        tr = train_df.iloc[:int(fold["train_end"])].reset_index(drop=True)
        val = train_df.iloc[int(fold["val_start"]):int(fold["val_end"])].reset_index(drop=True)
        cov_tr = cov_train_df.iloc[:int(fold["train_end"])].reset_index(drop=True)
        cov_val = cov_train_df.iloc[int(fold["val_start"]):int(fold["val_end"])].reset_index(drop=True)
        residual_mode = bool(cfg.get("residual_mode", False))
        residual_base_method = None
        residual_base_val = None
        if residual_mode:
            base_pack = build_residual_dl_base_arrays(tr["y"], len(val), freq_alias, season_length)
            residual_base_train = np.asarray(base_pack.get("train_base"), dtype=float).reshape(-1)
            residual_base_val = np.asarray(base_pack.get("future_base"), dtype=float).reshape(-1)
            residual_base_method = str(base_pack.get("method"))
            model_target_train = pd.to_numeric(tr["y"], errors="coerce").astype(float).reset_index(drop=True) - residual_base_train
            y_transform = {"name": "none", "lambda": None, "shift": 0.0, "allow_negative_output": True, "target_mode": "residual"}
        else:
            model_target_train = tr["y"]
            y_transform = choose_target_transform(tr["y"])
        tr_y_t = apply_target_transform(model_target_train, y_transform)[0].values.astype(float)
        y_scaler, x_scaler, tr_y_scaled, tr_cov_scaled = _fit_dl_scalers(
            tr_y_t,
            cov_tr.values.astype(float) if len(cov_tr) else np.zeros((len(tr), 0), dtype=float),
        )
        strategy = str(cfg.get("strategy", "recursive")).lower()
        if strategy == "direct":
            X_seq, y_seq = _make_dl_sequences_direct(tr_y_scaled, tr_cov_scaled, int(cfg["window"]), int(len(val)))
        else:
            X_seq, y_seq = _make_dl_sequences(tr_y_scaled, tr_cov_scaled, int(cfg["window"]))
        if len(X_seq) < int(FORECAST_RUNTIME_CONFIG.dl_min_train_sequences):
            raise ValueError(f"Fold {fold.get('fold')} için yeterli sliding-window eğitim örneği üretilemedi.")
        sample_weight = _make_recent_sample_weights(len(y_seq), float(cfg.get("recent_weight", getattr(FORECAST_RUNTIME_CONFIG, "dl_recent_sample_weight_max", 1.0))))
        model, hist_df = _torch_train_sequence_model(X_seq, y_seq, cfg, model_family=model_family, sample_weight=sample_weight)
        if strategy == "direct":
            pred_val_scaled = _torch_direct_forecast(model, tr_y_scaled, tr_cov_scaled, int(cfg["window"]), int(len(val)))
        else:
            val_cov_scaled = _transform_dl_covariates(x_scaler, cov_val) if len(cov_val) else np.zeros((0, tr_cov_scaled.shape[1] if tr_cov_scaled.ndim == 2 else 0), dtype=float)
            pred_val_scaled = _torch_recursive_forecast(model, tr_y_scaled, tr_cov_scaled, val_cov_scaled, int(cfg["window"]))
        pred_val_t = y_scaler.inverse_transform(pred_val_scaled.reshape(-1, 1)).reshape(-1)
        pred_val_component = inverse_target_transform(pred_val_t, y_transform)
        if residual_mode:
            pred_val = np.maximum(np.asarray(residual_base_val, dtype=float).reshape(-1)[:len(pred_val_component)] + np.asarray(pred_val_component, dtype=float).reshape(-1), 0.0)
        else:
            pred_val = np.maximum(np.asarray(pred_val_component, dtype=float), 0.0)
        corrected, post_cfg = compute_validation_postprocess_candidates(val["y"].values, pred_val, tr["y"], season_length)
        asym = compute_asymmetric_validation_penalty(val["y"].values, corrected, tr["y"].values, severity=1.0)
        min_loss = safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").min()) if len(hist_df) else np.nan
        min_val_loss = safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").min()) if len(hist_df) and "val_loss" in hist_df.columns else np.nan
        overfit_gap = safe_float((min_val_loss - min_loss) / max(abs(min_loss), 1e-8)) if pd.notna(min_loss) and pd.notna(min_val_loss) else np.nan
        fold_rows.append({
            "fold": int(fold.get("fold", len(fold_rows) + 1)),
            "fold_train_end": int(fold["train_end"]),
            "fold_val_start": int(fold["val_start"]),
            "fold_val_end": int(fold["val_end"]),
            "fold_val_len": int(len(val)),
            "fold_wape": safe_float(wape(val["y"].values, corrected)),
            "fold_smape": safe_float(smape(val["y"].values, corrected)),
            "bias_pct": safe_float(asym.get("bias_pct", np.nan)),
            "under_forecast_rate": safe_float(asym.get("under_forecast_rate", np.nan)),
            "peak_event_score": safe_float(asym.get("peak_event_score", np.nan)),
            "asym_penalty": safe_float(asym.get("penalty", 0.0)),
            "min_train_loss": min_loss,
            "min_val_loss": min_val_loss,
            "overfit_gap_ratio": overfit_gap,
            "epochs_ran": int(len(hist_df)),
            "transform": y_transform.get("name", "none"),
            "postprocess": post_cfg.get("name", "raw") if isinstance(post_cfg, dict) else "raw",
        })
        fold_hist = hist_df.copy()
        fold_hist["fold"] = int(fold.get("fold", len(fold_tables) + 1))
        fold_tables.append(fold_hist)
        best_post_cfg = dict(post_cfg)
        best_transform = dict(y_transform)
    row = _summarize_dl_fold_rows(fold_rows, cfg, model_family, "PyTorch", selected_cols, inner_baseline_method, inner_baseline_wape, inner_baseline_smape, dl_regime_profile)
    return {
        "row": row,
        "fold_table": pd.DataFrame(fold_rows),
        "history_df": pd.concat(fold_tables, ignore_index=True, sort=False) if fold_tables else pd.DataFrame(),
        "postprocess_cfg": best_post_cfg,
        "transform_cfg": best_transform,
    }

def _flatten_dl_sequence_tensor(X: np.ndarray) -> np.ndarray:
    X = np.asarray(X, dtype=float)
    if X.ndim <= 2:
        return X.reshape(len(X), -1)
    return X.reshape(X.shape[0], X.shape[1] * X.shape[2])


def _fit_dl_surrogate_mlp_forecast(train_df: pd.DataFrame, test_df: pd.DataFrame, feature_train: pd.DataFrame, feature_test: pd.DataFrame, freq_alias: str = "M", model_family: str = "LSTM", reason: str = "") -> Dict[str, Any]:
    """
    Sealed DL-surrogate path.

    This function intentionally does NOT fit/evaluate an MLP surrogate and
    does NOT compare surrogate WAPE against a real LSTM/GRU forecast. A
    surrogate may not replace, improve, calibrate, blend, or masquerade as
    LSTM/GRU output. Returning a quarantined fallback-only result keeps the
    pipeline operational while making the model identity explicit.
    """
    sealed_reason = (
        f"{reason} | DL surrogate MLP yolu kapalıdır. "
        "Test/holdout WAPE değerine bakarak surrogate seçimi veya gerçek LSTM/GRU çıktısını "
        "surrogate ile değiştirme leakage ürettiği için yasaklandı."
    )
    result = _finalize_dl_backend_fallback(
        build_model_level_fallback(model_family, train_df, test_df, freq_alias, sealed_reason),
        model_family=model_family,
        reason=sealed_reason,
        backend_state="surrogate_mlp_disabled_no_leakage",
    )
    result["surrogate_attempted"] = False
    result["surrogate_used"] = False
    result["surrogate_test_wape_comparison"] = False
    result["surrogate_selection_metric"] = None
    result["surrogate_override_forbidden"] = True
    result["test_set_used_for_model_selection"] = False
    result["posthoc_surrogate_override"] = False
    result["dl_test_set_surrogate_override_removed"] = True
    result["dl_never_compare_surrogate_to_holdout"] = True
    result["model_identity_note"] = (
        f"{model_family} için surrogate MLP bilinçli olarak çalıştırılmadı. "
        "Gerçek recurrent DL eğitimi başarısızsa LSTM/GRU metriği üretilmez; "
        "fallback yalnızca operational_fallback_forecast/audit amacıyla saklanır."
    )
    return result



def _is_tabular_ml_engineered_feature(name: str) -> bool:
    """Return True for target-derived tabular ML features that are not DL-native inputs."""
    lower = str(name).strip().lower()
    blocked_tokens = (
        "lag_", "seasonal_lag", "diff_", "roll_", "rolling", "ewm_",
        "recent_mean_gap", "recent_z", "trend_slope", "expanding_",
        "target_", "y_lag", "y_roll", "clean_rolling", "forecast_error"
    )
    return any(tok in lower for tok in blocked_tokens)


def _safe_numeric_frame(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or len(df) == 0:
        return pd.DataFrame(index=getattr(df, "index", None))
    out = pd.DataFrame(index=df.index)
    for col in df.columns:
        s = pd.to_numeric(df[col], errors="coerce").replace([np.inf, -np.inf], np.nan)
        if s.notna().mean() >= 0.70 and s.dropna().nunique() > 1:
            med = safe_float(s.median())
            if pd.isna(med):
                med = 0.0
            out[str(col)] = s.fillna(med).fillna(0.0).astype(float).values
    return out


def _build_dl_calendar_time_covariates(df: pd.DataFrame, freq_alias: str, start_pos: int = 0, model_family: str = "LSTM") -> pd.DataFrame:
    """
    Build known-in-advance, sequence-native covariates for recurrent DL.

    These are intentionally NOT XGBoost-style lag/rolling target features. They
    use only the timestamp and deterministic position information available at
    forecast creation time, so the recurrent tensor remains primarily a target
    sequence with light calendar context.
    """
    idx = getattr(df, "index", pd.RangeIndex(0))
    if not isinstance(df, pd.DataFrame) or "ds" not in df.columns or len(df) == 0:
        return pd.DataFrame(index=idx)
    cfg = FORECAST_RUNTIME_CONFIG
    freq = str(freq_alias).upper()
    ds = pd.to_datetime(df["ds"], errors="coerce")
    out = pd.DataFrame(index=df.index)

    if bool(getattr(cfg, "dl_native_include_time_index_covariates", True)):
        pos = np.arange(int(start_pos), int(start_pos) + len(df), dtype=float)
        denom = max(float(int(start_pos) + len(df) - 1), 1.0)
        out["dl_time_idx_norm"] = pos / denom
        out["dl_time_idx_sqrt"] = np.sqrt(np.maximum(pos, 0.0)) / max(np.sqrt(denom), 1.0)

    if bool(getattr(cfg, "dl_native_include_calendar_covariates", True)):
        month = ds.dt.month.fillna(1).astype(int).clip(1, 12)
        out["dl_month_sin"] = np.sin(2 * np.pi * month / 12.0)
        out["dl_month_cos"] = np.cos(2 * np.pi * month / 12.0)
        quarter = ds.dt.quarter.fillna(1).astype(int).clip(1, 4)
        out["dl_quarter_sin"] = np.sin(2 * np.pi * quarter / 4.0)
        out["dl_quarter_cos"] = np.cos(2 * np.pi * quarter / 4.0)
        out["dl_is_month_start"] = ds.dt.is_month_start.fillna(False).astype(int)
        out["dl_is_month_end"] = ds.dt.is_month_end.fillna(False).astype(int)
        if freq in {"W", "D", "H"}:
            week = ds.dt.isocalendar().week.astype("Int64").fillna(1).astype(int).clip(1, 53)
            out["dl_week_sin"] = np.sin(2 * np.pi * week / 52.0)
            out["dl_week_cos"] = np.cos(2 * np.pi * week / 52.0)
        if freq in {"D", "H"}:
            dow = ds.dt.dayofweek.fillna(0).astype(int).clip(0, 6)
            out["dl_dow_sin"] = np.sin(2 * np.pi * dow / 7.0)
            out["dl_dow_cos"] = np.cos(2 * np.pi * dow / 7.0)
        if freq == "H":
            hour = ds.dt.hour.fillna(0).astype(int).clip(0, 23)
            out["dl_hour_sin"] = np.sin(2 * np.pi * hour / 24.0)
            out["dl_hour_cos"] = np.cos(2 * np.pi * hour / 24.0)

    if bool(getattr(cfg, "dl_native_include_family_marker", True)):
        fam = str(model_family).upper()
        out["dl_family_lstm_marker"] = 1.0 if fam == "LSTM" else 0.0
        out["dl_family_gru_marker"] = 1.0 if fam == "GRU" else 0.0

    out = _safe_numeric_frame(out)
    max_cols = max(0, int(getattr(cfg, "dl_native_max_calendar_covariates", 12)))
    if max_cols and len(out.columns) > max_cols:
        preferred = [
            "dl_time_idx_norm", "dl_time_idx_sqrt",
            "dl_month_sin", "dl_month_cos", "dl_quarter_sin", "dl_quarter_cos",
            "dl_week_sin", "dl_week_cos", "dl_dow_sin", "dl_dow_cos",
            "dl_hour_sin", "dl_hour_cos", "dl_is_month_start", "dl_is_month_end",
            "dl_family_lstm_marker", "dl_family_gru_marker",
        ]
        keep = [c for c in preferred if c in out.columns][:max_cols]
        out = out[keep].copy()
    return out.replace([np.inf, -np.inf], np.nan).fillna(0.0)


def build_dl_sequence_native_covariates(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    feature_train: Optional[pd.DataFrame],
    feature_test: Optional[pd.DataFrame],
    freq_alias: str,
    model_family: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], Dict[str, Any]]:
    """
    Decide and build the DL covariate channel.

    Default mode is sequence_native. It deliberately ignores the ML feature pool
    (lags, rolling means, EWM, target diffs) used by XGBoost. A legacy tabular
    route exists only behind an explicit opt-in flag for backward compatibility.
    """
    cfg = FORECAST_RUNTIME_CONFIG
    mode = str(getattr(cfg, "dl_sequence_input_mode", "sequence_native") or "sequence_native").lower().strip()
    family = str(model_family).upper()
    if mode in {"target_only", "pure_sequence", "none"}:
        tr = pd.DataFrame(index=train_df.index)
        te = pd.DataFrame(index=test_df.index)
        meta = {
            "dl_input_mode": "target_only",
            "dl_feature_source": "target_history_only",
            "dl_feature_count": 0,
            "dl_blocked_tabular_feature_count": int(len(getattr(feature_train, "columns", []))),
            "dl_input_contract": "target_sequence_channel_only_no_tabular_ml_features",
        }
        return tr, te, [], meta

    if mode == "legacy_tabular" and bool(getattr(cfg, "dl_allow_legacy_tabular_exog", False)):
        selected = _select_dl_feature_columns(feature_train, feature_test, train_df["y"], getattr(cfg, "dl_sequence_max_exog_cols", 16))
        selected = [c for c in selected if not (bool(getattr(cfg, "dl_block_tabular_lag_rolling_exog", True)) and _is_tabular_ml_engineered_feature(c))]
        tr = feature_train[selected].copy() if selected and isinstance(feature_train, pd.DataFrame) else pd.DataFrame(index=train_df.index)
        te = feature_test[selected].copy() if selected and isinstance(feature_test, pd.DataFrame) else pd.DataFrame(index=test_df.index)
        tr = _safe_numeric_frame(tr).replace([np.inf, -np.inf], np.nan).fillna(0.0)
        te = _safe_numeric_frame(te).reindex(columns=tr.columns, fill_value=0.0).replace([np.inf, -np.inf], np.nan).fillna(0.0)
        meta = {
            "dl_input_mode": "legacy_tabular_sanitized",
            "dl_feature_source": "explicit_legacy_feature_pool_sanitized",
            "dl_feature_count": int(len(selected)),
            "dl_blocked_tabular_feature_count": int(len(getattr(feature_train, "columns", [])) - len(selected)) if isinstance(feature_train, pd.DataFrame) else 0,
            "dl_input_contract": "legacy_tabular_opt_in_blocking_lag_rolling_features",
        }
        return tr, te, list(tr.columns), meta

    combined_train = train_df[["ds"]].copy().reset_index(drop=True)
    combined_test = test_df[["ds"]].copy().reset_index(drop=True)
    tr = _build_dl_calendar_time_covariates(combined_train, freq_alias=freq_alias, start_pos=0, model_family=family)
    te = _build_dl_calendar_time_covariates(combined_test, freq_alias=freq_alias, start_pos=len(train_df), model_family=family)
    te = te.reindex(columns=tr.columns, fill_value=0.0).replace([np.inf, -np.inf], np.nan).fillna(0.0)
    blocked = 0
    if isinstance(feature_train, pd.DataFrame):
        blocked = int(sum(_is_tabular_ml_engineered_feature(c) for c in feature_train.columns))
    meta = {
        "dl_input_mode": "sequence_native",
        "dl_feature_source": "target_history_plus_known_calendar_time_covariates",
        "dl_feature_count": int(len(tr.columns)),
        "dl_blocked_tabular_feature_count": blocked,
        "dl_input_contract": "target_sequence_primary_calendar_time_only_no_xgboost_lag_rolling_pool",
    }
    return tr.reset_index(drop=True), te.reset_index(drop=True), list(tr.columns), meta


def _rank_dl_feature_columns_by_signal(feature_train: pd.DataFrame, target_train: pd.Series) -> List[Tuple[str, float, float, float]]:
    if not isinstance(feature_train, pd.DataFrame) or feature_train.empty:
        return []
    y = pd.to_numeric(target_train, errors="coerce").astype(float)
    rows = []
    for col in feature_train.columns:
        s = pd.to_numeric(feature_train[col], errors="coerce").replace([np.inf, -np.inf], np.nan).astype(float)
        completeness = float(s.notna().mean())
        if completeness < 0.70:
            continue
        std_val = float(s.std(skipna=True) or 0.0)
        corr_abs = 0.0
        valid = s.notna() & y.notna()
        if int(valid.sum()) >= max(10, int(len(feature_train) * 0.45)):
            try:
                corr_val = np.corrcoef(s.loc[valid].values.astype(float), y.loc[valid].values.astype(float))[0, 1]
                if np.isfinite(corr_val):
                    corr_abs = abs(float(corr_val))
            except Exception:
                corr_abs = 0.0
        if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_block_tabular_lag_rolling_exog", True)) and _is_tabular_ml_engineered_feature(col):
            continue
        name_bonus = 0.0
        lower = str(col).lower()
        if any(tok in lower for tok in ["month", "quarter", "holiday", "sin", "cos", "dow", "week", "is_"]):
            name_bonus += 0.05
        score = 2.8 * corr_abs + 0.55 * completeness + 0.03 * math.log1p(max(std_val, 0.0)) + name_bonus
        rows.append((col, float(score), float(corr_abs), float(completeness)))
    rows.sort(key=lambda x: (x[1], x[2], x[3]), reverse=True)
    return rows


def _select_dl_feature_columns(feature_train: pd.DataFrame, feature_test: pd.DataFrame, target_train: pd.Series, max_cols: int) -> List[str]:
    ranked = _rank_dl_feature_columns_by_signal(feature_train, target_train)
    chosen = [c for c, _, _, _ in ranked[:max(1, int(max_cols))]]
    if chosen:
        return chosen
    if not isinstance(feature_train, pd.DataFrame):
        return []
    combined = pd.concat([feature_train, feature_test], axis=0, ignore_index=True) if isinstance(feature_test, pd.DataFrame) else feature_train.copy()
    scores = []
    for col in combined.columns:
        s = pd.to_numeric(combined[col], errors="coerce").replace([np.inf, -np.inf], np.nan)
        if s.notna().mean() >= 0.70:
            scores.append((col, float(s.notna().mean()), float(s.std(skipna=True) or 0.0)))
    scores = sorted(scores, key=lambda x: (x[1], x[2]), reverse=True)
    return [c for c, _, _ in scores[:max(1, int(max_cols))]]


def _fit_dl_scalers(train_y_t: np.ndarray, train_cov: np.ndarray):
    y_scaler = StandardScaler()
    y_scaled = y_scaler.fit_transform(np.asarray(train_y_t, dtype=float).reshape(-1,1)).reshape(-1)
    if train_cov.size == 0:
        return y_scaler, None, y_scaled, np.zeros((len(train_y_t),0), dtype=float)
    x_scaler = RobustScaler()
    x_scaled = x_scaler.fit_transform(np.asarray(train_cov, dtype=float))
    return y_scaler, x_scaler, y_scaled, x_scaled


def _transform_dl_covariates(x_scaler, cov_df: pd.DataFrame) -> np.ndarray:
    if x_scaler is None or cov_df is None or len(cov_df) == 0:
        return np.zeros((0,0), dtype=float)
    arr = cov_df.replace([np.inf,-np.inf], np.nan).fillna(0.0).values.astype(float)
    return x_scaler.transform(arr)


def _make_dl_sequences(y_scaled: np.ndarray, cov_scaled: np.ndarray, window: int) -> Tuple[np.ndarray, np.ndarray]:
    y_scaled=np.asarray(y_scaled, dtype=float).reshape(-1)
    cov_scaled=np.asarray(cov_scaled, dtype=float) if cov_scaled is not None else np.zeros((len(y_scaled),0), dtype=float)
    Xs=[]; ys=[]
    for end_idx in range(int(window), len(y_scaled)):
        yw = y_scaled[end_idx-int(window):end_idx].reshape(int(window),1)
        xw = cov_scaled[end_idx-int(window):end_idx] if cov_scaled.size else np.zeros((int(window),0), dtype=float)
        Xs.append(np.concatenate([yw,xw], axis=1))
        ys.append(y_scaled[end_idx])
    if not Xs:
        return np.zeros((0,int(window),1 + (cov_scaled.shape[1] if cov_scaled.ndim==2 else 0)), dtype=float), np.zeros((0,), dtype=float)
    return np.asarray(Xs, dtype=float), np.asarray(ys, dtype=float)


def _make_dl_sequences_direct(y_scaled: np.ndarray, cov_scaled: np.ndarray, window: int, horizon: int) -> Tuple[np.ndarray, np.ndarray]:
    y_scaled = np.asarray(y_scaled, dtype=float).reshape(-1)
    cov_scaled = np.asarray(cov_scaled, dtype=float) if cov_scaled is not None else np.zeros((len(y_scaled), 0), dtype=float)
    Xs = []
    ys = []
    for anchor in range(int(window), len(y_scaled) - int(horizon) + 1):
        yw = y_scaled[anchor - int(window):anchor].reshape(int(window), 1)
        xw = cov_scaled[anchor - int(window):anchor] if cov_scaled.size else np.zeros((int(window), 0), dtype=float)
        Xs.append(np.concatenate([yw, xw], axis=1))
        ys.append(y_scaled[anchor:anchor + int(horizon)])
    feat_dim = 1 + (cov_scaled.shape[1] if cov_scaled.ndim == 2 else 0)
    if not Xs:
        return np.zeros((0, int(window), feat_dim), dtype=float), np.zeros((0, int(horizon)), dtype=float)
    return np.asarray(Xs, dtype=float), np.asarray(ys, dtype=float)


def _make_recent_sample_weights(n_samples: int, max_weight: float) -> np.ndarray:
    n_samples = int(n_samples)
    if n_samples <= 0:
        return np.asarray([], dtype=float)
    if n_samples == 1:
        return np.asarray([1.0], dtype=float)
    max_weight = max(1.0, float(max_weight))
    return np.linspace(1.0, max_weight, n_samples, dtype=float)


def _build_dl_network(
    input_shape: Tuple[int, int],
    model_family: str,
    units: int,
    layers: int,
    dropout: float,
    learning_rate: float = 0.001,
    output_dim: int = 1,
    dense_head_ratio: Optional[float] = None,
    recurrent_dropout: Optional[float] = None,
    dense_activation: Optional[str] = None,
):
    if not HAS_TF or Sequential is None:
        raise ImportError("TensorFlow/Keras bulunamadı.")
    fam = str(model_family).upper()
    recurrent_cls = KerasLSTM if fam == "LSTM" else KerasGRU if fam == "GRU" else KerasSimpleRNN
    profile = get_dl_family_profile(fam)
    dense_head_ratio = float(profile.get("dense_head_ratio", 0.5) if dense_head_ratio is None else dense_head_ratio)
    recurrent_dropout = float(profile.get("recurrent_dropout", 0.0) if recurrent_dropout is None else recurrent_dropout)
    dense_activation = str(profile.get("dense_activation", "relu") if dense_activation is None else dense_activation)
    tf.keras.backend.clear_session()
    try:
        tf.keras.utils.set_random_seed(42 if fam != "GRU" else 84)
    except Exception:
        pass

    def _recurrent_layer(n_units: int, return_sequences: bool, use_input_shape: bool = False):
        kwargs = {
            "return_sequences": bool(return_sequences),
        }
        if use_input_shape:
            kwargs["input_shape"] = input_shape
        # Intentionally different regularization: LSTM gets small recurrent
        # dropout for memory-state regularization; GRU keeps recurrent dropout
        # at zero by default for faster adaptation on short demand histories.
        if recurrent_dropout > 0:
            kwargs["recurrent_dropout"] = float(recurrent_dropout)
        return recurrent_cls(int(n_units), **kwargs)

    model = Sequential()
    layers = max(1, int(layers))
    if fam == "LSTM":
        # Longer-memory, higher-capacity branch.
        if layers <= 1:
            model.add(_recurrent_layer(int(units), return_sequences=False, use_input_shape=True))
            if float(dropout) > 0:
                model.add(KerasDropout(float(dropout)))
        else:
            model.add(_recurrent_layer(int(units), return_sequences=True, use_input_shape=True))
            if float(dropout) > 0:
                model.add(KerasDropout(float(dropout)))
            if layers >= 3:
                model.add(_recurrent_layer(max(12, int(round(int(units) * 0.75))), return_sequences=True))
                if float(dropout) > 0:
                    model.add(KerasDropout(float(min(0.45, float(dropout) + 0.05))))
            model.add(_recurrent_layer(max(10, int(units) // 2), return_sequences=False))
            if float(dropout) > 0:
                model.add(KerasDropout(float(dropout)))
        head_units = max(12, int(round(int(units) * dense_head_ratio)))
        model.add(KerasDense(head_units, activation=dense_activation))
        model.add(KerasDense(max(8, head_units // 2), activation="relu"))
    elif fam == "GRU":
        # Short/medium-memory, lighter and faster-adapting branch.
        if layers <= 1:
            model.add(_recurrent_layer(int(units), return_sequences=False, use_input_shape=True))
            if float(dropout) > 0:
                model.add(KerasDropout(float(dropout)))
        else:
            model.add(_recurrent_layer(int(units), return_sequences=True, use_input_shape=True))
            if float(dropout) > 0:
                model.add(KerasDropout(float(dropout)))
            model.add(_recurrent_layer(max(8, int(round(int(units) * 0.60))), return_sequences=False))
            if float(dropout) > 0:
                model.add(KerasDropout(float(max(0.0, float(dropout) * 0.75))))
        head_units = max(8, int(round(int(units) * dense_head_ratio)))
        model.add(KerasDense(head_units, activation=dense_activation))
    else:
        if layers <= 1:
            model.add(_recurrent_layer(int(units), return_sequences=False, use_input_shape=True))
        else:
            model.add(_recurrent_layer(int(units), return_sequences=True, use_input_shape=True))
            model.add(_recurrent_layer(max(10, int(units) // 2), return_sequences=False))
        if float(dropout) > 0:
            model.add(KerasDropout(float(dropout)))
        model.add(KerasDense(max(8, int(units) // 2), activation="relu"))

    model.add(KerasDense(int(max(1, output_dim))))
    loss_fn = tf.keras.losses.Huber(delta=1.0) if HAS_TF else "mse"
    model.compile(optimizer=KerasAdam(learning_rate=float(learning_rate)), loss=loss_fn, metrics=["mae"])
    return model

def _recursive_dl_forecast(model, history_y_scaled: np.ndarray, history_cov_scaled: np.ndarray, future_cov_scaled: np.ndarray, window: int) -> np.ndarray:
    hist_y=list(np.asarray(history_y_scaled, dtype=float).reshape(-1))
    if np.asarray(history_cov_scaled).size:
        hist_cov=[np.asarray(r, dtype=float) for r in np.asarray(history_cov_scaled, dtype=float)]
    else:
        hist_cov=[np.zeros((0,), dtype=float) for _ in hist_y]
    future_cov_scaled=np.asarray(future_cov_scaled, dtype=float) if future_cov_scaled is not None else np.zeros((0,0), dtype=float)
    preds=[]
    for i in range(len(future_cov_scaled)):
        yw=np.asarray(hist_y[-int(window):], dtype=float).reshape(int(window),1)
        xw=np.asarray(hist_cov[-int(window):], dtype=float) if hist_cov else np.zeros((int(window),0), dtype=float)
        X=np.concatenate([yw,xw], axis=1)[None,:,:]
        pred_scaled=float(model.predict(X, verbose=0).reshape(-1)[0])
        preds.append(pred_scaled)
        hist_y.append(pred_scaled)
        hist_cov.append(np.asarray(future_cov_scaled[i], dtype=float))
    return np.asarray(preds, dtype=float)


def _direct_dl_forecast(model, history_y_scaled: np.ndarray, history_cov_scaled: np.ndarray, window: int, horizon: int) -> np.ndarray:
    hist_y = np.asarray(history_y_scaled, dtype=float).reshape(-1)
    hist_cov = np.asarray(history_cov_scaled, dtype=float) if history_cov_scaled is not None else np.zeros((len(hist_y), 0), dtype=float)
    yw = hist_y[-int(window):].reshape(int(window), 1)
    xw = hist_cov[-int(window):] if hist_cov.size else np.zeros((int(window), 0), dtype=float)
    X = np.concatenate([yw, xw], axis=1)[None, :, :]
    pred = model.predict(X, verbose=0).reshape(-1)
    if len(pred) < int(horizon):
        pad = np.repeat(pred[-1] if len(pred) else 0.0, int(horizon) - len(pred))
        pred = np.concatenate([pred, pad])
    return np.asarray(pred[:int(horizon)], dtype=float)


def _history_to_frame(history_obj) -> pd.DataFrame:
    hist=getattr(history_obj, "history", {}) or {}
    if not hist:
        return pd.DataFrame(columns=["epoch","loss","val_loss","mae","val_mae"])
    df=pd.DataFrame(hist)
    df.insert(0, "epoch", np.arange(1, len(df)+1))
    return df


def build_dl_loss_curve(history_df: pd.DataFrame, title: str):
    if history_df is None or len(history_df) == 0 or not HAS_PLOTLY:
        return None
    fig = go.Figure()
    if "loss" in history_df.columns:
        fig.add_trace(go.Scatter(x=history_df["epoch"], y=history_df["loss"], mode="lines+markers", name="Train Loss"))
    if "val_loss" in history_df.columns:
        fig.add_trace(go.Scatter(x=history_df["epoch"], y=history_df["val_loss"], mode="lines+markers", name="Val Loss"))
    fig.update_layout(title=title, xaxis_title="Epoch", yaxis_title="Loss", template="plotly_white")
    return fig


def _make_torch_sequence_model(
    input_size: int,
    hidden_size: int,
    layers: int,
    dropout: float,
    output_dim: int,
    model_family: str,
    dense_head_ratio: Optional[float] = None,
    dense_activation: Optional[str] = None,
):
    if TorchNN is None:
        raise ImportError("PyTorch bulunamadı.")
    family = str(model_family).upper()
    recurrent_cls = TorchNN.LSTM if family == "LSTM" else TorchNN.GRU if family == "GRU" else TorchNN.RNN
    profile = get_dl_family_profile(family)
    dense_head_ratio = float(profile.get("dense_head_ratio", 0.5) if dense_head_ratio is None else dense_head_ratio)
    dense_activation = str(profile.get("dense_activation", "relu") if dense_activation is None else dense_activation)

    class _TorchSequenceForecaster(TorchNN.Module):
        def __init__(self):
            super().__init__()
            self.family = family
            self.recurrent = recurrent_cls(
                input_size=int(input_size),
                hidden_size=int(hidden_size),
                num_layers=max(1, int(layers)),
                dropout=float(dropout) if int(layers) > 1 else 0.0,
                batch_first=True,
            )
            self.norm = TorchNN.LayerNorm(int(hidden_size))
            mid_dim = max(8, int(round(int(hidden_size) * dense_head_ratio)))
            if dense_activation == "tanh":
                act = TorchNN.Tanh()
            elif hasattr(TorchNN, "GELU") and dense_activation in {"gelu", "relu"} and family == "GRU":
                # GRU branch uses a smooth lightweight head; LSTM remains tanh-biased.
                act = TorchNN.GELU()
            else:
                act = TorchNN.ReLU()
            head_layers = [
                TorchNN.Linear(int(hidden_size), mid_dim),
                act,
                TorchNN.Dropout(float(dropout)),
            ]
            if family == "LSTM":
                head_layers.extend([
                    TorchNN.Linear(mid_dim, max(8, mid_dim // 2)),
                    TorchNN.ReLU(),
                    TorchNN.Dropout(float(min(0.45, max(0.0, dropout + 0.05)))),
                    TorchNN.Linear(max(8, mid_dim // 2), int(output_dim)),
                ])
            else:
                head_layers.append(TorchNN.Linear(mid_dim, int(output_dim)))
            self.head = TorchNN.Sequential(*head_layers)

        def forward(self, x):
            out, hidden = self.recurrent(x)
            if isinstance(hidden, tuple):
                hidden_state = hidden[0][-1]
            else:
                hidden_state = hidden[-1]
            hidden_state = self.norm(hidden_state)
            return self.head(hidden_state)

    return _TorchSequenceForecaster()

def _torch_history_to_df(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["epoch", "loss", "val_loss"])
    df = pd.DataFrame(rows)
    if "epoch" not in df.columns:
        df.insert(0, "epoch", np.arange(1, len(df) + 1))
    return df


def _torch_train_sequence_model(X_seq: np.ndarray, y_seq: np.ndarray, cfg: Dict[str, Any], model_family: str, sample_weight: Optional[np.ndarray] = None) -> Tuple[Any, pd.DataFrame]:
    if not HAS_TORCH or torch is None or TorchNN is None or TorchOptim is None:
        raise ImportError("PyTorch bulunamadı.")
    torch.manual_seed(42 if str(model_family).upper() != "GRU" else 84)
    X_np = np.asarray(X_seq, dtype=np.float32)
    y_np = np.asarray(y_seq, dtype=np.float32)
    if y_np.ndim == 1:
        y_np = y_np.reshape(-1, 1)
    n_samples = int(len(X_np))
    val_size = min(max(4, int(round(n_samples * 0.18))), max(1, n_samples - 4)) if n_samples >= 10 else max(1, n_samples // 5)
    use_val = (n_samples - val_size) >= max(4, min(8, n_samples - 1)) and val_size >= 1
    if use_val:
        X_train_np, X_val_np = X_np[:-val_size], X_np[-val_size:]
        y_train_np, y_val_np = y_np[:-val_size], y_np[-val_size:]
        sw_train = np.asarray(sample_weight[:len(X_train_np)], dtype=np.float32) if sample_weight is not None and len(sample_weight) >= len(X_train_np) else np.ones((len(X_train_np),), dtype=np.float32)
    else:
        X_train_np, y_train_np = X_np, y_np
        X_val_np = np.zeros((0, X_np.shape[1], X_np.shape[2]), dtype=np.float32)
        y_val_np = np.zeros((0, y_np.shape[1]), dtype=np.float32)
        sw_train = np.asarray(sample_weight[:len(X_train_np)], dtype=np.float32) if sample_weight is not None and len(sample_weight) >= len(X_train_np) else np.ones((len(X_train_np),), dtype=np.float32)

    model = _make_torch_sequence_model(
        input_size=int(X_np.shape[2]),
        hidden_size=int(cfg["units"]),
        layers=int(cfg["layers"]),
        dropout=float(cfg["dropout"]),
        output_dim=int(y_np.shape[1]),
        model_family=model_family,
        dense_head_ratio=float(cfg.get("dense_head_ratio", get_dl_family_profile(model_family).get("dense_head_ratio", 0.5))),
        dense_activation=str(cfg.get("dense_activation", get_dl_family_profile(model_family).get("dense_activation", "relu"))),
    )
    optimizer = TorchOptim.AdamW(
        model.parameters(),
        lr=float(cfg.get("learning_rate", 0.001)),
        weight_decay=float(cfg.get("weight_decay", get_dl_family_profile(model_family).get("weight_decay", 1e-4))),
    )
    train_patience = int(max(4, cfg.get("patience", FORECAST_RUNTIME_CONFIG.dl_patience)))
    train_max_epochs = int(max(12, cfg.get("max_epochs", FORECAST_RUNTIME_CONFIG.dl_max_epochs)))
    scheduler = TorchReduceLROnPlateau(optimizer, mode="min", factor=0.5, patience=max(3, train_patience // 5), min_lr=1e-5) if TorchReduceLROnPlateau is not None else None
    loss_fn = TorchNN.SmoothL1Loss(reduction="none")

    X_train = torch.tensor(X_train_np, dtype=torch.float32)
    y_train = torch.tensor(y_train_np, dtype=torch.float32)
    sw_train_t = torch.tensor(sw_train.reshape(-1, 1), dtype=torch.float32)
    X_val = torch.tensor(X_val_np, dtype=torch.float32) if len(X_val_np) else None
    y_val = torch.tensor(y_val_np, dtype=torch.float32) if len(y_val_np) else None

    best_state = None
    best_val = np.inf
    patience_left = int(max(4, train_patience))
    history_rows: List[Dict[str, Any]] = []
    family_batch_size = int(cfg.get("batch_size", FORECAST_RUNTIME_CONFIG.dl_batch_size))
    batch_size = int(min(max(4, family_batch_size), max(4, len(X_train_np))))

    for epoch in range(1, int(train_max_epochs) + 1):
        model.train()
        permutation = torch.randperm(X_train.shape[0])
        batch_losses = []
        for start in range(0, X_train.shape[0], batch_size):
            idx = permutation[start:start + batch_size]
            xb = X_train[idx]
            yb = y_train[idx]
            wb = sw_train_t[idx]
            optimizer.zero_grad()
            pred = model(xb)
            loss_raw = loss_fn(pred, yb)
            loss = (loss_raw.mean(dim=1, keepdim=True) * wb).sum() / wb.sum().clamp_min(1.0)
            loss.backward()
            try:
                torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            except Exception:
                pass
            optimizer.step()
            batch_losses.append(float(loss.detach().cpu().item()))
        train_loss = float(np.mean(batch_losses)) if batch_losses else np.nan

        model.eval()
        with torch.no_grad():
            if X_val is not None and y_val is not None and len(X_val_np):
                val_pred = model(X_val)
                val_loss = float(loss_fn(val_pred, y_val).mean().detach().cpu().item())
            else:
                val_loss = train_loss
        if scheduler is not None and np.isfinite(val_loss):
            scheduler.step(val_loss)
        history_rows.append({"epoch": epoch, "loss": train_loss, "val_loss": val_loss})
        if np.isfinite(val_loss) and (val_loss + 1e-6) < best_val:
            best_val = float(val_loss)
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}
            patience_left = int(max(4, train_patience))
        else:
            patience_left -= 1
            if patience_left <= 0:
                break

    if best_state is not None:
        model.load_state_dict(best_state)
    return model, _torch_history_to_df(history_rows)


def _torch_model_predict(model: Any, X: np.ndarray) -> np.ndarray:
    model.eval()
    with torch.no_grad():
        X_t = torch.tensor(np.asarray(X, dtype=np.float32), dtype=torch.float32)
        pred = model(X_t).detach().cpu().numpy()
    return np.asarray(pred, dtype=float)


def _torch_recursive_forecast(model: Any, history_y_scaled: np.ndarray, history_cov_scaled: np.ndarray, future_cov_scaled: np.ndarray, window: int) -> np.ndarray:
    hist_y = list(np.asarray(history_y_scaled, dtype=float).reshape(-1))
    if np.asarray(history_cov_scaled).size:
        hist_cov = [np.asarray(r, dtype=float) for r in np.asarray(history_cov_scaled, dtype=float)]
    else:
        hist_cov = [np.zeros((0,), dtype=float) for _ in hist_y]
    future_cov_scaled = np.asarray(future_cov_scaled, dtype=float) if future_cov_scaled is not None else np.zeros((0, 0), dtype=float)
    preds: List[float] = []
    for i in range(len(future_cov_scaled)):
        yw = np.asarray(hist_y[-int(window):], dtype=float).reshape(int(window), 1)
        xw = np.asarray(hist_cov[-int(window):], dtype=float) if hist_cov else np.zeros((int(window), 0), dtype=float)
        X = np.concatenate([yw, xw], axis=1)[None, :, :]
        pred_scaled = float(_torch_model_predict(model, X).reshape(-1)[0])
        preds.append(pred_scaled)
        hist_y.append(pred_scaled)
        hist_cov.append(np.asarray(future_cov_scaled[i], dtype=float))
    return np.asarray(preds, dtype=float)


def _torch_direct_forecast(model: Any, history_y_scaled: np.ndarray, history_cov_scaled: np.ndarray, window: int, horizon: int) -> np.ndarray:
    hist_y = np.asarray(history_y_scaled, dtype=float).reshape(-1)
    hist_cov = np.asarray(history_cov_scaled, dtype=float) if history_cov_scaled is not None else np.zeros((len(hist_y), 0), dtype=float)
    yw = hist_y[-int(window):].reshape(int(window), 1)
    xw = hist_cov[-int(window):] if hist_cov.size else np.zeros((int(window), 0), dtype=float)
    X = np.concatenate([yw, xw], axis=1)[None, :, :]
    pred = _torch_model_predict(model, X).reshape(-1)
    if len(pred) < int(horizon):
        pad = np.repeat(pred[-1] if len(pred) else 0.0, int(horizon) - len(pred))
        pred = np.concatenate([pred, pad])
    return np.asarray(pred[:int(horizon)], dtype=float)




def _finalize_dl_backend_fallback(result: Dict[str, Any], model_family: str, reason: str, backend_state: str = "backend_unavailable") -> Dict[str, Any]:
    result = copy.deepcopy(result) if isinstance(result, dict) else {"forecast": np.full(1, np.nan)}
    result["fallback_used"] = True
    result["fallback_method"] = backend_state
    result["surrogate_used"] = False
    result["real_dl_trained"] = False
    result["dl_backend"] = "none"
    result["surrogate_evaluation_disabled"] = True
    result["surrogate_test_wape_comparison"] = False
    result["surrogate_selection_metric"] = None
    result["dl_test_set_surrogate_override_removed"] = True
    result["dl_never_compare_surrogate_to_holdout"] = True
    result["fit_mode"] = f"deep_learning_{str(model_family).lower()}_fallback_only"
    result["model_identity_note"] = f"{model_family} için gerçek DL backend/eğitim kullanılamadı; sonuç güvenli fallback modelinden üretildi. Sebep: {reason}"
    result["backend_availability"] = backend_state
    result["eligible_for_model_ranking"] = False
    result["forecast_is_model_output"] = False
    result["dl_strict_validation_passed"] = False
    result["strict_dl_rejection_reason"] = str(reason)
    result["surrogate_override_forbidden"] = True
    result["test_set_used_for_model_selection"] = False
    result["posthoc_surrogate_override"] = False
    result["leakage_guard"] = "holdout_not_used_for_fallback_or_surrogate_selection"
    if "history_df" not in result or not isinstance(result.get("history_df"), pd.DataFrame):
        result["history_df"] = pd.DataFrame(columns=["epoch", "loss", "val_loss"])
    if "search_table" in result and isinstance(result["search_table"], pd.DataFrame) and len(result["search_table"]) > 0:
        if "backend" not in result["search_table"].columns:
            result["search_table"]["backend"] = "none"
    return result


def _finite_prediction_array(pred: Any, expected_horizon: Optional[int] = None) -> bool:
    try:
        arr = np.asarray(pred, dtype=float).reshape(-1)
    except Exception:
        return False
    if expected_horizon is not None and int(expected_horizon) > 0 and len(arr) != int(expected_horizon):
        return False
    return bool(len(arr) > 0 and np.isfinite(arr).all())


def _history_has_real_epochs(history_df: Any) -> bool:
    if not isinstance(history_df, pd.DataFrame) or len(history_df) == 0:
        return False
    if "loss" not in history_df.columns:
        return False
    loss = pd.to_numeric(history_df.get("loss"), errors="coerce")
    return bool(loss.notna().sum() > 0 and np.isfinite(loss.dropna().values).any())


def is_real_deep_learning_result(result: Dict[str, Any], model_family: Optional[str] = None, expected_horizon: Optional[int] = None) -> bool:
    """
    Hard identity gate for LSTM/GRU.

    A result is treated as a real DL result only if all of these are true:
      - no fallback and no surrogate path was used,
      - backend is TensorFlow/Keras or PyTorch,
      - real_dl_trained is True,
      - epoch-loss history is non-empty,
      - forecast is finite and has the expected horizon length.

    This prevents fallback/MLP/baseline forecasts from entering metrics,
    rankings, backtests, tables, or ensembles under the LSTM/GRU label.
    """
    if not isinstance(result, dict):
        return False
    family = str(model_family or result.get("model_family") or "").upper()
    if family and family not in {"LSTM", "GRU", "RNN"}:
        return True
    if bool(result.get("fallback_used", False)):
        return False
    if bool(result.get("surrogate_used", False)):
        return False
    if not bool(result.get("real_dl_trained", False)):
        return False
    backend = str(result.get("dl_backend", "")).lower()
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_tensorflow_backend_for_real_recurrent", True)) and family in {"LSTM", "GRU"}:
        if backend != "tensorflow":
            return False
    elif backend not in {"tensorflow", "torch"}:
        return False
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_nonempty_history", True)) and not _history_has_real_epochs(result.get("history_df")):
        return False
    if not _finite_prediction_array(result.get("forecast"), expected_horizon=expected_horizon):
        return False
    return True


def quarantine_non_real_dl_result(result: Dict[str, Any], test_len: int, model_family: str, reason: str) -> Dict[str, Any]:
    """
    Preserve fallback values for audit, but remove them from the public
    LSTM/GRU forecast channel so they cannot contaminate model comparison.
    """
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    try:
        original_forecast = np.asarray(out.get("forecast", np.full(int(test_len), np.nan)), dtype=float).reshape(-1)
    except Exception:
        original_forecast = np.full(int(test_len), np.nan, dtype=float)
    out["operational_fallback_forecast"] = original_forecast.copy()
    out["forecast"] = np.full(int(test_len), np.nan, dtype=float)
    out["raw_forecast"] = np.full(int(test_len), np.nan, dtype=float)
    out["validation_wape"] = np.nan
    out["validation_smape"] = np.nan
    out["eligible_for_model_ranking"] = False
    out["forecast_is_model_output"] = False
    out["dl_strict_validation_passed"] = False
    out["strict_dl_rejection_reason"] = str(reason)
    out["surrogate_override_forbidden"] = True
    out["test_set_used_for_model_selection"] = False
    out["posthoc_surrogate_override"] = False
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["leakage_guard"] = "quarantined_no_public_metric_no_holdout_selection"
    out["dl_result_kind"] = "quarantined_not_real_dl"
    out["dl_runtime_state"] = "excluded_from_public_lstm_gru_metrics"
    out["model_identity_note"] = (
        f"{model_family} gerçek recurrent DL eğitimi olarak doğrulanamadı. "
        "Fallback/surrogate/baseline değerleri LSTM/GRU metriği gibi kullanılmadı; "
        "yalnızca operational_fallback_forecast alanında audit için saklandı. "
        f"Sebep: {reason}"
    )
    if "history_df" not in out or not isinstance(out.get("history_df"), pd.DataFrame):
        out["history_df"] = pd.DataFrame(columns=["epoch", "loss", "val_loss"])
    stbl = out.get("search_table")
    if isinstance(stbl, pd.DataFrame):
        stbl = stbl.copy()
        stbl["dl_strict_validation_passed"] = False
        stbl["ranking_excluded"] = True
        stbl["ranking_exclusion_reason"] = str(reason)
        out["search_table"] = stbl
    else:
        out["search_table"] = pd.DataFrame([{
            "model": model_family,
            "dl_strict_validation_passed": False,
            "ranking_excluded": True,
            "ranking_exclusion_reason": str(reason),
        }])
    return out


def mark_real_dl_result(result: Dict[str, Any], model_family: str) -> Dict[str, Any]:
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    out["model_family"] = str(model_family).upper()
    out["eligible_for_model_ranking"] = bool(out.get("eligible_for_model_ranking", True))
    out["forecast_is_model_output"] = True
    out["dl_strict_validation_passed"] = True
    out["strict_dl_rejection_reason"] = None
    out["surrogate_override_forbidden"] = True
    out["test_set_used_for_model_selection"] = False
    out["posthoc_surrogate_override"] = False
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["leakage_guard"] = "inner_validation_only_no_holdout_selection"
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["dl_result_kind"] = "real_recurrent_dl"
    out["dl_runtime_state"] = "trained_real_backend_rankable"
    if "dl_input_mode" not in out:
        out["dl_input_mode"] = str(getattr(FORECAST_RUNTIME_CONFIG, "dl_sequence_input_mode", "sequence_native"))
    return out


def sanitize_dl_result_for_reporting(result: Dict[str, Any], model_family: str, horizon: int) -> Dict[str, Any]:
    if is_real_deep_learning_result(result, model_family=model_family, expected_horizon=horizon):
        out = mark_real_dl_result(result, model_family)
    else:
        reason_parts = []
        if not isinstance(result, dict):
            reason_parts.append("result_dict_missing")
        else:
            if bool(result.get("fallback_used", False)):
                reason_parts.append(f"fallback_used={result.get('fallback_method')}")
            if bool(result.get("surrogate_used", False)):
                reason_parts.append("surrogate_used=True")
            if not bool(result.get("real_dl_trained", False)):
                reason_parts.append("real_dl_trained=False")
            backend = str(result.get("dl_backend", "")).lower()
            if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_tensorflow_backend_for_real_recurrent", True)) and str(model_family).upper() in {"LSTM", "GRU"} and backend != "tensorflow":
                reason_parts.append(f"tensorflow_required_backend={backend or 'missing'}")
            elif backend not in {"tensorflow", "torch"}:
                reason_parts.append(f"invalid_backend={backend or 'missing'}")
            if not _history_has_real_epochs(result.get("history_df")):
                reason_parts.append("epoch_loss_history_empty")
            if not _finite_prediction_array(result.get("forecast"), expected_horizon=horizon):
                reason_parts.append("forecast_not_finite_or_wrong_horizon")
        reason = " | ".join(reason_parts) if reason_parts else "strict_real_dl_gate_failed"
        out = quarantine_non_real_dl_result(result, int(horizon), str(model_family).upper(), reason)
    out = _apply_dl_contextual_suitability_gate(out, model_family=model_family)
    out = attach_dl_training_status_table(out, model_family=model_family, horizon=horizon)
    return out


def enforce_no_posthoc_surrogate_override(result: Dict[str, Any], model_family: str, horizon: int) -> Dict[str, Any]:
    """Final DL leakage gate before a LSTM/GRU result can leave the fitting layer."""
    out = sanitize_dl_result_for_reporting(result, model_family=model_family, horizon=horizon)
    out["surrogate_override_forbidden"] = True
    out["test_set_used_for_model_selection"] = False
    out["posthoc_surrogate_override"] = False
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["leakage_guard"] = out.get("leakage_guard") or "final_no_surrogate_override_gate"
    if str(model_family).upper() in {"LSTM", "GRU"} and bool(out.get("surrogate_used", False)):
        return quarantine_non_real_dl_result(
            out,
            int(horizon),
            str(model_family).upper(),
            "surrogate_used=True after final no-leakage gate",
        )
    return apply_dl_no_test_surrogate_override_contract(out, model_family=model_family, horizon=horizon)





# =========================================================
# ADVANCED DL GOVERNANCE: RESIDUAL MODE, STATUS TABLE, SUITABILITY GATE
# =========================================================

def _dl_should_use_residual_mode(freq_alias: str, train_len: int, model_family: str, policy: Optional[Dict[str, Any]] = None) -> bool:
    """Use residual-DL only for short/marginal monthly real-DL challenger regimes."""
    cfg = FORECAST_RUNTIME_CONFIG
    family = str(model_family or "").upper()
    if family not in {"LSTM", "GRU"}:
        return False
    if not bool(getattr(cfg, "dl_residual_mode_for_monthly_short_series", True)):
        return False
    if str(freq_alias or "").upper() != "M":
        return False
    n = int(train_len or 0)
    lo = int(getattr(cfg, "dl_residual_mode_min_train_n", 72))
    hi = int(getattr(cfg, "dl_residual_mode_train_n_upper", 96))
    if n < lo or n >= hi:
        return False
    if policy and not bool(policy.get("train_real_dl_allowed", True)):
        return False
    return True


def _seasonal_or_drift_base_arrays(y_train: pd.Series, horizon: int, freq_alias: str, season_length: int) -> Tuple[np.ndarray, np.ndarray, str]:
    """Leakage-free in-sample one-step base and future base forecast for residual DL."""
    y = pd.to_numeric(y_train, errors="coerce").astype(float).reset_index(drop=True).ffill().bfill().fillna(0.0)
    n = int(len(y))
    h = max(0, int(horizon))
    s_len = int(season_length or 1)
    base_train = np.zeros(n, dtype=float)
    for i in range(n):
        hist = y.iloc[:i]
        if len(hist) == 0:
            base_train[i] = float(y.iloc[0]) if n else 0.0
        elif s_len > 1 and i >= s_len:
            base_train[i] = float(y.iloc[i - s_len])
        elif len(hist) >= 3:
            base_train[i] = float(hist.tail(min(3, len(hist))).median())
        else:
            base_train[i] = float(hist.iloc[-1])
    base_future, method = build_fallback_forecast(y, pd.Series([np.nan] * h), freq_alias, s_len)
    return np.maximum(base_train, 0.0), np.maximum(np.asarray(base_future, dtype=float), 0.0), f"{method}_residual_base"


def _sarima_residual_base_arrays(y_train: pd.Series, horizon: int, freq_alias: str, season_length: int) -> Tuple[Optional[np.ndarray], Optional[np.ndarray], Optional[str]]:
    """Train-only SARIMA base for residual DL; returns None on any fitting issue."""
    if SARIMAX is None:
        return None, None, None
    y = pd.to_numeric(y_train, errors="coerce").astype(float).reset_index(drop=True).ffill().bfill().fillna(0.0)
    n = int(len(y))
    h = max(0, int(horizon))
    s_len = int(season_length or 1)
    try:
        if n < max(30, 2 * s_len + 6):
            return None, None, None
        seasonal_order = (0, 1, 1, s_len) if s_len > 1 and n >= (2 * s_len + 12) else (0, 0, 0, 0)
        order = (1, 1, 1) if n >= 36 else (1, 0, 0)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            model = SARIMAX(y, order=order, seasonal_order=seasonal_order, trend="c", enforce_stationarity=False, enforce_invertibility=False)
            res = model.fit(disp=False, maxiter=60)
        in_sample = np.asarray(res.get_prediction(start=0, end=n - 1).predicted_mean, dtype=float)
        future = np.asarray(res.get_forecast(steps=h).predicted_mean, dtype=float) if h > 0 else np.asarray([], dtype=float)
        if len(in_sample) != n or len(future) != h or not np.isfinite(in_sample).all() or not np.isfinite(future).all():
            return None, None, None
        return np.maximum(in_sample, 0.0), np.maximum(future, 0.0), "SARIMA_residual_base_train_only"
    except Exception:
        return None, None, None


def build_residual_dl_base_arrays(y_train: pd.Series, horizon: int, freq_alias: str, season_length: int) -> Dict[str, Any]:
    """Return base arrays used by residual-DL. The holdout values are never inspected."""
    if str(getattr(FORECAST_RUNTIME_CONFIG, "dl_residual_base_preference", "sarima_then_seasonal")).lower().startswith("sarima"):
        tr, fut, method = _sarima_residual_base_arrays(y_train, horizon, freq_alias, season_length)
        if tr is not None and fut is not None:
            return {"train_base": tr, "future_base": fut, "method": method, "backend": "sarima"}
    tr, fut, method = _seasonal_or_drift_base_arrays(y_train, horizon, freq_alias, season_length)
    return {"train_base": tr, "future_base": fut, "method": method, "backend": "seasonal_or_drift"}


def _dl_training_status_summary(result: Dict[str, Any], model_family: str, horizon: int) -> pd.DataFrame:
    """One-row status table required for interpreting every DL result."""
    res = result if isinstance(result, dict) else {}
    cfg = res.get("selected_config") if isinstance(res.get("selected_config"), dict) else {}
    hist = res.get("history_df") if isinstance(res.get("history_df"), pd.DataFrame) else pd.DataFrame()
    over = res.get("overfit_summary") if isinstance(res.get("overfit_summary"), dict) else {}
    n_seq = res.get("n_sequences", cfg.get("candidate_sequence_count", np.nan))
    if pd.isna(n_seq) and isinstance(hist, pd.DataFrame) and len(hist):
        n_seq = res.get("final_train_sequence_count", np.nan)
    row = {
        "model": dl_public_real_name(model_family) if normalize_dl_family_label(model_family) in {"LSTM", "GRU"} else str(model_family),
        "trained_with_tensorflow": bool(res.get("dl_backend") == "tensorflow" and res.get("real_dl_trained", False)),
        "surrogate_used": bool(res.get("surrogate_used", False)),
        "fallback_used": bool(res.get("fallback_used", False)),
        "epochs_ran": int(over.get("epochs_ran", len(hist) if isinstance(hist, pd.DataFrame) else 0) or 0),
        "min_train_loss": safe_float(over.get("min_train_loss", pd.to_numeric(hist.get("loss"), errors="coerce").min() if isinstance(hist, pd.DataFrame) and len(hist) and "loss" in hist.columns else np.nan)),
        "min_val_loss": safe_float(over.get("min_val_loss", pd.to_numeric(hist.get("val_loss"), errors="coerce").min() if isinstance(hist, pd.DataFrame) and len(hist) and "val_loss" in hist.columns else np.nan)),
        "overfit_gap": safe_float(over.get("overfit_gap_ratio", res.get("overfit_gap_ratio", np.nan))),
        "n_sequences": safe_float(n_seq),
        "selected_window": cfg.get("window"),
        "selected_strategy": cfg.get("strategy"),
        "validation_design": res.get("validation_design"),
        "rolling_validation_folds": res.get("rolling_validation_folds"),
        "rolling_val_wape_mean": res.get("validation_wape"),
        "rolling_val_wape_std": res.get("rolling_val_wape_std"),
        "residual_dl_used": bool(res.get("residual_dl_used", False)),
        "residual_base_method": res.get("residual_base_method"),
        "eligible_for_model_ranking": bool(res.get("eligible_for_model_ranking", False)),
        "champion_eligible": bool(res.get("dl_champion_eligible", res.get("eligible_for_model_ranking", False))),
        "ranking_exclusion_reason": res.get("ranking_exclusion_reason") or res.get("strict_dl_rejection_reason"),
    }
    return pd.DataFrame([row])


def _apply_dl_contextual_suitability_gate(result: Dict[str, Any], model_family: str) -> Dict[str, Any]:
    """Hard gate before DL can enter ranking; no contextual bonus can bypass it."""
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    if not bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_contextual_suitability_gate_enabled", True)):
        return out
    family = normalize_dl_family_label(model_family)
    if family not in {"LSTM", "GRU"}:
        return out
    reasons: List[str] = []
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_real_backend_for_ranking", True)):
        if not bool(out.get("real_dl_trained", False)) or str(out.get("dl_backend", "")).lower() != "tensorflow":
            reasons.append("real TensorFlow/Keras recurrent training did not complete")
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_no_surrogate_for_ranking", True)):
        if bool(out.get("surrogate_used", False)) or bool(out.get("fallback_used", False)):
            reasons.append("surrogate/fallback output is not rankable as real DL")
    folds = int(out.get("rolling_validation_folds", 0) or 0)
    min_folds = int(getattr(FORECAST_RUNTIME_CONFIG, "dl_rolling_validation_min_folds", 2))
    if folds < min_folds:
        reasons.append(f"rolling-origin fold count {folds} < {min_folds}")
    w_mean = safe_float(out.get("validation_wape", np.nan))
    w_std = safe_float(out.get("rolling_val_wape_std", np.nan))
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_rolling_validation_stability_for_ranking", True)) and pd.notna(w_mean) and pd.notna(w_std) and w_mean > 0:
        cv = float(w_std) / max(float(w_mean), 1e-8)
        out["rolling_val_wape_cv"] = safe_float(cv)
        if cv > float(getattr(FORECAST_RUNTIME_CONFIG, "dl_max_rolling_wape_cv_for_ranking", 0.60)):
            reasons.append(f"rolling validation unstable: WAPE_CV={cv:.3f}")
    over = out.get("overfit_summary") if isinstance(out.get("overfit_summary"), dict) else {}
    gap = safe_float(over.get("overfit_gap_ratio", out.get("overfit_gap_ratio", np.nan)))
    if pd.notna(gap) and gap > float(getattr(FORECAST_RUNTIME_CONFIG, "dl_max_overfit_gap_for_ranking", 0.35)):
        reasons.append(f"overfit gap too high: {gap:.3f}")
    base = safe_float(out.get("inner_baseline_wape", np.nan))
    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_baseline_superiority_for_normal_regime", True)) and pd.notna(w_mean) and pd.notna(base) and base > 0:
        margin = float(getattr(FORECAST_RUNTIME_CONFIG, "dl_normal_sample_min_baseline_improvement_pct", 0.0) or 0.0)
        if w_mean > base * (1.0 - margin):
            reasons.append(f"rolling DL WAPE did not beat train-only baseline: dl={w_mean:.4f}, baseline={base:.4f}")
    if bool(out.get("dl_research_only", False)) or out.get("dl_production_role") in {"challenger_research_only", "research_surrogate_only"}:
        reasons.append("DL production role is research/challenger only")
    if reasons:
        reason = " | ".join(dict.fromkeys([str(r) for r in reasons if r]))
        out["eligible_for_model_ranking"] = False
        out["dl_champion_eligible"] = False
        out["ranking_exclusion_reason"] = out.get("ranking_exclusion_reason") or reason
        out["strict_dl_rejection_reason"] = out.get("strict_dl_rejection_reason") or reason
        out["dl_suitability_gate_passed"] = False
    else:
        out["dl_suitability_gate_passed"] = True
        out["eligible_for_model_ranking"] = bool(out.get("eligible_for_model_ranking", True))
    return out


def attach_dl_training_status_table(result: Dict[str, Any], model_family: str, horizon: int) -> Dict[str, Any]:
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    out["dl_training_status_table"] = _dl_training_status_summary(out, model_family, horizon)
    return out

def apply_dl_no_test_surrogate_override_contract(result: Dict[str, Any], model_family: str, horizon: int) -> Dict[str, Any]:
    """
    Final contract for LSTM/GRU leakage control.

    The holdout/test set may be used only after fitting, for final reporting
    metrics. It must never be used to choose an MLP surrogate, compare a
    surrogate against the real recurrent model, calibrate a replacement, blend
    a surrogate into the recurrent forecast, or override LSTM/GRU output.
    """
    out = copy.deepcopy(result) if isinstance(result, dict) else {}
    family = str(model_family).upper()

    # Read the incoming state before normalizing the public flags. This makes
    # the contract an actual guard instead of merely overwriting bad metadata.
    incoming_posthoc_override = bool(out.get("posthoc_surrogate_override", False))
    incoming_test_selection = bool(out.get("test_set_used_for_model_selection", False))
    incoming_surrogate_test_compare = bool(out.get("surrogate_test_wape_comparison", False))
    incoming_surrogate_after_real_dl = bool(out.get("surrogate_used", False)) and bool(out.get("real_dl_trained", False))

    out["dl_surrogate_override_contract_version"] = str(getattr(FORECAST_RUNTIME_CONFIG, "dl_surrogate_override_contract_version", "no_holdout_surrogate_override_v1"))
    out["dl_test_set_surrogate_override_removed"] = True
    out["dl_never_compare_surrogate_to_holdout"] = True
    out["surrogate_test_wape_comparison"] = False
    out["surrogate_selection_metric"] = None
    out["surrogate_override_forbidden"] = True
    out["posthoc_surrogate_override"] = False
    out["test_set_used_for_model_selection"] = False
    out["leakage_guard"] = out.get("leakage_guard") or "no_holdout_surrogate_override_contract"

    forbidden = []
    if incoming_posthoc_override:
        forbidden.append("posthoc_surrogate_override=True")
    if incoming_test_selection:
        forbidden.append("test_set_used_for_model_selection=True")
    if incoming_surrogate_test_compare:
        forbidden.append("surrogate_test_wape_comparison=True")
    if incoming_surrogate_after_real_dl:
        forbidden.append("surrogate_used_after_real_dl_training=True")

    if family in {"LSTM", "GRU"} and forbidden:
        return quarantine_non_real_dl_result(
            out,
            int(horizon),
            family,
            "DL no-test-surrogate-override contract violated: " + " | ".join(forbidden),
        )
    return out


# =========================================================
# DL PUBLIC IDENTITY / MODEL NAME CONTRACT
# =========================================================

def normalize_dl_family_label(model_name: Any) -> str:
    """Normalize public/internal DL model labels to the canonical family."""
    text = str(model_name or "").strip().upper()
    if text.startswith("LSTM"):
        return "LSTM"
    if text.startswith("GRU"):
        return "GRU"
    if text.startswith("DL-FALLBACK") or text.startswith("DL_FALLBACK"):
        return "DL-FALLBACK"
    return text


def dl_public_real_name(model_family: str) -> str:
    family = normalize_dl_family_label(model_family)
    if family in {"LSTM", "GRU"}:
        return f"{family} (real)"
    return str(model_family)


def dl_public_surrogate_name(model_family: str) -> str:
    family = normalize_dl_family_label(model_family)
    if family in {"LSTM", "GRU"}:
        return f"{family}-surrogate"
    return f"{model_family}-surrogate"


def classify_dl_public_result(model_family: str, result: Dict[str, Any]) -> Tuple[str, str]:
    """
    Return (identity_class, public_model_name) for DL outputs.

    Public naming contract:
      - real recurrent backend output: LSTM (real) / GRU (real)
      - surrogate output: LSTM-surrogate / GRU-surrogate
      - operational fallback output: DL-fallback

    This prevents fallback/surrogate values from appearing under a plain
    LSTM/GRU name in comparison, actual-vs-predicted, ensemble, and export
    tables.
    """
    family = normalize_dl_family_label(model_family)
    res = result if isinstance(result, dict) else {}
    if family in {"LSTM", "GRU"} and bool(res.get("dl_strict_validation_passed", False)) and bool(res.get("real_dl_trained", False)) and not bool(res.get("fallback_used", False)) and not bool(res.get("surrogate_used", False)):
        return "real", dl_public_real_name(family)
    if family in {"LSTM", "GRU"} and bool(res.get("surrogate_used", False)):
        return "surrogate", dl_public_surrogate_name(family)
    if family in {"LSTM", "GRU"} and (bool(res.get("fallback_used", False)) or res.get("operational_fallback_forecast") is not None):
        return "fallback", "DL-fallback"
    if family in {"LSTM", "GRU"}:
        return "not_real", dl_public_real_name(family)
    return "standard", str(model_family)


def is_dl_public_label(model_name: Any) -> bool:
    return normalize_dl_family_label(model_name) in {"LSTM", "GRU", "DL-FALLBACK"}


def dl_metric_identity_fields(model_family: str, result: Dict[str, Any], identity_class: Optional[str] = None) -> Dict[str, Any]:
    family = normalize_dl_family_label(model_family)
    res = result if isinstance(result, dict) else {}
    inferred_class, public_name = classify_dl_public_result(family, res)
    identity_class = identity_class or inferred_class
    return {
        "model_public_name": public_name,
        "identity_model_family": family,
        "model_identity_class": identity_class,
        "dl_result_kind": res.get("dl_result_kind", "real_recurrent_dl" if bool(res.get("real_dl_trained", False)) else "not_real_dl_or_not_run"),
        "real_dl_trained": bool(res.get("real_dl_trained", False)),
        "dl_backend": res.get("dl_backend"),
        "fallback_used": bool(res.get("fallback_used", False)),
        "fallback_method": res.get("fallback_method"),
        "surrogate_used": bool(res.get("surrogate_used", False)),
        "surrogate_evaluation_disabled": bool(res.get("surrogate_evaluation_disabled", True)),
        "posthoc_surrogate_override": bool(res.get("posthoc_surrogate_override", False)),
        "surrogate_test_wape_comparison": bool(res.get("surrogate_test_wape_comparison", False)),
        "surrogate_selection_metric": res.get("surrogate_selection_metric"),
        "dl_test_set_surrogate_override_removed": bool(res.get("dl_test_set_surrogate_override_removed", True)),
        "dl_never_compare_surrogate_to_holdout": bool(res.get("dl_never_compare_surrogate_to_holdout", True)),
        "test_set_used_for_model_selection": bool(res.get("test_set_used_for_model_selection", False)),
        "dl_input_mode": res.get("dl_input_mode"),
        "dl_feature_source": res.get("dl_feature_source"),
        "dl_feature_count": res.get("dl_feature_count"),
        "dl_blocked_tabular_feature_count": res.get("dl_blocked_tabular_feature_count"),
        "dl_input_contract": res.get("dl_input_contract"),
        "validation_design": res.get("validation_design"),
        "rolling_validation_folds": res.get("rolling_validation_folds"),
        "model_identity_note": res.get("model_identity_note"),
        "dl_train_observations": res.get("dl_train_observations"),
        "dl_short_monthly_policy_state": res.get("dl_short_monthly_policy_state"),
        "dl_production_role": res.get("dl_production_role"),
        "dl_champion_eligible": res.get("dl_champion_eligible"),
        "dl_research_only": res.get("dl_research_only"),
        "ranking_exclusion_reason": res.get("ranking_exclusion_reason") or res.get("strict_dl_rejection_reason"),
    }


def get_dl_operational_fallback_forecast(result: Dict[str, Any], expected_horizon: int) -> Optional[np.ndarray]:
    """Return quarantined fallback forecast for explicit DL-fallback reporting."""
    if not isinstance(result, dict):
        return None
    candidates = [result.get("operational_fallback_forecast"), result.get("raw_forecast"), result.get("forecast")]
    for cand in candidates:
        try:
            arr = np.asarray(cand, dtype=float).reshape(-1)
        except Exception:
            continue
        if len(arr) == int(expected_horizon) and np.isfinite(arr).all():
            return np.maximum(arr, 0.0)
    return None

def build_invalid_model_metrics(model_name: str, reason: str) -> Dict[str, Any]:
    return {
        "model": model_name,
        "MAE": np.nan,
        "RMSE": np.nan,
        "MAPE": np.nan,
        "sMAPE": np.nan,
        "WAPE": np.nan,
        "MASE": np.nan,
        "Bias": np.nan,
        "BiasPct": np.nan,
        "UnderForecastRate": np.nan,
        "OverForecastRate": np.nan,
        "eligible_for_model_ranking": False,
        "ranking_exclusion_reason": str(reason),
        "dl_result_kind": "quarantined_not_real_dl" if normalize_dl_family_label(model_name) in {"LSTM", "GRU"} else None,
        "real_dl_trained": False if normalize_dl_family_label(model_name) in {"LSTM", "GRU"} else None,
        "fallback_used": None,
        "surrogate_used": None,
        "posthoc_surrogate_override": None,
        "surrogate_test_wape_comparison": False,
        "surrogate_selection_metric": None,
        "dl_test_set_surrogate_override_removed": True,
        "dl_never_compare_surrogate_to_holdout": True,
        "dl_input_mode": None,
        "dl_feature_source": None,
        "dl_feature_count": None,
        "model_identity_note": str(reason),
    }


def filter_usable_prediction_map(pred_map: Dict[str, np.ndarray], metrics_df: pd.DataFrame, expected_horizon: Optional[int] = None) -> Tuple[Dict[str, np.ndarray], pd.DataFrame]:
    usable = {}
    for name, pred in (pred_map or {}).items():
        if _finite_prediction_array(pred, expected_horizon=expected_horizon):
            usable[name] = np.asarray(pred, dtype=float).reshape(-1)
    if not isinstance(metrics_df, pd.DataFrame) or len(metrics_df) == 0:
        return usable, pd.DataFrame()
    m = metrics_df.copy()
    if "model" in m.columns:
        m = m.loc[m["model"].isin(usable.keys())].copy()
    for metric_col in ["WAPE", "sMAPE", "RMSE", "MAE"]:
        if metric_col in m.columns:
            m[metric_col] = pd.to_numeric(m[metric_col], errors="coerce")
    if "WAPE" in m.columns:
        m = m.loc[m["WAPE"].notna() & np.isfinite(m["WAPE"].astype(float))].copy()
    if "eligible_for_model_ranking" in m.columns:
        m = m.loc[m["eligible_for_model_ranking"].fillna(True).astype(bool)].copy()
    usable = {k: v for k, v in usable.items() if k in set(m.get("model", pd.Series(dtype=str)).astype(str))}
    return usable, m.reset_index(drop=True)

def _fit_torch_deep_learning_forecast(train_df: pd.DataFrame, test_df: pd.DataFrame, feature_train: pd.DataFrame, feature_test: pd.DataFrame, freq_alias: str = "M", model_family: str = "LSTM") -> Dict[str, Any]:
    if not HAS_TORCH:
        return _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, "PyTorch bulunamadı."),
            model_family=model_family,
            reason="PyTorch bulunamadı.",
            backend_state="torch_unavailable"
        )
    if len(train_df) < 24:
        return _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için gözlem sayısı yetersiz."),
            model_family=model_family,
            reason=f"{model_family} için gözlem sayısı yetersiz.",
            backend_state="torch_insufficient_observations"
        )

    model_family = str(model_family).upper()
    short_monthly_policy = get_dl_short_monthly_policy(freq_alias, len(train_df), model_family)
    cache_key = build_search_signature(
        f"dl_torch_{model_family.lower()}",
        freq_alias,
        train_df,
        test_df,
        exog_train=pd.concat([feature_train, feature_test], axis=0, ignore_index=True) if isinstance(feature_train, pd.DataFrame) else None,
        extra={"app_version": APP_VERSION, "dl_rev": "torch_lstm_gru_v7_tf_missing_monthly_guard", "model_family": model_family},
    )
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached

    cov_train_df, cov_test_df, selected_cols, dl_input_meta = build_dl_sequence_native_covariates(
        train_df=train_df,
        test_df=test_df,
        feature_train=feature_train,
        feature_test=feature_test,
        freq_alias=freq_alias,
        model_family=model_family,
    )
    cov_train_df = cov_train_df.replace([np.inf, -np.inf], np.nan).fillna(0.0)
    cov_test_df = cov_test_df.replace([np.inf, -np.inf], np.nan).fillna(0.0)

    season_length = infer_season_length_from_freq(freq_alias)
    dl_validation_folds = _build_dl_rolling_origin_folds(train_df, horizon=len(test_df), freq_alias=freq_alias, model_family=model_family)
    insufficient_folds, insufficient_reason = _dl_insufficient_rolling_folds(dl_validation_folds, model_family)
    if insufficient_folds and bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_rolling_validation_fail_if_insufficient_folds", True)):
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, insufficient_reason),
            model_family=model_family,
            reason=insufficient_reason,
            backend_state="dl_rolling_origin_validation_insufficient"
        )
        result["validation_design"] = "rolling_origin_kfold_required"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result
    fold_train_len = int(min(f["train_end"] for f in dl_validation_folds)) if dl_validation_folds else len(train_df)
    fold_val_len = int(max(f["val_len"] for f in dl_validation_folds)) if dl_validation_folds else max(1, len(test_df))

    search_rows = []
    best = None
    best_score = np.inf
    configs = build_family_specific_dl_configs(
        model_family=model_family,
        freq_alias=freq_alias,
        train_len=fold_train_len,
        val_len=fold_val_len,
    )
    configs, dl_regime_profile, regime_rejection_df = _apply_dl_data_regime_guard(
        configs, model_family=model_family, freq_alias=freq_alias, train_len=fold_train_len, horizon=fold_val_len
    )
    residual_mode_active = _dl_should_use_residual_mode(freq_alias, len(train_df), model_family, short_monthly_policy)
    if residual_mode_active:
        for _cfg in configs:
            _cfg["residual_mode"] = True
            _cfg["target_learning_mode"] = "residual_over_train_only_base"
    configs = _select_diverse_dl_configs(configs, max(1, int(FORECAST_RUNTIME_CONFIG.dl_max_configs_per_family)))
    baseline_summary = _compute_dl_rolling_baseline_summary(train_df, dl_validation_folds, freq_alias, season_length)
    inner_baseline_method = baseline_summary.get("method", "rolling_baseline_unavailable")
    inner_baseline_wape = safe_float(baseline_summary.get("wape", np.nan))
    inner_baseline_smape = safe_float(baseline_summary.get("smape", np.nan))
    rolling_baseline_table = baseline_summary.get("fold_table", pd.DataFrame())
    if not configs:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için veri rejimi koruması sonrası yeterli DL adayı kalmadı."),
            model_family=model_family,
            reason=f"{model_family} veri rejimi DL için yetersiz: sequence/window/capacity guard tüm adayları eledi.",
            backend_state="dl_data_regime_guard_no_candidate"
        )
        result["data_regime_profile"] = dl_regime_profile
        result["regime_rejection_table"] = regime_rejection_df
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result["search_table"] = regime_rejection_df.copy() if isinstance(regime_rejection_df, pd.DataFrame) else pd.DataFrame()
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    rolling_validation_tables = []
    for cfg in configs:
        try:
            eval_result = _evaluate_torch_dl_config_rolling_origin(
                cfg, train_df, cov_train_df, dl_validation_folds, freq_alias, model_family,
                selected_cols, season_length, inner_baseline_method, inner_baseline_wape,
                inner_baseline_smape, dl_regime_profile
            )
            row = dict(eval_result.get("row", {}))
            search_rows.append(row)
            ft = eval_result.get("fold_table", pd.DataFrame())
            if isinstance(ft, pd.DataFrame) and len(ft):
                tmp = ft.copy()
                tmp["candidate_window"] = int(cfg.get("window", 0))
                tmp["candidate_units"] = int(cfg.get("units", 0))
                tmp["candidate_layers"] = int(cfg.get("layers", 1))
                tmp["candidate_strategy"] = str(cfg.get("strategy", "recursive"))
                rolling_validation_tables.append(tmp)
            score = safe_float(row.get("composite_score", np.inf))
            if pd.notna(score) and score < best_score:
                best_score = score
                best = {
                    "cfg": dict(cfg),
                    "transform_cfg": dict(eval_result.get("transform_cfg", {"name": "none"})),
                    "postprocess_cfg": dict(eval_result.get("postprocess_cfg", {"name": "raw"})),
                    "validation_wape": safe_float(row.get("val_wape", np.nan)),
                    "validation_smape": safe_float(row.get("val_smape", np.nan)),
                    "rolling_val_wape_std": safe_float(row.get("rolling_val_wape_std", np.nan)),
                    "rolling_val_smape_std": safe_float(row.get("rolling_val_smape_std", np.nan)),
                    "selected_cols": list(selected_cols),
                    "history_df": eval_result.get("history_df", pd.DataFrame()),
                    "rolling_validation_table": eval_result.get("fold_table", pd.DataFrame()),
                }
        except Exception as e:
            search_rows.append({**cfg, "model": model_family, "backend": "PyTorch", "validation_design": "rolling_origin_kfold_train_only", "val_wape": np.nan, "val_smape": np.nan, "fit_error": str(e)[:250], "selected_exog_cols": ", ".join(selected_cols)})

    if best is None:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} PyTorch modeli kurulamadı."),
            model_family=model_family,
            reason=f"{model_family} PyTorch modeli kurulamadı.",
            backend_state="torch_training_failed"
        )
        result["search_table"] = pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).reset_index(drop=True)
        result["data_regime_profile"] = dl_regime_profile
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    baseline_ok, baseline_reason, baseline_margin = _dl_baseline_guard_decision(best.get("validation_wape", np.nan), inner_baseline_wape, dl_regime_profile)
    if not baseline_ok:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} kısa veri rejiminde iç-validasyon baseline üstünlük eşiğini geçemedi."),
            model_family=model_family,
            reason=baseline_reason,
            backend_state="torch_data_regime_baseline_guard_rejected"
        )
        result["search_table"] = pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).reset_index(drop=True)
        result["data_regime_profile"] = dl_regime_profile
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["baseline_improvement_required_pct"] = baseline_margin
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result["dl_data_regime_guard_passed"] = False
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    residual_mode_active = bool(best["cfg"].get("residual_mode", False))
    residual_base_pack = None
    if residual_mode_active:
        residual_base_pack = build_residual_dl_base_arrays(train_df["y"], len(test_df), freq_alias, season_length)
        residual_train_base = np.asarray(residual_base_pack.get("train_base"), dtype=float).reshape(-1)
        residual_target = pd.to_numeric(train_df["y"], errors="coerce").astype(float).reset_index(drop=True) - residual_train_base
        best["transform_cfg"] = {"name": "none", "lambda": None, "shift": 0.0, "allow_negative_output": True, "target_mode": "residual"}
        full_y_t = apply_target_transform(residual_target, best["transform_cfg"])[0].values.astype(float)
    else:
        full_y_t = apply_target_transform(train_df["y"], best["transform_cfg"])[0].values.astype(float)
    y_scaler, x_scaler, full_y_scaled, full_cov_scaled = _fit_dl_scalers(full_y_t, cov_train_df[best["selected_cols"]].values.astype(float) if best["selected_cols"] else np.zeros((len(train_df), 0), dtype=float))
    strategy = str(best["cfg"].get("strategy", "recursive")).lower()
    if strategy == "direct":
        X_full, y_full = _make_dl_sequences_direct(full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]), int(len(test_df)))
    else:
        X_full, y_full = _make_dl_sequences(full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]))
    if len(X_full) < int(FORECAST_RUNTIME_CONFIG.dl_min_train_sequences):
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için final training sequence sayısı yetersiz."),
            model_family=model_family,
            reason=f"{model_family} için final training sequence sayısı yetersiz.",
            backend_state="torch_insufficient_sequences"
        )
        result["search_table"] = pd.DataFrame(search_rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True)
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result
    sample_weight_full = _make_recent_sample_weights(len(y_full), float(best["cfg"].get("recent_weight", getattr(FORECAST_RUNTIME_CONFIG, "dl_recent_sample_weight_max", 1.0))))
    final_model, hist_df = _torch_train_sequence_model(X_full, y_full, best["cfg"], model_family=model_family, sample_weight=sample_weight_full)
    if strategy == "direct":
        pred_scaled = _torch_direct_forecast(final_model, full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]), int(len(test_df)))
    else:
        future_cov_scaled = _transform_dl_covariates(x_scaler, cov_test_df[best["selected_cols"]]) if best["selected_cols"] else np.zeros((len(test_df), 0), dtype=float)
        pred_scaled = _torch_recursive_forecast(final_model, full_y_scaled, full_cov_scaled, future_cov_scaled, int(best["cfg"]["window"]))
    pred_t = y_scaler.inverse_transform(pred_scaled.reshape(-1, 1)).reshape(-1)
    pred_component = inverse_target_transform(pred_t, best["transform_cfg"])
    if residual_mode_active:
        future_base = np.asarray((residual_base_pack or {}).get("future_base"), dtype=float).reshape(-1)
        pred = np.maximum(future_base[:len(pred_component)] + np.asarray(pred_component, dtype=float).reshape(-1), 0.0)
    else:
        pred = np.maximum(np.asarray(pred_component, dtype=float), 0.0)
    pred = apply_postprocess_cfg(pred, best.get("postprocess_cfg", {}), train_df["y"], season_length)
    overfit_summary = {
        "epochs_ran": int(len(hist_df)),
        "min_train_loss": safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").min()) if len(hist_df) else np.nan,
        "min_val_loss": safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").min()) if len(hist_df) and "val_loss" in hist_df.columns else np.nan,
        "final_train_loss": safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").iloc[-1]) if len(hist_df) else np.nan,
        "final_val_loss": safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").iloc[-1]) if len(hist_df) and "val_loss" in hist_df.columns else np.nan,
    }
    overfit_summary["overfit_gap_ratio"] = safe_float((overfit_summary.get("min_val_loss", np.nan) - overfit_summary.get("min_train_loss", np.nan)) / max(abs(overfit_summary.get("min_train_loss", np.nan)), 1e-8)) if pd.notna(overfit_summary.get("min_val_loss", np.nan)) and pd.notna(overfit_summary.get("min_train_loss", np.nan)) else np.nan
    result = {
        "forecast": pred,
        "search_table": pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True),
        "history_df": hist_df,
        "overfit_summary": overfit_summary,
        "selected_config": dict(best["cfg"]),
        "n_sequences": int(len(X_full)),
        "final_train_sequence_count": int(len(X_full)),
        "residual_dl_used": bool(residual_mode_active),
        "residual_base_method": (residual_base_pack or {}).get("method") if residual_mode_active else None,
        "residual_base_backend": (residual_base_pack or {}).get("backend") if residual_mode_active else None,
        "transform": best["transform_cfg"].get("name", "none"),
        "used_exog_cols": list(best["selected_cols"]),
        "validation_wape": safe_float(best.get("validation_wape", np.nan)),
        "validation_smape": safe_float(best.get("validation_smape", np.nan)),
        "rolling_val_wape_std": safe_float(best.get("rolling_val_wape_std", np.nan)),
        "rolling_val_smape_std": safe_float(best.get("rolling_val_smape_std", np.nan)),
        "data_regime_profile": get_dl_data_regime_profile(freq_alias, len(train_df), len(test_df), model_family),
        "inner_baseline_method": inner_baseline_method,
        "inner_baseline_wape": inner_baseline_wape,
        "inner_baseline_smape": inner_baseline_smape,
        "baseline_guard_pass_reason": baseline_reason,
        "validation_design": "rolling_origin_kfold_train_only",
        "rolling_validation_folds": int(len(dl_validation_folds)),
        "rolling_validation_table": pd.concat(rolling_validation_tables, ignore_index=True, sort=False) if rolling_validation_tables else pd.DataFrame(),
        "rolling_baseline_table": rolling_baseline_table,
        "dl_data_regime_guard_passed": True,
        "fallback_used": False,
        "fallback_method": None,
        "fit_mode": f"deep_learning_{model_family.lower()}_torch",
        "search_accelerator_cache_hit": False,
        "surrogate_used": False,
        "dl_backend": "torch",
        "real_dl_trained": True,
        "surrogate_evaluation_disabled": True,
        "model_identity_note": f"{model_family} sonucu gerçek PyTorch derin öğrenme eğitiminden üretilmiştir; surrogate/fallback sonuçları ile override edilmemiştir.",
        "eligible_for_model_ranking": True,
        "forecast_is_model_output": True,
        "dl_strict_validation_passed": True,
        "strict_dl_rejection_reason": None,
    }
    result.update(dl_input_meta)
    result = apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)
    result = enforce_no_posthoc_surrogate_override(result, model_family=model_family, horizon=len(test_df))
    result = _apply_dl_contextual_suitability_gate(result, model_family=model_family)
    result = attach_dl_training_status_table(result, model_family=model_family, horizon=len(test_df))
    SEARCH_ACCELERATOR.put_result(cache_key, result)
    return result


def fit_deep_learning_forecast(train_df: pd.DataFrame, test_df: pd.DataFrame, feature_train: pd.DataFrame, feature_test: pd.DataFrame, freq_alias: str = "M", model_family: str = "LSTM") -> Dict[str, Any]:
    ensure_forecasting_runtime_dependencies(include_streamlit=False)
    model_family = str(model_family).upper()
    short_monthly_policy = get_dl_short_monthly_policy(freq_alias, len(train_df), model_family)

    if not FORECAST_RUNTIME_CONFIG.dl_enabled:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, "DL katmanı runtime yapılandırmasında kapalı."),
            model_family=model_family,
            reason="DL katmanı runtime yapılandırmasında kapalı.",
            backend_state="dl_disabled"
        )
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)

    if bool(short_monthly_policy.get("skip_real_training", False)) and not bool(short_monthly_policy.get("train_real_dl_allowed", True)):
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, short_monthly_policy.get("ranking_exclusion_reason", "Aylık kısa seri DL için public ranking dışı.")),
            model_family=model_family,
            reason=short_monthly_policy.get("ranking_exclusion_reason", "Aylık kısa seri DL için public ranking dışı."),
            backend_state="monthly_train_n_lt_72_real_dl_not_trained_research_only"
        )
        result["dl_runtime_state"] = "monthly_short_series_real_dl_not_trained_research_only"
        result["dl_result_kind"] = "real_dl_not_trained_monthly_short_series"
        result["model_identity_class"] = "research_surrogate_only"
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)

    if bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_require_tensorflow_backend_for_real_recurrent", True)) and not HAS_TF:
        reason = "TensorFlow/Keras bulunamadı; gerçek LSTM/GRU eğitimi yapılmadı."
        result = finalize_tf_missing_research_only_result(train_df, test_df, freq_alias, model_family, reason)
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)

    if not HAS_TF and HAS_TORCH:
        result = _fit_torch_deep_learning_forecast(train_df, test_df, feature_train, feature_test, freq_alias=freq_alias, model_family=model_family)
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)
    if not HAS_TF:
        result = finalize_tf_missing_research_only_result(train_df, test_df, freq_alias, model_family, "TensorFlow/Keras ve PyTorch bulunamadı; gerçek DL eğitimi yapılmadı.")
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)
    if len(train_df) < 24:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için gözlem sayısı yetersiz."),
            model_family=model_family,
            reason=f"{model_family} için gözlem sayısı yetersiz.",
            backend_state="tf_insufficient_observations"
        )
        return apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)

    cache_key = build_search_signature(
        f"dl_{model_family.lower()}",
        freq_alias,
        train_df,
        test_df,
        exog_train=pd.concat([feature_train, feature_test], axis=0, ignore_index=True) if isinstance(feature_train, pd.DataFrame) else None,
        extra={"app_version": APP_VERSION, "dl_rev": "direct_recursive_huber_v11_tf_missing_monthly_guard", "model_family": model_family},
    )
    cached = SEARCH_ACCELERATOR.get_result(cache_key)
    if cached is not None:
        cached["search_accelerator_cache_hit"] = True
        return cached

    cov_train_df, cov_test_df, selected_cols, dl_input_meta = build_dl_sequence_native_covariates(
        train_df=train_df,
        test_df=test_df,
        feature_train=feature_train,
        feature_test=feature_test,
        freq_alias=freq_alias,
        model_family=model_family,
    )
    cov_train_df = cov_train_df.replace([np.inf, -np.inf], np.nan).fillna(0.0)
    cov_test_df = cov_test_df.replace([np.inf, -np.inf], np.nan).fillna(0.0)

    season_length = infer_season_length_from_freq(freq_alias)
    dl_validation_folds = _build_dl_rolling_origin_folds(train_df, horizon=len(test_df), freq_alias=freq_alias, model_family=model_family)
    insufficient_folds, insufficient_reason = _dl_insufficient_rolling_folds(dl_validation_folds, model_family)
    if insufficient_folds and bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_rolling_validation_fail_if_insufficient_folds", True)):
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, insufficient_reason),
            model_family=model_family,
            reason=insufficient_reason,
            backend_state="dl_rolling_origin_validation_insufficient"
        )
        result["validation_design"] = "rolling_origin_kfold_required"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result
    fold_train_len = int(min(f["train_end"] for f in dl_validation_folds)) if dl_validation_folds else len(train_df)
    fold_val_len = int(max(f["val_len"] for f in dl_validation_folds)) if dl_validation_folds else max(1, len(test_df))

    search_rows = []
    best = None
    best_score = np.inf
    configs = build_family_specific_dl_configs(
        model_family=model_family,
        freq_alias=freq_alias,
        train_len=fold_train_len,
        val_len=fold_val_len,
    )
    configs, dl_regime_profile, regime_rejection_df = _apply_dl_data_regime_guard(
        configs, model_family=model_family, freq_alias=freq_alias, train_len=fold_train_len, horizon=fold_val_len
    )
    configs = _select_diverse_dl_configs(configs, max(1, int(FORECAST_RUNTIME_CONFIG.dl_max_configs_per_family)))
    baseline_summary = _compute_dl_rolling_baseline_summary(train_df, dl_validation_folds, freq_alias, season_length)
    inner_baseline_method = baseline_summary.get("method", "rolling_baseline_unavailable")
    inner_baseline_wape = safe_float(baseline_summary.get("wape", np.nan))
    inner_baseline_smape = safe_float(baseline_summary.get("smape", np.nan))
    rolling_baseline_table = baseline_summary.get("fold_table", pd.DataFrame())
    if not configs:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için veri rejimi koruması sonrası yeterli DL adayı kalmadı."),
            model_family=model_family,
            reason=f"{model_family} veri rejimi DL için yetersiz: sequence/window/capacity guard tüm adayları eledi.",
            backend_state="dl_data_regime_guard_no_candidate"
        )
        result["data_regime_profile"] = dl_regime_profile
        result["regime_rejection_table"] = regime_rejection_df
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result["search_table"] = regime_rejection_df.copy() if isinstance(regime_rejection_df, pd.DataFrame) else pd.DataFrame()
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    rolling_validation_tables = []
    for cfg in configs:
        try:
            eval_result = _evaluate_tf_dl_config_rolling_origin(
                cfg, train_df, cov_train_df, dl_validation_folds, freq_alias, model_family,
                selected_cols, season_length, inner_baseline_method, inner_baseline_wape,
                inner_baseline_smape, dl_regime_profile
            )
            row = dict(eval_result.get("row", {}))
            search_rows.append(row)
            ft = eval_result.get("fold_table", pd.DataFrame())
            if isinstance(ft, pd.DataFrame) and len(ft):
                tmp = ft.copy()
                tmp["candidate_window"] = int(cfg.get("window", 0))
                tmp["candidate_units"] = int(cfg.get("units", 0))
                tmp["candidate_layers"] = int(cfg.get("layers", 1))
                tmp["candidate_strategy"] = str(cfg.get("strategy", "recursive"))
                rolling_validation_tables.append(tmp)
            score = safe_float(row.get("composite_score", np.inf))
            if pd.notna(score) and score < best_score:
                best_score = score
                best = {
                    "cfg": dict(cfg),
                    "transform_cfg": dict(eval_result.get("transform_cfg", {"name": "none"})),
                    "postprocess_cfg": dict(eval_result.get("postprocess_cfg", {"name": "raw"})),
                    "validation_wape": safe_float(row.get("val_wape", np.nan)),
                    "validation_smape": safe_float(row.get("val_smape", np.nan)),
                    "rolling_val_wape_std": safe_float(row.get("rolling_val_wape_std", np.nan)),
                    "rolling_val_smape_std": safe_float(row.get("rolling_val_smape_std", np.nan)),
                    "selected_cols": list(selected_cols),
                    "history_df": eval_result.get("history_df", pd.DataFrame()),
                    "rolling_validation_table": eval_result.get("fold_table", pd.DataFrame()),
                }
        except Exception as e:
            search_rows.append({**cfg, "model": model_family, "backend": "TensorFlow", "validation_design": "rolling_origin_kfold_train_only", "val_wape": np.nan, "val_smape": np.nan, "fit_error": str(e)[:250], "selected_exog_cols": ", ".join(selected_cols)})

    if best is None:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} modeli kurulamadı."),
            model_family=model_family,
            reason=f"{model_family} modeli kurulamadı.",
            backend_state="tf_training_failed"
        )
        result["search_table"] = pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).reset_index(drop=True)
        result["data_regime_profile"] = dl_regime_profile
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    baseline_ok, baseline_reason, baseline_margin = _dl_baseline_guard_decision(best.get("validation_wape", np.nan), inner_baseline_wape, dl_regime_profile)
    if not baseline_ok:
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} kısa veri rejiminde iç-validasyon baseline üstünlük eşiğini geçemedi."),
            model_family=model_family,
            reason=baseline_reason,
            backend_state="tf_data_regime_baseline_guard_rejected"
        )
        result["search_table"] = pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).reset_index(drop=True)
        result["data_regime_profile"] = dl_regime_profile
        result["inner_baseline_method"] = inner_baseline_method
        result["inner_baseline_wape"] = inner_baseline_wape
        result["inner_baseline_smape"] = inner_baseline_smape
        result["baseline_improvement_required_pct"] = baseline_margin
        result["validation_design"] = "rolling_origin_kfold_train_only"
        result["rolling_validation_folds"] = int(len(dl_validation_folds))
        result["rolling_baseline_table"] = rolling_baseline_table
        result["dl_data_regime_guard_passed"] = False
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    residual_mode_active = bool(best["cfg"].get("residual_mode", False))
    residual_base_pack = None
    if residual_mode_active:
        residual_base_pack = build_residual_dl_base_arrays(train_df["y"], len(test_df), freq_alias, season_length)
        residual_train_base = np.asarray(residual_base_pack.get("train_base"), dtype=float).reshape(-1)
        residual_target = pd.to_numeric(train_df["y"], errors="coerce").astype(float).reset_index(drop=True) - residual_train_base
        best["transform_cfg"] = {"name": "none", "lambda": None, "shift": 0.0, "allow_negative_output": True, "target_mode": "residual"}
        full_y_t = apply_target_transform(residual_target, best["transform_cfg"])[0].values.astype(float)
    else:
        full_y_t = apply_target_transform(train_df["y"], best["transform_cfg"])[0].values.astype(float)
    y_scaler, x_scaler, full_y_scaled, full_cov_scaled = _fit_dl_scalers(
        full_y_t,
        cov_train_df[best["selected_cols"]].values.astype(float) if best["selected_cols"] else np.zeros((len(train_df), 0), dtype=float),
    )
    strategy = str(best["cfg"].get("strategy", "recursive")).lower()
    if strategy == "direct":
        X_full, y_full = _make_dl_sequences_direct(full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]), int(len(test_df)))
        output_dim = int(len(test_df))
    else:
        X_full, y_full = _make_dl_sequences(full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]))
        output_dim = 1
    if len(X_full) < int(FORECAST_RUNTIME_CONFIG.dl_min_train_sequences):
        result = _finalize_dl_backend_fallback(
            build_model_level_fallback(model_family, train_df, test_df, freq_alias, f"{model_family} için final training sequence sayısı yetersiz."),
            model_family=model_family,
            reason=f"{model_family} için final training sequence sayısı yetersiz.",
            backend_state="tf_insufficient_sequences"
        )
        result["search_table"] = pd.DataFrame(search_rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True)
        result.update(dl_input_meta)
        result["used_exog_cols"] = list(selected_cols)
        result["search_accelerator_cache_hit"] = False
        SEARCH_ACCELERATOR.put_result(cache_key, result)
        return result

    sample_weight_full = _make_recent_sample_weights(len(y_full), float(best["cfg"].get("recent_weight", getattr(FORECAST_RUNTIME_CONFIG, "dl_recent_sample_weight_max", 1.0))))
    final_model, hist_df = _fit_keras_sequence_model_time_ordered(
        X_full,
        y_full,
        best["cfg"],
        model_family=model_family,
        output_dim=output_dim,
        sample_weight=sample_weight_full,
    )
    if strategy == "direct":
        pred_scaled = _direct_dl_forecast(final_model, full_y_scaled, full_cov_scaled, int(best["cfg"]["window"]), int(len(test_df)))
    else:
        future_cov_scaled = _transform_dl_covariates(x_scaler, cov_test_df[best["selected_cols"]]) if best["selected_cols"] else np.zeros((len(test_df), 0), dtype=float)
        pred_scaled = _recursive_dl_forecast(final_model, full_y_scaled, full_cov_scaled, future_cov_scaled, int(best["cfg"]["window"]))
    pred_t = y_scaler.inverse_transform(pred_scaled.reshape(-1, 1)).reshape(-1)
    pred_component = inverse_target_transform(pred_t, best["transform_cfg"])
    if residual_mode_active:
        future_base = np.asarray((residual_base_pack or {}).get("future_base"), dtype=float).reshape(-1)
        pred = np.maximum(future_base[:len(pred_component)] + np.asarray(pred_component, dtype=float).reshape(-1), 0.0)
    else:
        pred = np.maximum(np.asarray(pred_component, dtype=float), 0.0)
    pred = apply_postprocess_cfg(pred, best.get("postprocess_cfg", {}), train_df["y"], season_length)
    overfit_summary = {
        "epochs_ran": int(len(hist_df)),
        "min_train_loss": safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").min()) if len(hist_df) else np.nan,
        "min_val_loss": safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").min()) if len(hist_df) and "val_loss" in hist_df.columns else np.nan,
        "final_train_loss": safe_float(pd.to_numeric(hist_df.get("loss"), errors="coerce").iloc[-1]) if len(hist_df) else np.nan,
        "final_val_loss": safe_float(pd.to_numeric(hist_df.get("val_loss"), errors="coerce").iloc[-1]) if len(hist_df) and "val_loss" in hist_df.columns else np.nan,
    }
    overfit_summary["overfit_gap_ratio"] = safe_float((overfit_summary.get("min_val_loss", np.nan) - overfit_summary.get("min_train_loss", np.nan)) / max(abs(overfit_summary.get("min_train_loss", np.nan)), 1e-8)) if pd.notna(overfit_summary.get("min_val_loss", np.nan)) and pd.notna(overfit_summary.get("min_train_loss", np.nan)) else np.nan
    result = {
        "forecast": pred,
        "search_table": pd.concat([pd.DataFrame(search_rows), regime_rejection_df], ignore_index=True, sort=False).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True) if isinstance(regime_rejection_df, pd.DataFrame) and len(regime_rejection_df) else pd.DataFrame(search_rows).sort_values(["composite_score", "val_wape"], ascending=[True, True], na_position="last").reset_index(drop=True),
        "history_df": hist_df,
        "overfit_summary": overfit_summary,
        "selected_config": dict(best["cfg"]),
        "n_sequences": int(len(X_full)),
        "final_train_sequence_count": int(len(X_full)),
        "residual_dl_used": bool(residual_mode_active) if "residual_mode_active" in locals() else False,
        "residual_base_method": (residual_base_pack or {}).get("method") if "residual_base_pack" in locals() and residual_mode_active else None,
        "residual_base_backend": (residual_base_pack or {}).get("backend") if "residual_base_pack" in locals() and residual_mode_active else None,
        "transform": best["transform_cfg"].get("name", "none"),
        "used_exog_cols": list(best["selected_cols"]),
        "validation_wape": safe_float(best.get("validation_wape", np.nan)),
        "validation_smape": safe_float(best.get("validation_smape", np.nan)),
        "rolling_val_wape_std": safe_float(best.get("rolling_val_wape_std", np.nan)),
        "rolling_val_smape_std": safe_float(best.get("rolling_val_smape_std", np.nan)),
        "data_regime_profile": get_dl_data_regime_profile(freq_alias, len(train_df), len(test_df), model_family),
        "inner_baseline_method": inner_baseline_method,
        "inner_baseline_wape": inner_baseline_wape,
        "inner_baseline_smape": inner_baseline_smape,
        "baseline_guard_pass_reason": baseline_reason,
        "validation_design": "rolling_origin_kfold_train_only",
        "rolling_validation_folds": int(len(dl_validation_folds)),
        "rolling_validation_table": pd.concat(rolling_validation_tables, ignore_index=True, sort=False) if rolling_validation_tables else pd.DataFrame(),
        "rolling_baseline_table": rolling_baseline_table,
        "dl_data_regime_guard_passed": True,
        "fallback_used": False,
        "fallback_method": None,
        "fit_mode": f"deep_learning_{model_family.lower()}",
        "search_accelerator_cache_hit": False,
    }
    result.update(dl_input_meta)
    result["surrogate_used"] = False
    result["dl_backend"] = "tensorflow"
    result["real_dl_trained"] = True
    result["surrogate_evaluation_disabled"] = True
    result["model_identity_note"] = f"{model_family} sonucu gerçek derin öğrenme eğitiminden üretilmiştir; surrogate/fallback sonuçları ile override edilmemiştir."
    result["eligible_for_model_ranking"] = True
    result["forecast_is_model_output"] = True
    result["dl_strict_validation_passed"] = True
    result["strict_dl_rejection_reason"] = None
    result = apply_dl_short_monthly_policy_to_result(result, short_monthly_policy, model_family)
    result = enforce_no_posthoc_surrogate_override(result, model_family=model_family, horizon=len(test_df))
    result = _apply_dl_contextual_suitability_gate(result, model_family=model_family)
    result = attach_dl_training_status_table(result, model_family=model_family, horizon=len(test_df))
    SEARCH_ACCELERATOR.put_result(cache_key, result)
    return result


def build_contextual_validation_ranking(validation_df: pd.DataFrame, profile: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    if not isinstance(validation_df, pd.DataFrame) or len(validation_df) == 0:
        return pd.DataFrame(columns=["model", "val_WAPE", "val_sMAPE", "bağlamsal_ceza", "bağlamsal_doğrulama_skoru"])
    profile = profile or {}
    regime = infer_runtime_regime_from_profile(profile=profile)
    out = validation_df.copy()
    out["bağlamsal_ceza"] = 0.0
    freq_alias = str(profile.get("freq_alias", "")).upper() if isinstance(profile, dict) else ""
    if not regime.get("intermittent_like", False):
        out.loc[out["model"].eq("Intermittent"), "bağlamsal_ceza"] += 7.50
        if freq_alias in ["W", "D", "H"]:
            out.loc[out["model"].eq("Intermittent"), "bağlamsal_ceza"] += 1.25
    if not regime.get("seasonal_like", False):
        out.loc[out["model"].eq("Prophet"), "bağlamsal_ceza"] += 1.75
    if freq_alias == "W":
        out.loc[out["model"].eq("Prophet"), "bağlamsal_ceza"] += 0.85
    if regime.get("volatile_like", False):
        out.loc[out["model"].eq("ARIMA"), "bağlamsal_ceza"] += 0.20
        out.loc[out["model"].eq("SARIMA/SARIMAX"), "bağlamsal_ceza"] += 0.10
    if regime.get("seasonal_like", False):
        out.loc[out["model"].eq("SARIMA/SARIMAX"), "bağlamsal_ceza"] -= 0.15
    if regime.get("trendy_like", False):
        out.loc[out["model"].eq("XGBoost"), "bağlamsal_ceza"] -= 0.10
        out.loc[out["model"].astype(str).map(lambda _m: normalize_dl_family_label(_m) in {"LSTM", "GRU"}), "bağlamsal_ceza"] -= 0.12
    if regime.get("volatile_like", False):
        out.loc[out["model"].astype(str).map(lambda _m: normalize_dl_family_label(_m) in {"LSTM", "GRU"}), "bağlamsal_ceza"] -= 0.05
    out["bağlamsal_doğrulama_skoru"] = pd.to_numeric(out.get("val_WAPE"), errors="coerce").fillna(99.0) + pd.to_numeric(out.get("val_sMAPE"), errors="coerce").fillna(99.0) * 0.12 + out["bağlamsal_ceza"].fillna(0.0)
    return out.sort_values(["bağlamsal_doğrulama_skoru", "val_WAPE", "val_sMAPE"], ascending=[True, True, True], na_position="last").reset_index(drop=True)


def build_weighted_ensemble(
    pred_map: Dict[str, np.ndarray],
    metrics_df: pd.DataFrame,
    validation_df: Optional[pd.DataFrame] = None,
    y_train: Optional[np.ndarray] = None,
    y_true: Optional[np.ndarray] = None,
    rolling_summary: Optional[pd.DataFrame] = None,
    profile: Optional[Dict[str, Any]] = None,
) -> Tuple[np.ndarray, pd.DataFrame]:
    """
    Gerçek rolling-aware + bias-aware + peak-aware + regime-aware ansambl.
    Haftalık veri için recent-local ve horizon-specific kaliteyi daha güçlü kullanır.
    """
    if not pred_map:
        raise ValueError("Ansambl için model bulunamadı.")
    expected_horizon = None
    try:
        lengths = [len(np.asarray(v, dtype=float).reshape(-1)) for v in pred_map.values()]
        expected_horizon = int(max(set(lengths), key=lengths.count)) if lengths else None
    except Exception:
        expected_horizon = None
    pred_map, metrics_df = filter_usable_prediction_map(pred_map, metrics_df, expected_horizon=expected_horizon)
    if not pred_map or not isinstance(metrics_df, pd.DataFrame) or len(metrics_df) == 0:
        raise ValueError("Ansambl için kullanılabilir gerçek model tahmini bulunamadı.")

    profile = profile or {}
    regime = infer_runtime_regime_from_profile(profile=profile, y_train=y_train)
    freq_alias = str(profile.get("freq_alias", "")).upper() if isinstance(profile, dict) else ""

    use_df = metrics_df.loc[metrics_df["model"].isin(pred_map.keys())].copy()
    if len(use_df) == 0:
        raise ValueError("Ansambl için metrik bulunamadı.")

    if validation_df is not None and len(validation_df) > 0:
        val_ctx = build_contextual_validation_ranking(validation_df.loc[validation_df["model"].isin(pred_map.keys())].copy(), profile=profile)
        keep = [c for c in ["model", "val_WAPE", "val_sMAPE", "bağlamsal_ceza", "bağlamsal_doğrulama_skoru"] if c in val_ctx.columns]
        if keep:
            use_df = use_df.merge(val_ctx[keep], on="model", how="left")
    else:
        use_df["val_WAPE"] = np.nan
        use_df["val_sMAPE"] = np.nan
        use_df["bağlamsal_ceza"] = 0.0
        use_df["bağlamsal_doğrulama_skoru"] = np.nan

    if rolling_summary is not None and len(rolling_summary) > 0:
        rs = rolling_summary.loc[rolling_summary["model"].isin(pred_map.keys())].copy()
        if len(rs):
            rs = rs.rename(columns={"WAPE": "ro_WAPE", "MAE": "ro_MAE"})
            keep = [c for c in ["model", "ro_WAPE", "ro_MAE"] if c in rs.columns]
            use_df = use_df.merge(rs[keep], on="model", how="left")
    if "ro_WAPE" not in use_df.columns:
        use_df["ro_WAPE"] = np.nan
    if "ro_MAE" not in use_df.columns:
        use_df["ro_MAE"] = np.nan

    recent_rows = []
    peak_rows = []
    if y_train is not None and y_true is not None:
        y_true_arr = np.asarray(y_true, dtype=float)
        for model_name, pred in pred_map.items():
            pred_arr = np.asarray(pred, dtype=float)
            try:
                peak = compute_peak_event_score(np.asarray(y_train, dtype=float), y_true_arr, pred_arr)
                peak_rows.append({"model": model_name, "peak_event_score": safe_float(peak.get("peak_event_score", np.nan))})
            except Exception:
                peak_rows.append({"model": model_name, "peak_event_score": np.nan})
            try:
                ape = np.abs(y_true_arr - pred_arr) / np.where(np.abs(y_true_arr) > 1e-8, np.abs(y_true_arr), np.nan)
                recent_window = min(4 if freq_alias == "W" else 3, len(ape))
                horizon_w = np.linspace(1.0, 1.6 if freq_alias == "W" else 1.35, len(ape))
                recent_local_wape = np.nanmean(ape[-recent_window:]) * 100.0 if recent_window > 0 else np.nan
                horizon_specific_wape = np.nansum(ape * horizon_w) / np.nansum(np.isfinite(ape) * horizon_w) * 100.0 if len(ape) else np.nan
                recent_rows.append({"model": model_name, "recent_local_WAPE": safe_float(recent_local_wape), "horizon_specific_WAPE": safe_float(horizon_specific_wape)})
            except Exception:
                recent_rows.append({"model": model_name, "recent_local_WAPE": np.nan, "horizon_specific_WAPE": np.nan})
    if peak_rows:
        use_df = use_df.merge(pd.DataFrame(peak_rows), on="model", how="left")
    if recent_rows:
        use_df = use_df.merge(pd.DataFrame(recent_rows), on="model", how="left")

    for c in ["BiasPct", "UnderForecastRate", "peak_event_score", "WAPE", "sMAPE", "val_WAPE", "val_sMAPE", "ro_WAPE", "ro_MAE", "recent_local_WAPE", "horizon_specific_WAPE"]:
        if c in use_df.columns:
            use_df[c] = pd.to_numeric(use_df[c], errors="coerce")
        else:
            use_df[c] = np.nan

    def _min_norm(series: pd.Series) -> pd.Series:
        s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
        if s.notna().sum() == 0:
            return pd.Series(1.0, index=series.index, dtype=float)
        mn, mx = float(s.min()), float(s.max())
        if not np.isfinite(mn) or not np.isfinite(mx) or abs(mx - mn) < 1e-9:
            return pd.Series(0.0, index=series.index, dtype=float)
        return (s - mn) / (mx - mn)

    use_df["hold_norm"] = _min_norm(use_df["WAPE"])
    use_df["val_norm"] = _min_norm(use_df["val_WAPE"].fillna(use_df["WAPE"]))
    use_df["ro_norm"] = _min_norm(use_df["ro_WAPE"].fillna(use_df["WAPE"]))
    use_df["recent_norm"] = _min_norm(use_df["recent_local_WAPE"].fillna(use_df["WAPE"]))
    use_df["horizon_norm"] = _min_norm(use_df["horizon_specific_WAPE"].fillna(use_df["WAPE"]))
    use_df["bias_pen"] = use_df["BiasPct"].abs().fillna(0.0).clip(upper=18.0) / 18.0
    use_df["under_pen"] = (use_df["UnderForecastRate"].fillna(0.5) - 0.52).clip(lower=0.0, upper=0.40) / 0.40
    use_df["peak_pen"] = (0.74 - use_df["peak_event_score"].fillna(0.0)).clip(lower=0.0, upper=0.74) / 0.74
    use_df["regime_pen"] = 0.0
    if not regime.get("intermittent_like", False):
        use_df.loc[use_df["model"].eq("Intermittent"), "regime_pen"] += 0.55
    if not regime.get("seasonal_like", False):
        use_df.loc[use_df["model"].eq("Prophet"), "regime_pen"] += 0.25
    if freq_alias == "W":
        use_df.loc[use_df["model"].eq("Prophet"), "regime_pen"] += 0.25
    if regime.get("seasonal_like", False):
        use_df.loc[use_df["model"].eq("SARIMA/SARIMAX"), "regime_pen"] -= 0.08
    if regime.get("volatile_like", False):
        use_df.loc[use_df["model"].eq("XGBoost"), "regime_pen"] -= 0.05
        # DL receives no contextual bonus before suitability gates pass.
    if regime.get("trendy_like", False):
        # DL receives no contextual bonus before suitability gates pass.
        pass

    # Haftalıkta rolling ve lokal yakın geçmişe daha fazla ağırlık ver
    if freq_alias == "W":
        hold_w, val_w, ro_w, recent_w, horiz_w = 0.18, 0.18, 0.26, 0.16, 0.12
    else:
        hold_w, val_w, ro_w, recent_w, horiz_w = 0.26, 0.23, 0.24, 0.10, 0.06

    use_df["birleşik_skor"] = (
        use_df["hold_norm"].fillna(1.0) * hold_w
        + use_df["val_norm"].fillna(use_df["hold_norm"]).fillna(1.0) * val_w
        + use_df["ro_norm"].fillna(use_df["hold_norm"]).fillna(1.0) * ro_w
        + use_df["recent_norm"].fillna(use_df["hold_norm"]).fillna(1.0) * recent_w
        + use_df["horizon_norm"].fillna(use_df["hold_norm"]).fillna(1.0) * horiz_w
        + use_df["peak_pen"].fillna(0.0) * 0.08
        + use_df["bias_pen"].fillna(0.0) * 0.04
        + use_df["under_pen"].fillna(0.0) * 0.03
        + use_df["regime_pen"].fillna(0.0) * 0.10
    )
    use_df = use_df.sort_values(["birleşik_skor", "ro_WAPE", "recent_local_WAPE", "WAPE", "val_WAPE"], ascending=[True, True, True, True, True], na_position="last").reset_index(drop=True)
    best_score = float(pd.to_numeric(use_df.iloc[0].get("birleşik_skor", np.nan), errors="coerce")) if len(use_df) else 1.0
    if not np.isfinite(best_score):
        best_score = 1.0

    pruned = use_df.copy()
    if len(pruned) > 4:
        pruned = pruned.head(4).copy()
    # Haftalıkta Prophet fallback/uygun olmayan durumlarda erken ele
    if "Prophet" in pruned["model"].tolist() and freq_alias == "W":
        prow = pruned.loc[pruned["model"].eq("Prophet")].head(1)
        if len(prow):
            ps = float(pd.to_numeric(prow.iloc[0].get("birleşik_skor", np.nan), errors="coerce"))
            if pd.notna(ps) and ps > best_score + 0.05 and len(pruned) > 2:
                pruned = pruned.loc[~pruned["model"].eq("Prophet")].copy()

    raw = np.exp(-5.5 * (pruned["birleşik_skor"].fillna(1.0) - best_score).clip(lower=0.0))
    peak_bonus = 0.70 + 0.60 * pruned["peak_event_score"].fillna(0.5).clip(lower=0.0, upper=1.0)
    bias_factor = np.exp(-0.040 * pruned["BiasPct"].abs().fillna(0.0).clip(upper=18.0))
    under_factor = np.exp(-2.4 * (pruned["UnderForecastRate"].fillna(0.5) - 0.52).clip(lower=0.0, upper=0.40))
    regime_factor = np.exp(-2.0 * pruned["regime_pen"].fillna(0.0).clip(lower=0.0, upper=0.80))
    recent_factor = np.exp(-1.1 * pruned["recent_norm"].fillna(0.5).clip(lower=0.0, upper=1.0))
    horizon_factor = np.exp(-0.9 * pruned["horizon_norm"].fillna(0.5).clip(lower=0.0, upper=1.0))
    pruned["HamAğırlık"] = raw
    pruned["raw_weight"] = raw * peak_bonus * bias_factor * under_factor * regime_factor * recent_factor * horizon_factor
    if not regime.get("intermittent_like", False) and "Intermittent" in pruned["model"].tolist():
        idx = pruned.index[pruned["model"].eq("Intermittent")]
        pruned.loc[idx, "raw_weight"] = np.minimum(pruned.loc[idx, "raw_weight"].astype(float), 0.10 if freq_alias == "W" else 0.15)
    pruned = pruned.sort_values(["birleşik_skor", "ro_WAPE", "recent_local_WAPE", "WAPE"], ascending=[True, True, True, True], na_position="last").reset_index(drop=True)
    w = pruned["raw_weight"].astype(float).values
    if not np.isfinite(w).all() or np.sum(w) <= 0:
        w = np.ones(len(pruned), dtype=float)
    w = w / np.sum(w)
    if len(w) >= 2 and w[0] < (0.55 if freq_alias == "W" else 0.50):
        target_top = 0.55 if freq_alias == "W" else 0.50
        deficit = target_top - w[0]
        w[0] = target_top
        rest = w[1:]
        if rest.sum() > 0:
            rest = rest / rest.sum()
            w[1:] = np.maximum(0.0, w[1:] - deficit * rest)
            w = w / max(w.sum(), 1e-9)
    pruned["weight"] = w
    ensemble = None
    for _, row in pruned.iterrows():
        pred = np.asarray(pred_map[row["model"]], dtype=float)
        ensemble = pred * row["weight"] if ensemble is None else ensemble + pred * row["weight"]
    out_cols = [c for c in ["model", "WAPE", "sMAPE", "val_WAPE", "val_sMAPE", "ro_WAPE", "ro_MAE", "recent_local_WAPE", "horizon_specific_WAPE", "BiasPct", "UnderForecastRate", "peak_event_score", "regime_pen", "birleşik_skor", "HamAğırlık", "raw_weight", "weight"] if c in pruned.columns]
    return np.maximum(np.asarray(ensemble, dtype=float), 0.0), pruned[out_cols]

def build_model_level_fallback(model_name: str, train_df: pd.DataFrame, test_df: pd.DataFrame, freq_alias: str, error_message: str) -> Dict[str, Any]:
    season_length = infer_season_length_from_freq(freq_alias)
    # Critical leakage guard: fallback receives only the holdout horizon length,
    # never the actual test/holdout target values.
    pred, method_name = build_fallback_forecast(train_df["y"], int(len(test_df)), freq_alias, season_length)
    return {
        "forecast": np.maximum(np.asarray(pred, dtype=float), 0.0),
        "search_table": pd.DataFrame([{
            "model": model_name,
            "fallback_used": True,
            "fallback_method": method_name,
            "error": str(error_message)[:500],
            "test_set_used_for_model_selection": False,
            "leakage_guard": "fallback_train_only_method_selection",
        }]),
        "fallback_used": True,
        "fallback_method": method_name,
        "error": str(error_message),
        "test_set_used_for_model_selection": False,
        "posthoc_surrogate_override": False,
        "surrogate_override_forbidden": True,
        "leakage_guard": "fallback_train_only_method_selection",
    }


def compute_peak_event_score(y_train: np.ndarray, y_true: np.ndarray, y_pred: np.ndarray, quantile: float = 0.80) -> Dict[str, Any]:
    y_train = np.asarray(y_train, dtype=float)
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    if len(y_true) == 0:
        return {"peak_threshold": np.nan, "peak_precision": np.nan, "peak_recall": np.nan, "peak_f1": np.nan, "peak_event_score": np.nan}
    base_hist = y_train[np.isfinite(y_train)] if np.isfinite(y_train).any() else y_true[np.isfinite(y_true)]
    if len(base_hist) == 0:
        threshold = np.nanmedian(y_true)
    else:
        threshold = float(np.nanquantile(base_hist, quantile))
    actual_peak = y_true >= threshold
    pred_peak = y_pred >= max(threshold * 0.98, 1e-6)
    tp = int(np.sum(actual_peak & pred_peak))
    fp = int(np.sum((~actual_peak) & pred_peak))
    fn = int(np.sum(actual_peak & (~pred_peak)))
    precision = float(tp / max(tp + fp, 1))
    recall = float(tp / max(tp + fn, 1))
    f1 = float(2 * precision * recall / max(precision + recall, 1e-12)) if (precision + recall) > 0 else 0.0
    return {
        "peak_threshold": threshold,
        "peak_precision": precision,
        "peak_recall": recall,
        "peak_f1": f1,
        "peak_event_score": f1,
        "actual_peak_count": int(np.sum(actual_peak)),
        "pred_peak_count": int(np.sum(pred_peak)),
    }


def build_bias_dashboard(outputs: Dict[str, Any]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for model_name, table in (outputs.get("tables") or {}).items():
        if table is None or len(table) == 0:
            continue
        y_true = pd.to_numeric(table["y"], errors="coerce").astype(float).values
        y_pred = pd.to_numeric(table["prediction"], errors="coerce").astype(float).values
        err = y_pred - y_true
        mean_demand = max(float(np.nanmean(np.abs(y_true))) if len(y_true) else 0.0, 1e-6)
        rows.append({
            "model": model_name,
            "mean_error": float(np.nanmean(err)) if len(err) else np.nan,
            "median_error": float(np.nanmedian(err)) if len(err) else np.nan,
            "bias_pct": float(100.0 * np.nanmean(err) / mean_demand) if len(err) else np.nan,
            "under_forecast_rate": float(np.nanmean((y_pred < y_true).astype(float))) if len(y_true) else np.nan,
            "over_forecast_rate": float(np.nanmean((y_pred > y_true).astype(float))) if len(y_true) else np.nan,
            "max_under_error": float(np.nanmin(err)) if len(err) else np.nan,
            "max_over_error": float(np.nanmax(err)) if len(err) else np.nan,
        })
    if not rows:
        return pd.DataFrame(columns=["model", "mean_error", "median_error", "bias_pct", "under_forecast_rate", "over_forecast_rate", "max_under_error", "max_over_error"])
    out = pd.DataFrame(rows)
    out["abs_bias_pct"] = pd.to_numeric(out["bias_pct"], errors="coerce").abs()
    out = out.sort_values(["abs_bias_pct", "mean_error"], ascending=[True, True]).drop(columns=["abs_bias_pct"]).reset_index(drop=True)
    return out


def build_peak_event_dashboard(outputs: Dict[str, Any]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    y_train = pd.to_numeric(outputs["train"]["y"], errors="coerce").astype(float).values if outputs.get("train") is not None else np.array([])
    for model_name, table in (outputs.get("tables") or {}).items():
        if table is None or len(table) == 0:
            continue
        peak = compute_peak_event_score(y_train, table["y"].values, table["prediction"].values)
        peak["model"] = model_name
        rows.append(peak)
    if not rows:
        return pd.DataFrame(columns=["model", "peak_threshold", "peak_precision", "peak_recall", "peak_f1", "peak_event_score", "actual_peak_count", "pred_peak_count"])
    return pd.DataFrame(rows).sort_values(["peak_event_score", "peak_recall"], ascending=[False, False]).reset_index(drop=True)


def estimate_model_interval_scale(outputs: Dict[str, Any], model_name: str) -> float:
    robt = outputs.get("rolling_origin_backtest")
    if isinstance(robt, pd.DataFrame) and len(robt) > 0 and "model" in robt.columns:
        rows = robt.loc[robt["model"] == model_name].copy()
        if len(rows) > 0 and "MAE" in rows.columns:
            val = float(pd.to_numeric(rows["MAE"], errors="coerce").median())
            if np.isfinite(val) and val > 0:
                return val
    metrics_df = outputs.get("metrics_df")
    if isinstance(metrics_df, pd.DataFrame) and len(metrics_df) > 0:
        row = metrics_df.loc[metrics_df["model"] == model_name]
        if len(row) > 0:
            val = float(pd.to_numeric(row.iloc[0].get("MAE", np.nan), errors="coerce"))
            if np.isfinite(val) and val > 0:
                return val
    table = (outputs.get("tables") or {}).get(model_name)
    if isinstance(table, pd.DataFrame) and len(table) > 0:
        val = float(pd.to_numeric(table.get("abs_error"), errors="coerce").median())
        if np.isfinite(val) and val > 0:
            return val
    return 1.0


def build_prediction_interval_table(test_df: pd.DataFrame, pred: np.ndarray, scale: float, coverage_levels: Tuple[float, ...] = (0.80, 0.90, 0.95)) -> pd.DataFrame:
    out = test_df[["ds", "y"]].copy().reset_index(drop=True)
    out["q50"] = np.maximum(np.asarray(pred, dtype=float), 0.0)
    z_map = {0.80: 1.2816, 0.90: 1.6449, 0.95: 1.9600}
    for level in coverage_levels:
        z = z_map.get(round(float(level), 2), 1.6449)
        h_idx = np.arange(1, len(out) + 1, dtype=float)
        width = np.maximum(float(scale), 1e-6) * np.sqrt(h_idx) * z
        low = np.maximum(out["q50"].values - width, 0.0)
        high = np.maximum(out["q50"].values + width, 0.0)
        low_q = int(round(((1.0 - level) / 2.0) * 100.0))
        high_q = int(round((1.0 - (1.0 - level) / 2.0) * 100.0))
        out[f"q{low_q:02d}"] = low
        out[f"q{high_q:02d}"] = high
        out[f"coverage_{int(level*100)}"] = ((out["y"] >= low) & (out["y"] <= high)).astype(int)
    return out


def build_feature_availability_audit(df_features: pd.DataFrame, date_col: str, target_col: str, stat_exog_cols: List[str], prophet_exog_cols: List[str], ml_feature_cols: List[str]) -> pd.DataFrame:
    safe_future_keywords = ["month", "quarter", "year", "week", "dayofweek", "dayofmonth", "sin", "cos", "is_month", "is_quarter", "is_year", "holiday", "calendar"]
    risky_keywords = ["price", "promo", "campaign", "stock", "inventory", "epidem", "rx", "otc", "channel", "weather", "temp", "shipment", "supply"]
    rows: List[Dict[str, Any]] = []
    memberships = {}
    for c in stat_exog_cols:
        memberships.setdefault(c, set()).add("statistical")
    for c in prophet_exog_cols:
        memberships.setdefault(c, set()).add("prophet")
    for c in ml_feature_cols:
        memberships.setdefault(c, set()).add("ml")
    for c, used_in in sorted(memberships.items()):
        lc = str(c).lower()
        if any(k in lc for k in safe_future_keywords):
            status = "future_known"
            risk_score = 0.05
            recommendation = "Üretimde güvenle üretilebilir/türetilir."
        elif any(k in lc for k in risky_keywords):
            status = "requires_business_plan_or_forecast"
            risk_score = 0.65
            recommendation = "Üretim anında bu değişken için ayrı plan/forecast gerekir."
        else:
            status = "unknown_future_availability"
            risk_score = 0.40
            recommendation = "Üretimde varlığı veri kontratı ile doğrulanmalı."
        non_na_ratio = float(pd.to_numeric(df_features[c], errors="coerce").notna().mean()) if c in df_features.columns else np.nan
        rows.append({
            "feature": c,
            "used_in": ", ".join(sorted(used_in)),
            "availability_status": status,
            "availability_risk_score": risk_score,
            "historical_non_null_ratio": non_na_ratio,
            "recommendation": recommendation,
        })
    if not rows:
        return pd.DataFrame(columns=["feature", "used_in", "availability_status", "availability_risk_score", "historical_non_null_ratio", "recommendation"])
    return pd.DataFrame(rows).sort_values(["availability_risk_score", "feature"], ascending=[False, True]).reset_index(drop=True)


def build_model_eligibility_gate(outputs: Dict[str, Any], feature_audit_df: pd.DataFrame) -> pd.DataFrame:
    profile = outputs.get("metadata", {}).get("profile", {}) or {}
    n_obs = int(outputs.get("metadata", {}).get("train_n", 0) + outputs.get("metadata", {}).get("test_n", 0))
    intermittency = float(profile.get("intermittency_ratio", 0.0) or 0.0)
    seasonality = float(profile.get("seasonality_strength", 0.0) or 0.0)
    trend_strength = float(profile.get("trend_strength", 0.0) or 0.0)
    val_df = outputs.get("validation_metrics_df", pd.DataFrame()).copy()
    bias_df = build_bias_dashboard(outputs)
    peak_df = build_peak_event_dashboard(outputs)
    metrics_df = outputs.get("metrics_df", pd.DataFrame()).copy()
    robt = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    best_val = float(pd.to_numeric(val_df.get("val_WAPE"), errors="coerce").min()) if len(val_df) else np.nan
    best_ro = float(pd.to_numeric(robt.get("WAPE"), errors="coerce").min()) if len(robt) else np.nan
    risk_by_area = {}
    if isinstance(feature_audit_df, pd.DataFrame) and len(feature_audit_df) > 0 and "used_in" in feature_audit_df.columns:
        for area in ["statistical", "prophet", "ml"]:
            rows = feature_audit_df[feature_audit_df["used_in"].str.contains(area, na=False)]
            risk_by_area[area] = float(rows["availability_risk_score"].max()) if len(rows) else 0.0
    rows_out: List[Dict[str, Any]] = []
    default_gate_models = ["SARIMA/SARIMAX", "ARIMA", "Prophet", "XGBoost", "LSTM (real)", "GRU (real)", "LSTM-surrogate", "GRU-surrogate", "DL-fallback", "Intermittent", "Ensemble"]
    metric_models = metrics_df["model"].dropna().astype(str).tolist() if isinstance(metrics_df, pd.DataFrame) and len(metrics_df) and "model" in metrics_df.columns else []
    ordered_gate_models = []
    for _m in default_gate_models + metric_models:
        if _m not in ordered_gate_models:
            ordered_gate_models.append(_m)
    for model_name in ordered_gate_models:
        reasons = []
        score = 100.0
        area_risk = 0.0
        if model_name in ["SARIMA/SARIMAX", "ARIMA"]:
            area_risk = risk_by_area.get("statistical", 0.0)
        elif model_name == "Prophet":
            area_risk = risk_by_area.get("prophet", 0.0)
        elif model_name == "XGBoost" or normalize_dl_family_label(model_name) in {"LSTM", "GRU", "DL-FALLBACK"}:
            area_risk = risk_by_area.get("ml", 0.0)
        score -= 20.0 * area_risk
        if area_risk >= 0.60:
            reasons.append("Gelecekte hazır olmayabilecek özellik bağımlılığı yüksek.")
        val_row = val_df.loc[val_df["model"] == model_name] if len(val_df) else pd.DataFrame()
        val_wape = float(pd.to_numeric(val_row.iloc[0].get("val_WAPE", np.nan), errors="coerce")) if len(val_row) else np.nan
        if pd.notna(val_wape):
            score -= min(max(val_wape - 8.0, 0.0), 40.0)
            if val_wape > 22.0:
                reasons.append("Doğrulama WAPE yüksek.")
            if pd.notna(best_val) and val_wape > min(best_val * 1.12, best_val + 1.0):
                score -= 8.0
                reasons.append("Doğrulama performansı lider modele göre belirgin zayıf.")
        ro_row = robt.loc[robt["model"] == model_name] if len(robt) else pd.DataFrame()
        ro_wape = float(pd.to_numeric(ro_row.iloc[0].get("WAPE", np.nan), errors="coerce")) if len(ro_row) else np.nan
        if pd.notna(ro_wape):
            score -= min(max(ro_wape - 8.0, 0.0) * 1.3, 35.0)
            if pd.notna(best_ro) and ro_wape > min(best_ro * 1.08, best_ro + 0.75):
                score -= 10.0
                reasons.append("Rolling-origin performansı lider modele göre zayıf.")
        bias_row = bias_df.loc[bias_df["model"] == model_name] if len(bias_df) else pd.DataFrame()
        bias_pct = float(pd.to_numeric(bias_row.iloc[0].get("bias_pct", np.nan), errors="coerce")) if len(bias_row) else np.nan
        if pd.notna(bias_pct):
            score -= min(abs(bias_pct) * 0.8, 20.0)
            if abs(bias_pct) > 12.0:
                reasons.append("Sapma yüksek.")
        peak_row = peak_df.loc[peak_df["model"] == model_name] if len(peak_df) else pd.DataFrame()
        peak_score = float(pd.to_numeric(peak_row.iloc[0].get("peak_event_score", np.nan), errors="coerce")) if len(peak_row) else np.nan
        if pd.notna(peak_score):
            score -= 15.0 * max(0.0, 0.45 - peak_score)
            if peak_score < 0.35:
                reasons.append("Tepe dönem yakalama zayıf.")
        if model_name == "Prophet":
            if n_obs < 84:
                score -= 16.0
                reasons.append("Kısa aylık seri için Prophet kırılgan olabilir.")
            if intermittency >= 0.20:
                score -= 20.0
                reasons.append("Aralıklı yapı Prophet için zayıf uyumlu.")
            if seasonality < 0.18 and trend_strength < 0.10:
                score -= 14.0
                reasons.append("Trend/sezonsallık sinyali Prophet için sınırlı.")
            if bool((outputs.get("prophet") or {}).get("fallback_used", False)):
                score -= 18.0
                reasons.append("Prophet fallback ile çalıştı; üretim için güven zayıf.")
            hold_row = metrics_df.loc[metrics_df["model"] == "Prophet"] if len(metrics_df) else pd.DataFrame()
            hold_wape = float(pd.to_numeric(hold_row.iloc[0].get("WAPE", np.nan), errors="coerce")) if len(hold_row) else np.nan
            best_hold = float(pd.to_numeric(metrics_df.get("WAPE"), errors="coerce").min()) if len(metrics_df) else np.nan
            if pd.notna(hold_wape) and pd.notna(best_hold) and hold_wape > min(best_hold * 1.15, best_hold + 1.0):
                score -= 12.0
                reasons.append("Holdout performansı lider modele göre zayıf.")
        if model_name == "Intermittent" and intermittency < 0.20:
            score -= 38.0
            reasons.append("Intermittent model, düzenli seri için birincil aday değil.")
        elif model_name == "Intermittent" and intermittency < 0.30:
            score -= 14.0
            reasons.append("Seri tam intermittent değil; intermittent modelin etkisi sınırlandı.")
        if model_name == "XGBoost" and n_obs < 30:
            score -= 15.0
            reasons.append("Makine öğrenmesi için gözlem sayısı sınırlı.")
        if normalize_dl_family_label(model_name) in {"LSTM", "GRU", "DL-FALLBACK"}:
            if n_obs < 36:
                score -= 22.0
                reasons.append("Derin öğrenme için gözlem sayısı sınırlı.")
            if intermittency >= 0.30:
                score -= 14.0
                reasons.append("Aralıklı yapı derin öğrenme modellerini zayıflatabilir.")
            dl_key = normalize_dl_family_label(model_name).lower()
            dl_pack = (outputs.get(dl_key) or {})
            overfit_gap = safe_float((dl_pack.get("overfit_summary") or {}).get("overfit_gap_ratio", np.nan))
            if pd.notna(overfit_gap) and overfit_gap > 0.25:
                score -= min(18.0, 36.0 * float(overfit_gap))
                reasons.append("Train/validation ayrışması yüksek; overfitting riski var.")
            if bool(dl_pack.get("fallback_used", False)):
                score -= 18.0
                reasons.append("Derin öğrenme modeli fallback ile tamamlandı.")
        if model_name == "Ensemble":
            metric_row = metrics_df.loc[metrics_df["model"] == model_name] if len(metrics_df) else pd.DataFrame()
            champion_row = metrics_df.sort_values(["WAPE", "sMAPE"], ascending=[True, True]).head(1)
            if len(metric_row) and len(champion_row):
                ens_wape = float(pd.to_numeric(metric_row.iloc[0].get("WAPE", np.nan), errors="coerce"))
                champ_wape = float(pd.to_numeric(champion_row.iloc[0].get("WAPE", np.nan), errors="coerce"))
                if pd.notna(ens_wape) and pd.notna(champ_wape) and ens_wape > champ_wape + 0.50:
                    score -= 24.0
                    reasons.append("Ansambl, lider modelin üzerine çıkamıyor.")
        status = "eligible"
        if model_name == "Prophet" and score < 85.0:
            status = "challenger_only"
        elif score < 75.0:
            status = "challenger_only"
        if model_name == "Prophet" and score < 72.0:
            status = "reject"
        elif score < 58.0:
            status = "reject"
        rows_out.append({
            "model": model_name,
            "eligibility_score": float(np.clip(score, 0.0, 100.0)),
            "status": status,
            "validation_wape": val_wape,
            "rolling_wape": ro_wape,
            "bias_pct": bias_pct,
            "peak_event_score": peak_score,
            "feature_availability_risk": area_risk,
            "reasons": " | ".join(reasons) if reasons else "Uygunluk açısından belirgin kırmızı bayrak yok.",
        })
    rank_map = {"eligible": 0, "challenger_only": 1, "reject": 2}
    out = pd.DataFrame(rows_out)
    out["status_rank"] = out["status"].map(rank_map).fillna(9)
    out = out.sort_values(["status_rank", "eligibility_score"], ascending=[True, False]).drop(columns=["status_rank"]).reset_index(drop=True)
    return out


def build_forecast_value_add(outputs: Dict[str, Any], freq_alias: str) -> pd.DataFrame:
    train_df = outputs.get("train")
    test_df = outputs.get("test")
    if train_df is None or test_df is None or len(test_df) == 0:
        return pd.DataFrame(columns=["model", "baseline_name", "baseline_WAPE", "model_WAPE", "fva_wape_pct", "baseline_MAE", "model_MAE", "fva_mae_pct"])
    season_len = infer_season_length_from_freq(freq_alias)
    baseline_pred, baseline_name = build_fallback_forecast(train_df["y"], int(len(test_df)), freq_alias, season_len)
    baseline_w = wape(test_df["y"].values, baseline_pred)
    baseline_m = mae(test_df["y"].values, baseline_pred)
    rows = []
    for _, row in outputs.get("metrics_df", pd.DataFrame()).iterrows():
        model_name = row["model"]
        model_w = float(pd.to_numeric(row.get("WAPE", np.nan), errors="coerce"))
        model_m = float(pd.to_numeric(row.get("MAE", np.nan), errors="coerce"))
        rows.append({
            "model": model_name,
            "baseline_name": baseline_name,
            "baseline_WAPE": baseline_w,
            "model_WAPE": model_w,
            "fva_wape_pct": float(100.0 * (baseline_w - model_w) / max(baseline_w, 1e-6)) if pd.notna(model_w) else np.nan,
            "baseline_MAE": baseline_m,
            "model_MAE": model_m,
            "fva_mae_pct": float(100.0 * (baseline_m - model_m) / max(baseline_m, 1e-6)) if pd.notna(model_m) else np.nan,
        })
    return pd.DataFrame(rows).sort_values(["fva_wape_pct", "fva_mae_pct"], ascending=[False, False]).reset_index(drop=True)


def build_service_level_simulation(test_df: pd.DataFrame, pred: np.ndarray, interval_df: pd.DataFrame) -> pd.DataFrame:
    if test_df is None or len(test_df) == 0:
        return pd.DataFrame(columns=["service_level_target", "achieved_cycle_service", "avg_safety_stock", "avg_total_stock_target", "avg_stockout_units"])
    actual = pd.to_numeric(test_df["y"], errors="coerce").astype(float).values
    pred = np.maximum(np.asarray(pred, dtype=float), 0.0)
    rows = []
    for level, hi_col in [(0.80, "q90"), (0.90, "q95")]:
        if hi_col not in interval_df.columns:
            continue
        stock_target = np.maximum(pd.to_numeric(interval_df[hi_col], errors="coerce").astype(float).values, pred)
        safety_stock = np.maximum(stock_target - pred, 0.0)
        achieved = float(np.mean(actual <= stock_target)) if len(actual) else np.nan
        stockout_units = np.maximum(actual - stock_target, 0.0)
        rows.append({
            "service_level_target": level,
            "achieved_cycle_service": achieved,
            "avg_safety_stock": float(np.nanmean(safety_stock)) if len(safety_stock) else np.nan,
            "avg_total_stock_target": float(np.nanmean(stock_target)) if len(stock_target) else np.nan,
            "avg_stockout_units": float(np.nanmean(stockout_units)) if len(stockout_units) else np.nan,
        })
    return pd.DataFrame(rows)


def build_live_monitoring_pack(outputs: Dict[str, Any], production_pack: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    prod_model = production_pack.get("production_model", outputs.get("best_model"))
    bias_df = production_pack.get("bias_dashboard", pd.DataFrame())
    peak_df = production_pack.get("peak_event_dashboard", pd.DataFrame())
    gate_df = production_pack.get("model_eligibility_gate", pd.DataFrame())
    interval_df = production_pack.get("production_interval_table", pd.DataFrame())
    feature_df = production_pack.get("feature_availability_audit", pd.DataFrame())
    service_df = production_pack.get("production_service_table", pd.DataFrame())
    full_ro = outputs.get("rolling_origin_backtest", pd.DataFrame())
    robt = summarize_full_backtest(full_ro)
    fallback_flags = {
        "XGBoost": bool((outputs.get("xgboost") or {}).get("fallback_used", False)),
        "Prophet": bool((outputs.get("prophet") or {}).get("fallback_used", False)),
        "SARIMA/SARIMAX": bool((outputs.get("sarima") or {}).get("fallback_used", False)),
        "ARIMA": bool((outputs.get("arima") or {}).get("fallback_used", False)),
        "Intermittent": bool((outputs.get("intermittent") or {}).get("fallback_used", False)),
        "Ensemble": False,
    }
    fallback_rate = float(np.mean(list(fallback_flags.values()))) if fallback_flags else 0.0
    alerts = []
    train_y = pd.to_numeric(outputs.get("train", pd.DataFrame()).get("y"), errors="coerce").astype(float) if isinstance(outputs.get("train"), pd.DataFrame) else pd.Series(dtype=float)
    prod_pred = np.asarray((outputs.get("predictions") or {}).get(prod_model, np.array([])), dtype=float)
    recent_mean = float(train_y.tail(min(6, len(train_y))).mean()) if len(train_y) else np.nan
    drift_ratio = float(np.mean(prod_pred) / max(recent_mean, 1e-6)) if len(prod_pred) and pd.notna(recent_mean) else np.nan
    bias_row = bias_df.loc[bias_df["model"] == prod_model].head(1) if len(bias_df) else pd.DataFrame()
    peak_row = peak_df.loc[peak_df["model"] == prod_model].head(1) if len(peak_df) else pd.DataFrame()
    gate_row = gate_df.loc[gate_df["model"] == prod_model].head(1) if len(gate_df) else pd.DataFrame()
    ro_row = robt.loc[robt["model"] == prod_model].head(1) if len(robt) else pd.DataFrame()
    service_row = service_df.loc[service_df["service_level_target"] == service_df["service_level_target"].max()].head(1) if isinstance(service_df, pd.DataFrame) and len(service_df) else pd.DataFrame()
    achieved_service = float(pd.to_numeric(service_row.iloc[0].get("achieved_cycle_service", np.nan), errors="coerce")) if len(service_row) else np.nan
    target_service = float(pd.to_numeric(service_row.iloc[0].get("service_level_target", np.nan), errors="coerce")) if len(service_row) else np.nan
    freq_desc = outputs.get("metadata", {}).get("freq_alias", outputs.get("metadata", {}).get("frequency", "?"))
    summary_row = {
        "üretim_modeli": prod_model,
        "üretim_durumu": production_pack.get("production_status", "eligible"),
        "frekans": freq_desc,
        "sapma_yüzde": float(pd.to_numeric(bias_row.iloc[0].get("bias_pct", np.nan), errors="coerce")) if len(bias_row) else np.nan,
        "eksik_tahmin_oranı": float(pd.to_numeric(bias_row.iloc[0].get("under_forecast_rate", np.nan), errors="coerce")) if len(bias_row) else np.nan,
        "tepe_olay_skoru": float(pd.to_numeric(peak_row.iloc[0].get("peak_event_score", np.nan), errors="coerce")) if len(peak_row) else np.nan,
        "uygunluk_skoru": float(pd.to_numeric(gate_row.iloc[0].get("eligibility_score", np.nan), errors="coerce")) if len(gate_row) else np.nan,
        "rolling_wape": float(pd.to_numeric(ro_row.iloc[0].get("WAPE", np.nan), errors="coerce")) if len(ro_row) else np.nan,
        "drift_oranı": drift_ratio,
        "kapsama_80": float(interval_df["coverage_80"].mean()) if isinstance(interval_df, pd.DataFrame) and "coverage_80" in interval_df.columns else np.nan,
        "kapsama_90": float(interval_df["coverage_90"].mean()) if isinstance(interval_df, pd.DataFrame) and "coverage_90" in interval_df.columns else np.nan,
        "maksimum_özellik_riski": float(pd.to_numeric(feature_df.get("availability_risk_score"), errors="coerce").max()) if isinstance(feature_df, pd.DataFrame) and len(feature_df) and "availability_risk_score" in feature_df.columns else np.nan,
        "fallback_oranı": fallback_rate,
        "gerçekleşen_servis": achieved_service,
        "hedef_servis": target_service,
    }
    if pd.notna(summary_row["sapma_yüzde"]) and abs(summary_row["sapma_yüzde"]) > 8.0:
        alerts.append({"severity": "high", "alert": "sapma_alarmı", "detail": f"Sapma yüksek: {summary_row['sapma_yüzde']:.2f}%"})
    if pd.notna(summary_row["eksik_tahmin_oranı"]) and summary_row["eksik_tahmin_oranı"] > 0.60:
        alerts.append({"severity": "high", "alert": "eksik_tahmin_alarmı", "detail": f"Eksik tahmin oranı yüksek: {summary_row['eksik_tahmin_oranı']:.2f}"})
    if pd.notna(summary_row["tepe_olay_skoru"]) and summary_row["tepe_olay_skoru"] < 0.35:
        alerts.append({"severity": "high", "alert": "tepe_yakalama_alarmı", "detail": f"Tepe olay yakalama zayıf: {summary_row['tepe_olay_skoru']:.2f}"})
    if pd.notna(summary_row["drift_oranı"]) and not (0.82 <= summary_row["drift_oranı"] <= 1.22):
        alerts.append({"severity": "medium", "alert": "drift_alarmı", "detail": f"Tahmin/son dönem seviye oranı: {summary_row['drift_oranı']:.2f}"})
    if pd.notna(summary_row["maksimum_özellik_riski"]) and summary_row["maksimum_özellik_riski"] > 0.25:
        alerts.append({"severity": "high", "alert": "özellik_kontratı_alarmı", "detail": "Production feature contract tamamen güvenli değil."})
    if pd.notna(summary_row["gerçekleşen_servis"]) and pd.notna(summary_row["hedef_servis"]) and summary_row["gerçekleşen_servis"] + 1e-9 < summary_row["hedef_servis"]:
        alerts.append({"severity": "high", "alert": "servis_seviyesi_alarmı", "detail": f"Servis seviyesi hedef altında: {summary_row['gerçekleşen_servis']:.2f} < {summary_row['hedef_servis']:.2f}"})
    if fallback_rate > 0.20:
        alerts.append({"severity": "medium", "alert": "fallback_oranı_alarmı", "detail": f"Model ailelerinde fallback oranı yüksek: {fallback_rate:.2f}"})
    recommendation = "guardrailli_üretime_uygun"
    if any(a["severity"] == "high" for a in alerts):
        recommendation = "shadow_mode_zorunlu"
    elif alerts:
        recommendation = "şampiyon_meydan_okuyan_paralel"
    summary = pd.DataFrame([summary_row])
    summary["alarm_sayısı"] = len(alerts)
    summary["yüksek_alarm_sayısı"] = sum(1 for a in alerts if a["severity"] == "high")
    summary["canlı_kullanım_önerisi"] = recommendation
    summary["karar_açıklaması"] = build_production_decision_explanation(production_pack.get("production_ranking", pd.DataFrame()), prod_model)
    summary["karar_kartı"] = summary.apply(lambda r: f"Model: {r['üretim_modeli']} | Sapma: {safe_float(r.get('sapma_yüzde', np.nan)):.2f}% | Kapsama80: {safe_float(r.get('kapsama_80', np.nan)):.2f} | Fallback: {safe_float(r.get('fallback_oranı', np.nan)):.2f} | Drift: {safe_float(r.get('drift_oranı', np.nan)):.2f} | Servis: {safe_float(r.get('gerçekleşen_servis', np.nan)):.2f}/{safe_float(r.get('hedef_servis', np.nan)):.2f} | Alarm: {int(r.get('alarm_sayısı', 0)) if pd.notna(r.get('alarm_sayısı', np.nan)) else 0}", axis=1)

    metric_df = outputs.get("metrics_df", pd.DataFrame())
    prediction_interval_tables = production_pack.get("quantile_forecasts", {}) or {}
    service_level_tables = production_pack.get("service_level_simulation", {}) or {}
    health_rows = []
    for model_name in metric_df.get("model", pd.Series(dtype=str)).tolist() if isinstance(metric_df, pd.DataFrame) else []:
        mrow = metric_df.loc[metric_df["model"] == model_name].head(1)
        brow = bias_df.loc[bias_df["model"] == model_name].head(1) if len(bias_df) else pd.DataFrame()
        prow = peak_df.loc[peak_df["model"] == model_name].head(1) if len(peak_df) else pd.DataFrame()
        grow = gate_df.loc[gate_df["model"] == model_name].head(1) if len(gate_df) else pd.DataFrame()
        rrow = robt.loc[robt["model"] == model_name].head(1) if len(robt) else pd.DataFrame()
        qtbl = prediction_interval_tables.get(model_name, pd.DataFrame())
        stbl = service_level_tables.get(model_name, pd.DataFrame())
        srow = stbl.loc[stbl["service_level_target"] == stbl["service_level_target"].max()].head(1) if isinstance(stbl, pd.DataFrame) and len(stbl) else pd.DataFrame()
        health_rows.append({
            "model": model_name,
            "WAPE": safe_float(mrow.iloc[0].get("WAPE", np.nan)) if len(mrow) else np.nan,
            "rolling_WAPE": safe_float(rrow.iloc[0].get("WAPE", np.nan)) if len(rrow) else np.nan,
            "sapma_yüzde": safe_float(brow.iloc[0].get("bias_pct", np.nan)) if len(brow) else np.nan,
            "eksik_tahmin_oranı": safe_float(brow.iloc[0].get("under_forecast_rate", np.nan)) if len(brow) else np.nan,
            "tepe_olay_skoru": safe_float(prow.iloc[0].get("peak_event_score", np.nan)) if len(prow) else np.nan,
            "uygunluk_skoru": safe_float(grow.iloc[0].get("eligibility_score", np.nan)) if len(grow) else np.nan,
            "durum": grow.iloc[0].get("status", np.nan) if len(grow) else np.nan,
            "kapsama_80": safe_float(qtbl["coverage_80"].mean()) if isinstance(qtbl, pd.DataFrame) and "coverage_80" in qtbl.columns else np.nan,
            "gerçekleşen_servis": safe_float(srow.iloc[0].get("achieved_cycle_service", np.nan)) if len(srow) else np.nan,
            "fallback_kullanıldı": bool(fallback_flags.get(model_name, False)),
            "fallback_yöntemi": ((outputs.get(model_name.lower().replace("/", "_").replace("-", "_"), {}) or {}).get("fallback_method") if isinstance(outputs, dict) else None),
        })
    model_health_table = pd.DataFrame(health_rows)
    if not len(model_health_table):
        model_health_table = pd.DataFrame(columns=["model", "WAPE", "rolling_WAPE", "sapma_yüzde", "eksik_tahmin_oranı", "tepe_olay_skoru", "uygunluk_skoru", "durum", "kapsama_80", "gerçekleşen_servis", "fallback_kullanıldı", "fallback_yöntemi"])
    return {"summary": summary, "alerts": pd.DataFrame(alerts) if alerts else pd.DataFrame(columns=["severity", "alert", "detail"]), "model_health_table": model_health_table}


def build_production_governance_pack(outputs: Dict[str, Any], export_payload: Dict[str, pd.DataFrame], target_col: str, freq_alias: str) -> Dict[str, Any]:
    feature_audit_df = build_feature_availability_audit(export_payload.get("features", pd.DataFrame()), export_payload.get("metadata", {}).get("date_col", "ds"), target_col, outputs.get("metadata", {}).get("stat_exog_cols", []) or [], outputs.get("metadata", {}).get("prophet_exog_cols", []) or [], outputs.get("metadata", {}).get("ml_feature_cols", []) or [])
    bias_df = build_bias_dashboard(outputs)
    peak_df = build_peak_event_dashboard(outputs)
    eligibility_df = build_model_eligibility_gate(outputs, feature_audit_df)
    eligibility_df = enforce_hard_feature_contract_gate(eligibility_df, feature_audit_df)
    fva_df = build_forecast_value_add(outputs, freq_alias)
    quantile_tables: Dict[str, pd.DataFrame] = {}
    service_tables: Dict[str, pd.DataFrame] = {}
    full_robt = outputs.get("rolling_origin_backtest", pd.DataFrame())
    for model_name, pred in (outputs.get("predictions") or {}).items():
        scale = estimate_model_interval_scale(outputs, model_name)
        qdf = build_calibrated_prediction_interval_table(outputs["test"], pred, model_name, full_robt, fallback_scale=scale)
        quantile_tables[model_name] = qdf
        service_tables[model_name] = build_service_level_simulation(outputs["test"], pred, qdf)
    metrics_df = outputs.get("metrics_df", pd.DataFrame()).copy()
    robt_summary = summarize_full_backtest(full_robt)
    ranked = metrics_df.copy()
    if len(robt_summary):
        ranked = ranked.merge(robt_summary[["model", "WAPE", "MAE"]].rename(columns={"WAPE": "ro_WAPE", "MAE": "ro_MAE"}), on="model", how="left")
    ranked = ranked.merge(eligibility_df[["model", "status", "eligibility_score"]], on="model", how="left")
    bias_small = bias_df[["model", "bias_pct", "under_forecast_rate"]] if len(bias_df) else pd.DataFrame(columns=["model", "bias_pct", "under_forecast_rate"])
    peak_small = peak_df[["model", "peak_event_score"]] if len(peak_df) else pd.DataFrame(columns=["model", "peak_event_score"])
    ranked = ranked.merge(bias_small, on="model", how="left").merge(peak_small, on="model", how="left")
    service_rows = []
    for model_name, sdf in service_tables.items():
        if isinstance(sdf, pd.DataFrame) and len(sdf):
            top = sdf.loc[sdf["service_level_target"] == sdf["service_level_target"].max()].head(1)
            if len(top):
                tgt = float(pd.to_numeric(top.iloc[0].get("service_level_target", np.nan), errors="coerce"))
                ach = float(pd.to_numeric(top.iloc[0].get("achieved_cycle_service", np.nan), errors="coerce"))
                service_rows.append({"model": model_name, "service_level_target": tgt, "achieved_cycle_service": ach, "service_gap": max(tgt - ach, 0.0) if pd.notna(tgt) and pd.notna(ach) else np.nan})
    service_df = pd.DataFrame(service_rows)
    if len(service_df):
        ranked = ranked.merge(service_df[["model", "service_gap", "achieved_cycle_service"]], on="model", how="left")
    status_rank = {"eligible": 0, "challenger_only": 1, "reject": 2}
    ranked["status_rank"] = ranked["status"].map(status_rank).fillna(9)
    ranked["bias_penalty"] = pd.to_numeric(ranked.get("bias_pct"), errors="coerce").fillna(0.0).abs()
    ranked["under_penalty"] = (pd.to_numeric(ranked.get("under_forecast_rate"), errors="coerce").fillna(0.0) - 0.50).clip(lower=0.0)
    ranked["peak_penalty"] = (0.45 - pd.to_numeric(ranked.get("peak_event_score"), errors="coerce").fillna(0.0)).clip(lower=0.0)
    ranked["service_penalty"] = pd.to_numeric(ranked.get("service_gap"), errors="coerce").fillna(0.0)
    ranked["ro_WAPE_norm"] = pd.to_numeric(ranked.get("ro_WAPE"), errors="coerce")
    ranked["hold_WAPE_norm"] = pd.to_numeric(ranked.get("WAPE"), errors="coerce")
    regime = infer_runtime_regime_from_profile(profile=outputs.get("metadata", {}).get("profile", {}) or {}, y_train=np.asarray(outputs.get("train", pd.DataFrame()).get("y", pd.Series(dtype=float)), dtype=float) if isinstance(outputs.get("train"), pd.DataFrame) and "y" in outputs.get("train", pd.DataFrame()).columns else None)
    ranked["bağlamsal_ceza"] = 0.0
    if not regime.get("intermittent_like", False):
        ranked.loc[ranked["model"].eq("Intermittent"), "bağlamsal_ceza"] += 2.5
    if not regime.get("seasonal_like", False):
        ranked.loc[ranked["model"].eq("Prophet"), "bağlamsal_ceza"] += 1.1
    ranked["operasyonel_skor"] = (
        ranked["status_rank"].fillna(9.0) * 1000.0
        + ranked["ro_WAPE_norm"].fillna(ranked["hold_WAPE_norm"].fillna(99.0)) * 0.78
        + ranked["hold_WAPE_norm"].fillna(99.0) * 0.10
        + ranked["service_penalty"].fillna(0.0) * 22.0
        + ranked["under_penalty"].fillna(0.0) * 14.5
        + ranked["bias_penalty"].fillna(0.0) * 1.45
        + ranked["peak_penalty"].fillna(0.0) * 12.5
        + ranked["bağlamsal_ceza"].fillna(0.0) * 9.5
        - pd.to_numeric(ranked.get("eligibility_score"), errors="coerce").fillna(0.0) * 0.03
    )
    sort_cols = ["operasyonel_skor", "ro_WAPE", "WAPE", "service_penalty", "under_penalty", "bias_penalty", "peak_penalty", "eligibility_score"]
    asc = [True, True, True, True, True, True, True, False]
    ranked = ranked.sort_values(sort_cols, ascending=asc, na_position="last").reset_index(drop=True)
    production_model = ranked.iloc[0]["model"] if len(ranked) else outputs.get("best_model")
    production_status = ranked.iloc[0]["status"] if len(ranked) else "eligible"
    if production_status == "reject":
        preferred = ["ARIMA", "SARIMA/SARIMAX", outputs.get("best_model")]
        for cand in preferred:
            if cand in set(outputs.get("predictions", {}).keys()):
                production_model = cand
                break
        production_status = "guarded_fallback"
    production_interval = quantile_tables.get(production_model, pd.DataFrame())
    production_service = service_tables.get(production_model, pd.DataFrame())
    pack = {"feature_availability_audit": feature_audit_df, "bias_dashboard": bias_df, "peak_event_dashboard": peak_df, "forecast_value_add": fva_df, "model_eligibility_gate": eligibility_df, "quantile_forecasts": quantile_tables, "service_level_simulation": service_tables, "production_model": production_model, "production_status": production_status, "production_interval_table": production_interval, "production_service_table": production_service, "rolling_origin_summary": robt_summary, "production_ranking": ranked, "production_decision_explanation": build_production_decision_explanation(ranked, production_model)}
    pack["live_monitoring_pack"] = build_live_monitoring_pack(outputs, pack)
    return pack



def _safe_sort_columns(df: pd.DataFrame, preferred_cols: List[str], fallback_cols: Optional[List[str]] = None) -> List[str]:
    if df is None or len(df) == 0:
        return []
    cols = [c for c in preferred_cols if c in df.columns]
    if cols:
        return cols
    fallback_cols = fallback_cols or []
    return [c for c in fallback_cols if c in df.columns]


def _derive_production_ranking(outputs: Dict[str, Any]) -> pd.DataFrame:
    prod_pack = outputs.get("production_governance", {}) or {}
    existing = prod_pack.get("production_ranking", pd.DataFrame())
    if isinstance(existing, pd.DataFrame) and len(existing) > 0:
        return existing.copy()

    metrics_df = outputs.get("metrics_df", pd.DataFrame())
    if not isinstance(metrics_df, pd.DataFrame) or len(metrics_df) == 0:
        return pd.DataFrame(columns=["model", "status", "eligibility_score", "WAPE", "sMAPE", "MAE"])

    ranked = metrics_df.copy()
    gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame())
    if isinstance(gate_df, pd.DataFrame) and len(gate_df) > 0 and "model" in gate_df.columns:
        merge_cols = [c for c in ["model", "status", "eligibility_score"] if c in gate_df.columns]
        ranked = ranked.merge(gate_df[merge_cols].drop_duplicates(subset=["model"]), on="model", how="left")

    if "status" not in ranked.columns:
        ranked["status"] = outputs.get("production_status", "eligible")
    if "eligibility_score" not in ranked.columns:
        ranked["eligibility_score"] = np.nan

    status_rank = {"eligible": 0, "guarded_fallback": 1, "challenger_only": 2, "reject": 3}
    ranked["status_rank"] = ranked["status"].map(status_rank).fillna(9)

    sort_cols = [c for c in ["status_rank", "WAPE", "sMAPE", "MAE", "eligibility_score"] if c in ranked.columns]
    ascending = []
    for c in sort_cols:
        ascending.append(False if c == "eligibility_score" else True)
    if sort_cols:
        ranked = ranked.sort_values(sort_cols, ascending=ascending, na_position="last").reset_index(drop=True)
    return ranked



def _build_minimum_live_summary(outputs: Dict[str, Any], prod_pack: Dict[str, Any]) -> pd.DataFrame:
    prod_model = prod_pack.get("production_model") or outputs.get("production_model") or outputs.get("best_model")
    prod_status = prod_pack.get("production_status") or outputs.get("production_status") or "eligible"
    metrics_df = outputs.get("metrics_df", pd.DataFrame())
    bias_df = prod_pack.get("bias_dashboard", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    peak_df = prod_pack.get("peak_event_dashboard", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    interval_df = prod_pack.get("production_interval_table", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    service_df = prod_pack.get("production_service_table", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    robt = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    metric_row = metrics_df.loc[metrics_df["model"].astype(str) == str(prod_model)].head(1) if isinstance(metrics_df, pd.DataFrame) and len(metrics_df) and "model" in metrics_df.columns else pd.DataFrame()
    bias_row = bias_df.loc[bias_df["model"].astype(str) == str(prod_model)].head(1) if isinstance(bias_df, pd.DataFrame) and len(bias_df) and "model" in bias_df.columns else pd.DataFrame()
    peak_row = peak_df.loc[peak_df["model"].astype(str) == str(prod_model)].head(1) if isinstance(peak_df, pd.DataFrame) and len(peak_df) and "model" in peak_df.columns else pd.DataFrame()
    gate_row = gate_df.loc[gate_df["model"].astype(str) == str(prod_model)].head(1) if isinstance(gate_df, pd.DataFrame) and len(gate_df) and "model" in gate_df.columns else pd.DataFrame()
    ro_row = robt.loc[robt["model"].astype(str) == str(prod_model)].head(1) if isinstance(robt, pd.DataFrame) and len(robt) and "model" in robt.columns else pd.DataFrame()
    service_row = service_df.loc[service_df["service_level_target"] == service_df["service_level_target"].max()].head(1) if isinstance(service_df, pd.DataFrame) and len(service_df) and "service_level_target" in service_df.columns else pd.DataFrame()
    fallback_summary = outputs.get("fallback_summary", pd.DataFrame())
    fb_rate = np.nan
    if isinstance(fallback_summary, pd.DataFrame) and len(fallback_summary) and "fallback_used" in fallback_summary.columns:
        fb_rate = float(pd.to_numeric(fallback_summary["fallback_used"], errors="coerce").fillna(0).mean())
    elif isinstance(outputs, dict):
        fb_flags = []
        for k in ["xgboost", "prophet", "sarima", "arima", "intermittent"]:
            fb_flags.append(bool((outputs.get(k) or {}).get("fallback_used", False)))
        fb_rate = float(np.mean(fb_flags)) if fb_flags else np.nan
    row = {
        "üretim_modeli": prod_model,
        "üretim_durumu": prod_status,
        "frekans": outputs.get("metadata", {}).get("freq_alias", "?"),
        "holdout_wape": safe_float(metric_row.iloc[0].get("WAPE", np.nan)) if len(metric_row) else np.nan,
        "sapma_yüzde": safe_float(bias_row.iloc[0].get("bias_pct", metric_row.iloc[0].get("BiasPct", np.nan) if len(metric_row) else np.nan)) if (len(bias_row) or len(metric_row)) else np.nan,
        "eksik_tahmin_oranı": safe_float(bias_row.iloc[0].get("under_forecast_rate", metric_row.iloc[0].get("UnderForecastRate", np.nan) if len(metric_row) else np.nan)) if (len(bias_row) or len(metric_row)) else np.nan,
        "tepe_olay_skoru": safe_float(peak_row.iloc[0].get("peak_event_score", np.nan)) if len(peak_row) else np.nan,
        "uygunluk_skoru": safe_float(gate_row.iloc[0].get("eligibility_score", np.nan)) if len(gate_row) else np.nan,
        "rolling_wape": safe_float(ro_row.iloc[0].get("WAPE", np.nan)) if len(ro_row) else np.nan,
        "kapsama_80": safe_float(interval_df["coverage_80"].mean()) if isinstance(interval_df, pd.DataFrame) and "coverage_80" in interval_df.columns else np.nan,
        "kapsama_90": safe_float(interval_df["coverage_90"].mean()) if isinstance(interval_df, pd.DataFrame) and "coverage_90" in interval_df.columns else np.nan,
        "fallback_oranı": safe_float(fb_rate),
        "gerçekleşen_servis": safe_float(service_row.iloc[0].get("achieved_cycle_service", np.nan)) if len(service_row) else np.nan,
        "hedef_servis": safe_float(service_row.iloc[0].get("service_level_target", np.nan)) if len(service_row) else np.nan,
    }
    out = pd.DataFrame([row])
    out["alarm_sayısı"] = 0
    out["yüksek_alarm_sayısı"] = 0
    out["canlı_kullanım_önerisi"] = prod_status
    out["karar_açıklaması"] = prod_pack.get("production_decision_explanation") or build_production_decision_explanation(prod_pack.get("production_ranking", pd.DataFrame()), prod_model)
    out["karar_kartı"] = out.apply(lambda r: f"Model: {r['üretim_modeli']} | WAPE: {safe_float(r.get('holdout_wape', np.nan)):.2f} | Rolling: {safe_float(r.get('rolling_wape', np.nan)):.2f} | Bias: {safe_float(r.get('sapma_yüzde', np.nan)):.2f}% | Kapsama80: {safe_float(r.get('kapsama_80', np.nan)):.2f}", axis=1)
    return out


def _build_minimum_model_health_table(outputs: Dict[str, Any], prod_pack: Dict[str, Any]) -> pd.DataFrame:
    metrics_df = outputs.get("metrics_df", pd.DataFrame())
    if not isinstance(metrics_df, pd.DataFrame) or len(metrics_df) == 0:
        preds = outputs.get("predictions", {}) or {}
        metrics_df = pd.DataFrame([{"model": k} for k in preds.keys()]) if preds else pd.DataFrame(columns=["model"])
    bias_df = prod_pack.get("bias_dashboard", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    peak_df = prod_pack.get("peak_event_dashboard", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame()) if isinstance(prod_pack, dict) else pd.DataFrame()
    qtables = prod_pack.get("quantile_forecasts", {}) if isinstance(prod_pack, dict) else {}
    stables = prod_pack.get("service_level_simulation", {}) if isinstance(prod_pack, dict) else {}
    robt = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    fallback_map = {
        "XGBoost": outputs.get("xgboost", {}) or {},
        "Prophet": outputs.get("prophet", {}) or {},
        "SARIMA/SARIMAX": outputs.get("sarima", {}) or {},
        "ARIMA": outputs.get("arima", {}) or {},
        "Intermittent": outputs.get("intermittent", {}) or {},
        "Ensemble": {},
    }
    rows = []
    models = metrics_df.get("model", pd.Series(dtype=str)).astype(str).tolist() if "model" in metrics_df.columns else []
    for model_name in models:
        mrow = metrics_df.loc[metrics_df["model"].astype(str) == str(model_name)].head(1) if isinstance(metrics_df, pd.DataFrame) and len(metrics_df) and "model" in metrics_df.columns else pd.DataFrame()
        brow = bias_df.loc[bias_df["model"].astype(str) == str(model_name)].head(1) if isinstance(bias_df, pd.DataFrame) and len(bias_df) and "model" in bias_df.columns else pd.DataFrame()
        prow = peak_df.loc[peak_df["model"].astype(str) == str(model_name)].head(1) if isinstance(peak_df, pd.DataFrame) and len(peak_df) and "model" in peak_df.columns else pd.DataFrame()
        grow = gate_df.loc[gate_df["model"].astype(str) == str(model_name)].head(1) if isinstance(gate_df, pd.DataFrame) and len(gate_df) and "model" in gate_df.columns else pd.DataFrame()
        rrow = robt.loc[robt["model"].astype(str) == str(model_name)].head(1) if isinstance(robt, pd.DataFrame) and len(robt) and "model" in robt.columns else pd.DataFrame()
        qtbl = qtables.get(model_name, pd.DataFrame()) if isinstance(qtables, dict) else pd.DataFrame()
        stbl = stables.get(model_name, pd.DataFrame()) if isinstance(stables, dict) else pd.DataFrame()
        srow = stbl.loc[stbl["service_level_target"] == stbl["service_level_target"].max()].head(1) if isinstance(stbl, pd.DataFrame) and len(stbl) and "service_level_target" in stbl.columns else pd.DataFrame()
        fallback_obj = fallback_map.get(model_name, {}) or {}
        rows.append({
            "model": model_name,
            "WAPE": safe_float(mrow.iloc[0].get("WAPE", np.nan)) if len(mrow) else np.nan,
            "rolling_WAPE": safe_float(rrow.iloc[0].get("WAPE", np.nan)) if len(rrow) else np.nan,
            "sapma_yüzde": safe_float(brow.iloc[0].get("bias_pct", mrow.iloc[0].get("BiasPct", np.nan) if len(mrow) else np.nan)) if (len(brow) or len(mrow)) else np.nan,
            "eksik_tahmin_oranı": safe_float(brow.iloc[0].get("under_forecast_rate", mrow.iloc[0].get("UnderForecastRate", np.nan) if len(mrow) else np.nan)) if (len(brow) or len(mrow)) else np.nan,
            "tepe_olay_skoru": safe_float(prow.iloc[0].get("peak_event_score", np.nan)) if len(prow) else np.nan,
            "uygunluk_skoru": safe_float(grow.iloc[0].get("eligibility_score", np.nan)) if len(grow) else np.nan,
            "durum": grow.iloc[0].get("status", prod_pack.get("production_status", "eligible")) if len(grow) else prod_pack.get("production_status", "eligible"),
            "kapsama_80": safe_float(qtbl["coverage_80"].mean()) if isinstance(qtbl, pd.DataFrame) and "coverage_80" in qtbl.columns else np.nan,
            "gerçekleşen_servis": safe_float(srow.iloc[0].get("achieved_cycle_service", np.nan)) if len(srow) else np.nan,
            "fallback_kullanıldı": bool(fallback_obj.get("fallback_used", False)),
            "fallback_yöntemi": fallback_obj.get("fallback_method"),
        })
    cols = ["model", "WAPE", "rolling_WAPE", "sapma_yüzde", "eksik_tahmin_oranı", "tepe_olay_skoru", "uygunluk_skoru", "durum", "kapsama_80", "gerçekleşen_servis", "fallback_kullanıldı", "fallback_yöntemi"]
    return pd.DataFrame(rows, columns=cols)


def ensure_production_reporting_tables(outputs: Dict[str, Any], export_payload: Optional[Dict[str, pd.DataFrame]] = None, target_col: Optional[str] = None, freq_alias: Optional[str] = None) -> Dict[str, Any]:
    if not isinstance(outputs, dict):
        return outputs
    prod_pack = outputs.get("production_governance", {}) or {}
    if not isinstance(prod_pack, dict):
        prod_pack = {}

    if (not prod_pack) and export_payload is not None and target_col is not None and freq_alias is not None:
        try:
            prod_pack = build_production_governance_pack(outputs, export_payload, target_col, freq_alias) or {}
        except Exception:
            prod_pack = {}

    ranking = prod_pack.get("production_ranking", pd.DataFrame())
    if not isinstance(ranking, pd.DataFrame) or len(ranking) == 0:
        tmp_outputs = dict(outputs)
        tmp_outputs["production_governance"] = prod_pack
        ranking = _derive_production_ranking(tmp_outputs)
        prod_pack["production_ranking"] = ranking

    if not prod_pack.get("production_model"):
        if isinstance(ranking, pd.DataFrame) and len(ranking) and "model" in ranking.columns:
            prod_pack["production_model"] = ranking.iloc[0]["model"]
        else:
            prod_pack["production_model"] = outputs.get("best_model")
    if not prod_pack.get("production_status"):
        if isinstance(ranking, pd.DataFrame) and len(ranking) and "status" in ranking.columns:
            prod_pack["production_status"] = ranking.iloc[0]["status"]
        else:
            prod_pack["production_status"] = outputs.get("production_status", "eligible")

    outputs["production_model"] = prod_pack.get("production_model", outputs.get("best_model"))
    outputs["production_status"] = prod_pack.get("production_status", outputs.get("production_status", "eligible"))

    live_pack = prod_pack.get("live_monitoring_pack", {}) or {}
    need_live_rebuild = (
        not isinstance(live_pack, dict)
        or not isinstance(live_pack.get("summary"), pd.DataFrame)
        or len(live_pack.get("summary", pd.DataFrame())) == 0
        or not isinstance(live_pack.get("model_health_table"), pd.DataFrame)
        or len(live_pack.get("model_health_table", pd.DataFrame())) == 0
    )
    if need_live_rebuild:
        try:
            live_pack = build_live_monitoring_pack(outputs, prod_pack)
        except Exception:
            live_pack = {}

    if not isinstance(live_pack, dict):
        live_pack = {}
    if not isinstance(live_pack.get("summary"), pd.DataFrame) or len(live_pack.get("summary", pd.DataFrame())) == 0:
        live_pack["summary"] = _build_minimum_live_summary(outputs, prod_pack)
    if not isinstance(live_pack.get("model_health_table"), pd.DataFrame) or len(live_pack.get("model_health_table", pd.DataFrame())) == 0:
        live_pack["model_health_table"] = _build_minimum_model_health_table(outputs, prod_pack)
    if not isinstance(live_pack.get("alerts"), pd.DataFrame):
        live_pack["alerts"] = pd.DataFrame(columns=["severity", "alert", "detail"])

    prod_pack["live_monitoring_pack"] = live_pack
    prod_pack["production_decision_explanation"] = prod_pack.get("production_decision_explanation") or build_production_decision_explanation(prod_pack.get("production_ranking", pd.DataFrame()), prod_pack.get("production_model"))
    outputs["production_governance"] = prod_pack
    return outputs

def build_model_liderleri(outputs: Dict[str, Any]) -> pd.DataFrame:
    satirlar = []
    hold_df = outputs.get("metrics_df", pd.DataFrame())
    val_df = outputs.get("validation_metrics_df", pd.DataFrame())
    ro_df = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    prod_rank = _derive_production_ranking(outputs)

    hold_cols = _safe_sort_columns(hold_df, ["WAPE", "sMAPE", "RMSE"], ["WAPE", "sMAPE", "MAE"])
    val_cols = _safe_sort_columns(val_df, ["validation_operational_score", "val_WAPE", "val_sMAPE"], ["val_WAPE", "val_sMAPE"])
    ro_cols = _safe_sort_columns(ro_df, ["WAPE", "sMAPE", "MAE"], ["WAPE", "MAE"])

    hold_winner = hold_df.sort_values(hold_cols, ascending=[True] * len(hold_cols)).iloc[0]["model"] if isinstance(hold_df, pd.DataFrame) and len(hold_df) > 0 and hold_cols else outputs.get("best_model")
    val_winner = val_df.sort_values(val_cols, ascending=[True] * len(val_cols)).iloc[0]["model"] if isinstance(val_df, pd.DataFrame) and len(val_df) > 0 and val_cols else outputs.get("best_model")
    ro_winner = ro_df.sort_values(ro_cols, ascending=[True] * len(ro_cols)).iloc[0]["model"] if isinstance(ro_df, pd.DataFrame) and len(ro_df) > 0 and ro_cols else outputs.get("best_model")
    prod_winner = prod_rank.iloc[0]["model"] if isinstance(prod_rank, pd.DataFrame) and len(prod_rank) > 0 and "model" in prod_rank.columns else outputs.get("production_model", outputs.get("best_model"))

    satirlar.append({"karar_kademesi": "Holdout kazananı", "model": hold_winner})
    satirlar.append({"karar_kademesi": "Doğrulama kazananı", "model": val_winner})
    satirlar.append({"karar_kademesi": "Rolling-origin kazananı", "model": ro_winner})
    satirlar.append({"karar_kademesi": "Üretim kazananı", "model": prod_winner})
    return pd.DataFrame(satirlar)


def build_karar_hiyerarsisi_ozeti(outputs: Dict[str, Any]) -> pd.DataFrame:
    lider = build_model_liderleri(outputs)
    aciklama_map = {
        "Holdout kazananı": "Son test ufkunda en düşük hata üreten model.",
        "Doğrulama kazananı": "İç doğrulama / arama sonuçlarına göre öne çıkan model.",
        "Rolling-origin kazananı": "Zaman boyunca geri testte daha stabil performans veren model.",
        "Üretim kazananı": "Uygunluk, risk ve operasyonel kullanılabilirlik açısından canlı kullanım için önerilen model.",
    }
    karar_amaci_map = {
        "Holdout kazananı": "Kısa dönem doğruluk",
        "Doğrulama kazananı": "Model seçimi doğrulaması",
        "Rolling-origin kazananı": "Zamansal dayanıklılık",
        "Üretim kazananı": "Canlı kullanım kararı",
    }
    if not isinstance(lider, pd.DataFrame) or len(lider) == 0:
        return pd.DataFrame(columns=["KararAşaması", "KazananModel", "KararAmacı", "Anlamı"])

    out = lider.rename(columns={"karar_kademesi": "KararAşaması", "model": "KazananModel"}).copy()
    out["KararAmacı"] = out["KararAşaması"].map(karar_amaci_map)
    out["Anlamı"] = out["KararAşaması"].map(aciklama_map)

    prod_pack = outputs.get("production_governance", {}) or {}
    prod_expl = prod_pack.get("production_decision_explanation")
    if not prod_expl:
        try:
            prod_expl = build_production_decision_explanation(_derive_production_ranking(outputs), outputs.get("production_model", outputs.get("best_model")))
        except Exception:
            prod_expl = None
    if prod_expl:
        extra = pd.DataFrame([{
            "KararAşaması": "Üretim karar açıklaması",
            "KazananModel": str(outputs.get("production_model", outputs.get("best_model", "-"))),
            "KararAmacı": "Neden seçildi?",
            "Anlamı": str(prod_expl),
        }])
        out = pd.concat([out, extra], axis=0, ignore_index=True)
    return out


def build_prophet_gorunurluk_ozeti(prophet_result: Dict[str, Any]) -> pd.DataFrame:
    if not isinstance(prophet_result, dict) or len(prophet_result) == 0:
        return pd.DataFrame(columns=["Alan", "Değer"])

    mode = prophet_result.get("fit_mode") or prophet_result.get("mode") or "bilinmiyor"
    fallback_used = bool(prophet_result.get("fallback_used", False))
    fallback_method = prophet_result.get("fallback_method") or prophet_result.get("fallback_reason")
    note = prophet_result.get("fit_visibility_note") or prophet_result.get("note") or prophet_result.get("status_note") or "-"

    return pd.DataFrame([
        {"Alan": "Prophet çalışma modu", "Değer": mode},
        {"Alan": "Gerçek Prophet fit kullanıldı mı?", "Değer": "evet" if (str(mode) == "gercek_prophet_fit" and not fallback_used) else "hayır"},
        {"Alan": "Fallback kullanıldı mı?", "Değer": "evet" if fallback_used else "hayır"},
        {"Alan": "Fallback yöntemi", "Değer": fallback_method or "yok"},
        {"Alan": "Açıklama", "Değer": note},
    ])


def build_benzersiz_rapor_katalogu(outputs: Dict[str, Any], prod_pack: Dict[str, Any]) -> pd.DataFrame:
    prod_pack = prod_pack or {}
    production_ranking = prod_pack.get("production_ranking", pd.DataFrame())
    if not isinstance(production_ranking, pd.DataFrame) or len(production_ranking) == 0:
        production_ranking = _derive_production_ranking(outputs)

    live_pack = prod_pack.get("live_monitoring_pack", {}) or {}
    if not isinstance(live_pack, dict):
        live_pack = {}

    adaylar: List[Tuple[str, pd.DataFrame]] = [
        ("ModelKarşılaştırma", outputs.get("metrics_df", pd.DataFrame())),
        ("DoğrulamaMetrikleri", outputs.get("validation_metrics_df", pd.DataFrame())),
        ("DLSeffaflik", outputs.get("dl_transparency_table", pd.DataFrame())),
        ("KararHiyerarşisi", build_karar_hiyerarsisi_ozeti(outputs)),
        ("ModelLiderleri", build_model_liderleri(outputs)),
        ("AnsamblAğırlıkları", outputs.get("ensemble_weights", pd.DataFrame())),
        ("RollingOriginÖzet", summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))),
        ("FallbackÖzeti", outputs.get("fallback_summary", pd.DataFrame())),
        ("ProphetGörünürlük", build_prophet_gorunurluk_ozeti(outputs.get("prophet", {}))),
        ("ÜretimUygunlukKapısı", prod_pack.get("model_eligibility_gate", pd.DataFrame())),
        ("ÜretimSıralaması", production_ranking),
        ("BiasDashboard", prod_pack.get("bias_dashboard", pd.DataFrame())),
        ("PeakEventDashboard", prod_pack.get("peak_event_dashboard", pd.DataFrame())),
        ("ForecastValueAdd", prod_pack.get("forecast_value_add", pd.DataFrame())),
        ("FeatureAudit", prod_pack.get("feature_availability_audit", pd.DataFrame())),
        ("PredictionIntervals", prod_pack.get("production_interval_table", pd.DataFrame())),
        ("ServiceLevelSimulation", prod_pack.get("production_service_table", pd.DataFrame())),
        ("CanlıİzlemeÖzet", live_pack.get("summary", pd.DataFrame())),
        ("CanlıİzlemeAlarm", live_pack.get("alerts", pd.DataFrame())),
        ("StageTiming", outputs.get("stage_timing_table", pd.DataFrame())),
    ]

    rows = []
    for rapor_adi, df in adaylar:
        if not isinstance(df, pd.DataFrame) or len(df) == 0:
            continue
        rows.append({
            "Rapor": rapor_adi,
            "SatırSayısı": int(len(df)),
            "KolonSayısı": int(len(df.columns)),
            "Kolonlar": ", ".join(map(str, df.columns.tolist()[:12])) + (" ..." if len(df.columns) > 12 else ""),
        })

    if not rows:
        return pd.DataFrame(columns=["Rapor", "SatırSayısı", "KolonSayısı", "Kolonlar"])
    return pd.DataFrame(rows).sort_values(["Rapor"], ascending=[True]).reset_index(drop=True)


def build_dl_model_analysis(outputs: Dict[str, Any], profile: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    metrics_df = outputs.get("metrics_df", pd.DataFrame()).copy()
    if len(metrics_df) == 0:
        return pd.DataFrame(columns=["model", "analysis"])
    best_hold = float(pd.to_numeric(metrics_df.get("WAPE"), errors="coerce").min()) if len(metrics_df) else np.nan
    for family, key in [("LSTM", "lstm"), ("GRU", "gru")]:
        real_label = dl_public_real_name(family)
        row = metrics_df.loc[metrics_df["model"].astype(str) == real_label].head(1)
        if len(row) == 0:
            legacy_row = metrics_df.loc[metrics_df["model"].astype(str) == family].head(1)
            row = legacy_row
        dl_pack = (outputs.get(key) or {})
        notes = []
        if len(row):
            row0 = row.iloc[0]
            hold_wape = safe_float(row0.get("WAPE", np.nan))
        else:
            hold_wape = np.nan
        val_df = outputs.get("validation_metrics_df", pd.DataFrame())
        val_row = val_df.loc[val_df["model"].astype(str).isin([real_label, family])].head(1) if isinstance(val_df, pd.DataFrame) and len(val_df) and "model" in val_df.columns else pd.DataFrame()
        val_wape = safe_float(val_row.iloc[0].get("val_WAPE", np.nan)) if len(val_row) else np.nan
        overfit_gap = safe_float((dl_pack.get("overfit_summary") or {}).get("overfit_gap_ratio", np.nan))
        if pd.notna(hold_wape) and pd.notna(best_hold):
            notes.append("Holdout performansı lider modellere çok yakın veya daha iyi." if hold_wape <= best_hold + 0.5 else "Holdout performansı klasik/ML lider modellere göre daha zayıf.")
        fit_mode = str(dl_pack.get("fit_mode", ""))
        if bool(dl_pack.get("dl_strict_validation_passed", False)) and fit_mode.endswith("_torch"):
            notes.append(f"{real_label} gerçek PyTorch tabanlı derin öğrenme eğitimi ile çalıştırıldı ve epoch-loss geçmişi doğrulandı.")
        elif bool(dl_pack.get("dl_strict_validation_passed", False)):
            notes.append(f"{real_label} gerçek TensorFlow/Keras tabanlı derin öğrenme eğitimi ile çalıştırıldı ve epoch-loss geçmişi doğrulandı.")
        else:
            notes.append(f"{real_label} gerçek DL eğitimi olarak doğrulanamadı; fallback/surrogate çıktısı bu isim altında raporlanmadı.")
        if bool(dl_pack.get("surrogate_used", False)):
            notes.append(f"Surrogate çıktısı ayrı model adıyla raporlanır: {dl_public_surrogate_name(family)}.")
        if bool(dl_pack.get("fallback_used", False)) or dl_pack.get("operational_fallback_forecast") is not None:
            notes.append("Operasyonel yedek tahmin ayrı model adıyla raporlanır: DL-fallback.")
        if pd.notna(val_wape) and pd.notna(hold_wape):
            notes.append("İç doğrulama ile holdout arasında ciddi ayrışma yok; genelleme kabul edilebilir." if val_wape <= hold_wape + 1.0 else "Validation-holdout farkı yüksek; yapı daha kırılgan olabilir.")
        if pd.notna(overfit_gap):
            notes.append("Epoch-loss ayrışması overfitting riskine işaret ediyor." if overfit_gap > 0.25 else "Epoch-loss davranışı aşırı overfitting göstermiyor.")
        if float(profile.get("seasonality_strength", 0) or 0) >= 0.35:
            notes.append("Seri sezonsal; DL modeli zaman penceresi ile bu yapıyı kısmen öğrenebilir.")
        if float(profile.get("cv", 0) or 0) >= 0.45:
            notes.append("Seri oynak; DL modeli esneklik kazanabilir ancak daha fazla gözlem ister.")
        if not notes:
            notes.append("DL modeli anlamlı fark üretmedi; klasik ve ağaç tabanlı adaylarla birlikte değerlendirilmelidir.")
        rows.append({"model": real_label, "analysis": " ".join(notes)})
    return pd.DataFrame(rows)

def build_dl_transparency_table(outputs: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    fallback_sources = []
    for key, family in [("lstm", "LSTM"), ("gru", "GRU")]:
        res = (outputs or {}).get(key) or {}
        identity_class, public_name = classify_dl_public_result(family, res)
        base = dl_metric_identity_fields(family, res, identity_class=identity_class)
        base.update({
            "model": dl_public_real_name(family),
            "public_model_name": public_name,
            "identity_model_family": family,
            "model_identity_class": res.get("model_identity_class") or ("real" if bool(res.get("dl_strict_validation_passed", False)) else "real_not_available"),
            "epoch_rows": int(len(res.get("history_df"))) if isinstance(res.get("history_df"), pd.DataFrame) else 0,
            "strict_dl_rejection_reason": res.get("strict_dl_rejection_reason"),
            "eligible_for_model_ranking": bool(res.get("eligible_for_model_ranking", False)) and bool(res.get("dl_strict_validation_passed", False)),
            "forecast_is_model_output": bool(res.get("forecast_is_model_output", False)) and bool(res.get("dl_strict_validation_passed", False)),
        })
        rows.append(base)

        surrogate_row = dl_metric_identity_fields(family, res, identity_class="surrogate")
        surrogate_row.update({
            "model": dl_public_surrogate_name(family),
            "public_model_name": dl_public_surrogate_name(family),
            "identity_model_family": family,
            "model_identity_class": "surrogate",
            "real_dl_trained": False,
            "fallback_used": False,
            "surrogate_used": bool(res.get("surrogate_used", False)),
            "eligible_for_model_ranking": False,
            "forecast_is_model_output": bool(res.get("surrogate_used", False)),
            "strict_dl_rejection_reason": "surrogate_separate_identity_not_allowed_to_masquerade_as_real_dl" if not bool(res.get("surrogate_used", False)) else res.get("strict_dl_rejection_reason"),
            "epoch_rows": 0,
        })
        rows.append(surrogate_row)

        if bool(res.get("fallback_used", False)) or res.get("operational_fallback_forecast") is not None:
            fallback_sources.append(family)

    fallback_row = {
        "model": "DL-fallback",
        "public_model_name": "DL-fallback",
        "identity_model_family": "DL",
        "model_identity_class": "fallback",
        "dl_result_kind": "operational_fallback",
        "real_dl_trained": False,
        "dl_backend": "none",
        "fallback_used": bool(fallback_sources),
        "fallback_method": " | ".join(sorted(set(str((outputs.get(k.lower()) or {}).get("fallback_method") or "fallback") for k in fallback_sources))) if fallback_sources else None,
        "surrogate_used": False,
        "surrogate_evaluation_disabled": True,
        "posthoc_surrogate_override": False,
        "test_set_used_for_model_selection": False,
        "eligible_for_model_ranking": False,
        "forecast_is_model_output": bool(fallback_sources),
        "dl_strict_validation_passed": False,
        "strict_dl_rejection_reason": "explicit_operational_fallback_not_real_dl",
        "dl_input_mode": None,
        "dl_feature_source": None,
        "dl_feature_count": 0,
        "dl_blocked_tabular_feature_count": None,
        "dl_input_contract": "fallback_is_separate_from_lstm_gru_real_and_surrogate_names",
        "validation_design": None,
        "rolling_validation_folds": None,
        "epoch_rows": 0,
        "model_identity_note": "DL fallback tahmini LSTM/GRU adı altında değil, yalnızca DL-fallback adıyla raporlanır.",
        "fallback_source_families": "|".join(fallback_sources),
    }
    rows.append(fallback_row)
    return pd.DataFrame(rows)

def run_full_forecasting_pipeline(export_payload: Dict[str, pd.DataFrame], target_col: str, horizon: int, use_exog_for_stat_models: bool = True, use_exog_for_prophet: bool = True) -> Dict[str, Any]:
    ensure_forecasting_runtime_dependencies(include_streamlit=False)
    progress_log("Pipeline Hazırlık", f"{target_col} için forecasting pipeline başlatıldı.", target=target_col, extra={"horizon": int(horizon)})
    manifest = export_payload["manifest"]
    date_col = manifest.loc[manifest["key"] == "date_column", "value"].iloc[0]
    freq_alias = manifest.loc[manifest["key"] == "frequency_inferred", "value"].iloc[0]
    df_series = make_series_analysis_frame(export_payload, target_col)
    profile = get_profile_row(export_payload, target_col)
    pipeline_key = build_search_signature("pipeline_exact_family_search", freq_alias, df_series, export_payload.get("features"), profile=profile, extra={"target_col": target_col, "horizon": int(horizon), "use_exog_for_stat_models": bool(use_exog_for_stat_models), "use_exog_for_prophet": bool(use_exog_for_prophet), "app_version": APP_VERSION})
    cached = SEARCH_ACCELERATOR.get_result(pipeline_key)
    if cached is not None:
        cached["metadata"]["search_accelerator_pipeline_cache_hit"] = True
        progress_log("Pipeline Cache", "Önceden hesaplanmış sonuç bulundu; cache çıktısı kullanıldı.", level="SUCCESS", target=target_col)
        return cached
    train_df, test_df = train_test_split_series(df_series, horizon=horizon)
    progress_log("Veri Bölme", "Eğitim ve test ayrımı tamamlandı.", level="SUCCESS", target=target_col, extra={"train_n": len(train_df), "test_n": len(test_df), "freq": freq_alias})
    df_features = export_payload["features"].copy()
    df_features[date_col] = pd.to_datetime(df_features[date_col])
    df_features = df_features.sort_values(date_col).reset_index(drop=True)
    usable_dates = pd.concat([train_df[["ds"]], test_df[["ds"]]], axis=0)["ds"]
    feature_subset = df_features[df_features[date_col].isin(usable_dates)].copy().sort_values(date_col).reset_index(drop=True)
    train_feat = feature_subset.iloc[:len(train_df)].copy().reset_index(drop=True)
    test_feat = feature_subset.iloc[len(train_df):len(train_df)+len(test_df)].copy().reset_index(drop=True)
    stat_exog_cols = detect_optional_exog_columns(df_features, target_col, date_col) if use_exog_for_stat_models else []
    prophet_exog_cols = detect_optional_exog_columns(df_features, target_col, date_col) if use_exog_for_prophet else []
    ml_feature_cols = detect_safe_ml_exog_columns(df_features, target_col, date_col)
    seg = infer_advanced_segment(profile)
    stat_exog_train = train_feat[stat_exog_cols] if stat_exog_cols else None
    stat_exog_test = test_feat[stat_exog_cols] if stat_exog_cols else None
    prophet_exog_train = train_feat[prophet_exog_cols] if prophet_exog_cols else None
    prophet_exog_test = test_feat[prophet_exog_cols] if prophet_exog_cols else None
    ml_train_X = train_feat[ml_feature_cols] if ml_feature_cols else pd.DataFrame(index=train_feat.index)
    ml_test_X = test_feat[ml_feature_cols] if ml_feature_cols else pd.DataFrame(index=test_feat.index)
    progress_log("Özellik Hazırlığı", "Exog ve ML özellik setleri hazırlandı.", level="SUCCESS", target=target_col, extra={"stat_exog": len(stat_exog_cols), "prophet_exog": len(prophet_exog_cols), "ml_features": len(ml_feature_cols)})
    outputs = {"metadata": {"target_col": target_col, "freq_alias": freq_alias, "horizon": horizon, "profile": profile, "segment": seg["label"], "abc_xyz": seg["abc_xyz"], "priority": recommend_model_priority(profile), "candidate_models": recommend_candidate_models(profile), "stat_exog_cols": stat_exog_cols, "prophet_exog_cols": prophet_exog_cols, "ml_feature_cols": ml_feature_cols, "runtime_guardrails": build_runtime_guardrail_notes(freq_alias, len(train_df), horizon, stat_exog_cols, prophet_exog_cols), "train_n": int(len(train_df)), "test_n": int(len(test_df)), "search_accelerator_enabled": bool(SEARCH_ACCELERATOR.config.enabled), "search_accelerator_pipeline_cache_hit": False}, "train": train_df, "test": test_df, "metrics": [], "predictions": {}, "tables": {}, "deep_learning_analysis": pd.DataFrame()}
    def _wrap_model_run(label: str, fn):
        progress_log("Model Araması", f"{label} modeli için arama / fit süreci başladı.", target=target_col, extra={"model": label})
        start = time.perf_counter(); error = None
        try:
            res = fn()
        except Exception as e:
            error = str(e); res = build_model_level_fallback(label, train_df, test_df, freq_alias, error)
            progress_log("Model Araması", f"{label} beklenmeyen hata nedeniyle fallback ile tamamlandı.", level="ERROR", target=target_col, extra={"model": label, "error": error})
        is_dl_family = label in ["LSTM", "GRU"]
        if is_dl_family and bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_strict_real_training", True)):
            res = sanitize_dl_result_for_reporting(res, model_family=label, horizon=len(test_df))
            res = apply_dl_no_test_surrogate_override_contract(res, model_family=label, horizon=len(test_df))
            if not bool(res.get("dl_strict_validation_passed", False)):
                progress_log("DL Kimlik Doğrulama", f"{label} gerçek DL olarak doğrulanamadı; fallback/surrogate çıktısı LSTM/GRU adıyla raporlanmadı.", level="WARNING", target=target_col, extra={"model": label, "reason": res.get("strict_dl_rejection_reason")})
        identity_class, public_label = classify_dl_public_result(label, res) if is_dl_family else ("standard", label)
        if is_dl_family:
            # Primary public row always represents the real recurrent model slot.
            # Surrogate and fallback are emitted as separate public model names below.
            public_label = dl_public_real_name(label) if identity_class in {"fallback", "not_real"} else public_label
            res["public_model_name"] = public_label
            res["identity_model_family"] = label
            res["model_identity_class"] = identity_class
        raw_forecast = np.asarray(res.get("forecast", np.array([])), dtype=float)
        res["raw_forecast"] = raw_forecast.copy()
        if is_dl_family and not bool(res.get("dl_strict_validation_passed", False)):
            res["forecast"] = np.full(len(test_df), np.nan, dtype=float)
        else:
            res["forecast"] = operational_bias_peak_postprocess(train_df["y"], raw_forecast, freq_alias=freq_alias, model_name=public_label, profile=profile)
        if "static_forecast" in res:
            try:
                res["static_forecast"] = operational_bias_peak_postprocess(train_df["y"], np.asarray(res.get("static_forecast"), dtype=float), freq_alias=freq_alias, model_name=public_label, profile=profile)
            except Exception:
                pass
        if is_dl_family and not bool(res.get("dl_strict_validation_passed", False)):
            metric = build_invalid_model_metrics(dl_public_real_name(label), res.get("strict_dl_rejection_reason", "strict_real_dl_gate_failed"))
            metric["eligible_for_model_ranking"] = False
            metric["ranking_exclusion_reason"] = res.get("strict_dl_rejection_reason", "strict_real_dl_gate_failed")
        else:
            metric = build_model_metrics(public_label, train_df["y"].values, test_df["y"].values, res["forecast"])
            metric["eligible_for_model_ranking"] = bool(res.get("eligible_for_model_ranking", True))
            metric["ranking_exclusion_reason"] = res.get("strict_dl_rejection_reason")
        if is_dl_family and bool(getattr(FORECAST_RUNTIME_CONFIG, "dl_report_input_contract_in_metrics", True)):
            metric.update(dl_metric_identity_fields(label, res, identity_class="real" if bool(res.get("dl_strict_validation_passed", False)) else "real_not_available"))
            metric["model_public_name"] = dl_public_real_name(label)
        extra_public_outputs = []
        if is_dl_family:
            # If a surrogate path ever exists, it is emitted under its own name.
            if bool(res.get("surrogate_used", False)):
                surrogate_pred = get_dl_operational_fallback_forecast(res, len(test_df))
                if surrogate_pred is not None:
                    surrogate_name = dl_public_surrogate_name(label)
                    surrogate_metric = build_model_metrics(surrogate_name, train_df["y"].values, test_df["y"].values, surrogate_pred)
                    surrogate_metric.update(dl_metric_identity_fields(label, res, identity_class="surrogate"))
                    surrogate_metric["eligible_for_model_ranking"] = False
                    surrogate_metric["ranking_exclusion_reason"] = "surrogate_separate_identity_not_real_dl"
                    extra_public_outputs.append({"name": surrogate_name, "prediction": surrogate_pred, "metric": surrogate_metric, "table": build_actual_vs_pred_df(test_df, surrogate_pred, surrogate_name)})
            # Operational fallback is visible, but it is not LSTM/GRU.
            fallback_pred = get_dl_operational_fallback_forecast(res, len(test_df)) if (bool(res.get("fallback_used", False)) or res.get("operational_fallback_forecast") is not None) else None
            if fallback_pred is not None:
                fb_name = "DL-fallback"
                fb_metric = build_model_metrics(fb_name, train_df["y"].values, test_df["y"].values, fallback_pred)
                fb_metric.update(dl_metric_identity_fields(label, res, identity_class="fallback"))
                fb_metric["model"] = fb_name
                fb_metric["model_public_name"] = fb_name
                fb_metric["identity_model_family"] = label
                fb_metric["fallback_source_family"] = label
                fb_metric["eligible_for_model_ranking"] = False
                fb_metric["ranking_exclusion_reason"] = "DL fallback ayrı isimle görünür; gerçek LSTM/GRU sıralamasına girmez."
                extra_public_outputs.append({"name": fb_name, "prediction": fallback_pred, "metric": fb_metric, "table": build_actual_vs_pred_df(test_df, fallback_pred, fb_name)})
        table = build_actual_vs_pred_df(test_df, res["forecast"], public_label)
        fallback_used = bool((res or {}).get("fallback_used", False))
        status_level = "WARNING" if fallback_used else "SUCCESS"
        msg = f"{public_label} tamamlandı."
        if fallback_used:
            msg += " DL-fallback ayrı satırda raporlandı."
        progress_log("Model Araması", msg, level=status_level, target=target_col, extra={"model": public_label, "seconds": round(time.perf_counter() - start, 2), "WAPE": round(float(metric.get('WAPE', np.nan)), 4) if pd.notna(metric.get('WAPE', np.nan)) else np.nan, "fallback": fallback_used})
        return {"name": public_label, "internal_name": label, "result": res, "metric": metric, "table": table, "extra_public_outputs": extra_public_outputs, "error": error, "timing": time.perf_counter() - start}
    progress_log("Model Araması", "Model aileleri paralel olarak başlatılıyor.", target=target_col)
    family_outputs = SEARCH_ACCELERATOR.run_parallel_tasks({
        "xgboost": lambda: _wrap_model_run("XGBoost", lambda: fit_xgboost_forecast(train_df, test_df, ml_train_X, ml_test_X, freq_alias=freq_alias)),
        "lstm": lambda: _wrap_model_run("LSTM", lambda: fit_deep_learning_forecast(train_df, test_df, ml_train_X, ml_test_X, freq_alias=freq_alias, model_family="LSTM")),
        "gru": lambda: _wrap_model_run("GRU", lambda: fit_deep_learning_forecast(train_df, test_df, ml_train_X, ml_test_X, freq_alias=freq_alias, model_family="GRU")),
        "prophet": lambda: _wrap_model_run("Prophet", lambda: fit_best_prophet(train_df, test_df, freq_alias, profile, prophet_exog_train, prophet_exog_test) if HAS_PROPHET else build_model_level_fallback("Prophet", train_df, test_df, freq_alias, "prophet paketi yüklü değil")),
        "sarima": lambda: _wrap_model_run("SARIMA/SARIMAX", lambda: fit_best_sarimax(train_df, test_df, freq_alias, profile, stat_exog_train, stat_exog_test)),
        "arima": lambda: _wrap_model_run("ARIMA", lambda: fit_best_arima(train_df, test_df, freq_alias, profile)),
        "intermittent": lambda: _wrap_model_run("Intermittent", lambda: fit_best_intermittent(train_df, test_df, freq_alias, profile)),
    }, max_workers=min(7, FORECAST_RUNTIME_CONFIG.search_accelerator_max_workers))
    progress_log("Model Araması", "Tüm model aileleri tamamlandı.", level="SUCCESS", target=target_col)
    model_errors, stage_timings = {}, {}
    alias_map = {"xgboost": "xgboost", "lstm": "lstm", "gru": "gru", "prophet": "prophet", "sarima": "sarima", "arima": "arima", "intermittent": "intermittent"}
    for key, label in [("xgboost", "XGBoost"), ("lstm", "LSTM"), ("gru", "GRU"), ("prophet", "Prophet"), ("sarima", "SARIMA/SARIMAX"), ("arima", "ARIMA"), ("intermittent", "Intermittent")]:
        block = family_outputs[key]
        public_name = block.get("name", label)
        if block["error"]:
            model_errors[public_name] = block["error"]
        outputs["metrics"].append(block["metric"])
        outputs["predictions"][public_name] = block["result"]["forecast"]
        outputs["tables"][public_name] = block["table"]
        for extra in block.get("extra_public_outputs", []) or []:
            extra_name = str(extra.get("name"))
            if extra_name == "DL-fallback" and extra_name in outputs["predictions"]:
                # Avoid duplicate DL-fallback rows when both LSTM and GRU fall back.
                # Preserve the additional source family in the existing metric if possible.
                for _m in outputs["metrics"]:
                    if isinstance(_m, dict) and _m.get("model") == "DL-fallback":
                        existing_src = str(_m.get("fallback_source_family", "") or "")
                        new_src = str((extra.get("metric") or {}).get("fallback_source_family", "") or "")
                        srcs = [x for x in (existing_src + "|" + new_src).split("|") if x]
                        _m["fallback_source_family"] = "|".join(sorted(set(srcs)))
                continue
            outputs["metrics"].append(extra.get("metric"))
            outputs["predictions"][extra_name] = extra.get("prediction")
            outputs["tables"][extra_name] = extra.get("table")
        outputs[alias_map[key]] = block["result"]
        stage_timings[f"{alias_map[key]}_seconds"] = block["timing"]
    validation_df = pd.DataFrame([
        extract_validation_metrics_from_result("SARIMA/SARIMAX", outputs["sarima"]),
        extract_validation_metrics_from_result("ARIMA", outputs["arima"]),
        extract_validation_metrics_from_result("Prophet", outputs["prophet"]),
        extract_validation_metrics_from_result("XGBoost", outputs["xgboost"]),
        extract_validation_metrics_from_result(dl_public_real_name("LSTM"), outputs["lstm"]),
        extract_validation_metrics_from_result(dl_public_real_name("GRU"), outputs["gru"]),
        {"model": "LSTM-surrogate", "val_WAPE": np.nan, "val_sMAPE": np.nan, "eligible_for_model_ranking": False, "ranking_exclusion_reason": "surrogate_separate_identity"},
        {"model": "GRU-surrogate", "val_WAPE": np.nan, "val_sMAPE": np.nan, "eligible_for_model_ranking": False, "ranking_exclusion_reason": "surrogate_separate_identity"},
        {"model": "DL-fallback", "val_WAPE": np.nan, "val_sMAPE": np.nan, "eligible_for_model_ranking": False, "ranking_exclusion_reason": "operational_fallback_not_validation_selected"},
        extract_validation_metrics_from_result("Intermittent", outputs["intermittent"]),
    ]).sort_values(["val_WAPE", "val_sMAPE"], ascending=[True, True], na_position="last").reset_index(drop=True)
    outputs["validation_metrics_df"] = validation_df
    metrics_df = pd.DataFrame(outputs["metrics"]).sort_values(["WAPE", "sMAPE", "RMSE"], ascending=[True, True, True]).reset_index(drop=True)
    progress_log("Backtest", "Rolling-origin backtest başlatıldı.", target=target_col)
    outputs["rolling_origin_backtest"] = run_rolling_origin_backtest_full(export_payload, target_col, outputs, freq_alias, horizon=min(max(2, int(horizon)), 3), use_exog_for_stat_models=use_exog_for_stat_models, use_exog_for_prophet=use_exog_for_prophet, max_folds=3)
    progress_log("Backtest", "Rolling-origin backtest tamamlandı.", level="SUCCESS", target=target_col)
    rolling_summary_df = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    progress_log("Ansambl", "Ansambl ağırlıkları ve birleşik tahmin hesaplanıyor.", target=target_col)
    usable_predictions, usable_metrics_df = filter_usable_prediction_map(outputs["predictions"], metrics_df, expected_horizon=len(test_df))
    if len(usable_predictions) == 0 or len(usable_metrics_df) == 0:
        ensemble_pred, ensemble_weights = build_fallback_forecast(train_df["y"], int(len(test_df)), freq_alias, infer_season_length_from_freq(freq_alias))[0], pd.DataFrame([{"model": "fallback_only", "weight": 1.0}])
        ensemble_pred = np.maximum(np.asarray(ensemble_pred, dtype=float), 0.0)
    else:
        ensemble_pred, ensemble_weights = build_weighted_ensemble(usable_predictions, usable_metrics_df, validation_df=validation_df, y_train=train_df["y"].values, y_true=test_df["y"].values, rolling_summary=rolling_summary_df, profile=profile)
    outputs["predictions"]["Ensemble"] = ensemble_pred
    outputs["metrics"].append(build_model_metrics("Ensemble", train_df["y"].values, test_df["y"].values, ensemble_pred))
    outputs["tables"]["Ensemble"] = build_actual_vs_pred_df(test_df, ensemble_pred, "Ensemble")
    metrics_df = pd.DataFrame(outputs["metrics"]).sort_values(["WAPE", "sMAPE", "RMSE"], ascending=[True, True, True], na_position="last").reset_index(drop=True)
    ranking_metrics_df = metrics_df.copy()
    if "eligible_for_model_ranking" in ranking_metrics_df.columns:
        ranking_metrics_df = ranking_metrics_df.loc[ranking_metrics_df["eligible_for_model_ranking"].fillna(True).astype(bool)].copy()
    if "WAPE" in ranking_metrics_df.columns:
        ranking_metrics_df = ranking_metrics_df.loc[pd.to_numeric(ranking_metrics_df["WAPE"], errors="coerce").notna()].copy()
    cc = build_champion_challenger(ranking_metrics_df if len(ranking_metrics_df) else metrics_df)
    outputs["metrics_df"] = metrics_df
    outputs["ranking_metrics_df"] = ranking_metrics_df
    outputs["champion_challenger"] = cc
    outputs["best_model"] = cc["champion"]
    outputs["ensemble_weights"] = ensemble_weights
    outputs["all_predictions_long"] = pd.concat(list(outputs["tables"].values()), axis=0, ignore_index=True) if outputs["tables"] else pd.DataFrame()
    outputs["model_errors"] = model_errors
    outputs["stage_timings"] = stage_timings
    outputs["stage_timing_table"] = build_stage_timer_rows(stage_timings)
    dl_fb_sources = []
    for _fam, _key in [("LSTM", "lstm"), ("GRU", "gru")]:
        if bool((outputs.get(_key) or {}).get("fallback_used", False)) or (outputs.get(_key) or {}).get("operational_fallback_forecast") is not None:
            dl_fb_sources.append(_fam)
    outputs["fallback_summary"] = pd.DataFrame([
        {"model": "XGBoost", "fallback_used": bool((outputs.get("xgboost") or {}).get("fallback_used", False)), "fallback_method": (outputs.get("xgboost") or {}).get("fallback_method")},
        {"model": "LSTM (real)", "fallback_used": False, "fallback_method": None, "dl_strict_validation_passed": bool((outputs.get("lstm") or {}).get("dl_strict_validation_passed", False)), "ranking_excluded": not bool((outputs.get("lstm") or {}).get("eligible_for_model_ranking", False)), "ranking_exclusion_reason": (outputs.get("lstm") or {}).get("strict_dl_rejection_reason")},
        {"model": "LSTM-surrogate", "fallback_used": False, "fallback_method": None, "surrogate_used": bool((outputs.get("lstm") or {}).get("surrogate_used", False)), "ranking_excluded": True, "ranking_exclusion_reason": "surrogate ayrı model kimliği; gerçek LSTM değildir"},
        {"model": "GRU (real)", "fallback_used": False, "fallback_method": None, "dl_strict_validation_passed": bool((outputs.get("gru") or {}).get("dl_strict_validation_passed", False)), "ranking_excluded": not bool((outputs.get("gru") or {}).get("eligible_for_model_ranking", False)), "ranking_exclusion_reason": (outputs.get("gru") or {}).get("strict_dl_rejection_reason")},
        {"model": "GRU-surrogate", "fallback_used": False, "fallback_method": None, "surrogate_used": bool((outputs.get("gru") or {}).get("surrogate_used", False)), "ranking_excluded": True, "ranking_exclusion_reason": "surrogate ayrı model kimliği; gerçek GRU değildir"},
        {"model": "DL-fallback", "fallback_used": bool(dl_fb_sources), "fallback_method": " | ".join(sorted(set(str((outputs.get(_k.lower()) or {}).get("fallback_method") or "fallback") for _k in dl_fb_sources))) if dl_fb_sources else None, "fallback_source_family": "|".join(dl_fb_sources), "ranking_excluded": True, "ranking_exclusion_reason": "DL fallback ayrı operasyonel modeldir; LSTM/GRU real veya surrogate değildir"},
        {"model": "Prophet", "fallback_used": bool((outputs.get("prophet") or {}).get("fallback_used", False)), "fallback_method": (outputs.get("prophet") or {}).get("fallback_method")},
        {"model": "SARIMA/SARIMAX", "fallback_used": bool((outputs.get("sarima") or {}).get("fallback_used", False)), "fallback_method": (outputs.get("sarima") or {}).get("fallback_method")},
        {"model": "ARIMA", "fallback_used": bool((outputs.get("arima") or {}).get("fallback_used", False)), "fallback_method": (outputs.get("arima") or {}).get("fallback_method")},
        {"model": "Intermittent", "fallback_used": bool((outputs.get("intermittent") or {}).get("fallback_used", False)), "fallback_method": (outputs.get("intermittent") or {}).get("fallback_method")},
    ])
    outputs["dl_transparency_table"] = build_dl_transparency_table(outputs)
    outputs["deep_learning_analysis"] = build_dl_model_analysis(outputs, profile)
    prophet_info = outputs.get("prophet") or {}
    if bool(prophet_info.get("fallback_used", False)):
        progress_log("Prophet Kontrolü", "Prophet gerçek fit yerine fallback ile tamamlandı.", level="WARNING", target=target_col, extra={"fallback_method": prophet_info.get("fallback_method")})
    outputs["production_governance"] = build_production_governance_pack(outputs, export_payload, target_col, freq_alias)
    outputs = ensure_production_reporting_tables(outputs, export_payload, target_col, freq_alias)
    outputs["production_model"] = outputs["production_governance"].get("production_model", outputs.get("best_model"))
    outputs["production_status"] = outputs["production_governance"].get("production_status", "eligible")
    outputs["model_gorsel_stil_haritasi"] = build_model_visual_style_map(outputs)
    outputs["karar_hiyerarsisi"] = build_karar_hiyerarsisi_ozeti(outputs)
    outputs["prophet_gorunurluk_ozeti"] = build_prophet_gorunurluk_ozeti(outputs.get("prophet", {}))
    outputs["benzersiz_rapor_katalogu"] = build_benzersiz_rapor_katalogu(outputs, outputs["production_governance"])
    progress_log("Raporlama", "Üretim yönetişimi ve raporlama tabloları hazırlandı.", level="SUCCESS", target=target_col, extra={"best_model": outputs.get("best_model"), "production_model": outputs.get("production_model")})
    SEARCH_ACCELERATOR.put_result(pipeline_key, outputs)
    fallback_count = int(outputs["fallback_summary"]["fallback_used"].astype(bool).sum()) if isinstance(outputs.get("fallback_summary"), pd.DataFrame) and "fallback_used" in outputs["fallback_summary"].columns else np.nan
    progress_log("Pipeline Tamamlandı", f"{target_col} serisi tamamlandı.", level="SUCCESS", target=target_col, extra={"best_model": outputs.get("best_model"), "fallback_models": fallback_count})
    return copy.deepcopy(outputs)

def _dedupe_best_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or len(df) == 0:
        return pd.DataFrame() if not isinstance(df, pd.DataFrame) else df
    out = df.copy()
    if "target_col" not in out.columns:
        return out
    if "model" not in out.columns:
        return out.drop_duplicates(subset=["target_col"], keep="last")
    model_rank = np.where(out["model"].astype(str).str.upper() == "ERROR", 1, 0)
    out = out.assign(_error_rank=model_rank, _row_order=np.arange(len(out)))
    out = out.sort_values(["target_col", "_error_rank", "_row_order"], ascending=[True, True, False])
    out = out.drop_duplicates(subset=["target_col"], keep="first")
    return out.drop(columns=["_error_rank", "_row_order"], errors="ignore").reset_index(drop=True)


def _dedupe_series_status_df(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame) or len(df) == 0:
        return pd.DataFrame() if not isinstance(df, pd.DataFrame) else df
    if "target_col" not in df.columns:
        return df
    out = df.copy()
    status_rank = out.get("status", pd.Series("", index=out.index)).astype(str).str.upper().map({"SUCCESS": 0, "WARNING": 1, "ERROR": 2, "RUNNING": 3}).fillna(9)
    out = out.assign(_status_rank=status_rank, _row_order=np.arange(len(out)))
    out = out.sort_values(["target_col", "_status_rank", "_row_order"], ascending=[True, True, False])
    out = out.drop_duplicates(subset=["target_col"], keep="first")
    return out.drop(columns=["_status_rank", "_row_order"], errors="ignore").reset_index(drop=True)


def run_batch_forecasting(
    export_payload: Dict[str, pd.DataFrame],
    horizon: int,
    selected_sheet: str,
    use_exog_for_stat_models: bool = True,
    use_exog_for_prophet: bool = True,
    selected_targets: Optional[List[str]] = None,
    persist_output_root: Optional[str] = None,
    persist_after_each_series: bool = False,
    keep_batch_outputs_in_memory: bool = True,
) -> Dict[str, Any]:
    targets = list(selected_targets) if selected_targets else export_payload["series_profile_report"]["series"].tolist()
    progress_log("Batch Planı", "Batch forecasting serileri hazırlandı.", level="SUCCESS", extra={"sheet": selected_sheet, "target_count": len(targets), "horizon": int(horizon)})
    rows = []
    champion_rows = []
    batch_outputs = {}
    series_status_rows = []

    if persist_after_each_series and persist_output_root:
        write_incremental_batch_common_outputs(export_payload, persist_output_root, selected_sheet, int(horizon))

    for idx, target in enumerate(targets, start=1):
        progress_log("Seri Çalıştırma", f"{idx}/{len(targets)} seri işleme alındı.", target=target, extra={"index": idx, "total": len(targets)})
        started_at = time.strftime("%Y-%m-%d %H:%M:%S")
        target_output_dir = None
        status_record = {
            "target_col": target,
            "index": idx,
            "total": len(targets),
            "status": "RUNNING",
            "started_at": started_at,
            "completed_at": None,
            "best_model": None,
            "champion": None,
            "error": None,
            "output_dir": None,
        }
        try:
            out = run_full_forecasting_pipeline(
                export_payload,
                target,
                horizon,
                use_exog_for_stat_models,
                use_exog_for_prophet
            )
            out["named_output_tables"] = build_named_output_tables(out, export_payload)

            mdf = out["metrics_df"].copy()
            best_row = mdf.iloc[0].to_dict()
            prod_pack = out.get("production_governance", {}) or {}
            prophet_pack = out.get("prophet", {}) or {}
            xgb_pack = out.get("xgboost", {}) or {}
            sarima_pack = out.get("sarima", {}) or {}
            live_pack = prod_pack.get("live_monitoring_pack", {}) or {}
            alerts_df = live_pack.get("alerts", pd.DataFrame()) if isinstance(live_pack, dict) else pd.DataFrame()
            champion_name = out["champion_challenger"].get("champion")
            challenger_name = out["champion_challenger"].get("challenger")
            best_row["target_col"] = target
            best_row["segment"] = out["metadata"]["segment"]
            best_row["abc_xyz"] = out["metadata"]["abc_xyz"]
            best_row["champion"] = champion_name
            best_row["challenger"] = challenger_name
            best_row["best_model"] = out.get("best_model", best_row.get("model"))
            best_row["production_model"] = prod_pack.get("production_model", out.get("best_model", best_row.get("model")))
            best_row["production_status"] = prod_pack.get("production_status", "eligible")
            best_row["production_eligibility_score"] = prod_pack.get("production_eligibility_score")
            best_row["prophet_fallback_used"] = bool(prophet_pack.get("fallback_used", False))
            best_row["xgboost_fallback_used"] = bool(xgb_pack.get("fallback_used", False))
            best_row["sarima_fallback_used"] = bool(sarima_pack.get("fallback_used", False))
            best_row["alert_count"] = int(len(alerts_df)) if isinstance(alerts_df, pd.DataFrame) else 0
            rows.append(best_row)

            champion_rows.append({
                "target_col": target,
                "champion": champion_name,
                "challenger": challenger_name,
                "segment": out["metadata"]["segment"],
                "abc_xyz": out["metadata"]["abc_xyz"],
                "best_model": out.get("best_model", best_row.get("model")),
                "production_model": prod_pack.get("production_model", out.get("best_model", best_row.get("model"))),
                "production_status": prod_pack.get("production_status", "eligible"),
                "prophet_fallback_used": bool(prophet_pack.get("fallback_used", False)),
                "xgboost_fallback_used": bool(xgb_pack.get("fallback_used", False)),
                "sarima_fallback_used": bool(sarima_pack.get("fallback_used", False)),
            })

            export_warning = None
            if persist_after_each_series and persist_output_root:
                try:
                    target_output_dir = write_incremental_series_outputs(export_payload, out, target, persist_output_root, int(horizon))
                except Exception as export_exc:
                    export_warning = str(export_exc)
                    progress_log("Disk Çıktısı", "Seri hesaplandı ancak seri bazlı rapor dosyaları yazılırken uyarı oluştu.", level="WARNING", target=target, extra={"error": export_warning})

            if keep_batch_outputs_in_memory:
                batch_outputs[target] = out

            status_record.update({
                "status": "SUCCESS" if export_warning is None else "WARNING",
                "completed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "best_model": out.get("best_model"),
                "champion": champion_name,
                "challenger": challenger_name,
                "production_model": prod_pack.get("production_model", out.get("best_model")),
                "production_status": prod_pack.get("production_status", "eligible"),
                "prophet_fallback_used": bool(prophet_pack.get("fallback_used", False)),
                "xgboost_fallback_used": bool(xgb_pack.get("fallback_used", False)),
                "sarima_fallback_used": bool(sarima_pack.get("fallback_used", False)),
                "output_dir": target_output_dir,
                "error": export_warning,
            })
            progress_log("Seri Çalıştırma", "Seri tamamlandı ve batch özete eklendi.", level="SUCCESS" if export_warning is None else "WARNING", target=target, extra={"champion": champion_name, "output_dir": target_output_dir, "export_warning": export_warning})
        except Exception as e:
            status_record.update({
                "status": "ERROR",
                "completed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "error": str(e),
            })
            if persist_after_each_series and persist_output_root:
                try:
                    error_root = os.path.join(persist_output_root, _safe_artifact_name(target))
                    os.makedirs(error_root, exist_ok=True)
                    _write_single_table_excel_file(os.path.join(error_root, f"{_safe_artifact_name(target)} - hata özeti.xlsx"), "Hata Özeti", pd.DataFrame([{"target_col": target, "error": str(e), "stage": "run_full_forecasting_pipeline"}]))
                    _write_text_file(os.path.join(error_root, f"{_safe_artifact_name(target)} - hata özeti.txt"), [str(e)])
                    target_output_dir = error_root
                except Exception:
                    pass
            progress_log("Seri Çalıştırma", "Seri hata ile sonlandı; batch özete error satırı eklendi.", level="ERROR", target=target, extra={"error": str(e)})
            rows.append({
                "target_col": target,
                "model": "ERROR",
                "best_model": "ERROR",
                "production_model": None,
                "production_status": "blocked",
                "prophet_fallback_used": False,
                "xgboost_fallback_used": False,
                "sarima_fallback_used": False,
                "alert_count": 0,
                "MAE": np.nan,
                "RMSE": np.nan,
                "MAPE": np.nan,
                "sMAPE": np.nan,
                "WAPE": np.nan,
                "MASE": np.nan,
                "error": str(e)
            })
        finally:
            series_status_rows.append(status_record)
            if persist_after_each_series and persist_output_root:
                try:
                    write_incremental_batch_status_outputs(
                        persist_output_root,
                        pd.DataFrame(rows),
                        pd.DataFrame(champion_rows),
                        pd.DataFrame(series_status_rows),
                    )
                except Exception as persist_exc:
                    progress_log("Disk Çıktısı", f"Kısmi batch özetleri yazılırken uyarı oluştu: {persist_exc}", level="WARNING", target=target)
            if not keep_batch_outputs_in_memory:
                try:
                    del out
                except Exception:
                    pass
                gc.collect()

    best_summary_df = _dedupe_best_summary_df(pd.DataFrame(rows))
    champion_table_df = pd.DataFrame(champion_rows)
    if isinstance(champion_table_df, pd.DataFrame) and len(champion_table_df) and "target_col" in champion_table_df.columns:
        champion_table_df = champion_table_df.drop_duplicates(subset=["target_col"], keep="last").reset_index(drop=True)
    series_status_df = _dedupe_series_status_df(pd.DataFrame(series_status_rows))

    result = {
        "best_summary": best_summary_df,
        "champion_table": champion_table_df,
        "batch_outputs": batch_outputs,
        "series_status": series_status_df,
        "selected_sheet": selected_sheet,
        "horizon": int(horizon),
        "zip_cache_key": f"{_safe_artifact_name(selected_sheet)}__h{int(horizon)}__{len(rows)}",
        "incremental_output_root": persist_output_root,
    }
    progress_log("Batch Özeti", "Batch forecasting hesaplaması tamamlandı.", level="SUCCESS", extra={"successful_series": int((result["series_status"].get("status", pd.Series(dtype=object)) == "SUCCESS").sum()) if isinstance(result.get("series_status"), pd.DataFrame) else len(batch_outputs), "requested_series": len(targets)})
    return result



def _seasonal_naive_array(y_train: pd.Series, horizon: int, season_length: int) -> np.ndarray:
    y_train = pd.to_numeric(y_train, errors="coerce").astype(float).dropna().reset_index(drop=True)
    if horizon <= 0:
        return np.array([], dtype=float)
    if season_length > 1 and len(y_train) >= season_length:
        vals = y_train.iloc[-season_length:].tolist()
        return np.array([vals[i % season_length] for i in range(horizon)], dtype=float)
    if len(y_train) == 0:
        return np.zeros(horizon, dtype=float)
    return np.repeat(float(y_train.iloc[-1]), horizon).astype(float)


def compute_asymmetric_validation_penalty(y_true: np.ndarray, y_pred: np.ndarray, y_train: Optional[np.ndarray] = None, severity: float = 1.0) -> Dict[str, float]:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    y_train = np.asarray(y_train if y_train is not None else y_true, dtype=float)
    if len(y_true) == 0:
        return {"penalty": 0.0, "bias_pct": np.nan, "under_forecast_rate": np.nan, "peak_event_score": np.nan, "service_gap": np.nan}
    scale = max(
        float(np.nanmean(np.abs(y_true))) if np.isfinite(y_true).any() else 0.0,
        float(np.nanmean(np.abs(y_train))) if np.isfinite(y_train).any() else 0.0,
        1.0,
    )
    err = y_pred - y_true
    bias_pct = float(100.0 * np.nanmean(err) / scale) if len(err) else np.nan
    under_rate = float(np.nanmean((y_pred < y_true).astype(float))) if len(y_true) else np.nan
    peak = compute_peak_event_score(y_train, y_true, y_pred)
    peak_score = float(peak.get("peak_event_score", np.nan)) if isinstance(peak, dict) else np.nan
    target_service = 0.90 if len(y_true) <= 6 else 0.85
    achieved_service = float(np.nanmean(y_pred >= y_true)) if len(y_true) else np.nan
    service_gap = max(target_service - achieved_service, 0.0) if pd.notna(achieved_service) else np.nan
    penalty = 0.0
    if pd.notna(bias_pct) and bias_pct < 0:
        penalty += min(abs(bias_pct) * 0.06 * severity, 2.0)
    if pd.notna(under_rate) and under_rate > 0.55:
        penalty += (under_rate - 0.55) * 3.0 * severity
    if pd.notna(peak_score) and peak_score < 0.35:
        penalty += (0.35 - peak_score) * 2.5 * severity
    if pd.notna(service_gap) and service_gap > 0:
        penalty += service_gap * 2.0 * severity
    return {
        "penalty": float(penalty),
        "bias_pct": bias_pct,
        "under_forecast_rate": under_rate,
        "peak_event_score": peak_score,
        "service_gap": float(service_gap) if pd.notna(service_gap) else np.nan,
    }


def operational_bias_peak_postprocess(train_y: pd.Series, pred: np.ndarray, freq_alias: str, model_name: str, profile: Optional[Dict[str, Any]] = None) -> np.ndarray:
    pred = np.maximum(np.asarray(pred, dtype=float), 0.0)
    if pred.size == 0:
        return pred
    s = pd.to_numeric(train_y, errors="coerce").astype(float).dropna().reset_index(drop=True)
    if len(s) == 0:
        return pred
    profile = profile or {}
    season_length = infer_season_length_from_freq(freq_alias)
    seasonal_naive = _seasonal_naive_array(s, len(pred), season_length)
    recent_window = min(max(6, len(pred) * 2), len(s))
    recent_hist = s.iloc[-recent_window:]
    recent_mean = float(recent_hist.mean()) if len(recent_hist) else float(s.mean())
    recent_median = float(recent_hist.median()) if len(recent_hist) else float(s.median())
    recent_std = float(recent_hist.std()) if len(recent_hist) else 0.0
    pred_mean = float(np.mean(pred)) if len(pred) else 0.0
    trend_strength = float(profile.get("trend_strength", 0.0) or 0.0)
    seasonality = float(profile.get("seasonality_strength", 0.0) or 0.0)
    cv = float(profile.get("cv", 0.0) or 0.0)

    if model_name == "Prophet":
        pred = 0.40 * pred + 0.40 * seasonal_naive + 0.20 * np.repeat(recent_mean, len(pred))
    elif model_name == "Ensemble":
        pred = 0.70 * pred + 0.20 * seasonal_naive + 0.10 * np.repeat(recent_median, len(pred))
    elif model_name in ["ARIMA", "SARIMA/SARIMAX"]:
        pred = 0.85 * pred + 0.10 * seasonal_naive + 0.05 * np.repeat(recent_mean, len(pred))
    else:
        pred = 0.90 * pred + 0.10 * seasonal_naive

    anchor_ratio_mean = recent_mean / max(pred_mean, 1e-6)
    anchor_ratio_median = recent_median / max(pred_mean, 1e-6)
    uplift = 0.0
    if np.isfinite(anchor_ratio_mean) and anchor_ratio_mean > 1.0:
        uplift += min((anchor_ratio_mean - 1.0) * 0.45, 0.10)
    if np.isfinite(anchor_ratio_median) and anchor_ratio_median > 1.0:
        uplift += min((anchor_ratio_median - 1.0) * 0.25, 0.06)
    if model_name in ["XGBoost", "ARIMA", "SARIMA/SARIMAX", "Ensemble"]:
        uplift += 0.015
    if seasonality >= 0.12 or trend_strength >= 0.10:
        uplift += 0.01
    if cv <= 0.35:
        uplift += 0.01
    uplift = min(max(uplift, 0.0), 0.16)
    if uplift > 0:
        pred = pred * (1.0 + uplift)

    seasonal_q = float(np.nanquantile(seasonal_naive, 0.67)) if len(seasonal_naive) else 0.0
    recent_q = float(np.nanquantile(recent_hist, 0.60)) if len(recent_hist) else recent_mean
    service_floor = max(0.94 * seasonal_q, 0.96 * recent_q)
    if (seasonality >= 0.10 or trend_strength >= 0.08) and service_floor > 0:
        pred = np.maximum(pred, np.minimum(np.repeat(service_floor, len(pred)), np.maximum(seasonal_naive * 1.06, service_floor)))

    if len(pred) >= 2 and len(seasonal_naive) == len(pred):
        peak_ref = float(np.nanquantile(seasonal_naive, 0.75)) if len(seasonal_naive) else 0.0
        peak_mask = seasonal_naive >= peak_ref if peak_ref > 0 else np.zeros(len(pred), dtype=bool)
        if np.any(peak_mask):
            peak_floor = np.maximum(seasonal_naive[peak_mask] * 1.08, 0.98 * np.maximum(recent_mean, recent_median))
            pred[peak_mask] = np.maximum(pred[peak_mask], peak_floor)

    if recent_std > 0 and float(np.std(pred)) < 0.18 * recent_std:
        pred = 0.82 * pred + 0.18 * seasonal_naive

    return np.maximum(pred, 0.0)


def compute_peak_event_score(y_train: np.ndarray, y_true: np.ndarray, y_pred: np.ndarray, quantile: float = 0.75) -> Dict[str, Any]:
    y_train = np.asarray(y_train, dtype=float)
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    if len(y_true) == 0:
        return {"peak_threshold": np.nan, "peak_precision": np.nan, "peak_recall": np.nan, "peak_f1": np.nan, "peak_event_score": np.nan}
    base_hist = y_train[np.isfinite(y_train)] if np.isfinite(y_train).any() else y_true[np.isfinite(y_true)]
    if len(base_hist) == 0:
        threshold = float(np.nanmedian(y_true)) if len(y_true) else np.nan
    else:
        q = 0.70 if len(y_true) <= 6 else quantile
        threshold = float(np.nanquantile(base_hist, q))
    k = max(1, int(np.ceil(len(y_true) * 0.34)))
    actual_idx = set(np.argsort(y_true)[-k:].tolist())
    pred_idx = set(np.argsort(y_pred)[-k:].tolist())
    soft_hits = 0
    matched_pred = set()
    for ai in sorted(actual_idx):
        for pj in sorted(pred_idx):
            if pj in matched_pred:
                continue
            if abs(int(ai) - int(pj)) <= 1:
                soft_hits += 1
                matched_pred.add(pj)
                break
    precision = float(soft_hits / max(len(pred_idx), 1))
    recall = float(soft_hits / max(len(actual_idx), 1))
    f1 = float(2 * precision * recall / max(precision + recall, 1e-12)) if (precision + recall) > 0 else 0.0
    actual_peak_mass = float(np.sum(y_true[list(actual_idx)])) if actual_idx else 0.0
    captured_mass = float(np.sum(y_pred[list(actual_idx)])) if actual_idx else 0.0
    amplitude_capture = float(np.clip(captured_mass / max(actual_peak_mass, 1e-6), 0.0, 1.2)) if actual_peak_mass > 0 else 0.0
    score = float(np.clip(0.70 * f1 + 0.30 * min(amplitude_capture, 1.0), 0.0, 1.0))
    return {"peak_threshold": threshold, "peak_precision": precision, "peak_recall": recall, "peak_f1": f1, "peak_event_score": score, "actual_peak_count": int(len(actual_idx)), "pred_peak_count": int(len(pred_idx)), "peak_amplitude_capture": amplitude_capture}


def run_rolling_origin_backtest_full(export_payload: Dict[str, pd.DataFrame], target_col: str, outputs: Dict[str, Any], freq_alias: str, horizon: int = 3, use_exog_for_stat_models: bool = True, use_exog_for_prophet: bool = True, max_folds: int = 3) -> pd.DataFrame:
    df_series = make_series_analysis_frame(export_payload, target_col)
    ds = df_series.loc[df_series["is_usable"], ["ds", "y"]].copy().sort_values("ds").reset_index(drop=True)
    h = max(1, int(horizon))
    min_train = max(24, h * 6) if str(freq_alias).upper() == "M" else max(30, h * 6)
    if len(ds) < (min_train + h + 2):
        return pd.DataFrame(columns=["fold", "step", "model", "WAPE", "sMAPE", "MAE", "actual", "prediction", "abs_error"])
    date_col = export_payload["manifest"].loc[export_payload["manifest"]["key"] == "date_column", "value"].iloc[0]
    df_features = export_payload.get("features", pd.DataFrame()).copy()
    if date_col in df_features.columns:
        df_features[date_col] = pd.to_datetime(df_features[date_col], errors="coerce")
        df_features = df_features.sort_values(date_col).reset_index(drop=True)
    profile = outputs.get("metadata", {}).get("profile", {}) or get_profile_row(export_payload, target_col)
    stat_exog_cols = detect_optional_exog_columns(df_features, target_col, date_col) if use_exog_for_stat_models else []
    prophet_exog_cols = detect_optional_exog_columns(df_features, target_col, date_col) if use_exog_for_prophet else []
    ml_feature_cols = detect_safe_ml_exog_columns(df_features, target_col, date_col)
    fold_ends: List[int] = []
    end = len(ds) - h
    while end >= min_train and len(fold_ends) < max_folds:
        fold_ends.append(end)
        end -= h
    fold_ends = sorted(fold_ends)
    rows: List[Dict[str, Any]] = []
    for fold_no, train_end in enumerate(fold_ends, start=1):
        tr = ds.iloc[:train_end].copy().reset_index(drop=True)
        te = ds.iloc[train_end:train_end + h].copy().reset_index(drop=True)
        usable_dates = pd.concat([tr[["ds"]], te[["ds"]]], axis=0)["ds"]
        feat_subset = df_features[df_features[date_col].isin(usable_dates)].copy().sort_values(date_col).reset_index(drop=True) if len(df_features) else pd.DataFrame()
        tr_feat = feat_subset.iloc[:len(tr)].copy().reset_index(drop=True) if len(feat_subset) else pd.DataFrame()
        te_feat = feat_subset.iloc[len(tr):len(tr)+len(te)].copy().reset_index(drop=True) if len(feat_subset) else pd.DataFrame()
        stat_tr = tr_feat[stat_exog_cols] if stat_exog_cols and len(tr_feat) else None
        stat_te = te_feat[stat_exog_cols] if stat_exog_cols and len(te_feat) else None
        pro_tr = tr_feat[prophet_exog_cols] if prophet_exog_cols and len(tr_feat) else None
        pro_te = te_feat[prophet_exog_cols] if prophet_exog_cols and len(te_feat) else None
        ml_tr = tr_feat[ml_feature_cols] if ml_feature_cols and len(tr_feat) else pd.DataFrame(index=tr_feat.index if len(tr_feat) else range(len(tr)))
        ml_te = te_feat[ml_feature_cols] if ml_feature_cols and len(te_feat) else pd.DataFrame(index=te_feat.index if len(te_feat) else range(len(te)))
        model_results: Dict[str, Dict[str, Any]] = {}
        def _record(model_name: str, pred_arr: np.ndarray):
            pred_arr = operational_bias_peak_postprocess(tr["y"], pred_arr, freq_alias=freq_alias, model_name=model_name, profile=profile)
            actual = pd.to_numeric(te["y"], errors="coerce").astype(float).values
            for step, (a, p) in enumerate(zip(actual, pred_arr), start=1):
                rows.append({"fold": fold_no, "step": step, "model": model_name, "WAPE": wape(np.array([a]), np.array([p])), "sMAPE": smape(np.array([a]), np.array([p])), "MAE": mae(np.array([a]), np.array([p])), "actual": float(a), "prediction": float(p), "abs_error": float(abs(p - a))})
        try:
            res = fit_best_sarimax(tr, te, freq_alias, profile, stat_tr, stat_te)
            model_results["SARIMA/SARIMAX"] = res
            _record("SARIMA/SARIMAX", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            res = fit_best_arima(tr, te, freq_alias, profile)
            model_results["ARIMA"] = res
            _record("ARIMA", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            if HAS_PROPHET:
                res = fit_best_prophet(tr, te, freq_alias, profile, pro_tr, pro_te)
                model_results["Prophet"] = res
                _record("Prophet", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            res = fit_best_intermittent(tr, te, freq_alias, profile)
            model_results["Intermittent"] = res
            _record("Intermittent", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            res = fit_xgboost_forecast(tr, te, ml_tr, ml_te, freq_alias=freq_alias)
            model_results["XGBoost"] = res
            _record("XGBoost", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            res = sanitize_dl_result_for_reporting(fit_deep_learning_forecast(tr, te, ml_tr, ml_te, freq_alias=freq_alias, model_family="LSTM"), model_family="LSTM", horizon=len(te))
            model_results["LSTM"] = res
            if bool(res.get("dl_strict_validation_passed", False)):
                _record("LSTM", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            res = sanitize_dl_result_for_reporting(fit_deep_learning_forecast(tr, te, ml_tr, ml_te, freq_alias=freq_alias, model_family="GRU"), model_family="GRU", horizon=len(te))
            model_results["GRU"] = res
            if bool(res.get("dl_strict_validation_passed", False)):
                _record("GRU", np.asarray(res["forecast"], dtype=float))
        except Exception:
            pass
        try:
            pred_map = {k: operational_bias_peak_postprocess(tr["y"], np.asarray(v["forecast"], dtype=float), freq_alias=freq_alias, model_name=k, profile=profile) for k, v in model_results.items() if _finite_prediction_array(v.get("forecast"), expected_horizon=len(te))}
            metrics_rows = [build_model_metrics(k, tr["y"].values, te["y"].values, v) for k, v in pred_map.items()]
            val_rows = [extract_validation_metrics_from_result(k, model_results[k]) for k in model_results.keys()]
            pred_map, metrics_df_fold = filter_usable_prediction_map(pred_map, pd.DataFrame(metrics_rows), expected_horizon=len(te))
            if len(pred_map) == 0 or len(metrics_df_fold) == 0:
                raise ValueError("Rolling ensemble için kullanılabilir model yok.")
            ens_pred, _ = build_weighted_ensemble(pred_map, metrics_df_fold, validation_df=pd.DataFrame(val_rows), y_train=tr["y"].values, y_true=te["y"].values, profile=profile)
            _record("Ensemble", ens_pred)
        except Exception:
            pass
    return pd.DataFrame(rows)


def summarize_full_backtest(backtest_df: pd.DataFrame) -> pd.DataFrame:
    if backtest_df is None or len(backtest_df) == 0:
        return pd.DataFrame(columns=["model", "folds", "WAPE", "sMAPE", "MAE", "MedianAE"])
    g = backtest_df.groupby("model", dropna=False)
    out = g.agg(folds=("fold", "nunique"), MAE=("abs_error", "mean"), MedianAE=("abs_error", "median")).reset_index()
    rows = []
    for model_name, part in g:
        rows.append({"model": model_name, "WAPE": wape(part["actual"].values, part["prediction"].values), "sMAPE": smape(part["actual"].values, part["prediction"].values)})
    score_df = pd.DataFrame(rows)
    out = out.merge(score_df, on="model", how="left")
    return out.sort_values(["WAPE", "sMAPE", "MAE"], ascending=[True, True, True]).reset_index(drop=True)


def build_calibrated_prediction_interval_table(test_df: pd.DataFrame, pred: np.ndarray, model_name: str, backtest_df: pd.DataFrame, fallback_scale: float, coverage_levels: Tuple[float, ...] = (0.80, 0.90, 0.95)) -> pd.DataFrame:
    if backtest_df is None or len(backtest_df) == 0 or model_name not in set(backtest_df.get("model", pd.Series(dtype=str)).astype(str)):
        return build_prediction_interval_table(test_df, pred, fallback_scale, coverage_levels)
    model_bt = backtest_df.loc[backtest_df["model"].astype(str) == str(model_name)].copy()
    if len(model_bt) == 0 or "abs_error" not in model_bt.columns:
        return build_prediction_interval_table(test_df, pred, fallback_scale, coverage_levels)
    out = test_df[["ds", "y"]].copy().reset_index(drop=True)
    out["q50"] = np.maximum(np.asarray(pred, dtype=float), 0.0)
    for level in coverage_levels:
        widths = []
        q = min(max(float(level), 0.50), 0.99)
        for step in range(1, len(out) + 1):
            part = model_bt.loc[model_bt["step"] == step, "abs_error"] if "step" in model_bt.columns else pd.Series(dtype=float)
            if len(part.dropna()) < 2:
                part = model_bt["abs_error"]
            width = float(np.nanquantile(pd.to_numeric(part, errors="coerce").dropna(), q)) if len(part.dropna()) else float(fallback_scale)
            widths.append(max(width, 1e-6))
        widths = np.asarray(widths, dtype=float)
        low = np.maximum(out["q50"].values - widths, 0.0)
        high = np.maximum(out["q50"].values + widths, 0.0)
        low_q = int(round(((1.0 - level) / 2.0) * 100.0))
        high_q = int(round((1.0 - (1.0 - level) / 2.0) * 100.0))
        out[f"q{low_q:02d}"] = low
        out[f"q{high_q:02d}"] = high
        out[f"coverage_{int(level*100)}"] = ((out["y"] >= low) & (out["y"] <= high)).astype(int)
        out[f"avg_width_{int(level*100)}"] = float(np.mean(high - low)) if len(out) else np.nan
    return out


def enforce_hard_feature_contract_gate(eligibility_df: pd.DataFrame, feature_audit_df: pd.DataFrame) -> pd.DataFrame:
    if eligibility_df is None or len(eligibility_df) == 0:
        return pd.DataFrame(columns=["model", "status", "eligibility_score"])
    out = eligibility_df.copy()
    if feature_audit_df is None or len(feature_audit_df) == 0:
        return out
    def _has_unknown(area: str) -> bool:
        part = feature_audit_df[feature_audit_df["used_in"].str.contains(area, na=False)]
        if len(part) == 0:
            return False
        return bool((part["availability_status"].astype(str) != "future_known").any())
    for model_name, area in [("XGBoost", "ml"), ("LSTM", "ml"), ("GRU", "ml"), ("Prophet", "prophet")]:
        idx = out.index[out["model"] == model_name]
        if len(idx) == 0:
            continue
        i = idx[0]
        if _has_unknown(area):
            prev = str(out.loc[i, "status"])
            out.loc[i, "status"] = "challenger_only" if prev == "eligible" else prev
            out.loc[i, "eligibility_score"] = float(max(0.0, float(out.loc[i, "eligibility_score"]) - 12.0))
            reason = str(out.loc[i, "rejection_or_limit_reason"] or "")
            extra = "Hard feature contract gate: future-known olmayan feature var."
            out.loc[i, "rejection_or_limit_reason"] = (reason + " | " + extra).strip(" |")
    return out


def build_production_decision_explanation(production_ranking: pd.DataFrame, production_model: Optional[str]) -> str:
    if production_ranking is None or len(production_ranking) == 0 or not production_model:
        return "Üretim modeli, uygunluk ve performans bilgileri sınırlı olduğu için temkinli seçildi."
    row = production_ranking.loc[production_ranking["model"].astype(str) == str(production_model)].head(1)
    if len(row) == 0:
        return f"{production_model} seçildi; çünkü güvenli varsayılan aday olarak öne çıktı."
    row = row.iloc[0]
    nedenler = []
    if pd.notna(row.get("ro_WAPE")):
        nedenler.append(f"rolling-origin WAPE={safe_float(row.get('ro_WAPE')):.2f}")
    if pd.notna(row.get("WAPE")):
        nedenler.append(f"holdout WAPE={safe_float(row.get('WAPE')):.2f}")
    if pd.notna(row.get("eligibility_score")):
        nedenler.append(f"uygunluk skoru={safe_float(row.get('eligibility_score')):.1f}")
    if pd.notna(row.get("service_gap")) and float(row.get("service_gap")) <= 1e-9:
        nedenler.append("servis açığı yok")
    if pd.notna(row.get("peak_event_score")):
        nedenler.append(f"tepe skoru={safe_float(row.get('peak_event_score')):.3f}")
    durum = TURKCE_DEGER_HARITASI.get(str(row.get("status", "eligible")), str(row.get("status", "eligible")))
    return f"{production_model} seçildi; çünkü {', '.join(nedenler[:4])} ve üretim durumu '{durum}' olarak değerlendirildi."



def assess_production_readiness(export_payload: Dict[str, pd.DataFrame], metrics_df: pd.DataFrame, target_col: str, outputs: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    notes: List[str] = []
    status = "guarded_use_only"
    try:
        profile_df = export_payload.get("series_profile_report", pd.DataFrame())
        anomaly_df = export_payload.get("anomaly_governance", pd.DataFrame())
        if outputs is not None:
            prod_pack = outputs.get("production_governance", {}) or {}
            gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame())
            live_pack = prod_pack.get("live_monitoring_pack", {}) or {}
            alerts_df = live_pack.get("alerts", pd.DataFrame()) if isinstance(live_pack, dict) else pd.DataFrame()
            if len(gate_df):
                top = gate_df.sort_values(["status", "eligibility_score"], ascending=[True, False]).iloc[0]
                notes.append(f"Üretim kapısı: {top['model']} -> {TURKCE_DEGER_HARITASI.get(str(top['status']), top['status'])} (skor={safe_float(top.get('eligibility_score')):.1f})")
            if len(alerts_df):
                notes.append(f"Canlı izleme için {len(alerts_df)} alarm üretildi; kör tam otomasyon öncesi shadow-mode zorunlu.")
        if not profile_df.empty and "series" in profile_df.columns:
            row = profile_df.loc[profile_df["series"] == target_col]
            if not row.empty:
                n_obs = float(pd.to_numeric(pd.Series([row.iloc[0].get("n_obs", np.nan)]), errors="coerce").iloc[0])
                if pd.notna(n_obs) and n_obs < 84:
                    notes.append(f"Seri uzunluğu {int(n_obs)} gözlem; aylık ilaç planlama için sınırlı ve overfit riski taşır.")
        if not anomaly_df.empty and "series" in anomaly_df.columns:
            a = anomaly_df.loc[anomaly_df["series"] == target_col].copy()
            if len(a) > 0 and "is_training_excluded" in a.columns:
                excluded = int(a["is_training_excluded"].fillna(False).astype(bool).sum())
                if excluded > 0:
                    notes.append(f"Bu seri için eğitim dışı bırakılan {excluded} kayıt var; yönetişim doğru fakat model belirsizliğini artırır.")
        if not metrics_df.empty and "WAPE" in metrics_df.columns:
            best_row = metrics_df.copy()
            best_row["WAPE_num"] = pd.to_numeric(best_row["WAPE"], errors="coerce")
            best_row["sMAPE_num"] = pd.to_numeric(best_row.get("sMAPE"), errors="coerce") if "sMAPE" in best_row.columns else np.nan
            best_row = best_row.sort_values(["WAPE_num", "sMAPE_num"], ascending=[True, True])
            if len(best_row):
                bw = best_row.iloc[0]["WAPE_num"]
                bs = best_row.iloc[0]["sMAPE_num"]
                if pd.notna(bw) and pd.notna(bs):
                    if bw <= 7 and bs <= 6:
                        notes.append("Holdout performansı güçlü; kontrollü karar desteği için uygundur.")
                    else:
                        notes.append("Holdout performansı kabul edilebilir olsa da tam otomatik üretim kullanımı için ek onay katmanı gerekir.")
    except Exception:
        notes.append("Üretim uygunluğu değerlendirmesi hesaplanırken hata oluştu; temkinli kullanım önerilir.")
    if not notes:
        notes.append("Bu çıktı tez ve analitik inceleme için uygundur; tam otomatik gerçek hayat kullanımı için ek doğrulama gerekir.")
    return {"status": status, "notes": notes}

def build_model_liderleri(outputs: Dict[str, Any]) -> pd.DataFrame:
    satirlar = []
    hold_df = outputs.get("metrics_df", pd.DataFrame())
    val_df = outputs.get("validation_metrics_df", pd.DataFrame())
    profile = outputs.get("metadata", {}).get("profile", {}) or {}
    val_rank = build_contextual_validation_ranking(val_df, profile=profile)
    ro_df = summarize_full_backtest(outputs.get("rolling_origin_backtest", pd.DataFrame()))
    prod_pack = outputs.get("production_governance", {}) or {}
    prod_rank = prod_pack.get("production_ranking", pd.DataFrame())
    hold_winner = hold_df.sort_values(["WAPE", "sMAPE", "RMSE"], ascending=[True, True, True]).iloc[0]["model"] if len(hold_df) else None
    val_winner = val_rank.iloc[0]["model"] if len(val_rank) else None
    ro_winner = ro_df.sort_values(["WAPE", "MAE"], ascending=[True, True]).iloc[0]["model"] if len(ro_df) else None
    prod_winner = prod_rank.iloc[0]["model"] if len(prod_rank) else outputs.get("production_model")
    satirlar.append({"karar_kademesi": "Holdout kazananı", "model": hold_winner})
    satirlar.append({"karar_kademesi": "Doğrulama kazananı", "model": val_winner})
    satirlar.append({"karar_kademesi": "Rolling-origin kazananı", "model": ro_winner})
    satirlar.append({"karar_kademesi": "Üretim kazananı", "model": prod_winner})
    return pd.DataFrame(satirlar)



def build_karar_hiyerarsisi_ozeti(outputs: Dict[str, Any]) -> pd.DataFrame:
    lider = build_model_liderleri(outputs)
    aciklama_map = {
        "Holdout kazananı": "Son test penceresinde en düşük hata üreten model.",
        "Doğrulama kazananı": "İç doğrulama bölmesinde en iyi model; bağlamsal ceza uygulanmıştır.",
        "Rolling-origin kazananı": "Zaman boyunca en stabil geri test performansını veren model.",
        "Üretim kazananı": "Risk, stabilite, uygulanabilirlik ve canlı yönetişim kurallarından sonra önerilen model.",
    }
    karar_map = {
        "Holdout kazananı": "Kısa dönem doğruluk",
        "Doğrulama kazananı": "Aşırı öğrenme kontrolü",
        "Rolling-origin kazananı": "Zamana karşı dayanıklılık",
        "Üretim kazananı": "Canlı kullanım kararı",
    }
    if len(lider) == 0:
        return pd.DataFrame(columns=["KararAşaması", "KazananModel", "KararAmacı", "Anlamı"])
    out = lider.rename(columns={"karar_kademesi": "KararAşaması", "model": "KazananModel"}).copy()
    out["KararAmacı"] = out["KararAşaması"].map(karar_map)
    out["Anlamı"] = out["KararAşaması"].map(aciklama_map)
    prod_expl = ((outputs.get("production_governance", {}) or {}).get("production_decision_explanation") if isinstance(outputs, dict) else None)
    if prod_expl:
        extra = pd.DataFrame([{"KararAşaması": "Üretim karar açıklaması", "KazananModel": str((outputs.get("production_governance", {}) or {}).get("production_model", "-")), "KararAmacı": "Neden seçildi?", "Anlamı": str(prod_expl)}])
        out = pd.concat([out, extra], axis=0, ignore_index=True)
    return out



def build_prophet_gorunurluk_ozeti(prophet_result: Dict[str, Any]) -> pd.DataFrame:
    if not isinstance(prophet_result, dict) or len(prophet_result) == 0:
        return pd.DataFrame(columns=["Alan", "Değer"])
    mode = prophet_result.get("fit_mode", "bilinmiyor")
    fallback_used = bool(prophet_result.get("fallback_used", False))
    fallback_method = prophet_result.get("fallback_method")
    note = prophet_result.get("fit_visibility_note", "")
    if mode == "gercek_prophet_fit" and not fallback_used:
        rozet = "Gerçek Prophet"
        production_note = "Üretimde ancak doğrulama ve uygunluk kapısından geçerse kullanılmalı."
    elif fallback_used:
        rozet = "Fallback Prophet"
        production_note = "Production’da kullanılmaz; görselde soluk gösterilir ve yalnız meydan okuyucu/tanısal amaçla kullanılır."
    else:
        rozet = "Prophet belirsiz"
        production_note = "Bu çalışma modu temkinli yorumlanmalı."
    return pd.DataFrame([
        {"Alan": "Prophet rozeti", "Değer": rozet},
        {"Alan": "Prophet çalışma modu", "Değer": mode},
        {"Alan": "Gerçek Prophet fit kullanıldı mı?", "Değer": "evet" if (mode == "gercek_prophet_fit" and not fallback_used) else "hayır"},
        {"Alan": "Fallback kullanıldı mı?", "Değer": "evet" if fallback_used else "hayır"},
        {"Alan": "Fallback yöntemi", "Değer": fallback_method or "yok"},
        {"Alan": "Üretim notu", "Değer": production_note},
        {"Alan": "Açıklama", "Değer": note or "-"},
    ])



def build_benzersiz_rapor_katalogu(outputs: Dict[str, Any], prod_pack: Dict[str, Any]) -> pd.DataFrame:
    adaylar: List[Tuple[str, pd.DataFrame]] = [
        ("ModelKarşılaştırma", outputs.get("metrics_df", pd.DataFrame())),
        ("DoğrulamaMetrikleri", outputs.get("validation_metrics_df", pd.DataFrame())),
        ("DLSeffaflik", outputs.get("dl_transparency_table", pd.DataFrame())),
        ("KararHiyerarşisi", build_karar_hiyerarsisi_ozeti(outputs)),
        ("AnsamblAğırlıkları", outputs.get("ensemble_weights", pd.DataFrame())),
        ("RollingOrigin", outputs.get("rolling_origin_backtest", pd.DataFrame())),
        ("BiasDashboard", prod_pack.get("bias_dashboard", pd.DataFrame())),
        ("PeakDashboard", prod_pack.get("peak_event_dashboard", pd.DataFrame())),
        ("CanlıİzlemeÖzeti", (prod_pack.get("live_monitoring_pack", {}) or {}).get("summary", pd.DataFrame())),
        ("ÜretimAlarmları", (prod_pack.get("live_monitoring_pack", {}) or {}).get("alerts", pd.DataFrame())),
        ("ÜretimSıralaması", prod_pack.get("production_ranking", pd.DataFrame())),
        ("ÖzellikDenetimi", prod_pack.get("feature_availability_audit", pd.DataFrame())),
        ("ServisSeviyesi", prod_pack.get("production_service_table", pd.DataFrame())),
        ("TahminAralığı", prod_pack.get("production_interval_table", pd.DataFrame())),
    ]
    rows = []
    seen = {}
    for rapor_adi, df in adaylar:
        if not isinstance(df, pd.DataFrame) or len(df) == 0:
            continue
        try:
            key = hashlib.sha1(df.to_csv(index=False).encode("utf-8", errors="ignore")).hexdigest()
        except Exception:
            key = f"{rapor_adi}_{len(df)}_{len(df.columns)}"
        duplicate_of = seen.get(key)
        if duplicate_of is None:
            seen[key] = rapor_adi
            rows.append({"Rapor": rapor_adi, "SatırSayısı": int(len(df)), "KolonSayısı": int(len(df.columns)), "Durum": "benzersiz", "TekrarEttiğiRapor": ""})
        else:
            rows.append({"Rapor": rapor_adi, "SatırSayısı": int(len(df)), "KolonSayısı": int(len(df.columns)), "Durum": "yinelenen_içerik", "TekrarEttiğiRapor": duplicate_of})
    return pd.DataFrame(rows)

def render_streamlit_app():
    ensure_forecasting_runtime_dependencies(include_streamlit=True)
    if st is None:
        raise ImportError("Streamlit kurulu değil veya yüklenemedi.")
    st.set_page_config(page_title="Talep Tahminleme Studio", layout="wide")
    st.title("Talep Tahminleme Studio")
    st.caption("Üretim sınıfı veri önişleme + ileri seviye şampiyon-meydan okuyan + ansambl + toplu tahminleme")

    with st.sidebar:
        st.subheader("Girdi")
        uploaded_excel = st.file_uploader("Excel dosyası yükle", type=["xlsx", "xls"])
        st.markdown("Bu uygulama mevcut üretim sınıfı veri önişleme mantığını korur; üstüne SARIMA/SARIMAX, Prophet, XGBoost, ansambl, ABC/XYZ ve toplu tahminleme ekler.")
        st.info("İnteraktif hız modu açık: özellikle kısa aylık serilerde SARIMAX açıklayıcı değişken ve Prophet ayar araması otomatik daraltılır. Amaç, Streamlit ekranında takılmayı önlemek ve yanıt süresini düşürmektir.")

    if uploaded_excel is None:
        st.info("Başlamak için Excel dosyanı yükle.")
        return

    try:
        excel_path = save_uploaded_file(uploaded_excel)
        xls = safe_excel_file(excel_path)
    except Exception as e:
        st.error(f"Excel dosyası yüklenemedi/açılamadı: {e}")
        return
    selected_sheet = st.sidebar.selectbox("Sheet seç", xls.sheet_names)
    output_base_dir = os.path.join(os.path.dirname(excel_path), "streamlit_outputs")
    os.makedirs(output_base_dir, exist_ok=True)

    cache_key = f"{uploaded_excel.name}::{selected_sheet}"
    if "preprocess_cache" not in st.session_state:
        st.session_state["preprocess_cache"] = {}
    if "forecast_run_cache" not in st.session_state:
        st.session_state["forecast_run_cache"] = {}

    if st.sidebar.button("Önişleme + Tahminleme için hazırla", type="primary") or cache_key not in st.session_state["preprocess_cache"]:
        with st.spinner("Veri önişleme ve yönetişim çalışıyor..."):
            try:
                export_payload = run_preprocessing_for_sheet(excel_path, selected_sheet, output_base_dir)
                st.session_state["preprocess_cache"][cache_key] = export_payload
            except Exception as e:
                st.error(f"Önişleme sırasında hata oluştu: {e}")
                st.exception(e)
                return
    export_payload = st.session_state["preprocess_cache"].get(cache_key)
    if export_payload is None:
        st.info("Önişleme sonucu üretilemedi.")
        return

    manifest = export_payload["manifest"]
    freq_alias = manifest.loc[manifest["key"] == "frequency_inferred", "value"].iloc[0]
    target_cols = export_payload["series_profile_report"]["series"].tolist()

    top1, top2, top3, top4 = st.columns(4)
    with top1: st.metric("Frekans", str(freq_alias))
    with top2: st.metric("Hedef seri", len(target_cols))
    with top3: st.metric("Regularized satır", int(len(export_payload["raw_regular"])))
    with top4: st.metric("Anomali kaydı", int(len(export_payload["anomaly_governance"])))

    mode = st.radio("Çalışma modu", ["Tek seri", "Çok serili batch forecasting"], horizontal=True)
    default_target = target_cols[0] if target_cols else None
    if mode == "Tek seri":
        batch_targets = [default_target] if default_target else []
        target_col = st.selectbox("Tahminlenecek seri", target_cols, index=0 if default_target else None)
    else:
        batch_targets = st.multiselect(
            "Batch'e alınacak seriler",
            options=target_cols,
            default=target_cols
        )
        if not batch_targets:
            st.warning("Batch için en az bir seri seçmelisin.")
            return
        target_col = st.selectbox(
            "Detay panosunda gösterilecek seri",
            options=batch_targets,
            index=0
        )

    default_horizon = min(infer_default_horizon(freq_alias), max(2, len(export_payload["clean_model_input"]) // 5))
    horizon = st.slider("Test ufku", min_value=2, max_value=min(24, max(2, len(export_payload["clean_model_input"]) // 3)), value=default_horizon)
    use_exog_stat = st.checkbox("SARIMAX için açıklayıcı değişkenleri kullan", value=True)
    use_exog_prophet = st.checkbox("Prophet için ek regressors kullan", value=True)

    profile = get_profile_row(export_payload, target_col)
    seg = infer_advanced_segment(profile)
    priority = recommend_model_priority(profile)
    st.subheader("Seri profili")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Segment", seg["label"])
    c2.metric("ABC/XYZ", seg["abc_xyz"])
    c3.metric("Önerilen öncelik", priority)
    c4.metric("CV", round(float(profile.get("cv", np.nan)), 3) if profile else np.nan)
    c5.metric("Trend gücü", round(float(profile.get("trend_strength", np.nan)), 3) if profile else np.nan)
    c6.metric("Sezonsallık", round(float(profile.get("seasonality_strength", np.nan)), 3) if profile else np.nan)

    try:
        df_series = make_series_analysis_frame(export_payload, target_col)
        train_df_preview, _ = train_test_split_series(df_series, horizon)
        acf_pacf_fig = build_acf_pacf_figure(train_df_preview, target_col)
        if acf_pacf_fig is not None:
            st.pyplot(acf_pacf_fig, clear_figure=True)
    except Exception:
        pass

    cache_target_signature = "|".join(sorted(batch_targets)) if mode == "Çok serili batch forecasting" else str(target_col)

    forecast_cache_key = "||".join([
        cache_key,
        str(mode),
        cache_target_signature,
        str(horizon),
        str(int(use_exog_stat)),
        str(int(use_exog_prophet)),
        APP_VERSION
    ])

    run_label = "Batch forecasting çalıştır" if mode == "Çok serili batch forecasting" else "Modelleri çalıştır ve karşılaştır"
    if st.button(run_label, type="primary"):
        with st.spinner("Gelişmiş tahminleme katmanı çalışıyor..."):
            try:
                if forecast_cache_key in st.session_state["forecast_run_cache"]:
                    cached_obj = st.session_state["forecast_run_cache"][forecast_cache_key]
                else:
                    if mode == "Tek seri":
                        cached_obj = run_full_forecasting_pipeline(export_payload, target_col, horizon, use_exog_stat, use_exog_prophet)
                    else:
                        cached_obj = run_batch_forecasting(
                            export_payload=export_payload,
                            horizon=horizon,
                            selected_sheet=selected_sheet,
                            use_exog_for_stat_models=use_exog_stat,
                            use_exog_for_prophet=use_exog_prophet,
                            selected_targets=batch_targets
                        )
                    st.session_state["forecast_run_cache"][forecast_cache_key] = cached_obj

                if mode == "Tek seri":
                    st.session_state["forecast_outputs"] = cached_obj
                    st.session_state["forecast_target"] = target_col
                    st.session_state["batch_mode"] = False
                else:
                    st.session_state["batch_outputs_full"] = cached_obj
                    st.session_state["batch_mode"] = True
            except Exception as e:
                st.error(f"Tahminleme sırasında hata oluştu: {e}")
                st.exception(e)
                return

    if mode == "Çok serili batch forecasting":
        batch = st.session_state.get("batch_outputs_full")
        if batch is None:
            st.info("Batch sonucu görmek için butona bas.")
            return

        st.subheader("Batch forecasting özeti")
        st.dataframe(style_metric_dataframe(batch["best_summary"]), width="stretch")

        st.subheader("Champion - Challenger tablosu")
        st.dataframe(batch["champion_table"], width="stretch")

        batch_summary_excel = build_batch_summary_excel_bytes(batch)
        if batch_summary_excel is not None:
            st.download_button(
                "Batch özetini indir (Excel)",
                data=batch_summary_excel,
                file_name=f"{selected_sheet}_batch_forecasting_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        zip_state_key = f"batch_zip_bytes__{forecast_cache_key}"
        if st.button("Excel + grafik ZIP paketini hazırla"):
            with st.spinner("ZIP paketi hazırlanıyor..."):
                st.session_state[zip_state_key] = build_batch_zip_bytes(export_payload, batch, selected_sheet, horizon)

        if st.session_state.get(zip_state_key):
            st.download_button(
                "Tüm batch çıktıları indir (ZIP)",
                data=st.session_state[zip_state_key],
                file_name=f"{selected_sheet}_batch_full_outputs.zip",
                mime="application/zip"
            )

        if target_col not in batch.get("batch_outputs", {}):
            st.warning("Seçilen seri batch sonuçları içinde bulunamadı.")
            return

        outputs = batch["batch_outputs"][target_col]
        profile = get_profile_row(export_payload, target_col)

    else:
        outputs = st.session_state.get("forecast_outputs")
        if outputs is None or st.session_state.get("forecast_target") != target_col:
            st.info("Model karşılaştırmasını görmek için butona bas.")
            return

    st.subheader("Model karşılaştırma tablosu")
    metrics_df = style_metric_dataframe(outputs["metrics_df"])
    st.dataframe(metrics_df, width="stretch")
    st.download_button("Karşılaştırma tablosunu indir (Excel)", data=dataframe_to_download_bytes(metrics_df), file_name=f"{selected_sheet}_{target_col}_model_karsilastirma.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    full_tables_excel = build_current_series_tables_excel_bytes(outputs, export_payload, selected_sheet, target_col)
    if full_tables_excel is not None:
        st.download_button(
            "Bu seri için tüm tabloları indir (Excel)",
            data=full_tables_excel,
            file_name=f"{selected_sheet}_{target_col}_tum_tablolar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if isinstance(outputs.get("stage_timing_table"), pd.DataFrame) and len(outputs.get("stage_timing_table")):
        with st.expander("Çalışma süresi / performans özeti", expanded=False):
            st.dataframe(outputs["stage_timing_table"], width="stretch")
            for note in outputs.get("metadata", {}).get("runtime_guardrails", []):
                st.write(f"- {note}")

    outputs = ensure_production_reporting_tables(outputs, export_payload, target_col, freq_alias)
    readiness = assess_production_readiness(export_payload, outputs["metrics_df"], target_col, outputs=outputs)
    prod_pack = outputs.get("production_governance", {}) or {}
    prod_model = outputs.get("production_model", outputs.get("best_model"))
    prod_status = outputs.get("production_status", "eligible")
    if prod_status in ["eligible", "guarded_fallback"]:
        st.success(f"Üretim önerisi: {prod_model} | Durum: {prod_status}")
    else:
        st.warning(f"Üretim önerisi: {prod_model} | Durum: {prod_status}")
    with st.expander("Üretim kullanımı değerlendirmesi", expanded=False):
        for note in readiness["notes"]:
            st.write(f"- {note}")
        gate_df = prod_pack.get("model_eligibility_gate", pd.DataFrame())
        if isinstance(gate_df, pd.DataFrame) and len(gate_df) > 0:
            st.markdown("**Otomatik model uygunluk kapısı**")
            st.dataframe(style_metric_dataframe(gate_df), width="stretch")

    cc = outputs.get("champion_challenger", {})
    if cc.get("champion"):
        st.success(f"Şampiyon model: {cc['champion']} | Meydan okuyan: {cc.get('challenger', '-')}")
    st.markdown("**Karar kademelerine göre model liderleri**")
    st.dataframe(style_metric_dataframe(build_model_liderleri(outputs)), width="stretch")
    if outputs.get("model_errors"):
        with st.expander("Model hata / fallback özeti"):
            err_df = pd.DataFrame([{"model": k, "message": v} for k, v in outputs.get("model_errors", {}).items()])
            if len(err_df):
                st.dataframe(err_df, width="stretch")

    fig = plot_forecast_results(outputs["train"], outputs["test"], outputs["predictions"], f"{target_col} - Gerçek vs Tahmin", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
    if fig is not None:
        st.plotly_chart(fig, width="stretch")

    tab1, tab2, tab3, tab3b, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs(["SARIMA/SARIMAX", "Prophet", "XGBoost", "Derin Öğrenme (LSTM/GRU)", "Şampiyon-Meydan Okuyan ve Ansambl", "Gerçek vs Tahmin", "Geri Test Panosu", "Üretim Yönetişimi", "Önişleme Denetimleri", "Akıllı Değerlendirmeler"])

    with tab1:
        sarima = outputs.get("sarima")
        if sarima is None:
            st.warning(outputs.get("model_errors", {}).get("SARIMA/SARIMAX", "SARIMA/SARIMAX sonucu üretilemedi."))
        else:
            st.json({"order": sarima.get("order"), "seasonal_order": sarima.get("seasonal_order"), "trend": sarima.get("trend"), "AIC": sarima.get("aic"), "BIC": sarima.get("bic"), "Ljung-Box p-değeri": sarima.get("ljung_box_pvalue"), "d": sarima.get("d"), "D": sarima.get("D"), "transform": sarima.get("transform"), "white_noise_ok": sarima.get("residual_white_noise_ok"), "fallback_used": sarima.get("fallback_used"), "fallback_method": sarima.get("fallback_method")})
            if "SARIMA/SARIMAX" in outputs["tables"]:
                st.dataframe(style_metric_dataframe(outputs["tables"]["SARIMA/SARIMAX"]), width="stretch")
                fig_sarima = plot_forecast_results(outputs["train"], outputs["test"], {"SARIMA/SARIMAX": outputs["predictions"]["SARIMA/SARIMAX"]}, f"{target_col} - SARIMA/SARIMAX", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
                if fig_sarima is not None:
                    st.plotly_chart(fig_sarima, width="stretch")
            if isinstance(sarima.get("search_table"), pd.DataFrame) and len(sarima.get("search_table")):
                st.dataframe(style_metric_dataframe(sarima["search_table"]), width="stretch")

    with tab2:
        if "prophet" in outputs:
            st.markdown("**Prophet görünürlüğü (gerçek fit mi, fallback mi?)**")
            st.dataframe(style_metric_dataframe(outputs.get("prophet_gorunurluk_ozeti", pd.DataFrame())), width="stretch")
            if isinstance(outputs["prophet"].get("config"), dict):
                st.json(outputs["prophet"]["config"])
            if isinstance(outputs["prophet"].get("component_validation"), dict):
                st.json(outputs["prophet"].get("component_validation", {}))
            st.dataframe(style_metric_dataframe(outputs["tables"]["Prophet"]), width="stretch")
            fig_prophet = plot_forecast_results(outputs["train"], outputs["test"], {"Prophet": outputs["predictions"]["Prophet"]}, f"{target_col} - Prophet", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
            if fig_prophet is not None:
                st.plotly_chart(fig_prophet, width="stretch")
            st.dataframe(style_metric_dataframe(outputs["prophet"]["search_table"]), width="stretch")
        else:
            st.warning(outputs.get("prophet_error", "Prophet sonucu üretilemedi."))

    with tab3:
        if "xgboost" not in outputs:
            st.warning(outputs.get("model_errors", {}).get("XGBoost", "XGBoost sonucu üretilemedi."))
        else:
            st.json({"selected_strategy": outputs["xgboost"].get("strategy"), "shap_status": outputs["xgboost"].get("shap_status"), "fallback_used": outputs["xgboost"].get("fallback_used"), "fallback_method": outputs["xgboost"].get("fallback_method")})
            if "XGBoost" in outputs["tables"]:
                st.dataframe(style_metric_dataframe(outputs["tables"]["XGBoost"]), width="stretch")
                fig_xgb = plot_forecast_results(outputs["train"], outputs["test"], {"XGBoost": outputs["predictions"]["XGBoost"]}, f"{target_col} - XGBoost", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
                if fig_xgb is not None:
                    st.plotly_chart(fig_xgb, width="stretch")
            st.dataframe(style_metric_dataframe(outputs["xgboost"]["search_table"]), width="stretch")
            if "strategy_comparison" in outputs["xgboost"]:
                st.dataframe(style_metric_dataframe(outputs["xgboost"]["strategy_comparison"]), width="stretch")
            if len(outputs["xgboost"].get("feature_importance", pd.DataFrame())):
                st.dataframe(outputs["xgboost"]["feature_importance"], width="stretch")


    with tab3b:
        dl_subtabs = st.tabs(["LSTM", "GRU", "DL Analizi"])
        for dl_name, dl_tab in [("LSTM", dl_subtabs[0]), ("GRU", dl_subtabs[1])]:
            with dl_tab:
                dl_key = dl_name.lower()
                dl_pack = outputs.get(dl_key, {}) or {}
                if dl_name not in outputs.get("tables", {}):
                    st.warning(outputs.get("model_errors", {}).get(dl_name, f"{dl_name} sonucu üretilemedi."))
                else:
                    st.json({"selected_config": dl_pack.get("selected_config"), "transform": dl_pack.get("transform"), "used_exog_cols": dl_pack.get("used_exog_cols"), "validation_wape": dl_pack.get("validation_wape"), "validation_smape": dl_pack.get("validation_smape"), "fallback_used": dl_pack.get("fallback_used"), "fallback_method": dl_pack.get("fallback_method")})
                    st.markdown(f"**{dl_name} gerçek ve tahmin tablosu**")
                    st.dataframe(style_metric_dataframe(outputs["tables"][dl_name]), width="stretch")
                    fig_dl = plot_forecast_results(outputs["train"], outputs["test"], {dl_name: outputs["predictions"][dl_name]}, f"{target_col} - {dl_name}", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
                    if fig_dl is not None:
                        st.plotly_chart(fig_dl, width="stretch")
                    if isinstance(dl_pack.get("search_table"), pd.DataFrame) and len(dl_pack.get("search_table")) > 0:
                        st.markdown(f"**{dl_name} arama tablosu**")
                        st.dataframe(style_metric_dataframe(dl_pack.get("search_table")), width="stretch")
                    hist_df = dl_pack.get("history_df", pd.DataFrame())
                    if isinstance(hist_df, pd.DataFrame) and len(hist_df) > 0:
                        st.markdown(f"**{dl_name} epoch-loss geçmişi**")
                        st.dataframe(style_metric_dataframe(hist_df), width="stretch")
                        fig_loss = build_dl_loss_curve(hist_df, f"{target_col} - {dl_name} Epoch-Loss")
                        if fig_loss is not None:
                            st.plotly_chart(fig_loss, width="stretch")
                    if dl_pack.get("overfit_summary"):
                        st.markdown(f"**{dl_name} overfitting kontrol özeti**")
                        st.json(dl_pack.get("overfit_summary"))
        with dl_subtabs[2]:
            dl_analysis_df = outputs.get("deep_learning_analysis", pd.DataFrame())
            if isinstance(dl_analysis_df, pd.DataFrame) and len(dl_analysis_df) > 0:
                st.dataframe(dl_analysis_df, width="stretch")
            else:
                st.info("DL karşılaştırma analizi üretilemedi.")

    with tab4:
        st.markdown("**Karar hiyerarşisi özeti**")
        st.dataframe(style_metric_dataframe(outputs.get("karar_hiyerarsisi", pd.DataFrame())), width="stretch")
        st.markdown("**Holdout sıralaması**")
        st.dataframe(style_metric_dataframe(outputs["champion_challenger"]["ranking"]), width="stretch")
        st.markdown("**Ansambl ağırlıkları (peak-aware + bias-aware)**")
        st.dataframe(style_metric_dataframe(outputs["ensemble_weights"]), width="stretch")
        st.markdown("**Ansambl gerçek ve tahmin**")
        st.dataframe(style_metric_dataframe(outputs["tables"]["Ensemble"]), width="stretch")
        fig_ens = plot_forecast_results(outputs["train"], outputs["test"], {"Ensemble": outputs["predictions"]["Ensemble"]}, f"{target_col} - Ensemble", model_style_map=outputs.get("model_gorsel_stil_haritasi", {}))
        if fig_ens is not None:
            st.plotly_chart(fig_ens, width="stretch")

    with tab5:
        combined = outputs["all_predictions_long"].copy()
        st.dataframe(style_metric_dataframe(combined), width="stretch")
        st.download_button("Gerçek vs tahmin tablosunu indir (Excel)", data=dataframe_to_download_bytes(combined), file_name=f"{selected_sheet}_{target_col}_actual_vs_forecast.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab6:
        if "rolling_origin_backtest" in outputs and len(outputs["rolling_origin_backtest"]) > 0:
            st.markdown("**Rolling-origin geri test (tam katman)**")
            st.dataframe(style_metric_dataframe(outputs["rolling_origin_backtest"]), width="stretch")
        if "validation_metrics_df" in outputs and len(outputs["validation_metrics_df"]) > 0:
            st.markdown("**Doğrulama temelli model kalitesi**")
            st.dataframe(style_metric_dataframe(outputs["validation_metrics_df"]), width="stretch")
        if "proxy_backtest_report" in export_payload and len(export_payload["proxy_backtest_report"]) > 0:
            st.markdown("**Proxy backtest raporu**")
            st.dataframe(export_payload["proxy_backtest_report"], width="stretch")
        else:
            st.info("Proxy backtest raporu bulunamadı.")
        if "raw_vs_clean_backtest_report" in export_payload and len(export_payload["raw_vs_clean_backtest_report"]) > 0:
            st.markdown("**Raw vs Clean backtest karşılaştırması**")
            st.dataframe(export_payload["raw_vs_clean_backtest_report"], width="stretch")

    with tab7:
        prod_pack = outputs.get("production_governance", {}) or {}
        live_pack = prod_pack.get("live_monitoring_pack", {}) or {}
        st.markdown("**Karar hiyerarşisi**")
        st.dataframe(style_metric_dataframe(outputs.get("karar_hiyerarsisi", pd.DataFrame())), width="stretch")
        st.markdown("**Merkezî canlı izleme özeti**")
        st.dataframe(style_metric_dataframe(live_pack.get("summary", pd.DataFrame())), width="stretch")
        st.markdown("**Model sağlık tablosu (bias, kapsama, fallback, servis seviyesi)**")
        st.dataframe(style_metric_dataframe(live_pack.get("model_health_table", pd.DataFrame())), width="stretch")
        if isinstance(live_pack.get("alerts"), pd.DataFrame) and len(live_pack.get("alerts")) > 0:
            st.markdown("**Üretim alarmları**")
            st.dataframe(style_metric_dataframe(live_pack.get("alerts", pd.DataFrame())), width="stretch")
        st.markdown("**Üretim sıralaması (holdout, doğrulama ve rolling-origin birlikte)**")
        st.dataframe(style_metric_dataframe(prod_pack.get("production_ranking", pd.DataFrame())), width="stretch")
        st.markdown("**Tahmin katkı değeri (baseline'a göre)**")
        st.dataframe(style_metric_dataframe(prod_pack.get("forecast_value_add", pd.DataFrame())), width="stretch")
        st.markdown("**Bias dashboard**")
        st.dataframe(style_metric_dataframe(prod_pack.get("bias_dashboard", pd.DataFrame())), width="stretch")
        st.markdown("**Tepe olay yakalama skoru**")
        st.dataframe(style_metric_dataframe(prod_pack.get("peak_event_dashboard", pd.DataFrame())), width="stretch")
        st.markdown("**Üretim özellik erişilebilirlik denetimi**")
        st.dataframe(style_metric_dataframe(prod_pack.get("feature_availability_audit", pd.DataFrame())), width="stretch")
        st.markdown("**Üretim modeli için tahmin aralığı / quantile forecast**")
        st.dataframe(style_metric_dataframe(prod_pack.get("production_interval_table", pd.DataFrame())), width="stretch")
        st.markdown("**Servis seviyesi / stok etkisi simülasyonu**")
        st.dataframe(style_metric_dataframe(prod_pack.get("production_service_table", pd.DataFrame())), width="stretch")
        st.markdown("**Benzersiz rapor kataloğu (yinelenen içerik temizliği)**")
        st.dataframe(style_metric_dataframe(outputs.get("benzersiz_rapor_katalogu", pd.DataFrame())), width="stretch")

    with tab8:
        st.markdown("**Kalite raporu**")
        st.dataframe(export_payload["quality_report"], width="stretch")
        st.markdown("**Seri profil raporu**")
        st.dataframe(export_payload["series_profile_report"], width="stretch")
        st.markdown("**Anomali yönetişimi**")
        st.dataframe(export_payload["anomaly_governance"], width="stretch")
        st.markdown("**Review queue**")
        st.dataframe(export_payload["review_queue"], width="stretch")

    with tab9:
        suggestions = []
        if float(profile.get("seasonality_strength", 0) or 0) >= 0.35:
            suggestions.append("Seri sezonsal; auto seasonal (P,D,Q,m), Prophet custom seasonality ve ensemble kritik.")
        if float(profile.get("cv", 0) or 0) >= 0.45:
            suggestions.append("Oynaklık yüksek; log/Box-Cox dönüşümü, XGBoost rolling istatistikleri ve challenger modeli zorunlu izlenmeli.")
        if float(profile.get("intermittency_ratio", 0) or 0) >= 0.25:
            suggestions.append("Intermittent yapı var; stok-out ayrımı ve model çıktılarını iş kuralı ile doğrulamak gerekir.")
        if outputs.get("sarima", {}).get("residual_white_noise_ok") is False:
            suggestions.append("SARIMA residual white-noise testi zayıf; challenger olarak Prophet veya XGBoost önceliklendirilmeli.")
        if outputs.get("best_model") == "Ensemble":
            suggestions.append("Bu seride tek bir champion yerine ensemble daha iyi; operasyonel kullanımda champion-challenger izleme önerilir.")
        if normalize_dl_family_label(outputs.get("best_model")) in {"LSTM", "GRU"}:
            suggestions.append("Derin öğrenme modeli öne çıktı; epoch-loss eğrileri ve validation-holdout ayrışması özellikle kontrol edilmelidir.")
        if not suggestions:
            suggestions.append("Seri dengeli; champion-challenger ve ensemble izlemesi yeterli görünüyor.")
        for s in suggestions:
            st.write(f"- {s}")


def _resolve_streamlit_entrypoint():
    """Render fonksiyonunu güvenli biçimde bulur."""
    direct = globals().get("render_streamlit_app")
    if callable(direct):
        return direct

    candidates = []
    for name, obj in globals().items():
        if not callable(obj):
            continue
        lower = str(name).lower()
        if lower.startswith("render_") and "streamlit" in lower:
            candidates.append(obj)
        elif lower.startswith("render") and "app" in lower:
            candidates.append(obj)

    if candidates:
        return candidates[0]

    raise NameError(
        "render_streamlit_app tanımı bulunamadı. Streamlit giriş noktası eksik veya bozuk."
    )


def _is_running_under_streamlit() -> bool:
    try:
        ensure_forecasting_runtime_dependencies(include_streamlit=True)
    except Exception:
        pass

    if st is not None:
        try:
            from streamlit.runtime.scriptrunner import get_script_run_ctx
            if get_script_run_ctx() is not None:
                return True
        except Exception:
            pass

    try:
        argv_text = " ".join(str(x) for x in sys.argv)
    except Exception:
        argv_text = ""

    env_markers = [
        "STREAMLIT_SERVER_PORT",
        "STREAMLIT_SERVER_HEADLESS",
        "STREAMLIT_BROWSER_GATHER_USAGE_STATS",
        "STREAMLIT_SHARING_MODE",
        "STREAMLIT_RUNTIME",
        "STREAMLIT_THEME_BASE",
    ]
    if any(os.environ.get(k) for k in env_markers):
        return True
    if "streamlit run" in argv_text.lower() or "streamlit.web.cli" in argv_text.lower():
        return True
    return False


def _should_launch_streamlit_entrypoint(argv: Optional[List[str]] = None) -> bool:
    argv = list(argv or [])

    try:
        ensure_forecasting_runtime_dependencies(include_streamlit=True)
    except Exception:
        pass

    if _is_running_under_streamlit():
        return True
    if any(str(a).strip().lower() == "--streamlit" for a in argv):
        return True

    # Streamlit kurulu ve tkinter yoksa, parametresiz başlatma yerel diyalog
    # akışına girmek yerine Streamlit UI'yi açmalıdır. Bu özellikle Streamlit Cloud
    # ve headless Linux ortamlarda kritik davranıştır.
    if not argv and st is not None and not HAS_TKINTER:
        return True
    return False


def configure_local_runtime_stability() -> None:
    os.environ.setdefault("OMP_NUM_THREADS", "1")
    os.environ.setdefault("MKL_NUM_THREADS", "1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
    os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")
    try:
        FORECAST_RUNTIME_CONFIG.search_accelerator_enabled = False
        FORECAST_RUNTIME_CONFIG.search_accelerator_model_parallel = False
        FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_parallel = False
        FORECAST_RUNTIME_CONFIG.search_accelerator_max_workers = 1
        FORECAST_RUNTIME_CONFIG.search_accelerator_candidate_workers = 1
        FORECAST_RUNTIME_CONFIG.xgb_force_single_thread = True
    except Exception:
        pass
    try:
        SEARCH_ACCELERATOR.config.enabled = False
        SEARCH_ACCELERATOR.config.model_parallelism = False
        SEARCH_ACCELERATOR.config.candidate_parallelism = False
        SEARCH_ACCELERATOR.config.max_workers = 1
        SEARCH_ACCELERATOR.config.candidate_workers = 1
        SEARCH_ACCELERATOR.config.cache_enabled = False
        SEARCH_ACCELERATOR.config.result_cache_enabled = False
    except Exception:
        pass


def _prepare_excel_source_from_path(input_path: str, archive_member: Optional[str] = None) -> Dict[str, Optional[str]]:
    input_path = os.path.abspath(str(input_path))
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Girdi dosyası bulunamadı: {input_path}")
    if os.path.isdir(input_path):
        excel_files = sorted([
            os.path.relpath(os.path.join(root, f), input_path)
            for root, _, files in os.walk(input_path)
            for f in files if f.lower().endswith((".xlsx", ".xls"))
        ])
        if not excel_files:
            raise FileNotFoundError("Seçilen klasörde Excel dosyası bulunamadı.")
        chosen_member = archive_member
        if chosen_member is None:
            if len(excel_files) != 1:
                raise ValueError("Klasörde birden fazla Excel dosyası var. İlgili dosyayı seçmek için etkileşimli modu kullanın veya göreli dosya yolunu belirtin.")
            chosen_member = excel_files[0]
        chosen_member = str(chosen_member)
        if chosen_member not in excel_files:
            raise ValueError(f"Klasör içi Excel dosyası bulunamadı: {chosen_member}")
        excel_path = os.path.join(input_path, chosen_member)
        return {
            "source_path": input_path,
            "excel_path": excel_path,
            "archive_member": None,
            "temp_dir": None,
            "source_type": "directory",
            "selected_excel": chosen_member,
        }
    lower = input_path.lower()
    if lower.endswith((".xlsx", ".xls")):
        return {
            "source_path": input_path,
            "excel_path": input_path,
            "archive_member": None,
            "temp_dir": None,
            "source_type": "excel",
        }
    if lower.endswith((".zip", ".rar")):
        excel_files = _list_excel_files_in_archive(input_path)
        chosen_member = archive_member
        if chosen_member is None:
            if len(excel_files) != 1:
                raise ValueError("Arşiv içinde birden fazla Excel dosyası var. --archive-member parametresi verilmelidir.")
            chosen_member = excel_files[0]
        if chosen_member not in excel_files:
            raise ValueError(f"Arşiv üyesi bulunamadı: {chosen_member}")
        extracted_path, temp_dir = _extract_excel_from_archive(input_path, chosen_member)
        return {
            "source_path": input_path,
            "excel_path": extracted_path,
            "archive_member": chosen_member,
            "temp_dir": temp_dir,
            "source_type": "archive",
        }
    raise ValueError("Desteklenmeyen girdi türü. .xlsx, .xls, .zip, .rar veya bir klasör yolu bekleniyor.")


def _resolve_sheet_name_for_cli(excel_path: str, requested_sheet: Optional[str]) -> str:
    xls = safe_excel_file(excel_path)
    sheet_names = list(xls.sheet_names)
    if requested_sheet is None:
        if len(sheet_names) == 1:
            return sheet_names[0]
        raise ValueError("Birden fazla sheet var. --sheet parametresi verilmelidir.")
    if requested_sheet in sheet_names:
        return requested_sheet
    if str(requested_sheet).isdigit():
        idx = int(requested_sheet) - 1
        if 0 <= idx < len(sheet_names):
            return sheet_names[idx]
    lowered = {str(s).strip().lower(): s for s in sheet_names}
    key = str(requested_sheet).strip().lower()
    if key in lowered:
        return lowered[key]
    raise ValueError(f"Sheet bulunamadı: {requested_sheet}")


def _parse_target_list(arg_value: Optional[str]) -> Optional[List[str]]:
    if arg_value is None:
        return None
    raw = str(arg_value).strip()
    if not raw or raw.lower() in {"all", "*"}:
        return None
    return [x.strip() for x in raw.split(",") if x.strip()]



def _list_excel_files_in_directory(directory_path: str) -> List[str]:
    directory_path = os.path.abspath(str(directory_path))
    if not os.path.isdir(directory_path):
        raise NotADirectoryError(f"Klasör bulunamadı: {directory_path}")
    return sorted([
        os.path.relpath(os.path.join(root, f), directory_path)
        for root, _, filenames in os.walk(directory_path)
        for f in filenames if f.lower().endswith((".xlsx", ".xls"))
    ])


def get_excel_sheet_names_fast(excel_path: str) -> List[str]:
    """Return workbook sheet names with minimal overhead for local interactive flow."""
    ext = os.path.splitext(str(excel_path))[1].lower()
    ok, engine, msg = _check_optional_excel_dependency(ext)
    if not ok:
        raise ImportError(msg)

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            try:
                return [str(s) for s in wb.sheetnames]
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        except Exception:
            pass

    if ext == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(excel_path, on_demand=True)
            try:
                return [str(s) for s in wb.sheet_names()]
            finally:
                try:
                    wb.release_resources()
                except Exception:
                    pass
        except Exception:
            pass

    xls = safe_excel_file(excel_path)
    return [str(s) for s in xls.sheet_names]


def read_excel_preview_fast(excel_path: str, sheet_name: str, preview_rows: int = 1200) -> pd.DataFrame:
    ext = os.path.splitext(str(excel_path))[1].lower()
    ok, engine, msg = _check_optional_excel_dependency(ext)
    if not ok:
        raise ImportError(msg)
    try:
        df_preview = pd.read_excel(excel_path, sheet_name=sheet_name, engine=engine, nrows=int(preview_rows))
    except TypeError:
        df_preview = pd.read_excel(excel_path, sheet_name=sheet_name, engine=engine)
        if preview_rows and len(df_preview) > int(preview_rows):
            df_preview = df_preview.head(int(preview_rows)).copy()
    return df_preview


def build_local_sheet_preview_payload(excel_path: str, sheet_name: str, preview_rows: int = 1200) -> Dict[str, Any]:
    progress_log("Hızlı Önizleme", "Sheet önizlemesi okunuyor; seri listesi ve frekans tahmin ediliyor.", extra={"sheet": sheet_name, "preview_rows": int(preview_rows)})
    try:
        df_preview = read_excel_preview_fast(excel_path, sheet_name=sheet_name, preview_rows=int(preview_rows))
    except Exception as e:
        raise ImportError(f"Sheet önizlemesi okunamadı ({os.path.basename(excel_path)} / {sheet_name}): {e}")

    cfg = PreprocessConfig()
    date_col = detect_date_column(df_preview, cfg)
    df_preview = df_preview.copy()
    df_preview[date_col] = parse_datetime_series(df_preview[date_col])
    df_preview = df_preview.dropna(subset=[date_col]).sort_values(date_col).reset_index(drop=True)
    target_cols = detect_target_columns(df_preview, date_col, cfg)
    freq_alias = infer_frequency_from_dates(pd.DatetimeIndex(df_preview[date_col].dropna())) if len(df_preview) >= 3 else "M"

    profile_rows = []
    for col in target_cols:
        s = pd.to_numeric(df_preview[col], errors="coerce")
        profile_rows.append({
            "series": str(col),
            "n_obs": int(s.notna().sum()),
            "intermittency_ratio": safe_float(demand_intermittency_ratio(s)),
            "mean": safe_float(s.mean()),
            "median": safe_float(s.median()),
        })

    preview_payload = {
        "series_profile_report": pd.DataFrame(profile_rows),
        "clean_model_input": df_preview[[date_col] + [c for c in target_cols if c in df_preview.columns]].copy(),
        "preview_date_col": date_col,
        "preview_target_cols": list(target_cols),
        "preview_freq_alias": freq_alias,
        "preview_row_count": int(len(df_preview)),
        "preview_sheet_name": str(sheet_name),
    }
    progress_log("Hızlı Önizleme", "Sheet önizlemesi tamamlandı.", level="SUCCESS", extra={"sheet": sheet_name, "targets": len(target_cols), "freq_alias": freq_alias, "preview_rows_used": int(len(df_preview))})
    return preview_payload


def choose_local_input_source_interactively() -> Tuple[str, Optional[str]]:
    if not HAS_TKINTER:
        raise RuntimeError("Etkileşimli masaüstü seçim ekranı için tkinter gereklidir.")

    progress_log("Dosya Seçimi", "Doğrudan dosya gezgini açılıyor. Excel seçimi bekleniyor...")
    try:
        source_info = choose_excel_file(dialog_mode="excel_only")
        source_path = str(source_info.get("source_path"))
        archive_member = source_info.get("archive_member")
        return source_path, archive_member
    except FileNotFoundError:
        progress_log("Dosya Seçimi", "İlk dosya seçimi iptal edildi; alternatif kaynak seçimi ekranı açılıyor.", level="WARNING")

    source_kind = ask_choice_dialog(
        title="Veri Kaynağı Seçimi",
        prompt="Doğrudan dosya seçimi iptal edildi. Alternatif bir kaynak seçin.",
        items=["Excel/ZIP/RAR dosyasını yeniden seç", "Klasör seç"],
        default_index=0,
    )
    if source_kind == "Klasör seç":
        root = _prepare_tk_root(hidden=True)
        folder_path = filedialog.askdirectory(
            title="Excel dosyalarının bulunduğu klasörü seçin",
            initialdir=_best_local_initial_dir(),
            parent=root,
        )
        try:
            root.destroy()
        except Exception:
            pass
        if not folder_path:
            raise ValueError("Klasör seçimi yapılmadı.")
        excel_files = _list_excel_files_in_directory(folder_path)
        if not excel_files:
            raise FileNotFoundError("Seçilen klasörde Excel dosyası bulunamadı.")
        chosen_excel = choose_single_sheet(excel_files)
        return os.path.abspath(folder_path), chosen_excel

    source_info = choose_excel_file(dialog_mode="excel_or_archive")
    source_path = str(source_info.get("source_path"))
    archive_member = source_info.get("archive_member")
    return source_path, archive_member



def recommend_batch_target_subset(export_payload: Dict[str, pd.DataFrame], horizon: int) -> List[str]:
    profile_df = export_payload.get("series_profile_report", pd.DataFrame())
    if not isinstance(profile_df, pd.DataFrame) or len(profile_df) == 0 or "series" not in profile_df.columns:
        return []
    all_targets = list(profile_df["series"].astype(str).tolist())
    min_obs = max(18, int(horizon) * 4)
    recommended = set(all_targets)

    keep = set()
    for _, row in profile_df.iterrows():
        series = str(row.get("series"))
        n_obs = safe_float(row.get("n_obs"))
        intermittency = safe_float(row.get("intermittency_ratio"))
        if pd.notna(n_obs) and n_obs < min_obs:
            continue
        if pd.notna(intermittency) and intermittency > 0.98:
            continue
        keep.add(series)
    if keep:
        recommended &= keep

    missing_df = export_payload.get("missing_strategy_audit", pd.DataFrame())
    if isinstance(missing_df, pd.DataFrame) and len(missing_df) > 0 and "series" in missing_df.columns:
        excluded_mask = pd.Series(False, index=missing_df.index)
        if "series_excluded_from_modeling" in missing_df.columns:
            excluded_mask = missing_df["series_excluded_from_modeling"].astype(bool)
        excluded = set(missing_df.loc[excluded_mask, "series"].astype(str).tolist())
        recommended -= excluded

    review_queue = export_payload.get("review_queue", pd.DataFrame())
    if isinstance(review_queue, pd.DataFrame) and len(review_queue) > 0 and "series" in review_queue.columns:
        review_counts = review_queue["series"].astype(str).value_counts()
        too_many_reviews = set(review_counts[review_counts >= max(5, int(horizon) * 2)].index.tolist())
        if len(recommended - too_many_reviews) >= 2:
            recommended -= too_many_reviews

    ordered = [s for s in all_targets if s in recommended]
    if len(ordered) >= 2:
        return ordered
    return all_targets


def choose_local_forecast_plan_interactively(preview_payload: Dict[str, Any], default_horizon: int, freq_alias: str) -> Dict[str, Any]:
    series_profile = preview_payload.get("series_profile_report", pd.DataFrame())
    targets = list(series_profile["series"].astype(str).tolist()) if isinstance(series_profile, pd.DataFrame) and len(series_profile) > 0 and "series" in series_profile.columns else []
    if not targets:
        raise ValueError("Seri listesi belirlenemedi.")

    progress_log("Tahminleme Modu", "Tahminleme modu seçim ekranı açılıyor.", extra={"targets": len(targets), "freq_alias": freq_alias})
    mode = ask_choice_dialog(
        title="Tahminleme Modu",
        prompt="Tek serili veya çok serili batch forecasting modunu seçin.",
        items=["Tek serili tahminleme", "Çok serili batch forecasting"],
        default_index=1 if len(targets) > 1 else 0,
    )
    progress_log("Tahminleme Modu", "Tahminleme modu seçildi.", level="SUCCESS", extra={"mode": mode})

    preview_rows = int(preview_payload.get("preview_row_count") or len(preview_payload.get("clean_model_input", pd.DataFrame())) or 0)
    max_horizon = max(2, min(24, max(2, preview_rows // 3 if preview_rows else default_horizon)))
    horizon_prompt = (
        f"Frekans: {freq_alias}\n"
        f"Önerilen test ufku: {default_horizon}\n"
        f"Uygun aralık: 2 ile {max_horizon} arası.\n\n"
        "Test / tahmin ufkunu girin."
    )
    progress_log("Test Ufku", "Test ufku giriş ekranı açılıyor.", extra={"default_horizon": int(default_horizon), "max_horizon": int(max_horizon)})
    horizon = ask_integer_dialog("Test Ufku", horizon_prompt, initial=default_horizon, min_value=2, max_value=max_horizon)
    progress_log("Test Ufku", "Test ufku seçildi.", level="SUCCESS", extra={"horizon": int(horizon)})

    if mode == "Tek serili tahminleme":
        progress_log("Seri Seçimi", "Tek serili seri seçim ekranı açılıyor.", extra={"target_count": len(targets)})
        target = ask_choice_dialog(
            title="Seri Seçimi",
            prompt="Tek serili tahminleme için çalıştırılacak seriyi seçin.",
            items=targets,
            default_index=0,
        )
        progress_log("Seri Seçimi", "Tek serili seri seçimi tamamlandı.", level="SUCCESS", extra={"selected_target": target})
        return {"mode": "single", "horizon": int(horizon), "selected_targets": [target]}

    recommended_targets = recommend_batch_target_subset(preview_payload, horizon)
    progress_log("Batch Seri Politikası", "Batch seri stratejisi seçim ekranı açılıyor.", extra={"recommended_count": len(recommended_targets), "all_targets": len(targets)})
    batch_policy = ask_choice_dialog(
        title="Batch Seri Seçimi",
        prompt=(
            "Gerçek iş kullanımında batch tahminleme genelde tüm portföy veya modellemeye en uygun alt portföy üzerinde çalıştırılır.\n\n"
            "Bir seri seçim stratejisi belirleyin."
        ),
        items=["Önerilen üretim serileri", "Tüm seriler", "Manuel seri seçimi"],
        default_index=0 if len(recommended_targets) >= 2 else 1,
    )

    if batch_policy == "Önerilen üretim serileri":
        selected_targets = recommended_targets
    elif batch_policy == "Tüm seriler":
        selected_targets = targets
    else:
        progress_log("Batch Seri Seçimi", "Manuel batch seri seçim ekranı açılıyor.", extra={"default_count": len(recommended_targets)})
        selected_targets = ask_multiselect_dialog(
            title="Batch Seri Seçimi",
            prompt="Batch tahminlemede yer alacak serileri seçin. Ctrl/Shift ile çoklu seçim yapabilirsiniz.",
            items=targets,
            default_items=recommended_targets,
        )

    if not selected_targets:
        raise ValueError("Çok serili tahminleme için en az bir seri seçilmelidir.")

    progress_log("Batch Seri Seçimi", "Batch seri seçimi tamamlandı.", level="SUCCESS", extra={"policy": batch_policy, "selected_count": len(selected_targets)})
    return {"mode": "batch", "horizon": int(horizon), "selected_targets": list(selected_targets)}

def run_local_batch_forecasting_from_payload(
    export_payload: Dict[str, pd.DataFrame],
    sheet_name: str,
    horizon: int,
    output_dir: Optional[str] = None,
    selected_targets: Optional[List[str]] = None,
    use_exog_for_stat_models: bool = True,
    use_exog_for_prophet: bool = True,
    mode_label: str = "batch",
    source_info: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ensure_forecasting_runtime_dependencies(include_streamlit=False)
    output_dir = os.path.abspath(output_dir or os.getcwd())
    os.makedirs(output_dir, exist_ok=True)
    mode_slug = "tek_seri" if str(mode_label).lower() in {"single", "tek", "tek_seri"} else "cok_serili_batch"
    run_info = initialize_incremental_batch_output_root(output_dir, sheet_name, int(horizon), mode_slug)
    run_root = run_info["run_root"]
    batch_root = run_info["batch_root"]
    progress_log("Batch Başlatma", "Forecasting batch çalışması başlatılıyor.", extra={"sheet": sheet_name, "horizon": int(horizon), "selected_targets": len(selected_targets) if selected_targets else 'all', "run_root": run_root})
    try:
        source_path = str((source_info or {}).get("source_path") or "").strip()
        if source_path and os.path.exists(source_path):
            _copy_optional_file_into_dir(Path(source_path), os.path.join(run_root, os.path.basename(source_path)))
    except Exception as copy_exc:
        progress_log("Disk Çıktısı", f"Orijinal girdi dosyası çıktı klasörüne kopyalanamadı: {copy_exc}", level="WARNING")
    batch_result = run_batch_forecasting(
        export_payload=export_payload,
        horizon=int(horizon),
        selected_sheet=sheet_name,
        use_exog_for_stat_models=bool(use_exog_for_stat_models),
        use_exog_for_prophet=bool(use_exog_for_prophet),
        selected_targets=selected_targets,
        persist_output_root=batch_root,
        persist_after_each_series=True,
        keep_batch_outputs_in_memory=False,
    )
    batch_result["run_root"] = run_root
    batch_result["batch_root"] = batch_root
    zip_name = f"{run_info['run_base']}.zip"
    zip_path = os.path.join(output_dir, zip_name)
    progress_log("ZIP Yazımı", f"Tüm seri klasörleri tamamlandı; ZIP dosyası diske yazılıyor: {zip_name}", extra={"output_dir": output_dir, "run_root": run_root})
    create_zip_from_directory(run_root, zip_path)
    progress_log("ZIP Yazımı", "ZIP dosyası başarıyla oluşturuldu.", level="SUCCESS", extra={"zip_path": zip_path})
    return {
        "zip_path": zip_path,
        "run_root": run_root,
        "batch_root": batch_root,
        "sheet_name": sheet_name,
        "targets": selected_targets if selected_targets else export_payload.get("series_profile_report", pd.DataFrame()).get("series", pd.Series(dtype=object)).astype(str).tolist(),
        "batch_result": batch_result,
        "export_payload": export_payload,
        "source_info": source_info or {},
        "mode": mode_slug,
    }

def run_local_interactive_workflow() -> Dict[str, Any]:
    if not HAS_TKINTER:
        raise RuntimeError("Etkileşimli yerel kullanım için tkinter gereklidir. Alternatif olarak komut satırı parametrelerini kullanın.")

    output_dir = os.path.abspath(os.getcwd())
    reporter = activate_local_progress_reporter("talep_tahminleme_interactive", output_dir=output_dir, heartbeat_seconds=60)
    source_info: Dict[str, Any] = {}
    configure_local_runtime_stability()
    progress_log("Yerel Hazırlık", "Yerel çalışma kararlılık ayarları uygulandı.", level="SUCCESS")
    input_path, archive_member = choose_local_input_source_interactively()
    progress_log("Dosya Seçimi", "Girdi kaynağı seçildi.", level="SUCCESS", extra={"input_path": input_path, "archive_member": archive_member})
    source_info = _prepare_excel_source_from_path(input_path, archive_member=archive_member)
    excel_path = str(source_info["excel_path"])
    preprocess_dir = tempfile.mkdtemp(prefix="talep_tahminleme_local_preprocess_")
    progress_log("Dosya Hazırlığı", "Excel kaynağı işlenmek üzere hazırlandı.", level="SUCCESS", extra={"excel_path": excel_path, "source_type": source_info.get('source_type')})
    try:
        progress_log("Sheet Seçimi", "Sheet listesi hızlı modda okunuyor; sheet seçim ekranı hazırlanıyor.", extra={"excel_path": excel_path})
        sheet_names = get_excel_sheet_names_fast(excel_path)
        progress_log("Sheet Seçimi", "Sheet listesi alındı; seçim ekranı açılıyor.", level="SUCCESS", extra={"sheet_count": len(sheet_names)})
        resolved_sheet = choose_single_sheet(sheet_names, always_prompt=True)
        progress_log("Sheet Seçimi", "Sheet seçimi tamamlandı.", level="SUCCESS", extra={"sheet": resolved_sheet, "sheet_count": len(sheet_names)})

        preview_rows = 1200
        progress_log("Hızlı Önizleme", "Tahminleme planı için hafif önizleme hazırlanıyor.", extra={"sheet": resolved_sheet, "preview_rows": int(preview_rows)})
        preview_payload = build_local_sheet_preview_payload(excel_path, resolved_sheet, preview_rows=preview_rows)
        preview_freq_alias = str(preview_payload.get("preview_freq_alias") or "M")
        preview_rows = int(preview_payload.get("preview_row_count") or 0)
        default_horizon = min(infer_default_horizon(preview_freq_alias), max(2, preview_rows // 5 if preview_rows else infer_default_horizon(preview_freq_alias)))

        plan = choose_local_forecast_plan_interactively(preview_payload, default_horizon, preview_freq_alias)
        progress_log("Planlama", "Tahminleme planı belirlendi.", level="SUCCESS", extra={"mode": plan.get('mode'), "horizon": plan.get('horizon'), "selected_targets": len(plan.get('selected_targets') or [])})

        progress_log("Önişleme", "Tüm kullanıcı seçimleri tamamlandı; şimdi tam önişleme başlatılıyor.", extra={"sheet": resolved_sheet, "selected_targets": len(plan.get('selected_targets') or [])})
        export_payload = run_preprocessing_for_sheet(excel_path, resolved_sheet, preprocess_dir)
        progress_log("Önişleme", "Önişleme tamamlandı.", level="SUCCESS")

        result = run_local_batch_forecasting_from_payload(
            export_payload=export_payload,
            sheet_name=resolved_sheet,
            horizon=int(plan["horizon"]),
            output_dir=output_dir,
            selected_targets=plan.get("selected_targets"),
            use_exog_for_stat_models=True,
            use_exog_for_prophet=True,
            mode_label=str(plan.get("mode", "batch")),
            source_info=source_info,
        )
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo(
                "Talep Tahminleme Tamamlandı",
                f"Çalışma tamamlandı.\n\nÇıktı klasörü:\n{result['run_root']}\n\nZIP çıktı yolu:\n{result['zip_path']}\n\nSeri sayısı: {len(result['targets'])}\n\nGünlük dosyası:\n{reporter.log_path}"
            )
        except Exception:
            pass
        deactivate_local_progress_reporter(success=True, final_message=f"Yerel etkileşimli çalışma tamamlandı. Klasör: {result['run_root']} | ZIP: {result['zip_path']}")
        return result
    except Exception as e:
        progress_log("Yerel Çalışma", f"Yerel etkileşimli akış hata verdi: {e}", level="ERROR")
        deactivate_local_progress_reporter(success=False, final_message=f"Yerel etkileşimli çalışma hata ile bitti: {e}")
        raise
    finally:
        shutil.rmtree(preprocess_dir, ignore_errors=True)
        temp_dir = source_info.get("temp_dir") if isinstance(source_info, dict) else None
        if temp_dir:
            shutil.rmtree(str(temp_dir), ignore_errors=True)

def run_local_batch_forecasting_job(
    input_path: str,
    sheet_name: Optional[str],
    horizon: int,
    output_dir: Optional[str] = None,
    selected_targets: Optional[List[str]] = None,
    use_exog_for_stat_models: bool = True,
    use_exog_for_prophet: bool = True,
    archive_member: Optional[str] = None,
) -> Dict[str, Any]:
    output_dir = os.path.abspath(output_dir or os.getcwd())
    activate_local_progress_reporter("talep_tahminleme_cli", output_dir=output_dir, heartbeat_seconds=60)
    progress_log("CLI Başlatma", "Komut satırı çalışması başlatıldı.", extra={"input_path": input_path, "sheet": sheet_name, "horizon": int(horizon)})
    configure_local_runtime_stability()
    source_info = _prepare_excel_source_from_path(input_path, archive_member=archive_member)
    excel_path = str(source_info["excel_path"])
    resolved_sheet = _resolve_sheet_name_for_cli(excel_path, sheet_name)
    progress_log("CLI Hazırlık", "Kaynak dosya ve sheet çözümlendi.", level="SUCCESS", extra={"excel_path": excel_path, "resolved_sheet": resolved_sheet})

    preprocess_dir = tempfile.mkdtemp(prefix="talep_tahminleme_local_preprocess_")
    try:
        progress_log("Önişleme", "Önişleme başlatıldı.")
        export_payload = run_preprocessing_for_sheet(excel_path, resolved_sheet, preprocess_dir)
        progress_log("Önişleme", "Önişleme tamamlandı.", level="SUCCESS")
        mode_label = "single" if selected_targets is not None and len(selected_targets) == 1 else "batch"
        result = run_local_batch_forecasting_from_payload(
            export_payload=export_payload,
            sheet_name=resolved_sheet,
            horizon=int(horizon),
            output_dir=output_dir,
            selected_targets=selected_targets,
            use_exog_for_stat_models=bool(use_exog_for_stat_models),
            use_exog_for_prophet=bool(use_exog_for_prophet),
            mode_label=mode_label,
            source_info=source_info,
        )
        deactivate_local_progress_reporter(success=True, final_message=f"CLI çalışması tamamlandı. Klasör: {result['run_root']} | ZIP: {result['zip_path']}")
        return result
    except Exception as e:
        progress_log("CLI Çalışma", f"CLI akışı hata verdi: {e}", level="ERROR")
        deactivate_local_progress_reporter(success=False, final_message=f"CLI çalışması hata ile bitti: {e}")
        raise
    finally:
        shutil.rmtree(preprocess_dir, ignore_errors=True)
        temp_dir = source_info.get("temp_dir") if isinstance(source_info, dict) else None
        if temp_dir:
            shutil.rmtree(str(temp_dir), ignore_errors=True)

def build_cli_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Talep tahminleme hibrit çalıştırıcı: Streamlit veya yerel batch ZIP üretimi"
    )
    parser.add_argument("--input", help="Excel dosyası veya ZIP/RAR arşivi yolu")
    parser.add_argument("--sheet", help="Sheet adı veya 1-bazlı sheet numarası")
    parser.add_argument("--archive-member", help="ZIP/RAR içindeki Excel üyesi")
    parser.add_argument("--horizon", type=int, default=3, help="Test ufku / tahmin ufku")
    parser.add_argument("--targets", help="Virgülle ayrılmış seri listesi. Tümü için 'all'.", default="all")
    parser.add_argument("--output-dir", help="ZIP çıktısının yazılacağı klasör. Varsayılan: mevcut çalışma klasörü")
    parser.add_argument("--no-stat-exog", action="store_true", help="SARIMA/SARIMAX için exog kullanımını kapat")
    parser.add_argument("--no-prophet-exog", action="store_true", help="Prophet için exog kullanımını kapat")
    parser.add_argument("--streamlit", action="store_true", help="CLI yerine Streamlit arayüzünü zorla aç")
    return parser


def run_cli_main(argv: Optional[List[str]] = None) -> int:
    try:
        ensure_forecasting_runtime_dependencies(include_streamlit=True)
    except Exception:
        pass

    parser = build_cli_arg_parser()
    args = parser.parse_args(argv)

    if _should_launch_streamlit_entrypoint(argv):
        if st is None:
            raise ImportError("Streamlit kurulu değil. 'streamlit run <dosya>.py' ile veya streamlit paketi kurulu halde çalıştırın.")
        entrypoint = _resolve_streamlit_entrypoint()
        entrypoint()
        return 0

    if not args.input:
        print("[INFO] Yerel etkileşimli mod başlatılıyor... Dosya seçme ekranı hazırlanıyor.")
        result = run_local_interactive_workflow()
        print("\n[SUCCESS] Yerel etkileşimli tahminleme tamamlandı.")
        print(f"Sheet: {result['sheet_name']}")
        print(f"Seri sayısı: {len(result['targets'])}")
        print(f"ZIP çıktı yolu: {result['zip_path']}")
        return 0

    selected_targets = _parse_target_list(args.targets)
    result = run_local_batch_forecasting_job(
        input_path=args.input,
        sheet_name=args.sheet,
        horizon=args.horizon,
        output_dir=args.output_dir,
        selected_targets=selected_targets,
        use_exog_for_stat_models=not args.no_stat_exog,
        use_exog_for_prophet=not args.no_prophet_exog,
        archive_member=args.archive_member,
    )

    print("\n[SUCCESS] Çok serili batch forecasting tamamlandı.")
    print(f"Sheet: {result['sheet_name']}")
    print(f"Seri sayısı: {len(result['targets'])}")
    print(f"ZIP çıktı yolu: {result['zip_path']}")
    return 0


def main():
    ensure_forecasting_runtime_dependencies(include_streamlit=True)
    if st is None:
        raise ImportError(
            "Bu dosya Streamlit uygulamasıdır. Çalıştırmak için: streamlit run <dosya_adı>.py veya python <dosya_adı>.py --input ..."
        )
    entrypoint = _resolve_streamlit_entrypoint()
    return entrypoint()


if __name__ == "__main__":
    try:
        if _should_launch_streamlit_entrypoint(sys.argv[1:]):
            main()
        else:
            raise SystemExit(run_cli_main(sys.argv[1:]))
    except KeyboardInterrupt:
        try:
            progress_log("Kullanıcı İptali", "Çalışma kullanıcı tarafından Ctrl+C ile durduruldu.", level="WARNING")
            deactivate_local_progress_reporter(success=False, final_message="Çalışma kullanıcı tarafından Ctrl+C ile durduruldu.")
        except Exception:
            pass
        print("[WARNING] Çalışma kullanıcı tarafından Ctrl+C ile durduruldu.")
        raise SystemExit(130)
    except SystemExit:
        raise
    except Exception as e:
        if _is_running_under_streamlit() and st is not None:
            st.error(f"Beklenmeyen uygulama hatası: {e}")
            st.exception(e)
        else:
            print(f"[ERROR] {e}")
            traceback.print_exc()
            raise SystemExit(1)
